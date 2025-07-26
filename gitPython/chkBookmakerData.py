from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import datetime
import sys
import smtplib
from email.mime.text import MIMEText
import json
import datetime
import inspect as isp
import csv
import time
from python_analytics import write_prediction_data
import pandas as pd
import re


def execute(playwright):
    loop_bef_seq = -1
    # 対象時間を設定
    start_time = datetime.time(23, 30)
    end_time = datetime.time(23, 59)
    # 現在時刻
    now = datetime.datetime.today()
    now_info = datetime.datetime(now.year,now.month,now.day,now.hour,now.minute,now.second)

    if now_info.weekday() == 1 or now_info.weekday() == 2 or now_info.weekday() == 3 or now_info.weekday() == 4 or now_info.weekday() == 5:
        if now.hour == 12 or now.hour == 13 or now.hour == 14 or now.hour == 15 or now.hour == 16:
            print(f"{now_info}  データ取得対象外の時間です。")

    print(f"{now_info}  データ取得対象の時間です。")

    # Chromiumブラウザを起動
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    # Googleをひらく
    page.goto("https://www.google.com")

    # flashscoreに遷移
    page.goto("https://flashscore.co.jp/soccer/")
    # ライブメニュー遷移
    try:
        live_page = page.locator("div.filters__tab").nth(1)  # CSSセレクタを使用
        live_page.click()

        page.wait_for_load_state("networkidle")

        # 非表示要素を表示に変更
        alter = page.locator("button[data-testid='wcl-accordionButton']")
        alter_count = alter.count()
        for i in range(alter_count):
            alter_label = alter.nth(i).get_attribute("aria-label")
            if alter_label == None or alter_label == "リーグ全試合 表示":
                alter.nth(i).click()

        # ライブ試合取得
        live_game_list = []
        live_game = page.locator("a[href*='#/match-summary']")
        live_game_list = live_game.evaluate_all("live_game => live_game.map(e => e.href)")
    except Exception as e:
        print("ライブ試合リンクの取得失敗", e)
        live_game_list = []

    # cron実行通知用にホームチーム,アウェイチーム,カテゴリを入れたリストを保持
    cron_notice_data_list = []
    for j in range(len(live_game_list)):
        notice_data_list = []
        # 試合用文字列のみ抜く
        live = live_game_list[j]
        live_data = live.replace("https://www.flashscore.co.jp/match/", "")
        random_game_str = live_data.replace("/#/match-summary", "")

        # FLASHSCOREに遷移
        page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/?isDetailPopup=true#/match-summary/match-summary")
        time.sleep(3)

        # 試合の国及びカテゴリを取得
        game_info_sub_info_text = ""
        try:
            team_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-2)
            game_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-1)
            #game_info_sub_info = game_info.locator("span.tournamentHeader__country")
            team_info_text = team_info.text_content()
            game_info_text = game_info.text_content()
            print(team_info_text)
            print(game_info_text)
        except Exception as e:
            print("チーム名の取得失敗", e)
            continue

        if team_info_text == "" or game_info_text == "":
            print("チーム名の取得失敗: 空")
            continue

        game_info_sub_info_text = team_info_text + ": " + game_info_text
        print(game_info_sub_info_text)

        #試合のカテゴリが以下の場合のみ取得
        contains_list = []
        contains_list.append("ケニア: プレミアリーグ")
        contains_list.append("コロンビア: プリメーラ A")
        contains_list.append("タンザニア: プレミアリーグ")
        contains_list.append("イングランド: プレミアリーグ")
        contains_list.append("イングランド: EFL チャンピオンシップ")
        contains_list.append("イングランド: EFL リーグ 1")
        contains_list.append("エチオピア: プレミアリーグ")
        contains_list.append("コスタリカ: リーガ FPD")
        contains_list.append("ジャマイカ: プレミアリーグ")
        contains_list.append("スペイン: ラ・リーガ")
        contains_list.append("ブラジル: セリエ A ベターノ")
        contains_list.append("ブラジル: セリエ B")
        contains_list.append("ドイツ: ブンデスリーガ")
        contains_list.append("韓国: K リーグ 1")
        contains_list.append("中国: 中国スーパーリーグ")
        contains_list.append("日本: J1 リーグ")
        contains_list.append("日本: J2 リーグ")
        contains_list.append("日本: J3 リーグ")
        contains_list.append("インドネシア: リーガ 1")
        contains_list.append("オーストラリア: A リーグ・メン")
        contains_list.append("チュニジア: チュニジア･プロリーグ")
        contains_list.append("ウガンダ: プレミアリーグ")
        contains_list.append("メキシコ: リーガ MX")
        contains_list.append("フランス: リーグ・アン")
        contains_list.append("スコットランド: プレミアシップ")
        contains_list.append("オランダ: エールディビジ")
        contains_list.append("アルゼンチン: トルネオ・ベターノ")
        contains_list.append("イタリア: セリエ A")
        contains_list.append("イタリア: セリエ B")
        contains_list.append("ポルトガル: リーガ・ポルトガル")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("セルビア: スーペルリーガ")
        contains_list.append("日本: WEリーグ")
        contains_list.append("ボリビア: LFPB")
        contains_list.append("ブルガリア: パルヴァ・リーガ")
        contains_list.append("カメルーン: エリート 1")
        contains_list.append("ペルー: リーガ 1")
        contains_list.append("エストニア: メスタリリーガ")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("ウクライナ: プレミアリーグ")
        contains_list.append("ベルギー: ジュピラー･プロリーグ")
        contains_list.append("エクアドル: リーガ・プロ")
        contains_list.append("日本: YBC ルヴァンカップ")
        contains_list.append("日本: 天皇杯")
       

        under_list = []
        under_list.append("U17")
        under_list.append("U18")
        under_list.append("U19")
        under_list.append("U20")
        under_list.append("U21")
        under_list.append("U22")
        under_list.append("U23")
        under_list.append("U24")
        under_list.append("U25")

        gender_list = []
        gender_list.append("女子")

        exp_list = []
        exp_list.append("ポルトガル: リーガ・ポルトガル 2")
        exp_list.append("イングランド: プレミアリーグ 2")
        exp_list.append("イングランド: プレミアリーグ U18")

        if any(item in game_info_sub_info_text for item in contains_list):
            pass
        else:
            print("pass1")
            continue

        if any(item in game_info_sub_info_text for item in under_list):
            print("pass2")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in gender_list):
            print("pass3")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in exp_list):
            print("pass4")
            continue
        else:
            pass


        # ホームチーム,アウェーチーム名を取得
        try:
            home_team = page.locator("div.duelParticipant__home")
            home_team_info = home_team.locator("div.participant__participantNameWrapper")
            home_team_sub_info = home_team_info.locator("div.participant__participantName")
            home_team_name = home_team_sub_info.text_content()

            away_team = page.locator("div.duelParticipant__away")
            away_team_info = away_team.locator("div.participant__participantNameWrapper")
            away_team_sub_info = away_team_info.locator("div.participant__participantName")
            away_team_name = away_team_sub_info.text_content()
        except Exception as e:
            print("チーム名の取得失敗", e)
            continue

        # 保持用リストに設定
        notice_data_list.append(game_info_sub_info_text)
        notice_data_list.append(home_team_name)
        notice_data_list.append(away_team_name)
        cron_notice_data_list.append(notice_data_list)

        #監督名(新規タブでスカッドに移動して取得)
        home_manager = ""
        away_manager = ""
        team_data_list = page.locator("a[class='participant__participantLink participant__participantLink--team']")
        tag_flg = 0
        tag_list = []
        manager_name_list = []
        for link in range(team_data_list.count()):
            links = team_data_list.nth(link).get_attribute("href")
            if not links:
                continue
            
            new_page = context.new_page()
            new_page.goto("https://flashscore.co.jp" + links)
            #スカッドに移動
            new_page.goto(new_page.url + "squad/")
            time.sleep(3)
            try:
                manager_name = new_page.locator("a[class='lineupTable__cell--name']")
                if manager_name.count() != 0:
                    if tag_flg == 0:
                        tag_list.append("home")
                    elif tag_flg == 1:
                        tag_list.append("away")
                        
                    name = str.strip(manager_name.nth(manager_name.count() - 1).text_content())
                    manager_name_list.append(name)
            except Exception as e:
                print("スカッドなし", e)

            tag_flg += 1

            new_page.close()

        if len(manager_name_list) != 0:
            if len(manager_name_list) == 1 and tag_list[0] == "home":
                home_manager = manager_name_list[0]
            elif len(manager_name_list) == 1 and tag_list[0] == "away":
                away_manager = manager_name_list[0]
            else:
                home_manager = manager_name_list[0]
                away_manager = manager_name_list[1]
            print(home_manager)
            print(away_manager)

        #審判名,スタジアム,収容人数,観客数,場所
        judge_member = ""
        studium = ""
        capacity = ""
        audience = ""
        location = ""
        try:
            wrapper = page.locator("div[class*='wcl-infoLabelWrapper']")
            wrapperVal = page.locator("div[class*='wcl-infoValue']")
            for ind in range(wrapper.count()):
                key = wrapper.nth(ind).locator("span[data-testid='wcl-scores-overline-02']")
                value = wrapperVal.nth(ind)
                if key.text_content() == "レフェリー:":
                    judge_member = value.text_content()
                elif key.text_content() == "開催地:":
                    studium = str.strip(value.text_content())
                elif key.text_content() == "収容人数:":
                    capacity = value.text_content()
                elif key.text_content() == "参加:":
                    audience = value.text_content()

                prefix = studium.find("(")
                suffix = studium.find(")")
                location = studium[prefix + 1:suffix]

            print(judge_member)
            print(studium)
            print(capacity)
            print(audience)
            print(location)
                
        except Exception as e:
            print("試合場所情報取得失敗", e)

        # スコア取得
        try:
            score = page.locator("div.detailScore__wrapper").text_content()
        except Exception as e:
            print("スコア取得失敗", e)
            continue

        data_list = []
        try:
            # ライブなら0、終了済なら1をnthに導入
            live_fin = page.locator("span[class='fixedHeaderDuel__detailStatus']").nth(1)
            live_fin_text = live_fin.text_content()
            if live_fin_text == '終了済':
                index = 1
            elif live_fin_text == 'ハーフタイム':
                index = 2
            else:
                index = 0

            # イベントタイムを取得
            if index == 0:
                event_time = page.locator("div.eventAndAddedTime")
                event_time_text = event_time.locator("span.eventTime").text_content()
                live_fin_text = event_time_text

            # もっと表示をクリック
            button_locator = page.locator("a[href='#/match-summary/match-statistics']")
            print("もっと表示クリック")
            print(button_locator.count())

            # 拾えなかった場合
            if button_locator.count() == 0:
                print("統計情報が存在しません")
            else:
                page.wait_for_selector("a[href='#/match-summary/match-statistics']")

                # ボタンをクリック
                button_locator.nth(button_locator.count() - 1).click()

                # クリック後、非表示だった要素が表示されるのを待つ
                page.wait_for_selector("strong[data-testid='wcl-scores-simpleText-01']")

                # 期待値を含む全ての情報を取得
                get_data = page.locator("strong[data-testid='wcl-scores-simpleText-01']")  # CSSセレクタを使用
                get_data_count = get_data.count()
                for i in range(get_data_count):
                    text = get_data.nth(i).text_content()
                    data_list.append(text)

                # データ取得不十分の場合スキップ
                if len(data_list) < 2:
                    print("データ取得できなかったためスキップします")
                    continue
    
        except Exception as e:
            print("選択失敗", e)

        # FLASHSCORE(順位表)に遷移
        rank = True
        try:
            reponse = page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/overall")
        except Exception as e:
            print("順位表に遷移できません")
            rank = False

        home_team_max_getting_scorer = ""
        away_team_max_getting_scorer = ""
        home_team_max_getting_scorer_participating = "未出場"
        away_team_max_getting_scorer_participating = "未出場"

        # 順位表メニューをクリック
        home_score = ""
        away_score = ""
        if rank == True:
            try:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/")

                first_team_info = page.locator("div.ui-table__row.table__row--selected").nth(0)
                second_team_info = page.locator("div.ui-table__row.table__row--selected").nth(1)
                first_team_info_all = first_team_info.text_content()
                second_team_info_all = second_team_info.text_content()
                first_team_info_rank = first_team_info.locator("div.tableCellRank").text_content()
                second_team_info_rank = second_team_info.locator("div.tableCellRank").text_content()
            except Exception as e:
                print("順位表選択失敗", e)
                continue

            # ホームチーム,アウェーチームの順位を特定する
            home_rank = first_team_info_rank if home_team_name in first_team_info_all else second_team_info_rank
            away_rank = second_team_info_rank if home_rank == first_team_info_rank else first_team_info_rank

            # スコア分割
            score_split = score.split("-")
            home_score = score_split[0]
            away_score = score_split[1]

            #チーム得点,失点取得
            home_team_home_score = ""
            home_team_home_lost = ""
            away_team_home_score = ""
            away_team_home_lost = ""
            home_team_away_score = ""
            home_team_away_lost = ""
            away_team_away_score = ""
            away_team_away_lost = ""
            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/home")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub1 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub2 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub1.count())
                print(team_info_sub2.count())
                home_total_1 = ""
                home_total_2 = ""
                for ind in range(team_info_sub1.count()):
                    if team_info_sub1.nth(ind).text_content() == home_team_name:
                        home_total_1 = team_info_sub2.nth(ind).text_content()
                    elif team_info_sub1.nth(ind).text_content() == away_team_name:
                        home_total_2 = team_info_sub2.nth(ind).text_content()

                    if home_total_1 != "" and home_total_2 != "":
                        break

                if home_total_1 != "":
                    home_total_sub_1 = home_total_1.split(':')
                    home_team_home_score = home_total_sub_1[0]
                    home_team_home_lost = home_total_sub_1[1]

                if home_total_2 != "":
                    away_total_sub_1 = home_total_2.split(':')
                    away_team_home_score = away_total_sub_1[0]
                    away_team_home_lost = away_total_sub_1[1]

            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/away")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub3 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub4 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub3.count())
                print(team_info_sub4.count())
                away_total_1 = ""
                away_total_2 = ""
                for ind in range(team_info_sub3.count()):
                    if team_info_sub3.nth(ind).text_content() == home_team_name:
                        away_total_1 = team_info_sub4.nth(ind).text_content()
                    elif team_info_sub3.nth(ind).text_content() == away_team_name:
                        away_total_2 = team_info_sub4.nth(ind).text_content()

                    if away_total_1 != "" and away_total_2 != "":
                        break

                if away_total_1 != "":
                    away_total_sub_1 = away_total_1.split(':')
                    home_team_away_score = away_total_sub_1[0]
                    home_team_away_lost = away_total_sub_1[1]

                if away_total_2 != "":
                    away_total_sub_2 = away_total_2.split(':')
                    away_team_away_score = away_total_sub_2[0]
                    away_team_away_lost = away_total_sub_2[1]

            print(home_team_home_score)
            print(home_team_home_lost)
            print(away_team_home_score)
            print(away_team_home_lost)
            print(home_team_away_score)
            print(home_team_away_lost)
            print(away_team_away_score)
            print(away_team_away_lost)

            # チーム得点王取得
            print("チーム得点王取得")
            get_score_flg = True
            try:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/top_scorers/") # CSSセレクタを使用

                #ページが存在しない場合スキップ(処理追加)
                dummy_page = "https://www.flashscore.co.jp/match/" + random_game_str + "/#/match-summary/match-statistics"
                if page.url != dummy_page:
            
                    top_scorer_member = ""
                    top_scorer_team = ""
                    home_scorer_flg = False
                    away_scorer_flg = False
                    #得点王表が完全表示されるまで待機
                    time.sleep(2)
                    top_scorer_graph = page.locator("div[class*='topScorers__row--selected']")

                    for graph_ind in range(top_scorer_graph.count()):
                        sub_top_scorer_graph = top_scorer_graph.nth(graph_ind).locator("div.topScorers__participantCell").locator("a[class*='topScorersParticipant']")
                        tmp = 0
                        for sub_graph_ind in range(sub_top_scorer_graph.count()):
                            if tmp == 0:
                                top_scorer_member = sub_top_scorer_graph.nth(sub_graph_ind).text_content()
                            elif tmp == 1:
                                top_scorer_team = sub_top_scorer_graph.nth(sub_graph_ind).text_content()

                            if top_scorer_team == home_team_name and home_team_max_getting_scorer == "":
                                home_team_max_getting_scorer = top_scorer_member
                                home_scorer_flg = True
                            if top_scorer_team == away_team_name and away_team_max_getting_scorer == "":
                                away_team_max_getting_scorer = top_scorer_member
                                away_scorer_flg = True

                            tmp += 1
                                           
                        if home_scorer_flg == True and away_scorer_flg == True:
                            break

                    print("home_team_max_getting_scorer:" + home_team_max_getting_scorer) #取得済み
                    print("away_team_max_getting_scorer:" + away_team_max_getting_scorer) #取得済み

                else:
                    get_score_flg = False

            except Exception as e:
                print("得点王選択失敗", e)
                get_score_flg = False

            # チーム得点王,フォーメーション取得
            home_formation = ""
            away_formation = ""
            if get_score_flg == True:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/match-summary/lineups")
                #表示待ち
                time.sleep(2)

                try:
                    formation = page.locator("span[class='lf__headerPart']")
                    page.wait_for_selector("span[class='lf__headerPart']", state="visible", timeout=2000)
                            
                    print(formation)
                    home_formation = formation.nth(0).text_content()
                    away_formation = formation.nth(1).text_content()
                    print(home_formation)
                    print(away_formation)
                except Exception as e:
                    print("フォーメーション取得失敗", e)
                
                home_lineup = page.locator("div.lf__side").nth(0)
                away_lineup = page.locator("div.lf__side").nth(1)
                if home_lineup.count() != 11 and away_lineup.count() != 11:
                    home_lineup = page.locator("div.lf__side").nth(2)
                    away_lineup = page.locator("div.lf__side").nth(3)
                    
                try:
                    home_lineup_box = home_lineup.locator("div[class='lf__participantNew']")
                    print(home_lineup_box.count())
                    for home_box_ind in range(home_lineup_box.count()):
                        try:
                            home_lineup_box_sub = home_lineup_box.nth(home_box_ind).locator("div[class='wcl-participant_QKIld wcl-hasNumber_g943f']")
                            home_lineup_box_sub_all = home_lineup_box_sub.locator("div[class='wcl-mainRow_Xi7Hi']")
                            home_lineup_box_sub_member = home_lineup_box_sub_all.locator("a[class='wcl-link_bLtj3 wcl-linkBase_CdaEq wcl-primary_C2HA0 wcl-nameWrapper_m-xnS']")
                            home_lineup_box_sub_member2 = home_lineup_box_sub_member.locator("strong[class='wcl-simpleText_Asp-0 wcl-scores-simpleText-01_pV2Wk wcl-bold_roH-0 wcl-name_IO3G1']")
                            
                            #出場しているか
                            print("**************************")
                            print(home_team_max_getting_scorer)
                            print(home_lineup_box_sub_member2.text_content())
                            print("**************************")
                            if home_team_max_getting_scorer == home_lineup_box_sub_member2.text_content() or home_team_max_getting_scorer in home_lineup_box_sub_member2.text_content():
                                home_team_max_getting_scorer_participating = "出場中"
                                break

                            if home_box_ind == 10:
                                break
                                    
                            #if home_team_max_getting_scorer_participating == "":
                                #交代選手の場合の取得
                            #    home_sub_lineup = page.locator("div.lf__side").nth(2)

                            #    home_lineup_subbox = home_sub_lineup.locator("div.lf__participantNew")
                            #    for home_box_ind in range(home_lineup_subbox.count()):
                            #        home_lineup_box_sub = home_lineup_subbox.nth(home_box_ind).locator("div[data-testid='wcl-lineupsParticipantGeneral-left']")
                            #        home_lineup_box_sub_all = home_lineup_box_sub.locator("div[class*='wcl-mainRow']")
                        
                            #        home_lineup_box_sub_member = home_lineup_box_sub_all.nth(0)
                            #        home_lineup_box_sub_member2 = home_lineup_box_sub_member.nth(0).locator("a[data-testid='wcl-textLink']").locator("strong[data-testid='wcl-scores-simpleText-01']")
                            
                                    #ベンチ入りしているか
                            #        if home_team_max_getting_scorer == home_lineup_box_sub_member2.nth(0).text_content() or home_team_max_getting_scorer in home_lineup_box_sub_member2.nth(0).text_content():
                            #            home_team_max_getting_scorer_participating = "交替済み"
                            #            break
                        except Exception as e:
                            print("取得エラー", e)
                            break

                    print("出場結果")
                    print(home_team_max_getting_scorer_participating)
                #except Exception as e:
                #    print("出場結果取得失敗", e)

                #try:
                    away_lineup_box = away_lineup.locator("div[class='lf__participantNew lf__isReversed']")
                    print(away_lineup_box.count())
                    for away_box_ind in range(away_lineup_box.count()):
                        try:
                            away_lineup_box_sub = away_lineup_box.nth(away_box_ind).locator("div[class='wcl-participant_QKIld wcl-rtl_245jA wcl-hasNumber_g943f']")
                            away_lineup_box_sub_all = away_lineup_box_sub.locator("div[class='wcl-mainRow_Xi7Hi']")
                            away_lineup_box_sub_member = away_lineup_box_sub_all.locator("a[class='wcl-link_bLtj3 wcl-linkBase_CdaEq wcl-primary_C2HA0 wcl-nameWrapper_m-xnS']")
                            away_lineup_box_sub_member2 = away_lineup_box_sub_member.locator("strong[class='wcl-simpleText_Asp-0 wcl-scores-simpleText-01_pV2Wk wcl-bold_roH-0 wcl-name_IO3G1']")

                            #出場しているか
                            print("**************************")
                            print(away_team_max_getting_scorer)
                            print(away_lineup_box_sub_member2.text_content())
                            print("**************************")
                            if away_team_max_getting_scorer == away_lineup_box_sub_member2.text_content() or away_team_max_getting_scorer in away_lineup_box_sub_member2.text_content():
                                away_team_max_getting_scorer_participating = "出場中"
                                break

                            if away_box_ind == 10:
                                break

                            #if away_team_max_getting_scorer_participating == "":
                                #交代選手の場合の取得
                            #    away_sub_lineup = page.locator("div.lf__side").nth(3)

                            #    away_lineup_subbox = away_sub_lineup.locator("div.lf__participantNew")
                            #    for away_box_ind in range(away_lineup_subbox.count()):
                            #        away_lineup_box_sub = away_lineup_subbox.nth(away_box_ind).locator("div[data-testid='wcl-lineupsParticipantGeneral-right']")
                            #        away_lineup_box_sub_all = away_lineup_box_sub.locator("div[class*='wcl-mainRow']")
                        
                            #        away_lineup_box_sub_member = away_lineup_box_sub_all.nth(0)
                            #        away_lineup_box_sub_member2 = away_lineup_box_sub_member.nth(0).locator("a[data-testid='wcl-textLink']").locator("strong[data-testid='wcl-scores-simpleText-01']")
                            
                                    #ベンチ入りしているか
                            #        if away_team_max_getting_scorer == away_lineup_box_sub_member2.nth(0).text_content() or away_team_max_getting_scorer in away_lineup_box_sub_member2.nth(0).text_content():
                            #            away_team_max_getting_scorer_participating = "交替済み"
                            #            break
                        except Exception as e:
                            print("取得エラー", e)
                            break

                    print("出場結果")
                    print(away_team_max_getting_scorer_participating)
                except Exception as e:
                    print("出場結果取得失敗", e)

        #天気,気温,湿度
        weather = ""
        temperature = ""
        humid = ""

        #country_sub = game_info_sub_info_text.split(":")
        #country = country_sub[0]
        #print(country)
        #print(location)

        #query = f"{country} {location} 天気"
        #url = f"https://www.google.com/search?={query}"
        #page.goto(url)

        #try:
        #    weather = page.locator("img#dimg_SvOeZ8_mDoOv0-kPt_Os4Qg_1").get_attribute("alt")
        #    temperature = page.locator("span#wob_tm").inner_text()
        #    humid = page.locator("span#wob_hm").inner_text()
        #    print(weather)
        #    print(temperature)
        #    print(humid)
        #except Exception as e:
        #    print("天気取得失敗", e)

        # エクセルに表記
        loop_bef_seq = makeExcelAndNotice(loop_bef_seq, home_rank, away_rank, home_score, away_score, home_team_name, away_team_name, weather, temperature, humid,
                           game_info_sub_info_text, live_fin_text, data_list, random_game_str, home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
                        home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost, 
                       home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating)

    #noticeCronMail(cron_notice_data_list)

    page.close()
    context.close()
    browser.close()


def makeExcelAndNotice(loop_bef_seq, home_rank, away_rank, home_score, away_score, home_team_name, away_team_name, weather, temperature, humid,
                       game_info_sub_info_text, live_fin_text, data_list, random_game_str, home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
                       home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
                       home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating):

    # データ整形
    edit_list = []
    for i in range(0, len(data_list), 3):
        edit_list.append(data_list[i: i + 3])

    event_time = ""
    home_exp = ""
    away_exp = ""
    home_domination = ""
    away_domination = ""
    home_shoot_all = ""
    away_shoot_all = ""
    home_shoot_in = ""
    away_shoot_in = ""
    home_shoot_out = ""
    away_shoot_out = ""
    home_block_shoot = ""
    away_block_shoot = ""
    home_big_chance = ""
    away_big_chance = ""
    home_corner = ""
    away_corner = ""
    home_shoot_box_in = ""
    away_shoot_box_in = ""
    home_shoot_box_out = ""
    away_shoot_box_out = ""
    home_goal_post = ""
    away_goal_post = ""
    home_head_goal = ""
    away_head_goal = ""
    home_keeper_save = ""
    away_keeper_save = ""
    home_free_kick = ""
    away_free_kick = ""
    home_offside = ""
    away_offside = ""
    home_foul = ""
    away_foul = ""
    home_yellow_card = ""
    away_yellow_card = ""
    home_red_card = ""
    away_red_card = ""
    home_slow_in = ""
    away_slow_in = ""
    home_box_touch = ""
    away_box_touch = ""
    home_pass = ""
    away_pass = ""
    home_final_pass = ""
    away_final_pass = ""
    home_cross = ""
    away_cross = ""
    home_tackle = ""
    away_tackle = ""
    home_clear = ""
    away_clear = ""
    home_intercept = ""
    away_intercept = ""
    for edit in edit_list:
        if edit[1] == "ゴール期待値（xG）":
            home_exp = edit[0]
            away_exp = edit[2]
        elif edit[1] == "ボール支配率":
            home_domination = edit[0]
            away_domination = edit[2]
        elif edit[1] == "シュート数":
            home_shoot_all = edit[0]
            away_shoot_all = edit[2]
        elif edit[1] == "枠内シュート":
            home_shoot_in = edit[0]
            away_shoot_in = edit[2]
        elif edit[1] == "枠外シュート":
            home_shoot_out = edit[0]
            away_shoot_out = edit[2]
        elif edit[1] == "ブロックト・シュート":
            home_block_shoot = edit[0]
            away_block_shoot = edit[2]
        elif edit[1] == "ビッグチャンス":
            home_big_chance = edit[0]
            away_big_chance = edit[2]
        elif edit[1] == "コーナーキック":
            home_corner = edit[0]
            away_corner = edit[2]
        elif edit[1] == "ボックス内からのシュート":
            home_shoot_box_in = edit[0]
            away_shoot_box_in = edit[2]
        elif edit[1] == "ボックス外からのシュート":
            home_shoot_box_out = edit[0]
            away_shoot_box_out = edit[2]
        elif edit[1] == "ゴール枠に当たる":
            home_goal_post = edit[0]
            away_goal_post = edit[2]
        elif edit[1] == "ヘディングによるゴール":
            home_head_goal = edit[0]
            away_head_goal = edit[2]
        elif edit[1] == "キーパーセーブ数":
            home_keeper_save = edit[0]
            away_keeper_save = edit[2]
        elif edit[1] == "フリーキック":
            home_free_kick = edit[0]
            away_free_kick = edit[2]
        elif edit[1] == "オフサイド":
            home_offside = edit[0]
            away_offside = edit[2]
        elif edit[1] == "ファウル":
            home_foul = edit[0]
            away_foul = edit[2]
        elif edit[1] == "イエローカード":
            home_yellow_card = edit[0]
            away_yellow_card = edit[2]
        elif edit[1] == "レッドカード":
            home_red_card = edit[0]
            away_red_card = edit[2]
        elif edit[1] == "スローイン":
            home_slow_in = edit[0]
            away_slow_in = edit[2]
        elif edit[1] == "相手ボックス内でのタッチ":
            home_box_touch = edit[0]
            away_box_touch = edit[2]
        elif edit[1] == "パス数":
            home_pass = edit[0]
            away_pass = edit[2]
        elif edit[1] == "ファイナルサードでのパス数":
            home_final_pass = edit[0]
            away_final_pass = edit[2]
        elif edit[1] == "クロス数":
            home_cross = edit[0]
            away_cross = edit[2]
        elif edit[1] == "タックル数":
            home_tackle = edit[0]
            away_tackle = edit[2]
        elif edit[1] == "クリアリング数":
            home_clear = edit[0]
            away_clear = edit[2]
        elif edit[1] == "インターセプト数":
            home_intercept = edit[0]
            away_intercept = edit[2]

    # 条件に該当するならメール通知
    chkResult = []
    chkResult = conditionChk(game_info_sub_info_text, home_rank, away_rank, home_score, away_score, home_exp, away_exp, weather, temperature, humid,
                             home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in, live_fin_text,  home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
                             home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
                             home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating)
    print(chkResult)

    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    # 通知フラグ
    notice_flg = "未通知"
    if chkResult[0] != "" and chkResult[1] != "":
        subject = "試合データです"
        # jsonデータ作成
        json_data = {
            "通知詳細:": chkResult[0],
            "試合カテゴリ": game_info_sub_info_text,
            "試合時間": live_fin_text,
            "ホーム": {
                "ホームチーム": home_team_name,
                "ホーム順位": home_rank,
                "ホームスコア": home_score,
                "ホーム支配率": home_domination,
                "ホームシュート": home_shoot_all,
                "ホーム枠内シュート": home_shoot_in,
            },
            "アウェイ": {
                "アウェイチーム": away_team_name,
                "アウェイ順位": away_rank,
                "アウェイスコア": away_score,
                "アウェイ支配率": away_domination,
                "アウェイシュート": away_shoot_all,
                "アウェイ枠内シュート": away_shoot_in,
            }
        }
        json_str = json.dumps(json_data, ensure_ascii=False, indent=4)

        # ボディ
        body = f"試合データ:\n\n{json_str}"
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = to_mail
        msg['To'] = to_mail

        # 送信
        try:
            mailobj = smtplib.SMTP(send_mail, 587)
            mailobj.ehlo()
            mailobj.starttls()
            mailobj.ehlo()
            mailobj.login(to_mail, password)
            mailobj.sendmail(to_mail, to_mail, msg.as_string())
            mailobj.quit()
        except Exception as e:
            print("送信失敗", e)

        notice_flg = "通知済"
        print("送信成功")

    print("****************")
    print(game_info_sub_info_text)
    print(home_team_name)
    print(away_team_name)
    print("****************")
    if ":" in game_info_sub_info_text:
        country, rest = game_info_sub_info_text.split(":", 1)
        league = rest.split("-")[0].strip()

    # データ取得時間
    get_data_time = datetime.datetime.now()
    # 出力するエクセルを決定する(存在するoutput_通番.xlsxの中で最大の通番+1を持つエクセルを設定する)
    home_play_style = ""
    away_play_style = ""
    goal_prob = ""
    match_time = ""
    output_file_path = ""
    try:
        sheet = ""
        if loop_bef_seq < 0:
            for seq in range(1, 1200):
                OUTPUT_EXCEL_NAME = "output_" + str(seq) + ".xlsx"
                OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
                if os.path.exists(OUTPUT_EXCEL_PATH):
                    loop_bef_seq = seq
                else:
                    write_bk = Workbook()
                    sheet = write_bk.active
                    loop_bef_seq = seq
                    break
                    
        OUTPUT_EXCEL_NAME = "output_" + str(loop_bef_seq) + ".xlsx"
        OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
        if os.path.exists(OUTPUT_EXCEL_PATH):
            write_bk = load_workbook(OUTPUT_EXCEL_PATH)
            sheet = write_bk.active
        else:
            write_bk = Workbook()
            sheet = write_bk.active
            
        max_row = sheet.max_row + 1

        #初期化
        goal_time = ""
        goal_member = ""
        judge = ""
        home_team_style = ""
        away_team_style = ""
        # データなしの場合ヘッダー追加
        header_data = [
            "ホーム順位", "試合国及びカテゴリ", "試合時間", "ホームチーム", "ホームスコア", "アウェイ順位", "アウェイチーム",
            "アウェイスコア", "ホーム期待値", "アウェイ期待値",
            "ホームボール支配率", "アウェイボール支配率", "ホームシュート数", "アウェイシュート数",
            "ホーム枠内シュート数", "アウェイ枠内シュート数", "ホーム枠外シュート数", "アウェイ枠外シュート数",
            "ホームブロックシュート", "アウェイブロックシュート", "ホームビッグチャンス", "アウェイビッグチャンス",
            "ホームコーナーキック", "アウェイコーナーキック", "ホームボックス内シュート", "アウェイボックス内シュート",
            "ホームボックス外シュート", "アウェイボックス外シュート",
            "ホームゴールポスト", "アウェイゴールポスト", "ホームヘディングゴール", "アウェイヘディングゴール",
            "ホームキーパーセーブ", "アウェイキーパーセーブ", "ホームフリーキック", "アウェイフリーキック",
            "ホームオフサイド", "アウェイオフサイド", "ホームファウル", "アウェイファウル",
            "ホームイエローカード", "アウェイイエローカード", "ホームレッドカード", "アウェイレッドカード", "ホームスローイン", "アウェイスローイン",
            "ホーム相手ボックスタッチ", "アウェイ相手ボックスタッチ", "ホームパス", "アウェイパス", "ホームファイナルサードパス", "アウェイファイナルサードパス",
            "ホームクロス", "アウェイクロス", "ホームタックル", "アウェイタックル", "ホームクリア", "アウェイクリア", "ホームインターセプト", "アウェイインターセプト",
            "スコア時間", "天気", "気温", "湿度", "審判名", "ホーム監督名", "アウェイ監督名", "ホームフォーメーション", "アウェーフォーメーション",
            "スタジアム", "収容人数", "観客数", "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者", "ホームチーム最大得点取得者出場状況", "アウェーチーム最大得点取得者出場状況",
            "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点", "ホームチームアウェー得点", "ホームチームアウェー失点",
            "アウェーチームアウェー得点", "アウェーチームアウェー失点", "通知フラグ", "試合リンク文字列", "ゴール時間", "選手名", "判定結果", "ホームチームスタイル", "アウェイチームスタイル",
            "ゴール確率", "得点予想時間"
        ]
        if max_row == 2:
            sheet.append(header_data)

        row_data = [
            home_rank, game_info_sub_info_text, live_fin_text, home_team_name, home_score, away_rank, away_team_name,
            away_score, home_exp, away_exp,
            home_domination, away_domination,
            home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in,
            home_shoot_out, away_shoot_out, home_block_shoot, away_block_shoot,
            home_big_chance, away_big_chance, home_corner, away_corner,
            home_shoot_box_in, away_shoot_box_in, home_shoot_box_out, away_shoot_box_out,
            home_goal_post, away_goal_post, home_head_goal, away_head_goal,
            home_keeper_save, away_keeper_save, home_free_kick, away_free_kick,
            home_offside, away_offside, home_foul, away_foul,
            home_yellow_card, away_yellow_card, home_red_card, away_red_card, home_slow_in, away_slow_in,
            home_box_touch, away_box_touch, home_pass, away_pass, home_final_pass, away_final_pass,
            home_cross, away_cross, home_tackle, away_tackle, home_clear, away_clear, home_intercept, away_intercept,
            get_data_time, weather, temperature, humid, judge_member, home_manager, away_manager, home_formation, away_formation,
            studium, capacity, audience, home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating,
            home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score,
            away_team_away_lost, notice_flg, random_game_str, goal_time, goal_member, judge
        ]

        home_play_style, away_play_style, goal_prob, match_time, output_file_path = write_prediction_data.copydata(row_data, country, league, home_team_name, away_team_name)
    except Exception as e:
        print("確率,時間,フォーメーション取得失敗", e)
        
    print("****確率,時間,フォーメーション****")
    print(goal_prob)
    print(match_time)
    print(home_play_style)
    print(away_play_style)
    print("****確率,時間,フォーメーション****")

    row_data = [
            home_rank, game_info_sub_info_text, live_fin_text, home_team_name, home_score, away_rank, away_team_name,
            away_score, home_exp, away_exp,
            home_domination, away_domination,
            home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in,
            home_shoot_out, away_shoot_out, home_block_shoot, away_block_shoot,
            home_big_chance, away_big_chance, home_corner, away_corner,
            home_shoot_box_in, away_shoot_box_in, home_shoot_box_out, away_shoot_box_out,
            home_goal_post, away_goal_post, home_head_goal, away_head_goal,
            home_keeper_save, away_keeper_save, home_free_kick, away_free_kick,
            home_offside, away_offside, home_foul, away_foul,
            home_yellow_card, away_yellow_card, home_red_card, away_red_card, home_slow_in, away_slow_in,
            home_box_touch, away_box_touch, home_pass, away_pass, home_final_pass, away_final_pass,
            home_cross, away_cross, home_tackle, away_tackle, home_clear, away_clear, home_intercept, away_intercept,
            get_data_time, weather, temperature, humid, judge_member, home_manager, away_manager, home_formation, away_formation,
            studium, capacity, audience, home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating,
            home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score,
            away_team_away_lost, notice_flg, random_game_str, goal_time, goal_member, judge, home_play_style, away_play_style, str(goal_prob), str(match_time)
    ]
    sheet.append(row_data)

    # 保存
    write_bk.save(OUTPUT_EXCEL_PATH)
    print("メイン書き込み完了")


    #差分データの通知
    chk_diff_data_send_mail(output_file_path, game_info_sub_info_text, live_fin_text, home_team_name, away_team_name, home_score, away_score,
                            header_data, row_data)

    #条件式バイナリファイル作成
    executeInspect()

    return loop_bef_seq

def conditionChk(game_info_sub_info_text, home_rank, away_rank, home_score, away_score, home_exp, away_exp, weather, temperature, humid,
                 home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in, live_fin_text,  home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
                 home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
                 home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating):
    home_rank_rep = home_rank.replace(".", "")
    away_rank_rep = away_rank.replace(".", "")
    chkResult = []
    time_flg = ""
    # 90+のような表記の場合
    if "+" in live_fin_text:
        live_fin_text = live_fin_text.replace("+", "")
        live_fin_text = live_fin_text.replace("'", "")
        if int(live_fin_text) < 20:
            time_flg = "OK"

    # 68'のような表記の場合
    if "'" in live_fin_text:
        live_fin_text = live_fin_text.replace("'", "")
        if int(live_fin_text) < 20:
            time_flg = "OK"

    # 45:13のような表記の場合
    if ":" in live_fin_text:
        live_fin_text = live_fin_text.replace(":", "")
        if int(live_fin_text) < 2000:
            time_flg = "OK"

    print(live_fin_text)

    # アフリカの以下の国は順位がホームが上であること
    if "ケニア" in game_info_sub_info_text or "タンザニア" in game_info_sub_info_text or "ウガンダ" in game_info_sub_info_text:
        if (int(home_rank_rep) > int(away_rank_rep)) and (int(home_rank_rep) < 4 and int(away_rank_rep) > 9):
            chkResult.append("ケニア,タンザニア,ウガンダのいずれかです")
            chkResult.append(time_flg)
            return chkResult
        else:
            if (home_shoot_in != "" and home_shoot_all != "" and away_shoot_in != "" and away_shoot_all != "" and (
                    (float(int(home_shoot_in) / int(home_shoot_all)) >= 0.6) or (
                    float(int(away_shoot_in) / int(away_shoot_all)) >= 0.6))):
                chkResult.append("ケニア,タンザニア,ウガンダのいずれかで期待値は0.6以上です")
                chkResult.append(time_flg)
                return chkResult

    # 得点した方が得点数と枠内シュート数が同数（ただし1点）
    if time_flg == "OK" and (home_score == "1" or away_score == "1") and (home_score == home_shoot_all or away_score == away_shoot_all):
        chkResult.append("得点した方が得点数と枠内シュート数が同数（ただし1点）です")
        chkResult.append(time_flg)
        return chkResult

    # シュートが3本以上,枠内シュート2本以上,アウェーはシュートが2本以上,枠内シュート1本以上
    if (home_shoot_all != "" and home_shoot_in != "") and (away_shoot_all != "" and away_shoot_in != "") and (
            int(home_shoot_all) >= 3 and int(home_shoot_in) >= 2) and (
            int(away_shoot_all) >= 2 and int(away_shoot_in) >= 1):
        if ((float(int(home_shoot_in) / int(home_shoot_all)) >= 0.6) or (
                float(int(away_shoot_in) / int(away_shoot_all)) >= 0.6)):
            chkResult.append("シュートが3本以上,枠内シュート2本以上,アウェーはシュートが2本以上,枠内シュート1本以上です")
            chkResult.append(time_flg)
            return chkResult
        elif ((home_score == "1" or away_score == "1") and (
                float(int(home_shoot_in) / int(home_shoot_all)) >= 0.4) and (
                      float(int(away_shoot_in) / int(away_shoot_all)) >= 0.4)):
            chkResult.append("片方が1点取っており,どちらも4割以上の枠内シュートです。")
            chkResult.append(time_flg)
            return chkResult

    # ハーフタイム時に期待値が1.10以上,お互いシュート数が3本以上
    if live_fin_text == 'ハーフタイム' and home_exp != "" and float(home_exp) >= 1.1 and int(home_shoot_all) >= 3 and int(
            away_shoot_all) >= 3:
        chkResult.append("ハーフタイム時に期待値が1.10以上,お互いシュート数が3本以上です")
        chkResult.append(time_flg)
        return chkResult

    chkResult.append("")
    chkResult.append("")
    return chkResult


def noticeCronMail(cron_notice_data_list):
    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "cronの定期実行通知"
    body = "cronが実行されました。以下のライブデータに対してスクレイピングを行いました。"
    strings = ""
    for k in range(len(cron_notice_data_list)):
        stringstmp = ""
        data_list = cron_notice_data_list[k]
        chk = 0
        for l in range(len(data_list)):
            stringstmp += data_list[l]
            if chk == 0:
                stringstmp += ":"
            elif chk == 1:
                stringstmp += " vs "

            chk += 1
        if len(stringstmp) > 0:
            strings += "\n"
        strings += stringstmp

    body += "\n"
    body += "\n"
    body += strings

    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)

    print("cron実行通知を送信")

def executeInspect():
    source = isp.getsource(conditionChk)
    OUTPUT_CONDITION_PATH = "/Users/shiraishitoshio/bookmaker/conditiondata/conditiondata.csv"
    if os.path.exists(OUTPUT_CONDITION_PATH):
        os.remove(OUTPUT_CONDITION_PATH)
        
    with open(OUTPUT_CONDITION_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        for line in source.splitlines():
            writer.writerow([line])

def read_prediction_data(file_path):
    try:
        df = pd.read_excel(file_path)
        df.fillna(0, inplace=True)
        return df.tail(2) if len(df) >= 2 else df
    except Exception as e:
        print(f"Excelファイルの読み込みに失敗しました: {e}")
        return pd.DataFrame()

def diff_record(df, header_data):
    diff_list = []
    try:
        if len(df) < 2:
            print("データが2行未満のため、差分を取得できません。")
            return diff_list
    except Exception as e:
        print(f"Excelファイルの読み込みに失敗しました: {e}")
        return diff_list

    except_list = [
        "ホーム順位", "試合国及びカテゴリ", "試合時間", "ホームチーム", "ホームスコア", "アウェイ順位", "アウェイチーム",
        "アウェイスコア", "スコア時間", "天気", "気温", "湿度", "審判名", "ホーム監督名", "アウェイ監督名", "ホームフォーメーション", "アウェーフォーメーション",
        "スタジアム", "収容人数", "観客数", "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者", "ホームチーム最大得点取得者出場状況", "アウェーチーム最大得点取得者出場状況",
        "通知フラグ", "試合リンク文字列", "ゴール時間", "選手名", "判定結果", "ホームチームスタイル", "アウェイチームスタイル",
        "ゴール確率", "得点予想時間"
    ]

    three_data_list = [
        "ホームパス", "アウェイパス",
        "ホームファイナルサードパス", "アウェイファイナルサードパス",
        "ホームクロス", "アウェイクロス",
        "ホームタックル", "アウェイタックル"
    ]

    exp_list = [
        "ホーム期待値", "アウェイ期待値"
    ]

    percentage_list = [
        "ホームボール支配率", "アウェイボール支配率"
    ]

    
    for header in header_data:
        diff = ""
        if header in df.columns:  # カラムが存在するか確認
            if header in except_list:
                diff_list.append(diff)
                continue
            
            elif header in exp_list:
                bef = float(df.iloc[0][header])
                af = float(df.iloc[1][header])
                diff = af - bef
                if diff >= 0:
                    diff = "+" + str(diff)
                else:
                    diff = str(diff)

            elif header in percentage_list:
                bef = convert_feature_data(df.iloc[0][header])
                af = convert_feature_data(df.iloc[1][header])
                diff = af - bef
                diff *= 100
                if diff >= 0:
                    diff = "+" + str(diff) + "%"
                else:
                    diff = str(diff) + "%"

            elif header in three_data_list:
                success_bef, attempt_bef = extract_success_and_attempt(df.iloc[0][header])
                success_af, attempt_af = extract_success_and_attempt(df.iloc[1][header])
                success_diff = int(success_af) - int(success_bef)
                attempt_diff = int(attempt_af) - int(attempt_bef)
                success_diff_new = str(success_diff)
                attempt_diff_new = str(attempt_diff)
                if success_diff >= 0:
                    success_diff_new = "+" + success_diff_new
                if attempt_diff >= 0:
                    attempt_diff_new = "+" + attempt_diff_new
                    
                diff = success_diff_new + "/" + attempt_diff_new
            
            else:
                bef = int(df.iloc[0][header])
                af = int(df.iloc[1][header])
                diff = af - bef
                if diff >= 0:
                    diff = "+" + str(diff)
                else:
                    diff = str(diff)

        diff_list.append(diff)

    return diff_list

def convert_feature_data(data):
    print(data)
    data = str(data)
    if "%" in data:
        data = data.replace("%", "")
        return float(data) / 100
    else:
        return float(data)

def extract_success_and_attempt(value):
    """ 34%(20/60) の形式を (成功=20, 試行=60) に分割する """
    match = re.search(r"(\d+)/(\d+)", str(value))
    if match:
        success, attempt = map(int, match.groups())
        return success, attempt
    return 0, 0  # データがない場合は0を返す

def chk_diff_data_send_mail(file_path, game_info_sub_info_text, time, home, away, home_score, away_score, header_data, now_data):
    print("chk_diff_data_send_mail")
    df = read_prediction_data(file_path)
    if len(df) == 0:
        return
    
    diff_list = diff_record(df, header_data)
    if len(diff_list) == 0:
        return
    
    print(header_data)
    print(len(header_data))
    print(now_data)
    print(len(now_data))
    print(diff_list)
    print(len(diff_list))

    #通知するのは特定の特徴量が増加した時
    json_data = {
        key: f"{now} ({diff})"
        for key, now, diff in zip(header_data, now_data, diff_list)
        if str(diff) not in ["0", "+0", "+0/+0", "", "+0.0", "+0.0%"]
    }
    print(json_data)

    if not json_data:
        return

    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "前データとの差分データです"
    json_str = json.dumps(json_data, ensure_ascii=False, indent=4)

    # ボディ
    body = f"試合データ:{game_info_sub_info_text}\n{time}: {home}vs{away}({home_score}-{away_score})\n\n{json_str}"
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)


    print("送信成功")


with sync_playwright() as playwright:
    execute(playwright)

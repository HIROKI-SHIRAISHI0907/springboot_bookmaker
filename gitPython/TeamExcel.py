from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import datetime
import sys
import smtplib
from email.mime.text import MIMEText
import json
import datetime
import inspect as isp
import csv
import time
from python_analytics import write_prediction_data
import pandas as pd
import re


def load_existing_match_keys():
    existing_keys = set()
    for seq in range(1, 3000):
        file_path = f"/Users/shiraishitoshio/bookmaker/team_member_{seq}.xlsx"
        if not os.path.exists(file_path):
            break
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                category = row[0]
                home_team = row[4]
                away_team = row[6]
                key = (category, home_team, away_team)
                existing_keys.add(key)
        except Exception as e:
            print(f"{file_path} 読み込み失敗: {e}")
    return existing_keys


def execute(playwright):
    #データ保持
    existing_keys = load_existing_match_keys()
    
    loop_bef_seq = -1
    # 対象時間を設定
    start_time = datetime.time(23, 30)
    end_time = datetime.time(23, 59)
    # 現在時刻
    now = datetime.datetime.today()
    now_info = datetime.datetime(now.year,now.month,now.day,now.hour,now.minute,now.second)

    if now_info.weekday() == 1 or now_info.weekday() == 2 or now_info.weekday() == 3 or now_info.weekday() == 4 or now_info.weekday() == 5:
        if now.hour == 12 or now.hour == 13 or now.hour == 14 or now.hour == 15 or now.hour == 16:
            print(f"{now_info}  データ取得対象外の時間です。")

    print(f"{now_info}  データ取得対象の時間です。")

    #試合のカテゴリが以下の場合のみ取得
    contains_list = []
    contains_list.append("ケニア: プレミアリーグ")
    contains_list.append("コロンビア: プリメーラ A")
    contains_list.append("タンザニア: プレミアリーグ")
    contains_list.append("イングランド: プレミアリーグ")
    contains_list.append("イングランド: EFL チャンピオンシップ")
    contains_list.append("イングランド: EFL リーグ 1")
    contains_list.append("エチオピア: プレミアリーグ")
    contains_list.append("コスタリカ: リーガ FPD")
    contains_list.append("ジャマイカ: プレミアリーグ")
    contains_list.append("スペイン: ラ・リーガ")
    contains_list.append("ブラジル: セリエ A ベターノ")
    contains_list.append("ブラジル: セリエ B")
    contains_list.append("ドイツ: ブンデスリーガ")
    contains_list.append("韓国: K リーグ 1")
    contains_list.append("中国: 中国スーパーリーグ")
    contains_list.append("日本: J1 リーグ")
    contains_list.append("日本: J2 リーグ")
    contains_list.append("日本: J3 リーグ")
    contains_list.append("インドネシア: リーガ 1")
    contains_list.append("オーストラリア: A リーグ・メン")
    contains_list.append("チュニジア: チュニジア･プロリーグ")
    contains_list.append("ウガンダ: プレミアリーグ")
    contains_list.append("メキシコ: リーガ MX")
    contains_list.append("フランス: リーグ・アン")
    contains_list.append("スコットランド: プレミアシップ")
    contains_list.append("オランダ: エールディビジ")
    contains_list.append("アルゼンチン: トルネオ・ベターノ")
    contains_list.append("イタリア: セリエ A")
    contains_list.append("イタリア: セリエ B")
    contains_list.append("ポルトガル: リーガ・ポルトガル")
    contains_list.append("トルコ: スュペル・リグ")
    contains_list.append("セルビア: スーペルリーガ")
    contains_list.append("日本: WEリーグ")
    contains_list.append("ボリビア: LFPB")
    contains_list.append("ブルガリア: パルヴァ・リーガ")
    contains_list.append("カメルーン: エリート 1")
    contains_list.append("ペルー: リーガ 1")
    contains_list.append("エストニア: メスタリリーガ")
    contains_list.append("トルコ: スュペル・リグ")
    contains_list.append("ウクライナ: プレミアリーグ")
    contains_list.append("ベルギー: ジュピラー･プロリーグ")
    contains_list.append("エクアドル: リーガ・プロ")
    contains_list.append("日本: YBC ルヴァンカップ")
    contains_list.append("日本: 天皇杯")
       

    under_list = []
    under_list.append("U17")
    under_list.append("U18")
    under_list.append("U19")
    under_list.append("U20")
    under_list.append("U21")
    under_list.append("U22")
    under_list.append("U23")
    under_list.append("U24")
    under_list.append("U25")

    gender_list = []
    gender_list.append("女子")

    exp_list = []
    exp_list.append("ポルトガル: リーガ・ポルトガル 2")
    exp_list.append("イングランド: プレミアリーグ 2")
    exp_list.append("イングランド: プレミアリーグ U18")


    # Chromiumブラウザを起動
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    # Googleをひらく
    page.goto("https://www.google.com")

    # flashscoreに遷移
    page.goto("https://flashscore.co.jp/soccer/")
    # 必要な要素がロードされるのを待つ
    page.wait_for_selector("div.lmc__block")

    # 国リストをメニューから取得
    country_blocks = page.query_selector_all("div.lmc__block")

    result = []
    for block in country_blocks:
        name_span = block.query_selector("span.lmc__elementName")
        link_a = block.query_selector("a.lmc__element.lmc__item")

        if name_span and link_a:
            country_name = name_span.inner_text().strip()
            href = link_a.get_attribute("href").strip()

            print(f"{country_name} => 開くクリック")
            # スクロール＋クリックで展開
            block.scroll_into_view_if_needed()
            block.click()
            page.wait_for_timeout(1000)

            # 展開された国名を含む opened_block を取得（block は無効になっている可能性あり）
            opened_blocks = page.query_selector_all("div.lmc__block.lmc__blockOpened")
            opened_block = None
            for ob in opened_blocks:
                title_span = ob.query_selector("span.lmc__elementName")
                if title_span and title_span.inner_text().strip() == country_name:
                    opened_block = ob
                    break

            if not opened_block:
                print(f"{country_name}: 展開されたブロックが見つかりません")
                continue

            # 展開されたブロック内からテンプレート（span）を取得
            template = block.query_selector("span.lmc__template")
            if not template:
                print(f"{country_name}: テンプレート要素なし")
                continue

            # .lmc__template 内のリーグリストを取得
            leagues = template.query_selector_all("a")

            for league in leagues:
                league_name = league.inner_text().strip()

                # 指定の国だけ処理
                key = f"{country_name}: {league_name}"
                if key not in contains_list:
                    continue
                
                print(f"{key}: pass1")
            
                league_href = league.get_attribute("href")
                print(f"   - {league_name} -> {league_href}")
                result.append((country_name, league_name, league_href))

            for data in result:
                country = data[0]
                league = data[1]
                href = data[2]

                # 両方が contains_list に含まれているなら続きの処理を行う
                if country in contains_list and league in contains_list:
                    print("pass2")
                    pass
                else:
                    continue

                if league in gender_list:
                    print("pass3")
                    continue
                else:
                    pass

                if league in exp_list:
                    print("pass4")
                    continue
                else:
                    pass

                # 国,リーグページに遷移
                page.goto("https://flashscore.co.jp/" + href + "/standing")

                # 全ての順位表の行を取得
                rows = page.query_selector_all(".ui-table__row")

                result = []
                for row in rows:
                    team_link = row.query_selector("a.tableCellParticipant__image")

                    if team_link:
                        team_title = team_link.get_attribute("title")
                        href = team_link.get_attribute("href")
        
                        result.append((team_title, href))
                        print(f"チーム名: {team_title} | リンク: {href}")
        
#    print(country_blocks)


        # エクセルに表記
#        loop_bef_seq = makeExcelAndNotice(loop_bef_seq, game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name,
#                                          away_team_name, random_game_str,
#                home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
#                                          home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost, 
#            home_team_max_getting_scorer, away_team_max_getting_scorer)

    #noticeCronMail(cron_notice_data_list)

    page.close()
    context.close()
    browser.close()


def makeExcelAndNotice(loop_bef_seq, game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name, away_team_name, random_game_str,
                home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
                                          home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost, 
            home_team_max_getting_scorer, away_team_max_getting_scorer):


    # データ取得時間
    get_data_time = datetime.datetime.now()
    # 出力するエクセルを決定する(存在するfuture_通番.xlsxの中で最大の通番+1を持つエクセルを設定する)
    home_play_style = ""
    away_play_style = ""
    goal_prob = ""
    match_time = ""
    output_file_path = ""
    try:
        sheet = ""
        if loop_bef_seq < 0:
            for seq in range(1, 3000):
                OUTPUT_EXCEL_NAME = "team_member__" + str(seq) + ".xlsx"
                OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
                if os.path.exists(OUTPUT_EXCEL_PATH):
                    loop_bef_seq = seq
                else:
                    write_bk = Workbook()
                    sheet = write_bk.active
                    loop_bef_seq = seq
                    break
                    
        OUTPUT_EXCEL_NAME = "team_member_" + str(loop_bef_seq) + ".xlsx"
        OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
        if os.path.exists(OUTPUT_EXCEL_PATH):
            write_bk = load_workbook(OUTPUT_EXCEL_PATH)
            sheet = write_bk.active
        else:
            write_bk = Workbook()
            sheet = write_bk.active

        #初期化
        goal_time = ""
        goal_member = ""
        judge = ""
        home_team_style = ""
        away_team_style = ""
        # データなしの場合ヘッダー追加
        header_data = [
            "試合国及びカテゴリ", "試合予定時間", "ホーム順位", "アウェー順位", "ホームチーム", "アウェーチーム",
            "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者",
            "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点", "ホームチームアウェー得点", "ホームチームアウェー失点",
            "アウェーチームアウェー得点", "アウェーチームアウェー失点", "試合リンク文字列", "データ取得時間"
        ]
        
        # ヘッダー未出力なら出力（A1がNone or 空なら）
        # ここで1行目の空白を削除（Workbook 作成時に勝手に入る）
        if sheet.max_row == 1:
            sheet.delete_rows(1)
            # 明示的にヘッダー行を書く
            sheet.append(header_data)

        

        row_data = [
            game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name, away_team_name,
            home_team_max_getting_scorer, away_team_max_getting_scorer,
            home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score,
            away_team_away_lost, random_game_str, get_data_time
        ]
        sheet.append(row_data)
    except Exception as e:
        print("確率,時間,フォーメーション取得失敗", e)

    # 保存
    write_bk.save(OUTPUT_EXCEL_PATH)
    print("メイン書き込み完了")

    return loop_bef_seq

def noticeCronMail(cron_notice_data_list):
    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "cronの定期実行通知"
    body = "cronが実行されました。以下のライブデータに対してスクレイピングを行いました。"
    strings = ""
    for k in range(len(cron_notice_data_list)):
        stringstmp = ""
        data_list = cron_notice_data_list[k]
        chk = 0
        for l in range(len(data_list)):
            stringstmp += data_list[l]
            if chk == 0:
                stringstmp += ":"
            elif chk == 1:
                stringstmp += " vs "

            chk += 1
        if len(stringstmp) > 0:
            strings += "\n"
        strings += stringstmp

    body += "\n"
    body += "\n"
    body += strings

    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)

    print("cron実行通知を送信")


with sync_playwright() as playwright:
    execute(playwright)

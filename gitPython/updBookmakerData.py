import sys
import glob

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import pandas as pd

def findMinBook():
    minSeq = 1200
    for path in glob.glob("/Users/shiraishitoshio/bookmaker/*"):
        if "output_" in path and ".xlsx" in path:
            tmpPath = path.replace("/Users/shiraishitoshio/bookmaker/","")
            tmpPath = tmpPath.replace(".xlsx","")
            tmpPath = tmpPath.replace("output_","")
            if minSeq > int(tmpPath):
                minSeq = int(tmpPath)

    print("最小のエクセル通番:" + str(minSeq))
    return minSeq

def convertMaxSeq():
    maxSeq = 1
    for path in glob.glob("/Users/shiraishitoshio/bookmaker/*"):
        if "output_" in path and ".csv" in path:
            tmpPath = path.replace("/Users/shiraishitoshio/bookmaker/","")
            tmpPath = tmpPath.replace(".csv","")
            tmpPath = tmpPath.replace("output_","")
            if maxSeq < int(tmpPath):
                maxSeq = int(tmpPath)

    print("最大の変換csv通番:" + str(maxSeq))
    return maxSeq + 1

def loadFile(OUTPUT_EXCEL_PATH):
    try:
        if os.path.exists(OUTPUT_EXCEL_PATH):
            pass
        else:
            print("ブックが存在しません。" + OUTPUT_EXCEL_PATH)
            return ""
    except Exception as e:
        print("ファイルが破損しています。" + OUTPUT_EXCEL_PATH)
        os.remove(OUTPUT_EXCEL_PATH)
        return ""

    #ファイル読み込み
    inputFile = OUTPUT_EXCEL_PATH
    try:
        input_book = pd.ExcelFile(inputFile)
    except Exception as e:
        print("ファイルが破損しています。" + OUTPUT_EXCEL_PATH)
        os.remove(OUTPUT_EXCEL_PATH)
        return ""

    input_sheet_df = input_book.parse('Sheet')
    input_sheet_df["連番"] = range(1, len(input_sheet_df) + 1)
    seq = input_sheet_df["連番"]
    goal_time = input_sheet_df["ゴール時間"]
    team_member = input_sheet_df["選手名"]

    empty = 0
    for i in range(len(seq)):
        if (pd.isna(goal_time[i]) or goal_time[i] == "") or (pd.isna(team_member[i]) or team_member[i] == ""):
            empty += 1

    if empty > 0:
        print("ゴール時間,選手名が空のデータがあります。")
        return "convNG"

    return "convOK"

def loadFileAndGetLink(loop_bef_seq, updChkCount):
    while True:
        OUTPUT_EXCEL_NAME = "output_" + str(loop_bef_seq) + ".xlsx"
        OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
        try: 
            if os.path.exists(OUTPUT_EXCEL_PATH):
                break
            else:
                print("ブックが存在しません。" + OUTPUT_EXCEL_NAME)
                return [], [], [], [], OUTPUT_EXCEL_PATH, updChkCount
        except Exception as e:
            print("ファイルが破損しています。" + OUTPUT_EXCEL_NAME)
            os.remove(OUTPUT_EXCEL_PATH)
            return [], [], [], [], OUTPUT_EXCEL_PATH, updChkCount
    
    #ファイル読み込み
    inputFile = OUTPUT_EXCEL_PATH
    try:
        input_book = pd.ExcelFile(inputFile)
    except Exception as e:
        print("ファイルが破損しています。" + OUTPUT_EXCEL_NAME)
        os.remove(OUTPUT_EXCEL_PATH)
        return [], [], [], [], OUTPUT_EXCEL_PATH, updChkCount
    input_sheet_df = input_book.parse('Sheet')
    input_sheet_df["連番"] = range(1, len(input_sheet_df) + 1)
    seq = input_sheet_df["連番"]
    game_link = input_sheet_df["試合リンク文字列"]
    home_score = input_sheet_df["ホームスコア"]
    away_score = input_sheet_df["アウェイスコア"]
    goal_time = input_sheet_df["ゴール時間"]
    team_member = input_sheet_df["選手名"]
    seq_list = []
    link_list = []
    home_score_list = []
    away_score_list = []
    for i in range(len(game_link)):
        if (pd.isna(goal_time[i]) or goal_time[i] == "") or (pd.isna(team_member[i]) or team_member[i] == ""):
            seq_list.append(seq[i])
            link_list.append(game_link[i])
            home_score_list.append(home_score[i])
            away_score_list.append(away_score[i])

    if len(seq_list) == 0:
        print("対象更新データが存在しません。" + str(OUTPUT_EXCEL_PATH))
        return [], [], [], [], OUTPUT_EXCEL_PATH, updChkCount

    updChkCount += 1

    return seq_list, link_list, home_score_list, away_score_list, OUTPUT_EXCEL_PATH, updChkCount

def chkStatistics(playwright, seq_list, link_list, home_score_list, away_score_list, OUTPUT_EXCEL_PATH):
    print("処理開始します。chkStatistics:" + OUTPUT_EXCEL_PATH)
    
    # Chromiumブラウザを起動
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()
    
    #class=smv__participantRowを全て取得
    for i in range(len(seq_list)):
        #FLASHSCOREに遷移
        page.goto("https://www.flashscore.co.jp/match/" + link_list[i] + "/#/match-summary/match-summary")
        #スコア取得
        score_time_info = "---"
        score_member_info = "---"
        no_team_mem = False
        try:
            score = page.locator("div.smv__participantRow")
            score_sub = score.locator("div.smv__incident")
            score_time = score_sub.locator("div.smv__timeBox")
            score_time_sub = score_sub.locator("div.smv__incidentIcon")
            #ゴール情報以外も含んで取得  
            score_all = score_time_sub.locator("div[class*='Score'], svg[class*='Card-ico'], svg[class*='warning']")
            #得点者(見つからない場合はスキップ)
            score_member_list = []
            try:
                score_sub_sub_div = score_sub.locator("div[class*='smv__playerName']")
                for n in range(score_sub_sub_div.count()):
                    score_member_list.append(score_sub_sub_div.nth(n).text_content())

                score_sub_sub_a = score_sub.locator("a[class='smv__playerName']")
                score_member = score_sub_sub_a.locator("div")
                for o in range(score_member.count()):
                    score_member_list.append(score_member.nth(o).text_content())
            except Exception as e:
                print("得点者が見つかりません")
                no_team_mem = True 

            #格納用リスト
            score_all_list = []
            score_time_list = []
                    
            #ホームスコア,アウェイスコア
            home_score_value = ""
            away_score_value = ""
            
            for j in range(score_all.count()):
                home_score = score_all.nth(j).text_content()
                score_all_list.append(home_score)
                score_time_list.append(score_time.nth(j).text_content())

            for k in range(len(score_all_list)):
                if ("-" in score_all_list[k]):
                    #-で分割
                    #1-1など
                    score_split = score_all_list[k].split("-")
                    home_score = score_split[0]
                    away_score = score_split[1].replace(" ","")

                    if (int(home_score_list[i]) == int(home_score)) and (int(away_score_list[i]) == int(away_score)):
                        home_score_value = home_score
                        away_score_value = away_score
                        score_time_info = "---"
                        score_time_info = score_time_list[k]
                        score_member_info = "---"
                        if len(score_member_list) != 0:
                            score_member_info = score_member_list[k]
                            print(score_member_info)
                            
                            
        except Exception as e:
            print("スコア取得失敗", e)
            score_time_info = "取得エラー"
            score_member_info = "取得エラー"

        
        updateExcel(seq_list[i], OUTPUT_EXCEL_PATH, score_time_info, score_member_info)

        print("upd情報記入しました。", OUTPUT_EXCEL_PATH)

    manSeq = convertMaxSeq()
    csvPath = "/Users/shiraishitoshio/bookmaker/output_" + str(manSeq) + ".csv"
    print("convfile:" + csvPath)

    #CSV変換
    excel_to_csv(OUTPUT_EXCEL_PATH, csvPath)

    print("処理終了します。chkStatistics:" + OUTPUT_EXCEL_PATH)

    page.close()
    context.close()
    browser.close()

    #更新レコードが130行以上ある場合updateしたエクセル通番以降の連続更新は行わない
    if len(seq_list) >= 130:
        print("レコードが多いため影響を考慮し処理終了します。chkStatistics:" + str(len(seq_list)))
        sys.exit()
        

def updateExcel(cellIndex, OUTPUT_EXCEL_PATH, score_time_info, score_member_info):
    print("処理開始します。updateExcel:" + OUTPUT_EXCEL_PATH)
    
    try:
        if os.path.exists(OUTPUT_EXCEL_PATH):
            write_bk = load_workbook(OUTPUT_EXCEL_PATH)
            sheet = write_bk.active
        else:
            write_bk = Workbook()
            sheet = write_bk.active

        cellIndex += 1

        if score_member_info == "":
            score_member_info = "---"

        print("ファイル:" + OUTPUT_EXCEL_PATH)
        print("スコア:" + score_time_info)
        print("チームメンバー:" + score_member_info)

        #保存
        sheet["CI" + str(cellIndex)] = score_time_info
        sheet["CJ" + str(cellIndex)] = score_member_info
        write_bk.save(OUTPUT_EXCEL_PATH)
        print("記入完了。updateExcel:" + str(cellIndex) + "行目")
    except Exception as e:
        print("書き込み失敗", e)
        return

    print("処理終了します。updateExcel:" + OUTPUT_EXCEL_PATH)

def excel_to_csv(excel_file, csv_file):
    # Excelファイルを読み込む
    df = pd.read_excel(excel_file)

    #全て記述できているか
    chk = loadFile(excel_file)
    if chk == "convNG":
        return
    
    # CSVファイルに書き込む
    df.to_csv(csv_file, index=False)

    # 原エクセルを削除
    os.remove(excel_file)
    

with sync_playwright() as playwright:
    updChkCount = 0
    loop_bef_seq = findMinBook()
    while True:
        OUTPUT_EXCEL_NAME = "output_" + str(loop_bef_seq) + ".xlsx"
        OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
        chk = loadFile(OUTPUT_EXCEL_PATH)
        if chk == "convOK":
            maxSeq = convertMaxSeq()
            OUTPUT_CSV_NAME = "output_" + str(maxSeq) + ".csv"
            OUTPUT_CSV_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_CSV_NAME
            excel_to_csv(OUTPUT_EXCEL_PATH, OUTPUT_CSV_PATH)
        
        seq_list, link_list, home_score_list, away_score_list, OUTPUT_EXCEL_PATH, updChkCount = loadFileAndGetLink(loop_bef_seq, updChkCount)
        if seq_list != []:
            chkStatistics(playwright, seq_list, link_list, home_score_list, away_score_list, OUTPUT_EXCEL_PATH)

        print(str(updChkCount) + "番目")

        loop_bef_seq += 1

        if updChkCount == 50:
            break

        if loop_bef_seq >= 1200:
            break
    







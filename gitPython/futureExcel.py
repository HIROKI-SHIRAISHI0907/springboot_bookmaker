from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os
import datetime
import sys
import smtplib
from email.mime.text import MIMEText
import json
import datetime
import inspect as isp
import csv
import time
from python_analytics import write_prediction_data
import pandas as pd
import re


def load_existing_match_keys():
    existing_keys = set()
    for seq in range(1, 3000):
        file_path = f"/Users/shiraishitoshio/bookmaker/future_{seq}.xlsx"
        if not os.path.exists(file_path):
            break
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                category = row[0]
                home_team = row[4]
                away_team = row[6]
                key = (category, home_team, away_team)
                existing_keys.add(key)
        except Exception as e:
            print(f"{file_path} 読み込み失敗: {e}")
    return existing_keys


def execute(playwright):
    #データ保持
    existing_keys = load_existing_match_keys()
    
    loop_bef_seq = -1
    # 対象時間を設定
    start_time = datetime.time(23, 30)
    end_time = datetime.time(23, 59)
    # 現在時刻
    now = datetime.datetime.today()
    now_info = datetime.datetime(now.year,now.month,now.day,now.hour,now.minute,now.second)

    if now_info.weekday() == 1 or now_info.weekday() == 2 or now_info.weekday() == 3 or now_info.weekday() == 4 or now_info.weekday() == 5:
        if now.hour == 12 or now.hour == 13 or now.hour == 14 or now.hour == 15 or now.hour == 16:
            print(f"{now_info}  データ取得対象外の時間です。")

    print(f"{now_info}  データ取得対象の時間です。")

    # Chromiumブラウザを起動
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    # Googleをひらく
    page.goto("https://www.google.com")

    # flashscoreに遷移
    page.goto("https://flashscore.co.jp/soccer/")
    # 開催予定メニュー遷移
    try:
        live_page = page.locator("div.filters__tab").nth(4)  # CSSセレクタを使用
        live_page.click()

        page.wait_for_load_state("networkidle")

        # 非表示要素を表示に変更
        alter = page.locator("button[data-testid='wcl-accordionButton']")
        alter_count = alter.count()
        for i in range(alter_count):
            alter_label = alter.nth(i).get_attribute("aria-label")
            if alter_label == None or alter_label == "リーグ全試合 表示":
                alter.nth(i).click()

        #開催予定試合取得
        live_game_list = []
        live_game = page.locator("a[href*='#/match-summary']")
        live_game_list = live_game.evaluate_all("live_game => live_game.map(e => e.href)")
    except Exception as e:
        print("ライブ試合リンクの取得失敗", e)
        live_game_list = []

    

    # cron実行通知用にホームチーム,アウェイチーム,カテゴリを入れたリストを保持
    cron_notice_data_list = []
    for j in range(len(live_game_list)):
        notice_data_list = []
        # 試合用文字列のみ抜く
        live = live_game_list[j]
        live_data = live.replace("https://www.flashscore.co.jp/match/", "")
        random_game_str = live_data.replace("/#/match-summary", "")

        # FLASHSCOREに遷移
        page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/?isDetailPopup=true#/match-summary/match-summary")
        time.sleep(3)

        # 試合の国及びカテゴリを取得
        game_info_sub_info_text = ""
        try:
            team_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-2)
            game_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-1)
            #game_info_sub_info = game_info.locator("span.tournamentHeader__country")
            team_info_text = team_info.text_content()
            game_info_text = game_info.text_content()
            print(team_info_text)
            print(game_info_text)
        except Exception as e:
            print("チーム名の取得失敗", e)
            continue

        if team_info_text == "" or game_info_text == "":
            print("チーム名の取得失敗: 空")
            continue

        game_info_sub_info_text = team_info_text + ": " + game_info_text
        print(game_info_sub_info_text)


        #試合のカテゴリが以下の場合のみ取得
        contains_list = []
        contains_list.append("ケニア: プレミアリーグ")
        contains_list.append("コロンビア: プリメーラ A")
        contains_list.append("タンザニア: プレミアリーグ")
        contains_list.append("イングランド: プレミアリーグ")
        contains_list.append("イングランド: EFL チャンピオンシップ")
        contains_list.append("イングランド: EFL リーグ 1")
        contains_list.append("エチオピア: プレミアリーグ")
        contains_list.append("コスタリカ: リーガ FPD")
        contains_list.append("ジャマイカ: プレミアリーグ")
        contains_list.append("スペイン: ラ・リーガ")
        contains_list.append("ブラジル: セリエ A ベターノ")
        contains_list.append("ブラジル: セリエ B")
        contains_list.append("ドイツ: ブンデスリーガ")
        contains_list.append("韓国: K リーグ 1")
        contains_list.append("中国: 中国スーパーリーグ")
        contains_list.append("日本: J1 リーグ")
        contains_list.append("日本: J2 リーグ")
        contains_list.append("日本: J3 リーグ")
        contains_list.append("インドネシア: リーガ 1")
        contains_list.append("オーストラリア: A リーグ・メン")
        contains_list.append("チュニジア: チュニジア･プロリーグ")
        contains_list.append("ウガンダ: プレミアリーグ")
        contains_list.append("メキシコ: リーガ MX")
        contains_list.append("フランス: リーグ・アン")
        contains_list.append("スコットランド: プレミアシップ")
        contains_list.append("オランダ: エールディビジ")
        contains_list.append("アルゼンチン: トルネオ・ベターノ")
        contains_list.append("イタリア: セリエ A")
        contains_list.append("イタリア: セリエ B")
        contains_list.append("ポルトガル: リーガ・ポルトガル")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("セルビア: スーペルリーガ")
        contains_list.append("日本: WEリーグ")
        contains_list.append("ボリビア: LFPB")
        contains_list.append("ブルガリア: パルヴァ・リーガ")
        contains_list.append("カメルーン: エリート 1")
        contains_list.append("ペルー: リーガ 1")
        contains_list.append("エストニア: メスタリリーガ")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("ウクライナ: プレミアリーグ")
        contains_list.append("ベルギー: ジュピラー･プロリーグ")
        contains_list.append("エクアドル: リーガ・プロ")
        contains_list.append("日本: YBC ルヴァンカップ")
        contains_list.append("日本: 天皇杯")
       

        under_list = []
        under_list.append("U17")
        under_list.append("U18")
        under_list.append("U19")
        under_list.append("U20")
        under_list.append("U21")
        under_list.append("U22")
        under_list.append("U23")
        under_list.append("U24")
        under_list.append("U25")

        gender_list = []
        gender_list.append("女子")

        exp_list = []
        exp_list.append("ポルトガル: リーガ・ポルトガル 2")
        exp_list.append("イングランド: プレミアリーグ 2")
        exp_list.append("イングランド: プレミアリーグ U18")

        if any(item in game_info_sub_info_text for item in contains_list):
            pass
        else:
            print("pass1")
            continue

        if any(item in game_info_sub_info_text for item in under_list):
            print("pass2")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in gender_list):
            print("pass3")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in exp_list):
            print("pass4")
            continue
        else:
            pass

        #試合予定時間取得
        game_future_time = page.locator("div.duelParticipant__startTime").text_content()
        print("試合予定時間")
        print(game_future_time)

        # ホームチーム,アウェーチーム名を取得
        try:
            home_team = page.locator("div.duelParticipant__home")
            home_team_info = home_team.locator("div.participant__participantNameWrapper")
            home_team_sub_info = home_team_info.locator("div.participant__participantName")
            home_team_name = home_team_sub_info.text_content()

            away_team = page.locator("div.duelParticipant__away")
            away_team_info = away_team.locator("div.participant__participantNameWrapper")
            away_team_sub_info = away_team_info.locator("div.participant__participantName")
            away_team_name = away_team_sub_info.text_content()
        except Exception as e:
            print("チーム名の取得失敗", e)
            continue

        match_key = (game_info_sub_info_text, home_team_name, away_team_name)
        if match_key in existing_keys:
            print("すでに同じ試合データが存在するためスキップ:", match_key)
            continue

        # FLASHSCORE(順位表)に遷移
        rank = True
        try:
            reponse = page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/overall")
        except Exception as e:
            print("順位表に遷移できません")
            rank = False

        home_team_max_getting_scorer = ""
        away_team_max_getting_scorer = ""

        # 順位表メニューをクリック
        if rank == True:
            try:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/")

                first_team_info = page.locator("div.ui-table__row.table__row--selected").nth(0)
                second_team_info = page.locator("div.ui-table__row.table__row--selected").nth(1)
                first_team_info_all = first_team_info.text_content()
                second_team_info_all = second_team_info.text_content()
                first_team_info_rank = first_team_info.locator("div.tableCellRank").text_content()
                second_team_info_rank = second_team_info.locator("div.tableCellRank").text_content()
            except Exception as e:
                print("順位表選択失敗", e)
                continue

            # ホームチーム,アウェーチームの順位を特定する
            home_rank = first_team_info_rank if home_team_name in first_team_info_all else second_team_info_rank
            away_rank = second_team_info_rank if home_rank == first_team_info_rank else first_team_info_rank

            #チーム得点,失点取得
            home_team_home_score = ""
            home_team_home_lost = ""
            away_team_home_score = ""
            away_team_home_lost = ""
            home_team_away_score = ""
            home_team_away_lost = ""
            away_team_away_score = ""
            away_team_away_lost = ""
            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/home")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub1 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub2 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub1.count())
                print(team_info_sub2.count())
                home_total_1 = ""
                home_total_2 = ""
                for ind in range(team_info_sub1.count()):
                    if team_info_sub1.nth(ind).text_content() == home_team_name:
                        home_total_1 = team_info_sub2.nth(ind).text_content()
                    elif team_info_sub1.nth(ind).text_content() == away_team_name:
                        home_total_2 = team_info_sub2.nth(ind).text_content()

                    if home_total_1 != "" and home_total_2 != "":
                        break

                if home_total_1 != "":
                    home_total_sub_1 = home_total_1.split(':')
                    home_team_home_score = home_total_sub_1[0]
                    home_team_home_lost = home_total_sub_1[1]

                if home_total_2 != "":
                    away_total_sub_1 = home_total_2.split(':')
                    away_team_home_score = away_total_sub_1[0]
                    away_team_home_lost = away_total_sub_1[1]

            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/away")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub3 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub4 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub3.count())
                print(team_info_sub4.count())
                away_total_1 = ""
                away_total_2 = ""
                for ind in range(team_info_sub3.count()):
                    if team_info_sub3.nth(ind).text_content() == home_team_name:
                        away_total_1 = team_info_sub4.nth(ind).text_content()
                    elif team_info_sub3.nth(ind).text_content() == away_team_name:
                        away_total_2 = team_info_sub4.nth(ind).text_content()

                    if away_total_1 != "" and away_total_2 != "":
                        break

                if away_total_1 != "":
                    away_total_sub_1 = away_total_1.split(':')
                    home_team_away_score = away_total_sub_1[0]
                    home_team_away_lost = away_total_sub_1[1]

                if away_total_2 != "":
                    away_total_sub_2 = away_total_2.split(':')
                    away_team_away_score = away_total_sub_2[0]
                    away_team_away_lost = away_total_sub_2[1]

            print(home_team_home_score)
            print(home_team_home_lost)
            print(away_team_home_score)
            print(away_team_home_lost)
            print(home_team_away_score)
            print(home_team_away_lost)
            print(away_team_away_score)
            print(away_team_away_lost)


            # チーム得点王取得
            print("チーム得点王取得")
            get_score_flg = True
            try:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/top_scorers/") # CSSセレクタを使用

                #ページが存在しない場合スキップ(処理追加)
                dummy_page = "https://www.flashscore.co.jp/match/" + random_game_str + "/#/match-summary/match-statistics"
                if page.url != dummy_page:
            
                    top_scorer_member = ""
                    top_scorer_team = ""
                    home_scorer_flg = False
                    away_scorer_flg = False
                    #得点王表が完全表示されるまで待機
                    time.sleep(2)
                    top_scorer_graph = page.locator("div[class*='topScorers__row--selected']")

                    for graph_ind in range(top_scorer_graph.count()):
                        sub_top_scorer_graph = top_scorer_graph.nth(graph_ind).locator("div.topScorers__participantCell").locator("a[class*='topScorersParticipant']")
                        tmp = 0
                        for sub_graph_ind in range(sub_top_scorer_graph.count()):
                            if tmp == 0:
                                top_scorer_member = sub_top_scorer_graph.nth(sub_graph_ind).text_content()
                            elif tmp == 1:
                                top_scorer_team = sub_top_scorer_graph.nth(sub_graph_ind).text_content()

                            if top_scorer_team == home_team_name and home_team_max_getting_scorer == "":
                                home_team_max_getting_scorer = top_scorer_member
                                home_scorer_flg = True
                            if top_scorer_team == away_team_name and away_team_max_getting_scorer == "":
                                away_team_max_getting_scorer = top_scorer_member
                                away_scorer_flg = True

                            tmp += 1
                                           
                        if home_scorer_flg == True and away_scorer_flg == True:
                            break

                    print("home_team_max_getting_scorer:" + home_team_max_getting_scorer) #取得済み
                    print("away_team_max_getting_scorer:" + away_team_max_getting_scorer) #取得済み

                else:
                    get_score_flg = False

            except Exception as e:
                print("得点王選択失敗", e)
                get_score_flg = False
        

        # エクセルに表記
        loop_bef_seq = makeExcelAndNotice(loop_bef_seq, game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name,
                                          away_team_name, random_game_str,
                home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
                                          home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost, 
            home_team_max_getting_scorer, away_team_max_getting_scorer)

    #noticeCronMail(cron_notice_data_list)

    page.close()
    context.close()
    browser.close()


def makeExcelAndNotice(loop_bef_seq, game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name, away_team_name, random_game_str,
                home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
                                          home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost, 
            home_team_max_getting_scorer, away_team_max_getting_scorer):


    # データ取得時間
    get_data_time = datetime.datetime.now()
    # 出力するエクセルを決定する(存在するfuture_通番.xlsxの中で最大の通番+1を持つエクセルを設定する)
    home_play_style = ""
    away_play_style = ""
    goal_prob = ""
    match_time = ""
    output_file_path = ""
    try:
        sheet = ""
        if loop_bef_seq < 0:
            for seq in range(1, 3000):
                OUTPUT_EXCEL_NAME = "future_" + str(seq) + ".xlsx"
                OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
                if os.path.exists(OUTPUT_EXCEL_PATH):
                    loop_bef_seq = seq
                else:
                    write_bk = Workbook()
                    sheet = write_bk.active
                    loop_bef_seq = seq
                    break
                    
        OUTPUT_EXCEL_NAME = "future_" + str(loop_bef_seq) + ".xlsx"
        OUTPUT_EXCEL_PATH = "/Users/shiraishitoshio/bookmaker/" + OUTPUT_EXCEL_NAME
        if os.path.exists(OUTPUT_EXCEL_PATH):
            write_bk = load_workbook(OUTPUT_EXCEL_PATH)
            sheet = write_bk.active
        else:
            write_bk = Workbook()
            sheet = write_bk.active

        #初期化
        goal_time = ""
        goal_member = ""
        judge = ""
        home_team_style = ""
        away_team_style = ""
        # データなしの場合ヘッダー追加
        header_data = [
            "試合国及びカテゴリ", "試合予定時間", "ホーム順位", "アウェー順位", "ホームチーム", "アウェーチーム",
            "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者",
            "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点", "ホームチームアウェー得点", "ホームチームアウェー失点",
            "アウェーチームアウェー得点", "アウェーチームアウェー失点", "試合リンク文字列", "データ取得時間"
        ]
        
        # ヘッダー未出力なら出力（A1がNone or 空なら）
        # ここで1行目の空白を削除（Workbook 作成時に勝手に入る）
        if sheet.max_row == 1:
            sheet.delete_rows(1)
            # 明示的にヘッダー行を書く
            sheet.append(header_data)

        

        row_data = [
            game_info_sub_info_text, game_future_time, home_rank, away_rank, home_team_name, away_team_name,
            home_team_max_getting_scorer, away_team_max_getting_scorer,
            home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score,
            away_team_away_lost, random_game_str, get_data_time
        ]
        sheet.append(row_data)
    except Exception as e:
        print("確率,時間,フォーメーション取得失敗", e)

    # 保存
    write_bk.save(OUTPUT_EXCEL_PATH)
    print("メイン書き込み完了")

    return loop_bef_seq

def noticeCronMail(cron_notice_data_list):
    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "cronの定期実行通知"
    body = "cronが実行されました。以下のライブデータに対してスクレイピングを行いました。"
    strings = ""
    for k in range(len(cron_notice_data_list)):
        stringstmp = ""
        data_list = cron_notice_data_list[k]
        chk = 0
        for l in range(len(data_list)):
            stringstmp += data_list[l]
            if chk == 0:
                stringstmp += ":"
            elif chk == 1:
                stringstmp += " vs "

            chk += 1
        if len(stringstmp) > 0:
            strings += "\n"
        strings += stringstmp

    body += "\n"
    body += "\n"
    body += strings

    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)

    print("cron実行通知を送信")


with sync_playwright() as playwright:
    execute(playwright)

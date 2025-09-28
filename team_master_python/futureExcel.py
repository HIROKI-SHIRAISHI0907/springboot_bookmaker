#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook, load_workbook
import os, sys, time, traceback, datetime

# ========= 設定 =========
BOOK_DIR = "/Users/shiraishitoshio/bookmaker"
OUTPUT_PREFIX = "future_"
MAX_SEQ = 3000
PROCESSED_FILE = os.path.join(BOOK_DIR, "processed_keys.tsv")  # 永続キー保存（TSV）

HEADLESS = True
USE_CHROME_CHANNEL = True

NAV_TIMEOUT = 30_000
CLICK_TIMEOUT = 10_000
WAIT_TIMEOUT = 30_000

CONTAINS_LIST = [
    "ケニア: プレミアリーグ","コロンビア: プリメーラ A","タンザニア: プレミアリーグ",
    "イングランド: プレミアリーグ","イングランド: EFL チャンピオンシップ","イングランド: EFL リーグ 1",
    "エチオピア: プレミアリーグ","コスタリカ: リーガ FPD","ジャマイカ: プレミアリーグ",
    "スペイン: ラ・リーガ","ブラジル: セリエ A ベターノ","ブラジル: セリエ B",
    "ドイツ: ブンデスリーガ","韓国: K リーグ 1","中国: 中国スーパーリーグ",
    "日本: J1 リーグ","日本: J2 リーグ","日本: J3 リーグ","インドネシア: リーガ 1",
    "オーストラリア: A リーグ・メン","チュニジア: チュニジア･プロリーグ","ウガンダ: プレミアリーグ",
    "メキシコ: リーガ MX","フランス: リーグ・アン","スコットランド: プレミアシップ",
    "オランダ: エールディビジ","アルゼンチン: トルネオ・ベターノ","イタリア: セリエ A",
    "イタリア: セリエ B","ポルトガル: リーガ・ポルトガル","トルコ: スュペル・リグ",
    "セルビア: スーペルリーガ","日本: WEリーグ","ボリビア: LFPB",
    "ブルガリア: パルヴァ・リーガ","カメルーン: エリート 1","ペルー: リーガ 1",
    "エストニア: メスタリリーガ","ウクライナ: プレミアリーグ","ベルギー: ジュピラー･プロリーグ",
    "エクアドル: リーガ・プロ","日本: YBC ルヴァンカップ","日本: 天皇杯",
]
UNDER_LIST = ["U17","U18","U19","U20","U21","U22","U23","U24","U25"]
GENDER_LIST = ["女子"]
EXCLUDE_LIST = ["ポルトガル: リーガ・ポルトガル 2","イングランド: プレミアリーグ 2","イングランド: プレミアリーグ U18"]

# ========= ユーティリティ =========
def log(msg: str):
    now = datetime.datetime.now().strftime("%F %T")
    print(f"{now}  {msg}", flush=True)

def retry(tries=3, delay=1.5, backoff=1.7, on_exception=(Exception,)):
    def deco(fn):
        def wrapper(*args, **kwargs):
            t, d = tries, delay
            while True:
                try:
                    return fn(*args, **kwargs)
                except on_exception as e:
                    t -= 1
                    if t <= 0:
                        raise
                    log(f"retry after error: {e.__class__.__name__}: {e} (remain={t})")
                    time.sleep(d); d *= backoff
        return wrapper
    return deco

def ensure_book_dir():
    os.makedirs(BOOK_DIR, exist_ok=True)

def load_existing_match_keys():
    ensure_book_dir()
    keys = set()
    for seq in range(1, MAX_SEQ):
        # ← 余計な '}' を削除
        path = os.path.join(BOOK_DIR, f"{OUTPUT_PREFIX}{seq}.xlsx")
        if not os.path.exists(path):
            break
        try:
            wb = load_workbook(path)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                category, home, away = row[0], row[4], row[6]
                keys.add((category, home, away))
        except Exception as e:
            log(f"{path} 読み込み失敗: {e}")
    log(f"既存キー: {len(keys)} 件")
    return keys

def get_or_create_output_sheet(loop_bef_seq: int):
    ensure_book_dir()
    sheet = None
    wb = None
    if loop_bef_seq < 0:
        for seq in range(1, MAX_SEQ):
            path = os.path.join(BOOK_DIR, f"{OUTPUT_PREFIX}{seq}.xlsx")
            if os.path.exists(path):
                loop_bef_seq = seq
            else:
                wb = Workbook(); sheet = wb.active; loop_bef_seq = seq; break
    name = f"{OUTPUT_PREFIX}{loop_bef_seq}.xlsx"
    path = os.path.join(BOOK_DIR, name)
    if os.path.exists(path) and wb is None:
        wb = load_workbook(path); sheet = wb.active
    elif wb is None:
        wb = Workbook(); sheet = wb.active

    header = [
        "試合国及びカテゴリ","試合予定時間","ホーム順位","アウェー順位","ホームチーム","アウェーチーム",
        "ホームチーム最大得点取得者","アウェーチーム最大得点取得者",
        "ホームチームホーム得点","ホームチームホーム失点","アウェーチームホーム得点","アウェーチームホーム失点",
        "ホームチームアウェー得点","ホームチームアウェー失点",
        "アウェーチームアウェー得点","アウェーチームアウェー失点","試合リンク文字列","データ取得時間"
    ]
    if sheet.max_row == 1:
        sheet.delete_rows(1); sheet.append(header)
    return wb, sheet, path, loop_bef_seq

def append_row(wb, sheet, path, row):
    sheet.append(row); wb.save(path)

def load_processed_keys() -> set[tuple[str, str, str]]:
    """前回までに処理した(カテゴリ, ホーム, アウェイ)キーをTSVから読み込み"""
    ensure_book_dir()
    s: set[tuple[str, str, str]] = set()
    if not os.path.exists(PROCESSED_FILE):
        return s
    try:
        with open(PROCESSED_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.rstrip("\n")
                if not line:
                    continue
                # カテゴリ \t ホーム \t アウェイ
                parts = line.split("\t")
                if len(parts) >= 3:
                    s.add((parts[0], parts[1], parts[2]))
    except Exception as e:
        log(f"processed_keys 読み込み失敗: {e}")
    log(f"永続キー: {len(s)} 件")
    return s

def append_processed_key(key: tuple[str, str, str]):
    """処理済みキーをTSVに1行追記（カテゴリ,ホーム,アウェイ）"""
    ensure_book_dir()
    try:
        with open(PROCESSED_FILE, "a", encoding="utf-8") as f:
            f.write("\t".join(key) + "\n")
    except Exception as e:
        log(f"processed_keys 追記失敗: {e}")


# ========= Playwright =========
def launch_browser(pw):
    args = [
        "--disable-blink-features=AutomationControlled",
        "--disable-gpu", "--no-sandbox",
    ]
    if USE_CHROME_CHANNEL:
        browser = pw.chromium.launch(channel="chrome", headless=HEADLESS, args=args)
    else:
        browser = pw.chromium.launch(headless=HEADLESS, args=args)
    ctx = browser.new_context(
        locale="ja-JP",
        user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/127.0.0.0 Safari/537.36")
    )
    ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    ctx.set_default_navigation_timeout(NAV_TIMEOUT)
    page = ctx.new_page()
    return browser, ctx, page

@retry(tries=3, delay=1.2, backoff=1.7)
def goto(page, url, wait="domcontentloaded", timeout=NAV_TIMEOUT):
    return page.goto(url, wait_until=wait, timeout=timeout)

def safe_wait(page, selector, timeout=WAIT_TIMEOUT, state="attached"):
    page.wait_for_selector(selector, timeout=timeout, state=state)

# ========= 新UI対応：リンク収集 & URL生成 =========
def expand_all_leagues(page):
    safe_wait(page, "div.filters__tab")
    page.locator("div.filters__tab").filter(has_text="開催予定").first.click(timeout=CLICK_TIMEOUT)
    safe_wait(page, "div.headerLeague__wrapper")
    # 必要なら展開（「表示」だけクリック）
    for _ in range(3):
        to_open = page.locator('button[data-testid="wcl-accordionButton"][aria-label*="表示"]')
        cnt = to_open.count()
        if cnt == 0: break
        for i in range(min(cnt, 80)):
            btn = to_open.nth(i)
            try:
                btn.scroll_into_view_if_needed(timeout=1500)
                btn.click(timeout=2500)
            except Exception:
                pass
        page.wait_for_timeout(400)

def get_live_game_links(page) -> list[str]:
    log("Flashscoreへ")
    goto(page, "https://www.flashscore.co.jp/soccer/")
    expand_all_leagues(page)

    links_set = set()
    for day in range(7):
        log(f"{day+1}/7日目のリンク収集")
        # 新UIのリンク
        try:
            safe_wait(page, "a.eventRowLink", state="attached")
            hrefs = page.locator("a.eventRowLink").evaluate_all("els => els.map(e => e.href)")
        except PWTimeout:
            hrefs = []

        # 念のため旧形式もフォールバック
        if not hrefs:
            try:
                safe_wait(page, "a[href*='/match/'][href*='/summary/']", state="attached")
                hrefs = page.locator("a[href*='/match/'][href*='/summary/']").evaluate_all("els => els.map(e => e.href)")
            except PWTimeout:
                hrefs = []

        for h in hrefs:
            if "/match/" in h and "/summary/" in h:
                links_set.add(h)

        # 翌日へ
        try:
            next_btn = page.locator("button[aria-label='翌日']")
            if next_btn.count() and next_btn.first.is_visible():
                next_btn.first.click(timeout=CLICK_TIMEOUT)
                page.wait_for_timeout(700)
            else:
                log("翌日ボタン非表示 → 打ち切り"); break
        except Exception as e:
            log(f"翌日移動できず: {e} → 打ち切り"); break

    return list(links_set)

def urls_from_summary(summary_url: str) -> dict:
    if "/summary/" not in summary_url:
        return {
            "summary": summary_url,
            "standings": summary_url,
            "table_home": summary_url,
            "table_away": summary_url,
            "top_scorers": summary_url,
        }
    base = summary_url.split("/summary/")[0]
    return {
        "summary": summary_url,
        "standings":      f"{base}/standings/",
        "table_home":     f"{base}/standings/table/home",
        "table_away":     f"{base}/standings/table/away",
        "top_scorers":    f"{base}/standings/top_scorers/",
    }

# ========= ロジック =========
def passes_filters(category_text: str) -> bool:
    if not any(x in category_text for x in CONTAINS_LIST): return False
    if any(x in category_text for x in UNDER_LIST): return False
    if any(x in category_text for x in GENDER_LIST): return False
    if any(x in category_text for x in EXCLUDE_LIST): return False
    return True

def scrape_match(page, urls: dict, existing_keys):
    # サマリー
    goto(page, urls["summary"])
    page.wait_for_timeout(900)

    # 国/カテゴリ（末尾2つ）
    team_info_text = ""; game_info_text = ""
    try:
        elems = page.locator("span[data-testid='wcl-scores-overline-03']")
        cnt = elems.count()
        if cnt >= 2:
            team_info_text = elems.nth(cnt-2).text_content()
            game_info_text = elems.nth(cnt-1).text_content()
    except Exception as e:
        log(f"カテゴリ取得失敗: {e}"); return None
    if not team_info_text or not game_info_text:
        log("カテゴリ取得失敗: 空"); return None

    category_text = f"{team_info_text}: {game_info_text}"
    log(f"カテゴリ: {category_text}")
    if not passes_filters(category_text):
        log("カテゴリ条件に合致せずスキップ"); return None

    # 予定時間
    try:
        game_future_time = page.locator("div.duelParticipant__startTime").first.text_content()
    except Exception:
        game_future_time = ""

    # チーム名
    try:
        home_team_name = page.locator("div.duelParticipant__home div.participant__participantName").first.text_content()
        away_team_name = page.locator("div.duelParticipant__away div.participant__participantName").first.text_content()
    except Exception as e:
        log(f"チーム名取得失敗: {e}"); return None

    match_key = (category_text, home_team_name, away_team_name)
    if match_key in existing_keys:
        log("既存データのためスキップ"); return None

    # 順位
    home_rank = away_rank = ""
    try:
        goto(page, urls["standings"]); page.wait_for_timeout(700)
        rows = page.locator("div.ui-table__row.table__row--selected")
        if rows.count() >= 2:
            first = rows.nth(0); second = rows.nth(1)
            first_all = first.text_content(); second_all = second.text_content()
            first_rank = first.locator("div.tableCellRank").text_content()
            second_rank = second.locator("div.tableCellRank").text_content()
            home_rank = first_rank if home_team_name in first_all else second_rank
            away_rank = second_rank if home_rank == first_rank else first_rank
    except Exception as e:
        log(f"順位取得失敗: {e}")

    # 得点/失点（home/away テーブル）
    home_team_home_score = home_team_home_lost = ""
    away_team_home_score = away_team_home_lost = ""
    home_team_away_score = home_team_away_lost = ""
    away_team_away_score = away_team_away_lost = ""

    def parse_table(which: str, url_key: str):
        nonlocal home_team_home_score, home_team_home_lost
        nonlocal away_team_home_score, away_team_home_lost
        nonlocal home_team_away_score, home_team_away_lost
        nonlocal away_team_away_score, away_team_away_lost

        goto(page, urls[url_key]); page.wait_for_timeout(900)
        rows = page.locator("div.table__row--selected")
        rc = rows.count()
        for i in range(rc):
            row = rows.nth(i)
            names = row.locator("a.tableCellParticipant__name")
            scores = row.locator("span.table__cell--score")
            n = names.count()
            for k in range(n):
                name = names.nth(k).text_content()
                if k < scores.count():
                    score_txt = scores.nth(k).text_content()
                else:
                    continue
                if ":" not in score_txt: continue
                a, b = score_txt.split(":")
                if name == home_team_name:
                    if which == "home":
                        home_team_home_score, home_team_home_lost = a, b
                    else:
                        home_team_away_score, home_team_away_lost = a, b
                elif name == away_team_name:
                    if which == "home":
                        away_team_home_score, away_team_home_lost = a, b
                    else:
                        away_team_away_score, away_team_away_lost = a, b

    try: parse_table("home", "table_home")
    except Exception as e: log(f"homeテーブル解析失敗: {e}")
    try: parse_table("away", "table_away")
    except Exception as e: log(f"awayテーブル解析失敗: {e}")

    # 得点王
    home_scorer = away_scorer = ""
    try:
        goto(page, urls["top_scorers"]); page.wait_for_timeout(900)
        if "/top_scorers" in page.url:
            rows = page.locator("div.topScorers__row--selected")
            rc = rows.count()
            for i in range(rc):
                row = rows.nth(i).locator("div.topScorers__participantCell").locator("a.topScorersParticipant")
                if row.count() >= 2:
                    member = row.nth(0).text_content(); team = row.nth(1).text_content()
                    if team == home_team_name and not home_scorer: home_scorer = member
                    if team == away_team_name and not away_scorer: away_scorer = member
                    if home_scorer and away_scorer: break
    except Exception as e:
        log(f"得点王取得失敗: {e}")

    row = [
        category_text, game_future_time, home_rank, away_rank,
        home_team_name, away_team_name, home_scorer, away_scorer,
        home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
        home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
        urls["summary"], datetime.datetime.now()
    ]
    return row

# ========= メイン =========
def main():
    now = datetime.datetime.now().replace(microsecond=0)
    if now.weekday() in (1,2,3,4,5) and now.hour in (12,13,14,15,16):
        log(f"{now}  データ取得対象外の時間です。")
    log(f"{now}  データ取得対象の時間です。")

    existing_keys = load_existing_match_keys()
    # ← ここを追加
    persisted = load_processed_keys()
    existing_keys |= persisted   # マージ

    loop_bef_seq = -1

    with sync_playwright() as pw:
        browser, ctx, page = launch_browser(pw)
        try:
            try:
                links = get_live_game_links(page)
            except Exception as e:
                log(f"ライブ試合リンクの取得失敗: {e}\n{traceback.format_exc()}")
                links = []

            log(f"リンク件数: {len(links)}")

            wb, sheet, path, loop_bef_seq = get_or_create_output_sheet(loop_bef_seq)

            for j, summary_url in enumerate(links, 1):
                try:
                    log(f"[{j}/{len(links)}] 処理: {summary_url}")
                    urls = urls_from_summary(summary_url)
                    row = scrape_match(page, urls, existing_keys)
                    if row is None: continue
                    append_row(wb, sheet, path, row)
                    match_key = (row[0], row[4], row[5])  # (カテゴリ, ホーム, アウェイ)
                    existing_keys.add(match_key)
                    log("書き込み完了")
                except Exception as e:
                    log(f"試合処理中に失敗: {e}\n{traceback.format_exc()}")
        finally:
            try: page.close(); ctx.close(); browser.close()
            except Exception: pass

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"致命的エラー: {e}\n{traceback.format_exc()}"); sys.exit(1)

import os
import re
import asyncio
import csv
import pandas as pd
import json
from pathlib import Path
from typing import List, Tuple, Set, Any
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PwTimeoutError


BASE_OUTPUT_URL = "/Users/shiraishitoshio/bookmaker"

# ✅ パス修正（スラッシュ忘れ防止）
TEAMS_BY_LEAGUE_DIR = os.path.join(BASE_OUTPUT_URL, "teams_by_league")
SEASON_XLSX = str(Path(TEAMS_BY_LEAGUE_DIR) / "season_data.xlsx")

BASE_URL = "https://www.flashscore.co.jp"
SEASON_YEAR_HEADER = "シーズン年"
ROUND_HEADER = "ラウンド数"
ICON_HEADER = "リーグアイコン"
B001_JSON_PATH = str(Path(BASE_OUTPUT_URL) / "json/b001/b001_country_league.json")

CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]
UNDER_LIST  = ["U17", "U18", "U19", "U20", "U21", "U22", "U23", "U24", "U25"]
GENDER_LIST = ["女子"]
EXP_LIST    = ["ポルトガル: リーガ・ポルトガル 2", "イングランド: プレミアリーグ 2", "イングランド: プレミアリーグ U18"]

# --- Excel ユーティリティ ---
def open_or_init_season_book(file_path: str = SEASON_XLSX):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "season"
        ws.append(["国", "リーグ", SEASON_YEAR_HEADER, "シーズン開始", "シーズン終了", ROUND_HEADER, "パス", ICON_HEADER])

        wb.save(file_path)
        return wb, ws

    wb = load_workbook(file_path)
    ws = wb["season"] if "season" in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    expected = ["国", "リーグ", SEASON_YEAR_HEADER, "シーズン開始", "シーズン終了", ROUND_HEADER, "パス", ICON_HEADER]
    if not header or header[:7] != expected:
        ws.delete_rows(1, ws.max_row if ws.max_row else 1)
        ws.append(["国", "リーグ", SEASON_YEAR_HEADER, "シーズン開始", "シーズン終了", ROUND_HEADER, "パス", ICON_HEADER])
        wb.save(file_path)
        return wb, ws

    # シーズン年列が無い場合
    if SEASON_YEAR_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=SEASON_YEAR_HEADER)
        header.append(SEASON_YEAR_HEADER)
        wb.save(file_path)

    if ROUND_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=ROUND_HEADER)
        header.append(ROUND_HEADER)
        wb.save(file_path)

    # アイコン列が無い場合
    if ICON_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=ICON_HEADER)
        wb.save(file_path)

    return wb, ws

def get_colmap(ws):
    header = [c.value for c in ws[1]]
    return {name: idx + 1 for idx, name in enumerate(header)}


def existing_paths(ws) -> Set[str]:
    """ season_data.xlsx に既にある 'パス' をセットで返す（重複追記防止） """
    col = get_colmap(ws)
    pcol = col["パス"]
    out = set()
    for r in range(2, (ws.max_row or 1) + 1):
        v = (ws.cell(row=r, column=pcol).value or "").strip()
        if v:
            out.add(v)
    return out

def extract_countries_and_leagues(json_path: str) -> tuple[list[str], list[tuple[str, str]]]:
    p = Path(json_path).expanduser()
    if not p.exists():
        return [], []

    data = json.loads(p.read_text(encoding="utf-8"))
    pairs: list[tuple[str, str]] = []

    if isinstance(data, dict):
        for c, v in data.items():
            country = str(c).strip()
            if not country:
                continue
            if isinstance(v, list):
                for lv in v:
                    league = str(lv).strip()
                    if league:
                        pairs.append((country, league))
            else:
                league = str(v).strip()
                if league:
                    pairs.append((country, league))

    countries = sorted({c for c, _ in pairs})
    return countries, pairs

# --- Playwright helpers ---
async def kill_onetrust(page):
    # 出る時だけ消す
    try:
        btn = page.locator("#onetrust-accept-btn-handler").first
        if await btn.count() and await btn.is_visible():
            await btn.click(timeout=1500, force=True)
    except:
        pass
    try:
        await page.evaluate("""
        () => {
          const ids = ["onetrust-consent-sdk", "onetrust-banner-sdk"];
          ids.forEach(id => document.getElementById(id)?.remove());
          document.querySelectorAll(".ot-sdk-container, .ot-sdk-row, .otOverlay, .ot-pc-footer, .ot-pc-header")
            .forEach(el => el.remove());
          const overlay = document.querySelector("#onetrust-consent-sdk, .otOverlay");
          if (overlay) overlay.style.pointerEvents = "none";
        }
        """)
    except:
        pass


async def goto_home(page):
    await page.goto(BASE_URL, wait_until="domcontentloaded")
    await kill_onetrust(page)


def normalize_country_name(s: str) -> str:
    s = (s or "").strip()
    s = s.replace("　", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_comp(s: str) -> str:
    # 比較用（空白除去）
    return norm(s).replace(" ", "")

def is_excluded(category: str) -> bool:
    cat = norm(category)
    # UNDER/GENDER は "含まれていたら除外"
    for w in UNDER_LIST + GENDER_LIST:
        if w and w in cat:
            return True
    # EXP_LIST は「文字列に含まれていたら除外」＝部分一致で除外
    for w in EXP_LIST:
        if w and w in cat:
            return True
    return False

def build_allowed_set_from_contains() -> set[str]:
    # CONTAINS_LIST は完全一致想定だが、表記ゆれ対策で空白無し比較も作る
    return {norm_comp(x) for x in CONTAINS_LIST}

def build_allowed_set_from_json(pairs: list[tuple[str, str]]) -> set[str]:
    # "国: リーグ" 形式で持つ
    return {norm_comp(f"{c}: {l}") for c, l in pairs}

def is_allowed_category(category: str, allowed_set: set[str]) -> bool:
    if is_excluded(category):
        return False
    return norm_comp(category) in allowed_set


async def open_country_via_leftmenu(page, country: str, country_map: dict) -> bool:
    """
    country_map: {表示名: href}  例 {"イングランド": "/soccer/england/"}
    """
    target = normalize_country_name(country)

    # 1) 完全一致
    if target in country_map:
        href = country_map[target]
        await page.goto(BASE_URL + href, wait_until="domcontentloaded")
        await kill_onetrust(page)
        return True

    # 2) 部分一致（例: 表示が「ｱﾝﾃｨｸﾞｱ･ﾊﾞｰﾌﾞｰﾀﾞ」みたいなケース）
    #    JSON側が正式表記、表示名が半角カナの可能性があるので、まずは「含む」だけやる
    for name, href in country_map.items():
        if target in normalize_country_name(name) or normalize_country_name(name) in target:
            await page.goto(BASE_URL + href, wait_until="domcontentloaded")
            await kill_onetrust(page)
            return True

    print(f"  ❌ country href not found in leftmenu: {country}")
    return False

async def expand_leftmenu_countries(page, max_clicks: int = 200) -> int:
    """
    左メニュー「国」の 'もっと表示' を、消える/増えなくなるまでクリックする。
    戻り値: 実クリック回数
    """
    clicks = 0
    menu = page.locator("#category-left-menu")
    await menu.wait_for(state="visible", timeout=15000)

    # 「国」セクションのコンテナ（この中に a.lmc__item が並ぶ）
    # lmc__menu 全体を対象にしても良いが、過検出を避けるため menu 内に限定
    while clicks < max_clicks:
        more = menu.locator("span.lmc__itemMore").first

        # more が無い/見えない → 押し切り完了
        if await more.count() == 0 or not await more.is_visible():
            break

        before = await menu.locator("a.lmc__element.lmc__item[href^='/soccer/']").count()

        await more.scroll_into_view_if_needed()
        try:
            await more.click(timeout=5000)
        except:
            # 広告overlay等で失敗することがあるので少し待って再評価
            await page.wait_for_timeout(300)
            break

        clicks += 1

        # クリック後：国リンク数が増える or moreが消えるまで待つ
        try:
            await page.wait_for_function(
                """
                ({rootSel, aSel, before}) => {
                  const root = document.querySelector(rootSel);
                  if (!root) return true;
                  const now = root.querySelectorAll(aSel).length;
                  const more = root.querySelector("span.lmc__itemMore");
                  const moreVisible = more && more.offsetParent !== null;
                  return (now > before) || !moreVisible;
                }
                """,
                arg={
                    "rootSel": "#category-left-menu",
                    "aSel": "a.lmc__element.lmc__item[href^='/soccer/']",
                    "before": before,
                },
                timeout=15000
            )
        except PwTimeoutError:
            # 増えない/重い場合は終了
            break

    return clicks

def pick_first_league_fallback(country: str, leagues_raw: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """
    CONTAINS で1件も取れない国の救済:
    leagues_raw から最初の「まともな」リーグを1件だけ返す
    """
    for league_name, path in leagues_raw:
        # 既存の除外条件（Uxx/女子/EXPなど）
        if is_excluded(f"{country}: {league_name}"):
            continue

        # 念のため、見出し/ナビっぽいのを弾く（国ページ抽出が緩い場合の保険）
        if re.search(r"(概要|順位表|結果|日程|ニュース|統計|選手|チーム)", league_name or ""):
            continue

        if path and league_name:
            return [(league_name, path)]

    return []

async def get_all_country_links_from_leftmenu(page) -> List[Tuple[str, str]]:
    """
    左メニュー「国」一覧から (表示名, href) を全取得する。
    事前に expand_leftmenu_countries() で押し切る前提でも、
    この関数内で押し切ってもOK。
    """
    await goto_home(page)          # あなたの既存
    await kill_onetrust(page)      # あなたの既存

    clicks = await expand_leftmenu_countries(page)
    print(f"[LEFTMENU] more clicked: {clicks}")

    menu = page.locator("#category-left-menu")
    items = menu.locator("a.lmc__element.lmc__item[href^='/soccer/']")
    n = await items.count()

    out: List[Tuple[str, str]] = []
    seen = set()

    for i in range(n):
        a = items.nth(i)
        href = (await a.get_attribute("href")) or ""
        name = (await a.locator("span.lmc__elementName").first.text_content()) or ""
        name = re.sub(r"\s+", " ", name).strip()

        # 念のため /soccer/<slug>/ だけに限定
        if not re.match(r"^/soccer/[^/]+/?$", href):
            continue
        if not name:
            continue

        key = (name, href)
        if key in seen:
            continue
        seen.add(key)
        out.append((name, href))

    print(f"[LEFTMENU] countries: {len(out)}")
    return out

async def scrape_league_links_on_country_page(page) -> List[Tuple[str, str]]:
    """
    国ページからリーグ（大会）リンクを抽出して (league_name, path) を返す。
    """
    # まずある程度描画待ち
    try:
        await page.wait_for_timeout(800)
    except:
        pass

    anchors = page.locator("a[href]")
    n = await anchors.count()

    leagues: List[Tuple[str, str]] = []
    seen = set()

    for i in range(n):
        a = anchors.nth(i)
        try:
            href = (await a.get_attribute("href")) or ""
            txt = (await a.text_content()) or ""
            txt = re.sub(r"\s+", " ", txt).strip()

            if not href or not txt:
                continue

            # パスに統一
            path = href.replace(BASE_URL, "")
            if not path.startswith("/"):
                continue

            # リーグページっぽいものだけ
            # /soccer/<country>/<league>/ を基本として拾う
            if not re.match(r"^/soccer/[^/]+/[^/]+/?$", path):
                continue

            # 明らかなナビゲーションは除外
            if re.search(r"(概要|順位表|結果|日程|ニュース|統計|選手|チーム)", txt):
                continue

            key = (txt, path)
            if key in seen:
                continue
            seen.add(key)
            leagues.append((txt, path))
        except:
            pass

    return leagues

def norm_name(s: str) -> str:
    """
    比較用に正規化（全角/半角や空白の揺れを吸収）
    """
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")          # 全角スペース→半角
    s = re.sub(r"\s+", " ", s).strip()    # 連続空白を1個に
    return s

async def fetch_timeline_and_icon_from_league(page, league_path: str) -> tuple[str, str, str, str, str, str]:
    """
    # return: (start_ddmm, end_ddmm, icon_url, season_year, league_name_page)
    timeline が取れない場合は fixtures から推定する
    """
    league_path = (league_path or "").strip()
    if not league_path:
        return "", "", "", "", "", ""

    # --- 初期化（UnboundLocalError 防止） ---
    start_ddmm, end_ddmm, icon_url, season_year = "", "", "", ""
    round_num = ""

    # URL組み立て
    if league_path.startswith("http"):
        url = league_path
    else:
        if not league_path.startswith("/"):
            league_path = "/" + league_path
        url = BASE_URL + league_path

    print(f"[LEAGUE GOTO] {url}")
    await page.goto(url, wait_until="domcontentloaded")
    await kill_onetrust(page)

    # standings 等の hash ページへ飛ばされたらトップへ戻す
    if "#" in page.url:
        await page.goto(url.split("#", 1)[0], wait_until="domcontentloaded")
        await kill_onetrust(page)

    try:
        await page.wait_for_load_state("networkidle", timeout=12000)
    except:
        pass

    print(f"[LEAGUE AT] {page.url}")

    # --- season year ---
    try:
        info = page.locator("div.heading__info").first
        if await info.count():
            season_year = (await info.text_content() or "").strip()
    except:
        pass

    league_name_page = await fetch_league_name_from_page(page)

    # --- timeline（取れたら優先） ---
    try:
        await page.wait_for_selector("#timeline span", timeout=8000)
        spans = page.locator("#timeline span")
        cnt = await spans.count()

        if cnt >= 2:
            start_ddmm = (await spans.nth(0).text_content() or "").strip()
            end_ddmm   = (await spans.nth(1).text_content() or "").strip()
        else:
            s = page.locator("#timeline span[class*='wcl-start']").first
            e = page.locator("#timeline span[class*='wcl-end']").first
            if await s.count():
                start_ddmm = (await s.text_content() or "").strip()
            if await e.count():
                end_ddmm = (await e.text_content() or "").strip()
    except:
        pass

    # --- icon ---
    try:
        img = page.locator("img.heading__logo.heading__logo--1").first
        if await img.count():
            src = await img.get_attribute("src")
            if src:
                icon_url = src.strip()

        if not icon_url:
            img2 = page.locator("img.heading__logo").first
            if await img2.count():
                src2 = await img2.get_attribute("src")
                if src2:
                    icon_url = src2.strip()
    except:
        pass

    # 正規化
    start_ddmm = (start_ddmm or "").replace(" ", "")
    end_ddmm   = (end_ddmm or "").replace(" ", "")

    # ✅ timeline が無い/空なら fixtures へフォールバック
    if not start_ddmm or not end_ddmm:
        fs, fe, round_num = await fetch_start_end_from_fixtures(page, league_path)
        if not start_ddmm:
            start_ddmm = fs
        if not end_ddmm:
            end_ddmm = fe

        # fixtures 側で icon を取り直す（保険）
        if not icon_url:
            try:
                img = page.locator("img.heading__logo.heading__logo--1").first
                if await img.count():
                    src = await img.get_attribute("src")
                    if src:
                        icon_url = src.strip()
            except:
                pass

        # fixtures 側で season_year を取り直す（保険）
        if not season_year:
            try:
                info = page.locator("div.heading__info").first
                if await info.count():
                    season_year = (await info.text_content() or "").strip()
            except:
                pass

        if not round_num:
            _, _, round_num2 = await fetch_start_end_from_fixtures(page, league_path)
            if round_num2:
                round_num = round_num2

    print(f"[GOT] season_year={season_year} start={start_ddmm} end={end_ddmm} icon={'Y' if icon_url else 'N'}")
    return start_ddmm, end_ddmm, icon_url, season_year, league_name_page, round_num

async def click_show_more_fixtures_until_end(
    page,
    button_text: str = "もっと試合を表示する",
    max_clicks: int = 120,
    wait_timeout: int = 15000,
) -> int:
    clicks = 0
    while clicks < max_clicks:
        link = page.locator("a[data-testid='wcl-buttonLink']").filter(has_text=button_text).first

        if await link.count() == 0 or not await link.is_visible():
            break

        before = await page.locator(".event__match").count()

        await link.scroll_into_view_if_needed()
        try:
            await link.click(timeout=5000)
        except:
            break

        clicks += 1

        try:
            await page.wait_for_function(
                """
                ({before}) => {
                  const now = document.querySelectorAll(".event__match").length;
                  const btn = document.querySelector("a[data-testid='wcl-buttonLink']");
                  const visible = btn && btn.offsetParent !== null;
                  return (now > before) || !visible;
                }
                """,
                arg={"before": before},
                timeout=wait_timeout,
            )
        except PwTimeoutError:
            break

    return clicks

async def fetch_league_name_from_page(page) -> str:
    """
    リーグ名をヘッダーから取得。
    優先: div.heading__title div.heading__name  (例: "J2 リーグ")
    fallback: span.headerLeague__title-text     (例: "J2 リーグ - 昇格戦")
    """
    # ① ここが本命（あなたが欲しいやつ）
    try:
        loc = page.locator("div.heading__title div.heading__name").first
        if await loc.count():
            txt = (await loc.text_content() or "").strip()
            if txt:
                return txt
    except:
        pass

    # ② 予備（従来のやつ）
    try:
        loc = page.locator("span.headerLeague__title-text").first
        if await loc.count():
            txt = (await loc.text_content() or "").strip()
            if txt:
                return txt
    except:
        pass

    return ""

async def fetch_start_end_from_fixtures(page, league_path: str) -> tuple[str, str, str]:
    league_path = (league_path or "").strip()
    if not league_path:
        return "", "", ""

    if not league_path.endswith("/"):
        league_path += "/"

    url = BASE_URL + league_path + "fixtures/"
    print(f"[FIXTURES GOTO] {url}")
    await page.goto(url, wait_until="domcontentloaded")
    await kill_onetrust(page)

    try:
        await page.wait_for_selector(".event__match", timeout=12000)
    except PwTimeoutError:
        print("[FIXTURES] .event__match not found")
        return "", "", ""

    try:
        await page.wait_for_load_state("networkidle", timeout=12000)
    except:
        pass

    # ✅ もっと試合を表示する を押し切る
    clicks = await click_show_more_fixtures_until_end(page)
    print(f"[FIXTURES] show more clicked: {clicks}")

    # =========================
    # 開始日・終了日
    # =========================
    pat_date = re.compile(r"(\d{1,2})\.(\d{1,2})\.")
    times = await page.locator(".event__match .event__time").all_text_contents()

    ddmms: list[str] = []
    for t in times:
        m = pat_date.search(t or "")
        if not m:
            continue
        dd = int(m.group(1))
        mm = int(m.group(2))
        ddmms.append(f"{dd:02d}.{mm:02d}.")

    start_ddmm = ddmms[0] if ddmms else ""
    end_ddmm   = ddmms[-1] if ddmms else ""

    # =========================
    # ✅ ラウンド数（最大ラウンド）
    # =========================
    round_num = ""

    try:
        rounds = await page.locator("div.event__round.event__round--static").all_text_contents()
        nums = []
        for txt in rounds:
            m = re.search(r"ラウンド\s*(\d+)", txt or "")
            if m:
                nums.append(int(m.group(1)))

        if nums:
            round_num = str(max(nums))  # 最後 or 最大でOK
    except:
        pass

    print(f"[FIXTURES GOT] start={start_ddmm} end={end_ddmm} round={round_num}")
    return start_ddmm, end_ddmm, round_num

# --- メイン ---
async def main():
    wb, ws = open_or_init_season_book(SEASON_XLSX)
    col = get_colmap(ws)

    print(f"[SEASON_XLSX] {SEASON_XLSX}")

    # --- 1) JSON を読む（あれば country+league 指定、なければ CONTAINS_LIST 運用） ---
    countries, country_league_pairs = extract_countries_and_leagues(B001_JSON_PATH)

    json_has_pairs = bool(country_league_pairs)
    if json_has_pairs:
        allowed_set = build_allowed_set_from_json(country_league_pairs)  # "国: リーグ" の正規化セット
        allowed_leagues_by_country: dict[str, set[str]] = {}
        for c, l in country_league_pairs:
            c2 = norm_name(c)
            l2 = norm_name(l)
            if c2 and l2:
                allowed_leagues_by_country.setdefault(c2, set()).add(l2)
        allowed_countries = set(allowed_leagues_by_country.keys())
        print(f"[FILTER] JSONあり: pairs {len(allowed_set)} / countries {len(allowed_countries)}")
        print("[countries]", sorted(allowed_countries))
        print("[pairs]", sorted(country_league_pairs))
    else:
        allowed_set = build_allowed_set_from_contains()
        allowed_countries = {norm(x.split(":", 1)[0]) for x in CONTAINS_LIST if ":" in x}
        allowed_leagues_by_country = {}
        print(f"[FILTER] JSONなし: contains {len(allowed_set)} / countries {len(allowed_countries)}")
        print("[countries]", sorted(allowed_countries))

    existed = existing_paths(ws)
    appended = 0
    updated_timeline = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(locale="ja-JP", timezone_id="Asia/Tokyo")
        page = await context.new_page()

        # --- 2) 左メニュー国一覧を全部展開→ (表示名, href) を取得 ---
        all_countries = await get_all_country_links_from_leftmenu(page)
        country_map = {name: href for name, href in all_countries}

        # --- 3) 国→リーグ抽出→許可リーグだけ Excel 追記（path列を作る） ---
        for country in sorted(allowed_countries):
            print(f"\n=== {country} ===")

            ok = await open_country_via_leftmenu(page, country, country_map)
            if not ok:
                continue

            await kill_onetrust(page)

            leagues_raw = await scrape_league_links_on_country_page(page)

            # category で allowed_set を通す（JSON/contains共通、除外語もここで弾く）
            filtered_1: list[tuple[str, str]] = []
            for league_name, path in leagues_raw:
                category = f"{country}: {league_name}"
                if is_allowed_category(category, allowed_set):
                    filtered_1.append((league_name, path))

            # ✅ フォールバック：CONTAINSで0件なら先頭リーグ1件だけ採用
            if (not json_has_pairs) and (len(filtered_1) == 0):
                fb = pick_first_league_fallback(country, leagues_raw)
                if fb:
                    filtered_1 = fb
                    print(f"  [FALLBACK] no match in CONTAINS → pick first league: {filtered_1[0][0]}")
                else:
                    print(f"  [FALLBACK] no usable league found")
        
            print(f"  leagues matched(category): {len(filtered_1)} / scraped: {len(leagues_raw)}")

            # JSONありなら、その国で指定されたリーグ名だけにさらに絞る
            if json_has_pairs:
                allowed_leagues = allowed_leagues_by_country.get(norm_name(country), set())

                def is_allowed_league_name(league_name: str) -> bool:
                    ln = norm_name(league_name)
                    if ln in allowed_leagues:
                        return True
                    ln_comp = ln.replace(" ", "")
                    for a in allowed_leagues:
                        if ln_comp == a.replace(" ", ""):
                            return True
                    return False

                leagues_final = [(name, path) for (name, path) in filtered_1 if is_allowed_league_name(name)]
                print(f"  leagues matched(JSON): {len(leagues_final)} / allowed={sorted(list(allowed_leagues))}")
            else:
                leagues_final = filtered_1

            # --- 3-1) 追記（重複は弾く） ---
            for league_name, path in leagues_final:
                if path in existed:
                    continue
                ws.append([country, league_name, "", "", "", "", path, ""])
                existed.add(path)
                appended += 1

        # --- 4) いまExcelにある path を順に見て、timeline(icon) を埋める ---
        # ※ ここは「今回追記した分だけ」更新したいなら appended 行だけ追う実装にしてもOK
        max_row = ws.max_row or 1
        for r in range(2, max_row + 1):
            path = (ws.cell(row=r, column=col["パス"]).value or "").strip()
            if not path:
                continue

            # 既に埋まってるならスキップ（必要なら条件を変えてOK）
            cur_year  = ws.cell(row=r, column=col[SEASON_YEAR_HEADER]).value
            cur_start = ws.cell(row=r, column=col["シーズン開始"]).value
            cur_end   = ws.cell(row=r, column=col["シーズン終了"]).value
            cur_icon  = ws.cell(row=r, column=col[ICON_HEADER]).value
            # どれか1つでも空なら取りに行く
            if (cur_year not in (None, "")) and (cur_start not in (None, "")) and (cur_end not in (None, "")) and (cur_icon not in (None, "")):
                continue

            start_ddmm, end_ddmm, icon_url, season_year, round_num = "", "", "", "", ""

            start_ddmm, end_ddmm, icon_url, season_year, league_name_page, round_num = await fetch_timeline_and_icon_from_league(page, path)
            # Excelのリーグ列をページのリーグ名で更新（取れた場合のみ）
            if league_name_page:
                ws.cell(row=r, column=col["リーグ"], value=league_name_page)

            if season_year and (cur_year in (None, "")):
                ws.cell(row=r, column=col[SEASON_YEAR_HEADER], value=season_year)
            if start_ddmm and (cur_start in (None, "")):
                ws.cell(row=r, column=col["シーズン開始"], value=start_ddmm)
            if end_ddmm and (cur_end in (None, "")):
                ws.cell(row=r, column=col["シーズン終了"], value=end_ddmm)
            cur_round = ws.cell(row=r, column=col[ROUND_HEADER]).value
            if round_num and (cur_round in (None, "")):
                ws.cell(row=r, column=col[ROUND_HEADER], value=round_num)
            if icon_url and (cur_icon in (None, "")):
                ws.cell(row=r, column=col[ICON_HEADER], value=icon_url)

            if start_ddmm or end_ddmm or icon_url or season_year:
                updated_timeline += 1

            # 連打しすぎない
            await page.wait_for_timeout(120)

        await context.close()
        await browser.close()

    wb.save(SEASON_XLSX)
    print(f"\n✅ 追記数: {appended}")
    print(f"✅ timeline/icon 更新行数: {updated_timeline}")

    # CSV出力
    csv_path = os.path.splitext(SEASON_XLSX)[0] + ".csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["国", "リーグ", SEASON_YEAR_HEADER, "シーズン開始", "シーズン終了", ROUND_HEADER, "パス", ICON_HEADER])
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            writer.writerow(list(row_vals))
    print(f"CSV出力: {csv_path}")

    excel_to_csv_and_delete(SEASON_XLSX, csv_path)

def excel_to_csv_and_delete(excel_file: str, csv_file: str) -> bool:
    try:
        os.remove(excel_file)
        print(f"[CONVERT] {excel_file} -> {csv_file}（xlsx削除）")
        return True
    except Exception as e:
        print(f"[ERROR] Excel->CSV失敗: {excel_file} ({e})")
        return False

if __name__ == "__main__":
    asyncio.run(main())
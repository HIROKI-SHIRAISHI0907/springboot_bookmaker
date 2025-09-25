# -*- coding: utf-8 -*-
"""
TeamMemberExcel_writer_queue_seq_players.py
- teams_by_league/*.xlsx を全読み
- 4チーム並列（TEAM_CONCURRENCY=4）
- 各チーム内の選手詳細取得は 1 スレッド（1 Page を使い回して逐次遷移）
- page.goto の実同時数と発火間隔を制御（GLOBAL_NAV_CONCURRENCY / NAV_GAP_MS）
- 不要リソース/広告/同意ドメインは遮断
- Excel 書込は専用 Writer タスクに集約（チーム単位で一括投入）
"""

import os
import re
import glob
import asyncio
import datetime
from typing import List, Tuple, Dict, Set, Optional

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

# =========================
# 設定
# =========================
BASE_DIR = "/Users/shiraishitoshio/bookmaker"
TEAMS_BY_LEAGUE_DIR = os.path.join(BASE_DIR, "teams_by_league")

EXCEL_BASE_PREFIX = "team_member_"
EXCEL_MAX_RECORDS = 50

# タイムアウト(ms)
OP_TIMEOUT = 5000
NAV_TIMEOUT = 12000
SEL_TIMEOUT = 20000

# 並列度（要件）
TEAM_CONCURRENCY       = 4      # 同時チーム（固定4）
GLOBAL_NAV_CONCURRENCY = 4      # 全体同時 goto 上限（=チーム並列と同じ）
NAV_GAP_MS             = 150    # goto の最小発火間隔

# Writer の保存ポリシー
SAVE_INTERVAL_SEC      = 2.0    # これ秒ごとに未保存があれば保存
SAVE_EVERY_N_ROWS      = 80     # これ行以上たまったら即保存
LOG_EVERY_ENQUEUE      = True   # チームごとに ENQUEUE ログを出す

# ブロックする外部ドメイン
BLOCKED_HOST_KEYWORDS = [
    "googletagmanager", "google-analytics", "doubleclick", "googlesyndication",
    "scorecardresearch", "criteo", "adsystem", "mathtag", "quantserve",
    "taboola", "facebook", "twitter", "hotjar", "onetrust", "cookiebot",
    "trustarc", "ioam.de", "amazon-adsystem", "adservice"
]

# =========================
# ユーティリティ
# =========================
def ensure_dirs():
    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(TEAMS_BY_LEAGUE_DIR, exist_ok=True)

def load_existing_player_keys_from_excels() -> Set[Tuple[str, str, str, str]]:
    """既存の team_member_*.xlsx から (国,リーグ,チーム,選手名) を集合化（重複除去用）"""
    s: Set[Tuple[str, str, str, str]] = set()
    for path in sorted(glob.glob(os.path.join(BASE_DIR, f"{EXCEL_BASE_PREFIX}*.xlsx"))):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                s.add((str(row[0]), str(row[1]), str(row[2]), str(row[3])))
        except Exception as e:
            print(f"[WARN] 既存Excel読込失敗: {path} / {e}")
    return s

def read_teams_from_file(path: str) -> List[Tuple[str, str, str, str]]:
    out: List[Tuple[str, str, str, str]] = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1))
        name_to_idx: Dict[str, int] = {}
        for idx, cell in enumerate(header):
            val = (str(cell.value) if cell.value is not None else "").strip()
            if val:
                name_to_idx[val] = idx
        idx_country = name_to_idx.get("国", 0)
        idx_league  = name_to_idx.get("リーグ", 1)
        idx_team    = name_to_idx.get("チーム", 2)
        idx_href    = name_to_idx.get("チームリンク", 3)
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            c = ((r[idx_country] if len(r) > idx_country else "") or "").strip()
            l = ((r[idx_league]  if len(r) > idx_league  else "") or "").strip()
            t = ((r[idx_team]    if len(r) > idx_team    else "") or "").strip()
            h = ((r[idx_href]    if len(r) > idx_href    else "") or "").strip()
            if c and l and t and h:
                out.append((c, l, t, h))
    except Exception as e:
        print(f"[WARN] チームExcel読込失敗: {path} / {e}")
    return out

# =========================
# Playwright ヘルパ
# =========================
global_nav_sema = asyncio.Semaphore(GLOBAL_NAV_CONCURRENCY)
_nav_gap_lock = asyncio.Lock()
_last_nav_ts = 0.0

async def _respect_nav_gap():
    global _last_nav_ts
    async with _nav_gap_lock:
        loop = asyncio.get_running_loop()
        now = loop.time()
        due = _last_nav_ts + (NAV_GAP_MS / 1000.0)
        if now < due:
            await asyncio.sleep(due - now)
        _last_nav_ts = loop.time()

async def make_context(browser):
    ctx = await browser.new_context(
        locale="ja-JP",
        timezone_id="Asia/Tokyo",
        user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"),
        viewport={"width": 1280, "height": 800},
        ignore_https_errors=True,
        java_script_enabled=True,
    )
    ctx.set_default_timeout(OP_TIMEOUT)
    ctx.set_default_navigation_timeout(NAV_TIMEOUT)

    async def _route(route):
        req = route.request
        rtype = req.resource_type
        url = req.url.lower()
        if rtype in {"image", "stylesheet", "font", "media", "beacon"}:
            return await route.abort()
        if any(k in url for k in BLOCKED_HOST_KEYWORDS):
            return await route.abort()
        return await route.continue_()

    await ctx.route("**/*", _route)
    return ctx

async def safe_goto(page, url: str, selector_to_wait: Optional[str],
                    nav_timeout=NAV_TIMEOUT, sel_timeout=SEL_TIMEOUT, tries=3) -> bool:
    last = None
    for i in range(tries):
        try:
            async with global_nav_sema:
                await _respect_nav_gap()
                await page.goto(url, wait_until="domcontentloaded", timeout=nav_timeout)
            if selector_to_wait:
                await page.wait_for_selector(selector_to_wait, timeout=sel_timeout, state="attached")
            return True
        except Exception as e:
            last = e
            await asyncio.sleep(0.5 * (i + 1))
    print(f"Page.goto 最終失敗: {url} / {last}")
    return False

# =========================
# Excel Writer（専用タスク）
# =========================
class ExcelWriter:
    """
    - Queue に [row, row, ...] の「チーム単位バッチ」を入れる
    - 受け取り次第、現在開いているワークブックに append
    - 50行超えたら自動ローテーション
    - SAVE_INTERVAL_SEC ごと or SAVE_EVERY_N_ROWS ごとに保存
    """
    def __init__(self, base_dir: str, prefix: str, max_records: int):
        self.base_dir = base_dir
        self.prefix = prefix
        self.max_records = max_records
        self.q: asyncio.Queue = asyncio.Queue()
        self._task: Optional[asyncio.Task] = None

        self.wb = None
        self.ws = None
        self.seq = -1
        self.rows_in_current = 0
        self._dirty_since_last_save = 0

    def _scan_next_seq(self) -> int:
        existing = sorted(glob.glob(os.path.join(self.base_dir, f"{self.prefix}[0-9]*.xlsx")))
        if not existing:
            return 1
        def num(p):
            m = re.search(rf"{re.escape(self.prefix)}(\d+)\.xlsx$", os.path.basename(p))
            return int(m.group(1)) if m else 0
        last = max(existing, key=num)
        return num(last) if num(last) > 0 else 1

    def _open_initial(self):
        start_seq = self._scan_next_seq()
        for seq in range(start_seq, start_seq + 2):
            path = os.path.join(self.base_dir, f"{self.prefix}{seq}.xlsx")
            if os.path.exists(path):
                wb = load_workbook(path)
                ws = wb.active
                used = max(ws.max_row - 1, 0)
                if used < self.max_records:
                    self.wb, self.ws, self.seq = wb, ws, seq
                    self.rows_in_current = used
                    return
        self._rotate_new(start_seq if not os.path.exists(os.path.join(self.base_dir, f"{self.prefix}{start_seq}.xlsx")) else start_seq + 1)

    def _rotate_new(self, seq_new: int):
        path = os.path.join(self.base_dir, f"{self.prefix}{seq_new}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append([
            "国","リーグ","所属チーム","選手名","ポジション","背番号",
            "得点数","年齢","誕生日","市場価値","ローン保有元",
            "契約期限","顔写真","故障情報","データ取得時間"
        ])
        self.wb, self.ws, self.seq = wb, ws, seq_new
        self.rows_in_current = 0
        self._save_now()
        print(f"[WRITER] Rotate -> {os.path.basename(path)}")

    def _save_now(self):
        path = os.path.join(self.base_dir, f"{self.prefix}{self.seq}.xlsx")
        self.wb.save(path)
        self._dirty_since_last_save = 0

    def enqueue_team_rows(self, rows: List[List[object]], label: str):
        self.q.put_nowait((rows, label))

    async def run(self):
        self._open_initial()
        while True:
            try:
                rows, label = await asyncio.wait_for(self.q.get(), timeout=SAVE_INTERVAL_SEC)
            except asyncio.TimeoutError:
                if self._dirty_since_last_save > 0:
                    self._save_now()
                continue

            if rows is None:   # STOP
                if self._dirty_since_last_save > 0:
                    self._save_now()
                break

            # 50件単位ローテーションに合わせて分割 append
            offset = 0
            while offset < len(rows):
                remain = self.max_records - self.rows_in_current
                chunk = rows[offset: offset + remain]
                for r in chunk:
                    self.ws.append(r)
                self.rows_in_current += len(chunk)
                self._dirty_since_last_save += len(chunk)
                offset += len(chunk)

                if self.rows_in_current >= self.max_records:
                    self._save_now()
                    self._rotate_new(self.seq + 1)

            if LOG_EVERY_ENQUEUE:
                print(f"[WRITER] APPEND {label}: +{len(rows)} rows -> file #{self.seq} (cur={self.rows_in_current})")

            if self._dirty_since_last_save >= SAVE_EVERY_N_ROWS:
                self._save_now()

    async def start(self):
        if self._task is None:
            self._task = asyncio.create_task(self.run())

    async def stop(self):
        await self.q.put((None, "STOP"))
        if self._task:
            await self._task

# =========================
# スクレイピング本体（チーム内は逐次）
# =========================
async def scrape_player_detail_on_same_page(page, full_url: str) -> Dict[str, str]:
    """同じ Page を使い回して逐次遷移"""
    age = birth = mv = contract = loan = img = inj = "N/A"
    ok = await safe_goto(page, full_url, "div.playerInfoItem")
    if not ok:
        return {
            "age_text": age, "birth_date": birth, "market_value": mv,
            "loan_info": loan, "contract_date": contract,
            "img_url": img, "injury_text": inj
        }
    try:
        # preloadされた画像リンクの2番目
        preloads = await page.query_selector_all('link[rel="preload"][as="image"]')
        img = (await preloads[1].get_attribute("href")) if len(preloads) >= 2 else None
        #print(f"img: {img}")

        # ケガアイコンのtitle
        inj_el = await page.query_selector('svg[data-testid="wcl-icon-incidents-injury"] > title')
        inj = (await inj_el.text_content()).strip() if inj_el else ""
        #print(f"inj: {inj}")
    except Exception:
        pass

    try:
        info_items = await page.query_selector_all("div.playerInfoItem")
        for it in info_items:
            lab_el = await it.query_selector("strong")
            lab = (await lab_el.text_content()).strip() if lab_el else ""
            vals = await it.query_selector_all("span") or []
            if "年齢" in lab:
                if len(vals) >= 1: age = (await vals[0].text_content()).strip()
                if len(vals) >= 2: birth = (await vals[1].text_content()).strip().strip("()")
            elif "契約期限" in lab and vals:
                contract = (await vals[0].text_content()).strip()
            elif "市場価値" in lab and vals:
                mv = (await vals[0].text_content()).strip()
            elif "ローン元" in lab:
                if len(vals) >= 1: loan = (await vals[0].text_content()).strip()
                if len(vals) >= 2:
                    contract = (await vals[1].text_content()).strip().replace("期限:", "").strip("() ").strip()
    except Exception as e:
        print(f"[PLAYER DETAIL] parse失敗: {full_url} / {e}")

    return {
        "age_text": age, "birth_date": birth, "market_value": mv,
        "loan_info": loan, "contract_date": contract,
        "img_url": img, "injury_text": inj
    }

async def process_team_seq(ctx,
                           existing_keys: Set[Tuple[str, str, str, str]],
                           claimed_keys: Set[Tuple[str, str, str, str]],
                           claim_lock: asyncio.Lock,
                           team_sema: asyncio.Semaphore,
                           writer: ExcelWriter,
                           country: str, league: str, team: str, href: str):
    """チーム1件（選手は単一Pageで逐次）。完了後 Writer に一括投入。"""
    async with team_sema:
        page = await ctx.new_page()
        try:
            team_url  = href if href.startswith("http") else f"https://flashscore.co.jp{href}"
            squad_url = team_url.rstrip("/") + "/squad/"

            ok = await safe_goto(page, squad_url, "div.lineupTable.lineupTable--soccer")
            if not ok:
                print(f"[{country}:{league}:{team}] スカッド取得失敗: goto failed")
                await page.close()
                return

            # スカッド抽出
            players: List[Tuple[str, str, str, str, str]] = []
            tables = await page.query_selector_all("div.lineupTable.lineupTable--soccer")
            for t in tables or []:
                pos_el = await t.query_selector("div.lineupTable__title")
                position = (await pos_el.text_content()).strip() if pos_el else "N/A"
                rows = await t.query_selector_all("div.lineupTable__row")
                for r in rows:
                    jersey_el = await r.query_selector(".lineupTable__cell--jersey")
                    jersey = (await jersey_el.text_content()).strip() if jersey_el else "N/A"
                    name_el = await r.query_selector(".lineupTable__cell--player .lineupTable__cell--name")
                    if not name_el:
                        continue
                    name = (await name_el.text_content()).strip()
                    href2 = await name_el.get_attribute("href")
                    if not href2:
                        continue
                    goal_el = await r.query_selector(".lineupTable__cell--goal")
                    goals = (await goal_el.text_content()).strip() if goal_el else "N/A"
                    players.append((position, jersey, name, href2, goals))

            if not players:
                print(f"[{country}:{league}:{team}] 選手0件")
                await page.close()
                return

            # 同じ page を使い回して逐次で詳細へ
            rows_to_write: List[List[object]] = []
            for position, jersey, name, href2, goals in players:
                key = (country, league, team, name)
                async with claim_lock:
                    if key in existing_keys or key in claimed_keys:
                        continue
                    claimed_keys.add(key)

                purl = href2 if href2.startswith("http") else f"https://flashscore.co.jp{href2}"
                detail = await scrape_player_detail_on_same_page(page, purl)

                rows_to_write.append([
                    country, league, team, name, position, jersey, goals,
                    detail.get("age_text","N/A"), detail.get("birth_date","N/A"),
                    detail.get("market_value","N/A"), detail.get("loan_info","N/A"),
                    detail.get("contract_date","N/A"), detail.get("img_url","N/A"),
                    detail.get("injury_text","N/A"), datetime.datetime.now()
                ])

            # チーム単位で一括投入
            if rows_to_write:
                if LOG_EVERY_ENQUEUE:
                    print(f"[ENQUEUE] {country} / {league} / {team}: {len(rows_to_write)} 件をWriterへ")
                writer.enqueue_team_rows(rows_to_write, f"{country}/{league}/{team}")

            print(f"[TEAM-DONE] {country} / {league} / {team}")

        except Exception as e:
            print(f"[{country}:{league}:{team}] 処理中エラー: {e}")
        finally:
            try:    await page.close()
            except: pass

# =========================
# エントリポイント
# =========================
async def main():
    ensure_dirs()

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{now}  TeamMemberExcel: 開始（4チーム並列・選手逐次）")

    existing_keys = load_existing_player_keys_from_excels()
    print(f"[INIT] 既取得キー: {len(existing_keys)} 件")

    team_rows: List[Tuple[str, str, str, str]] = []
    paths = sorted(glob.glob(os.path.join(TEAMS_BY_LEAGUE_DIR, "*.xlsx")))
    if not paths:
        print(f"[EXIT] チームExcelが見つかりません: {TEAMS_BY_LEAGUE_DIR}")
        return

    for p in paths:
        rows = read_teams_from_file(p)
        print(f"[LOAD] {os.path.basename(p)}: {len(rows)} 行")
        team_rows.extend(rows)

    # 完全重複除去
    seen: Set[Tuple[str, str, str, str]] = set()
    uniq_rows: List[Tuple[str, str, str, str]] = []
    for row in team_rows:
        if row not in seen:
            seen.add(row)
            uniq_rows.append(row)

    print(f"[PLAN] 処理チーム数（重複除去後）: {len(uniq_rows)}")

    # ライター起動
    writer = ExcelWriter(BASE_DIR, EXCEL_BASE_PREFIX, EXCEL_MAX_RECORDS)
    await writer.start()

    claim_lock = asyncio.Lock()
    team_sema = asyncio.Semaphore(TEAM_CONCURRENCY)
    claimed_keys: Set[Tuple[str, str, str, str]] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--disable-background-timer-throttling",
                "--disable-renderer-backgrounding",
                "--disable-backgrounding-occluded-windows",
                "--disable-breakpad",
                "--disable-features=TranslateUI,BackForwardCache,AcceptCHFrame,MediaRouter",
                "--blink-settings=imagesEnabled=false",
            ],
        )
        ctx = await make_context(browser)

        # 4 並列でチームを流す（同時に大量のタスクを投げず gather は控えめに）
        # キュー方式で確実に 4 ワーカーに制限
        team_q: asyncio.Queue = asyncio.Queue()
        for it in uniq_rows:
            team_q.put_nowait(it)
        for _ in range(TEAM_CONCURRENCY):
            team_q.put_nowait(None)

        async def worker(wid: int):
            while True:
                item = await team_q.get()
                if item is None:
                    team_q.task_done()
                    break
                country, league, team, href = item
                try:
                    await process_team_seq(ctx, existing_keys, claimed_keys, claim_lock,
                                           team_sema, writer, country, league, team, href)
                finally:
                    team_q.task_done()

        workers = [asyncio.create_task(worker(i+1)) for i in range(TEAM_CONCURRENCY)]
        await team_q.join()
        for w in workers:
            w.cancel()
        try:
            await asyncio.gather(*workers, return_exceptions=True)
        except:
            pass

        await ctx.close()
        await browser.close()

    # ライター停止（残りを保存して終了）
    await writer.stop()

    print("TeamMemberExcel: 完了（4チーム並列・選手逐次）")

if __name__ == "__main__":
    asyncio.run(main())

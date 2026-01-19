# -*- coding: utf-8 -*-
"""
FutureData_aws_S3.py (ECS + S3)
- Flashscore.co.jp の「サッカー > 開催予定」から FUTURE_DAYS 日分の試合を取得
- 各試合の順位表(standings)から「国/リーグ」「ホーム順位/アウェー順位」を埋める
- 対象リーグのみ抽出（CONTAINS_LIST / UNDER_LIST / GENDER_LIST / EXP_LIST）
- future_N.xlsx をローテーション保存（MAX_ROWS_PER_FILE_SCHEDULED 行ごと）
- future_*.xlsx -> future_*.csv 変換
- S3（aws-s3-future-csv）へアップロード
  - S3内の future_*.csv 最大連番を取得し、次の番号から採番して上書き回避
"""

from playwright.sync_api import sync_playwright
import time
import re, os
import datetime
from typing import List, Dict, Optional
import pandas as pd
from pathlib import Path
import glob

try:
    import openpyxl
except ImportError:
    raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。")

# ✅ S3
import boto3

# =========================
# 環境変数
# =========================
FUTURE_DAYS = int(os.environ.get("FUTURE_DAYS", "7"))
SAVE_DIR_SCHEDULED = os.environ.get("FUTURE_SAVE_DIR", "/tmp/bookmaker/future")

S3_BUCKET_FUTURE = os.environ.get("S3_BUCKET_FUTURE", "aws-s3-future-csv")
S3_PREFIX_FUTURE = os.environ.get("S3_PREFIX_FUTURE", "")  # 例: "future/"

# ============== 取得条件 ====================

HEADER_SCHEDULED = [
    "試合国及びカテゴリ", "試合予定時間", "ホーム順位", "アウェー順位", "ホームチーム", "アウェーチーム",
    "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者",
    "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点",
    "ホームチームアウェー得点", "ホームチームアウェー失点",
    "アウェーチームアウェー得点", "アウェーチームアウェー失点", "試合リンク文字列", "データ取得時間"
]

MAX_ROWS_PER_FILE_SCHEDULED = 10
SHEET_NAME_SCHEDULED = "Sheet1"
FILE_PREFIX_SCHEDULED = "future_"
FILE_SUFFIX_SCHEDULED = ".xlsx"

# S3の最大連番+1 を起点にする
SERIAL_BASE_SCHEDULED = 1

# ============== 共通ユーティリティ ==============

VERBOSE = True

def log(msg: str):
    if VERBOSE:
        print(msg)

def text_clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def extract_mid(s: str) -> Optional[str]:
    """URLから mid を抽出"""
    if not s:
        return None
    s = str(s).strip()
    m = re.search(r"[?&#]mid=([A-Za-z0-9]+)", s)
    if m:
        return m.group(1)
    m = re.search(r"/match/([A-Za-z0-9]{6,20})(?:/|$)", s)
    if m:
        return m.group(1)
    return None

# ==========================================
# ✅ 対象リーグフィルタリング用リスト
# ==========================================
CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]
UNDER_LIST  = ["U17", "U18", "U19", "U20", "U21", "U22", "U23", "U24", "U25"]
GENDER_LIST = ["女子"]
EXP_LIST    = ["ポルトガル: リーガ・ポルトガル 2", "イングランド: プレミアリーグ 2", "イングランド: プレミアリーグ U18"]

# ==========================================
# ✅ S3 連携（最大連番取得、アップロード）
# ==========================================

def _s3_client():
    return boto3.client("s3")

def get_max_future_csv_seq_in_s3(bucket: str, prefix: str = "") -> int:
    """
    S3上の prefix + future_*.csv の最大連番を返す（無ければ0）
    """
    s3 = _s3_client()
    max_seq = 0
    token = None

    pattern = re.compile(rf"^{re.escape(prefix)}{re.escape(FILE_PREFIX_SCHEDULED)}(\d+)\.csv$")

    while True:
        kwargs = {"Bucket": bucket}
        if prefix:
            kwargs["Prefix"] = prefix
        if token:
            kwargs["ContinuationToken"] = token

        resp = s3.list_objects_v2(**kwargs)
        for obj in resp.get("Contents", []):
            key = obj.get("Key", "")
            m = pattern.match(key)
            if m:
                try:
                    max_seq = max(max_seq, int(m.group(1)))
                except:
                    pass

        if resp.get("IsTruncated"):
            token = resp.get("NextContinuationToken")
        else:
            break

    return max_seq

def upload_file_to_s3(local_path: str, bucket: str, key: str):
    s3 = _s3_client()
    s3.upload_file(local_path, bucket, key)
    log(f"☁️ S3アップロード: s3://{bucket}/{key}")

def init_serial_base_from_s3():
    """
    S3上の future_*.csv 最大連番+1 を SERIAL_BASE_SCHEDULED にセット
    """
    global SERIAL_BASE_SCHEDULED
    try:
        max_seq = get_max_future_csv_seq_in_s3(S3_BUCKET_FUTURE, S3_PREFIX_FUTURE)
        SERIAL_BASE_SCHEDULED = max_seq + 1
        log(f"🔢 S3最大連番: {max_seq} → 今回開始: {SERIAL_BASE_SCHEDULED}")
    except Exception as e:
        log(f"⚠️ S3から連番取得失敗（{e}）。SERIAL_BASE=1で続行（上書き注意）")
        SERIAL_BASE_SCHEDULED = 1

# ==========================================
# ✅ Excel 逐次書き込み（開催予定）
# ==========================================

def _existing_serials_scheduled(output_dir: str) -> List[int]:
    p = Path(output_dir)
    nums = []
    for f in p.glob(f"{FILE_PREFIX_SCHEDULED}*{FILE_SUFFIX_SCHEDULED}"):
        m = re.match(rf"^{re.escape(FILE_PREFIX_SCHEDULED)}(\d+){re.escape(FILE_SUFFIX_SCHEDULED)}$", f.name)
        if m:
            nums.append(int(m.group(1)))
    return sorted(nums)

def _next_serial_scheduled(output_dir: str) -> int:
    nums = _existing_serials_scheduled(output_dir)
    if nums:
        return max(nums) + 1
    return SERIAL_BASE_SCHEDULED

def _current_file_path_scheduled(output_dir: str) -> Path:
    p = Path(output_dir)
    nums = _existing_serials_scheduled(output_dir)
    if not nums:
        return p / f"{FILE_PREFIX_SCHEDULED}{SERIAL_BASE_SCHEDULED}{FILE_SUFFIX_SCHEDULED}"
    return p / f"{FILE_PREFIX_SCHEDULED}{max(nums)}{FILE_SUFFIX_SCHEDULED}"

def _data_rows_in_scheduled(path: Path) -> int:
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME_SCHEDULED] if SHEET_NAME_SCHEDULED in wb.sheetnames else wb.active
    total = ws.max_row or 0
    wb.close()
    return max(0, total - 1)

def _create_new_workbook_scheduled(path: Path):
    df = pd.DataFrame(columns=HEADER_SCHEDULED)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=SHEET_NAME_SCHEDULED)

def append_scheduled_row_to_excel(
    row_dict: Dict[str, str],
    output_dir: str = SAVE_DIR_SCHEDULED,
    max_rows_per_file: int = MAX_ROWS_PER_FILE_SCHEDULED
):
    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)

    cur = _current_file_path_scheduled(str(output_dir_path))
    if not cur.exists():
        _create_new_workbook_scheduled(cur)

    current_rows = _data_rows_in_scheduled(cur)

    if current_rows >= max_rows_per_file:
        next_serial = _next_serial_scheduled(str(output_dir_path))
        cur = output_dir_path / f"{FILE_PREFIX_SCHEDULED}{next_serial}{FILE_SUFFIX_SCHEDULED}"
        _create_new_workbook_scheduled(cur)
        current_rows = 0

    df = pd.DataFrame([row_dict])
    for col in HEADER_SCHEDULED:
        if col not in df.columns:
            df[col] = ""
    df = df[HEADER_SCHEDULED]

    with pd.ExcelWriter(cur, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
        startrow = current_rows + 1
        df.to_excel(w, index=False, header=False, sheet_name=SHEET_NAME_SCHEDULED, startrow=startrow)

    log(f"💾 [FUTURE] 追記: {cur.name}（{current_rows}→{current_rows+1}）")

def save_scheduled_to_excel(match_results: List[Dict[str, str]], output_dir: str = SAVE_DIR_SCHEDULED):
    if not match_results:
        log("✋ Excelに書き込む開催予定データがありません")
        return
    os.makedirs(output_dir, exist_ok=True)

# ==========================================
# ✅ Flashscore「開催予定」ナビ
# ==========================================

def goto_football_top(page):
    """サッカー → 開催予定タブへ遷移"""
    log("🌐 Flashscore トップへアクセス...")
    page.goto("https://www.flashscore.co.jp/", timeout=45000, wait_until="domcontentloaded")

    # Cookieバナーなど
    try:
        page.locator("#onetrust-accept-btn-handler").click(timeout=2000)
        log("✅ Cookieバナーを閉じました")
    except:
        pass

    # サッカーが選ばれていない場合に備えて「サッカー」クリック（保険）
    try:
        soccer_btn = page.locator("a,button").filter(has_text="サッカー").first
        if soccer_btn and soccer_btn.count():
            soccer_btn.click(timeout=4000)
            time.sleep(0.8)
    except:
        pass

    # 「開催予定」タブをクリック
    try:
        tab = page.locator("div.filters__tab[data-analytics-alias='scheduled']").first
        if tab and tab.count():
            tab.click(timeout=4000)
        else:
            tab = page.locator("div.filters__tab").filter(has_text=re.compile(r"(開催予定)")).first
            tab.click(timeout=4000)
        log("✅ 『開催予定』タブに切り替えました")
    except Exception as e:
        log(f"⚠️ 開催予定タブ切り替え失敗: {e}")

    try:
        page.wait_for_timeout(1000)
        page.wait_for_load_state("networkidle", timeout=8000)
    except:
        pass

def expand_all_collapsed_leagues(page):
    """
    Flashscore『開催予定』タブで、折りたたまれているリーグをすべて展開する。
    """
    print("📂 折りたたみリーグ（非表示）を展開します...")

    btn_selector = (
        "button[data-testid='wcl-accordionButton']"
        "[aria-label='リーグ全試合 表示']"
    )

    max_loops = 200
    for _ in range(max_loops):
        btns = page.locator(btn_selector)
        count = btns.count()

        if count == 0:
            print("   ✅ すべての折りたたみリーグを展開しました")
            return

        print(f"   残り『表示』ボタン数: {count}")

        btn = btns.first
        try:
            btn.scroll_into_view_if_needed()
        except Exception:
            pass

        try:
            btn.click(timeout=2000)
        except Exception as e:
            print(f"   ⚠️ ボタンクリック失敗: {e}")
            break

        page.wait_for_timeout(200)

    print("   ⚠️ ループ上限に達しました。まだ非表示リーグが残っている可能性があります。")

# ==========================================
# ✅ 試合行の基本情報取得
# ==========================================

def _get_match_row_teams_and_time(row):
    """
    1つの試合行から
      - 試合予定時間
      - ホームチーム名
      - アウェーチーム名
    を抽出（旧UI / 新UI 両対応）
    """
    ktime = ""

    # 旧UI: .event__time
    try:
        ktime = text_clean(row.locator(".event__time").first.text_content() or "")
    except Exception:
        pass

    # 新UI: data-testid ベース
    if not ktime:
        try:
            ktime = text_clean(
                row.locator(
                    "[data-testid='wcl-time'], "
                    "[data-testid='wcl-start-time'], "
                    "[data-testid='wcl-time-status']"
                ).first.text_content() or ""
            )
        except Exception:
            pass

    home = ""
    away = ""

    # パターン①: 旧UI (.event__participant--home / --away)
    try:
        h = row.locator(".event__participant--home .event__participant--name").first
        a = row.locator(".event__participant--away .event__participant--name").first
        if h.count():
            home = text_clean(h.text_content() or "")
        if a.count():
            away = text_clean(a.text_content() or "")
    except Exception:
        pass

    # パターン②: 旧UI フォールバック (.event__participant)
    if not home or not away:
        try:
            ps = row.locator(".event__participant .event__participant--name")
            if ps.count() >= 2:
                if not home:
                    home = text_clean(ps.first.text_content() or "")
                if not away:
                    away = text_clean(ps.last.text_content() or "")
        except Exception:
            pass

    # パターン③: 新UI 明示的セレクタ
    if not home or not away:
        try:
            h = row.locator(
                ".event__homeParticipant span.wcl-name_jjfMf, "
                ".event__homeParticipant [data-testid='wcl-scores-simple-text-01']"
            ).first
            a = row.locator(
                ".event__awayParticipant span.wcl-name_jjfMf, "
                ".event__awayParticipant [data-testid='wcl-scores-simple-text-01']"
            ).first

            if h and h.count():
                if not home:
                    home = text_clean(h.text_content() or "")
            if a and a.count():
                if not away:
                    away = text_clean(a.text_content() or "")
        except Exception:
            pass

    # パターン④: 新UI ゆるめ
    if not home or not away:
        try:
            ps = row.locator(
                "[data-testid='wcl-matchRow-participant'] span.wcl-name_jjfMf, "
                "[data-testid='wcl-matchRow-participant'] [data-testid='wcl-scores-simple-text-01']"
            )
            n = ps.count()
            if n >= 2:
                if not home:
                    home = text_clean(ps.nth(0).text_content() or "")
                if not away:
                    away = text_clean(ps.nth(n - 1).text_content() or "")
        except Exception:
            pass

    # パターン⑤: <img alt="チーム名">
    if not home or not away:
        try:
            imgs = row.locator("[data-testid='wcl-matchRow-participant'] img[data-testid='wcl-participantLogo']")
            n_img = imgs.count()
            if n_img >= 2:
                if not home:
                    alt0 = imgs.nth(0).get_attribute("alt") or ""
                    home = text_clean(alt0)
                if not away:
                    alt1 = imgs.nth(n_img - 1).get_attribute("alt") or ""
                    away = text_clean(alt1)
        except Exception:
            pass

    if not home or not away:
        try:
            snippet = (row.inner_text() or "").strip().replace("\n", " ")[:200]
        except Exception:
            snippet = "<inner_text 取得失敗>"
        print(f"⚠️ チーム名取得失敗: time={ktime}, snippet={snippet}")

    return ktime, home, away

def _get_match_row_link(row) -> str:
    """試合行から matchURL を抽出"""
    try:
        a = row.locator("a.eventRowLink[href*='/match/'][href*='?mid=']").first
        if a and a.count():
            href = a.get_attribute("href") or ""
            if href.startswith("http"):
                return href
            return "https://www.flashscore.co.jp" + href
    except:
        pass
    return ""

# ==========================================
# ✅ 日付
# ==========================================

def get_current_match_date(page) -> Optional[datetime.date]:
    """
    Flashscore トップの日付ボタン（例: '05/12 金'）から datetime.date を作って返す。
    """
    try:
        btn = page.locator("button[data-testid='wcl-dayPickerButton']").first
        if not btn or not btn.count():
            return None

        txt = text_clean(btn.inner_text() or "")
        m = re.search(r"(\d{2})/(\d{2})", txt)
        if not m:
            return None

        day = int(m.group(1))
        month = int(m.group(2))

        # 年は今年（年またぎは簡易対応）
        year = datetime.datetime.now().year
        return datetime.date(year, month, day)

    except Exception:
        return None

# ==========================================
# ✅ 開催予定タブから基本情報だけ集める
# ==========================================

def collect_scheduled_matches_on_current_day(page) -> List[Dict[str, str]]:
    """
    現在表示中の日付（開催予定タブ）から試合情報を収集。
    ここでは「時間・ホーム・アウェー・リンク」だけを集め、
    国リーグ＆順位は後で試合ページから埋める。
    """
    try:
        page.wait_for_selector("div.event__match", timeout=12000)
    except:
        log("⚠️ event__match が見つからないまま続行")

    expand_all_collapsed_leagues(page)

    match_date = get_current_match_date(page)

    rows = page.locator("div.event__match.event__match--scheduled")
    if rows.count() == 0:
        rows = page.locator("div.event__match")

    n = rows.count()
    log(f"🎯 開催予定試合 行数: {n}")

    results: List[Dict[str, str]] = []
    seen_mids = set()

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i in range(n):
        row = rows.nth(i)
        try:
            ktime, home, away = _get_match_row_teams_and_time(row)
            link = _get_match_row_link(row)

            if match_date and ktime:
                match_dt_str = f"{match_date.strftime('%Y-%m-%d')} {ktime}"
            else:
                match_dt_str = ktime

            mid = extract_mid(link)
            if mid and mid in seen_mids:
                log(f"   ⏭️ 重複試合(mid={mid})をスキップ")
                continue
            if mid:
                seen_mids.add(mid)

            d = {k: "" for k in HEADER_SCHEDULED}
            d["試合国及びカテゴリ"] = ""  # 後で埋める
            d["試合予定時間"]       = match_dt_str
            d["ホームチーム"]       = home
            d["アウェーチーム"]     = away
            d["試合リンク文字列"]   = link
            d["データ取得時間"]     = now_str

            results.append(d)
            log(f"   [{i+1}/{n}] | {match_dt_str} | {home} vs {away} | mid={mid}")

        except Exception as e:
            log(f"   ⚠️ 行{i}でエラー: {e}")
            continue

    log(f"✅ 当日分 取得件数: {len(results)}")
    return results

def click_next_day(page) -> bool:
    """
    「翌日」ボタンをクリックして日付を1日進める。
    """
    try:
        btn = page.locator("button.wcl-arrow_YpdN4[data-day-picker-arrow='next']").first
        if not btn or not btn.count():
            log("⚠️ 翌日ボタンが見つかりませんでした")
            return False
        btn.click(timeout=3000)
        log("➡️ 『翌日』ボタンをクリックしました")
        time.sleep(1.0)
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except:
            pass
        return True
    except Exception as e:
        log(f"⚠️ 翌日ボタンクリック失敗: {e}")
        return False

# ==========================================
# ✅ 順位表 & 国リーグ取得
# ==========================================

def build_standings_url_from_match_url(match_url: str) -> str:
    """
    試合ページURLを順位表タブのURLに変換する。
    """
    if not match_url:
        return ""
    if "/standings/" in match_url:
        return match_url
    return re.sub(r"/(\?mid=)", r"/standings/\1", match_url, count=1)

def get_team_ranks_from_standings_table(page, home_name: str, away_name: str):
    """
    すでに「順位表」タブ（オーバーオール）が表示されている page から順位を取得。
    """
    home_name_norm = text_clean(home_name)
    away_name_norm = text_clean(away_name)

    home_rank = ""
    away_rank = ""

    rows = page.locator("div.ui-table__body > div.ui-table__row")
    n_rows = rows.count()

    for i in range(n_rows):
        row = rows.nth(i)
        try:
            team_elem = row.locator(".tableCellParticipant__name").first
            if not team_elem.count():
                continue
            team_name = text_clean(team_elem.text_content() or "")

            rank_elem = row.locator(".tableCellRank").first
            if not rank_elem.count():
                continue
            rank_text = text_clean(rank_elem.text_content() or "")
            rank_text = rank_text.rstrip(".").strip()

            if not home_rank and team_name == home_name_norm:
                home_rank = rank_text
            if not away_rank and team_name == away_name_norm:
                away_rank = rank_text

            if home_rank and away_rank:
                break
        except:
            continue

    return home_rank, away_rank

def get_country_and_league_from_match_page(page):
    """
    試合ページ（サマリー / 順位表タブ）のパンくずから国名/リーグ名を取得する。
    """
    country = ""
    league = ""

    try:
        try:
            page.wait_for_selector(
                "nav[data-testid='wcl-breadcrumbs'] span[data-testid='wcl-scores-overline-03']",
                timeout=3000
            )
        except Exception:
            pass

        spans = page.locator(
            "nav[data-testid='wcl-breadcrumbs'] span[data-testid='wcl-scores-overline-03']"
        )
        count = spans.count()

        if count == 0:
            return "", ""

        texts = []
        for i in range(count):
            txt = text_clean(spans.nth(i).text_content() or "")
            if txt:
                texts.append(txt)

        start_idx = 0
        if texts and texts[0] == "サッカー":
            start_idx = 1

        if len(texts) > start_idx:
            country = texts[start_idx]
        if len(texts) > start_idx + 1:
            league = texts[start_idx + 1]

    except Exception:
        pass

    return country, league

def fetch_ranks_for_match(page, match_url: str, home_name: str, away_name: str):
    """
    試合URLから順位表タブに飛び、
      - ホーム順位
      - アウェー順位
      - 国名
      - リーグ名（ラウンド付き）
    を取得する。
    """
    if not match_url:
        return "", "", "", ""

    standings_url = build_standings_url_from_match_url(match_url)

    try:
        log(f"   📊 順位表取得: {standings_url}")
        page.goto(standings_url, timeout=25000, wait_until="domcontentloaded")

        # 念のため standings タブ選択（UI変化に備える）
        try:
            standings_tab = page.locator(
                "a[data-analytics-alias='stats-detail'] button, "
                "a[href*='/standings/'] button"
            ).first
            if standings_tab and standings_tab.count():
                selected = standings_tab.get_attribute("data-selected")
                if selected != "true":
                    standings_tab.click(timeout=3000)
                    page.wait_for_timeout(500)
        except:
            pass

        country, league = get_country_and_league_from_match_page(page)

        page.wait_for_selector("div.ui-table__body div.ui-table__row", timeout=12000)

        home_rank, away_rank = get_team_ranks_from_standings_table(page, home_name, away_name)
        log(f"      → rank: home={home_rank}, away={away_rank}, country={country}, league={league}")
        return home_rank, away_rank, country, league

    except Exception as e:
        log(f"   ⚠️ 順位表取得失敗: {e}")
        return "", "", "", ""

# ==========================================
# ✅ 全試合に対して 国リーグ & 順位 & フィルタ
# ==========================================

def fill_ranks_for_matches(ctx, matches: List[Dict[str, str]]):
    """
    「開催予定」で収集した試合リストに対して、
    各試合ページの順位表から順位/国リーグを埋め、対象リーグだけ残す。
    """
    if not matches:
        return

    page = ctx.new_page()
    filtered: List[Dict[str, str]] = []

    for idx, m in enumerate(matches):
        url  = m.get("試合リンク文字列") or ""
        home = m.get("ホームチーム") or ""
        away = m.get("アウェーチーム") or ""

        if not url or not home or not away:
            log(f"⏭️ URL/チーム名不足のためスキップ: {home} vs {away}")
            continue

        log(f"=== 順位取得 {idx+1}/{len(matches)} ===")
        home_rank, away_rank, country, league = fetch_ranks_for_match(page, url, home, away)

        game_category = ""
        if country and league:
            game_category = f"{country}: {league}"
        elif country or league:
            game_category = country or league

        if not game_category:
            log("⏭️ スキップ対象: （カテゴリ取得失敗）")
            continue

        # 1) CONTAINS_LIST に含まれるリーグだけ対象
        if not any(c in game_category for c in CONTAINS_LIST):
            log(f"⏭️ スキップ対象: {game_category}（リスト外）")
            continue

        # 2) 年代（Uxx）・女子・例外リーグを含む場合はスキップ
        if (any(x in game_category for x in UNDER_LIST) or
            any(x in game_category for x in GENDER_LIST) or
            any(x in game_category for x in EXP_LIST)):
            log(f"🚫 除外対象: {game_category}")
            continue

        if home_rank:
            m["ホーム順位"] = home_rank
        if away_rank:
            m["アウェー順位"] = away_rank
        m["試合国及びカテゴリ"] = game_category

        # ✅ ここで即 Excel に1行追記
        append_scheduled_row_to_excel(m, output_dir=SAVE_DIR_SCHEDULED)

        filtered.append(m)
        log(f"✅ 対象試合: {game_category} | {home} vs {away}")

    page.close()
    matches[:] = filtered

# ==========================================
# ✅ メイン入口（開催予定取得）
# ==========================================

def fetch_scheduled_matches(days: int) -> List[Dict[str, str]]:
    """
    Flashscore 開催予定タブから、
      - 今日（表示中の日付）
      - 翌日以降（days-1回『翌日』をクリック）
    の試合情報を取得してリストで返す。

    days=2 → 今日＋翌日
    """
    all_results: List[Dict[str, str]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, slow_mo=70)
        ctx = browser.new_context(
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
        )
        page = ctx.new_page()

        goto_football_top(page)

        # ① 開催予定タブから全試合を集める（カテゴリはまだ空）
        for day_idx in range(days):
            if day_idx > 0:
                ok = click_next_day(page)
                if not ok:
                    log("⏭️ 翌日への遷移ができなかったため、以降の取得はスキップします")
                    break
            log(f"==================== 日付オフセット {day_idx} 日目 ====================")
            day_results = collect_scheduled_matches_on_current_day(page)
            all_results.extend(day_results)

        page.close()

        # ② 各試合ページの「順位表」から埋めつつ対象リーグのみ抽出
        fill_ranks_for_matches(ctx, all_results)

        browser.close()

    log(f"🎉 総取得件数: {len(all_results)}")
    return all_results

# ==========================================
# ✅ future_*.xlsx → future_*.csv 変換 + S3アップロード
# ==========================================

def get_existing_future_xlsx_seqs(base_dir: str = SAVE_DIR_SCHEDULED) -> List[int]:
    pattern = os.path.join(base_dir, "future_*.xlsx")
    xlsx_files = glob.glob(pattern)
    seqs: List[int] = []

    for path in xlsx_files:
        basename = os.path.basename(path)
        if not basename.startswith("future_") or not basename.endswith(".xlsx"):
            continue
        try:
            num = int(basename.replace("future_", "").replace(".xlsx", ""))
            seqs.append(num)
        except:
            pass

    return sorted(seqs)

def excel_to_csv_and_upload(excel_file: str, csv_file: str, delete_local: bool = True):
    """
    1つの Excel (future_N.xlsx) を CSV (future_N.csv) に変換し、S3へアップロード。
    成功したらローカルのExcel/CSVを削除（delete_local=True）。
    """
    df = pd.read_excel(excel_file)
    df.to_csv(csv_file, index=False)
    log(f"✅ Excel->CSV: {os.path.basename(excel_file)} → {os.path.basename(csv_file)}")

    key = f"{S3_PREFIX_FUTURE}{os.path.basename(csv_file)}"
    upload_file_to_s3(csv_file, S3_BUCKET_FUTURE, key)

    if delete_local:
        try:
            os.remove(excel_file)
            log(f"🗑 ローカルExcel削除: {os.path.basename(excel_file)}")
        except:
            pass
        try:
            os.remove(csv_file)
            log(f"🗑 ローカルCSV削除: {os.path.basename(csv_file)}")
        except:
            pass

def convert_all_future_excels_to_csv_and_upload(base_dir: str = SAVE_DIR_SCHEDULED):
    """
    base_dir 配下の future_*.xlsx をすべて CSV へ変換し、S3へアップロードする。
    """
    seq_list_all = get_existing_future_xlsx_seqs(base_dir)
    if not seq_list_all:
        log("変換対象の future_*.xlsx がありません")
        return

    for seq in seq_list_all:
        xlsx_path = os.path.join(base_dir, f"future_{seq}.xlsx")
        csv_path  = os.path.join(base_dir, f"future_{seq}.csv")
        if not os.path.exists(xlsx_path):
            continue
        excel_to_csv_and_upload(xlsx_path, csv_path, delete_local=True)

# ==========================================
# ✅ ECS起動エントリ
# ==========================================

def main():
    # S3の最大連番を見て採番起点を決定
    init_serial_base_from_s3()

    os.makedirs(SAVE_DIR_SCHEDULED, exist_ok=True)

    matches = fetch_scheduled_matches(days=FUTURE_DAYS)
    log(f"総件数: {len(matches)}")

    # 今回の実装は append_scheduled_row_to_excel で逐次書いているため、
    # save_scheduled_to_excel はログ用途/将来拡張用（現状は何もしない）
    save_scheduled_to_excel(matches, output_dir=SAVE_DIR_SCHEDULED)

    # 生成された future_*.xlsx をCSV化してS3へ
    convert_all_future_excels_to_csv_and_upload(base_dir=SAVE_DIR_SCHEDULED)

if __name__ == "__main__":
    main()

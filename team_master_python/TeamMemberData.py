# -*- coding: utf-8 -*-
"""
TeamMemberExcel_writer_queue.py（リーグ並列 + 1リーグ内10チーム並列 + 重複完全回避）
- 途中停止して再実行しても「既に出力に書いた選手」は二度と書かない
  => 既存CSV/XLSXからロード + Writerが書いた瞬間に existing_keys を更新

- teams_by_league/teamData_*.csv を全読み
- (国,リーグ) 単位でリーグ並列（LEAGUE_CONCURRENCY）
- 各リーグは 10チームずつ並列（TEAMS_PER_LEAGUE_CONCURRENCY=10）
- チーム内は PagePool(PLAYER_PER_TEAM) で選手詳細を並列取得
- goto 制御（GLOBAL_NAV_CONCURRENCY / NAV_GAP_MS）
- Writer集約（saveはto_thread）
- 最後に xlsx -> csv 変換して xlsx 削除
"""
import json
from pathlib import Path
import os
import re
import glob
import asyncio
import datetime
from typing import List, Tuple, Dict, Set, Optional
import csv

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

# =========================
# 設定
# =========================
BASE_DIR = "/Users/shiraishitoshio/bookmaker"
TEAMS_BY_LEAGUE_DIR = os.path.join(BASE_DIR, "teams_by_league")
B001_JSON_PATH = "/Users/shiraishitoshio/bookmaker/json/b001/b001_country_league.json"

TEAMDATA_PREFIX = "teamData_"
TEAMDATA_GLOB = os.path.join(TEAMS_BY_LEAGUE_DIR, f"{TEAMDATA_PREFIX}*.csv")

EXCEL_BASE_PREFIX = "teamMemberData_"
EXCEL_MAX_RECORDS = 50

CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]

OP_TIMEOUT = 5000
NAV_TIMEOUT = 12000
SEL_TIMEOUT = 20000

# ---- 並列設定 ----
LEAGUE_CONCURRENCY = 4
TEAMS_PER_LEAGUE_CONCURRENCY = 10

PLAYER_PER_TEAM = 3
GLOBAL_PLAYERDETAIL_CONCURRENCY = 8

GLOBAL_NAV_CONCURRENCY = 3
NAV_GAP_MS = 150

SAVE_INTERVAL_SEC = 2.0
SAVE_EVERY_N_ROWS = 80
LOG_EVERY_ENQUEUE = True

BLOCKED_HOST_KEYWORDS = [
    "googletagmanager", "google-analytics", "doubleclick", "googlesyndication",
    "scorecardresearch", "criteo", "adsystem", "mathtag", "quantserve",
    "taboola", "facebook", "twitter", "hotjar", "onetrust", "cookiebot",
    "trustarc", "ioam.de", "amazon-adsystem", "adservice"
]

# =========================
# ユーティリティ
# =========================
def ensure_dirs():
    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(TEAMS_BY_LEAGUE_DIR, exist_ok=True)

def load_existing_player_keys(base_dir: str) -> Set[Tuple[str, str, str, str]]:
    """
    既存の teamMemberData_*.csv / *.xlsx から (国,リーグ,チーム,選手名) を集合化
    """
    s: Set[Tuple[str, str, str, str]] = set()

    for path in sorted(glob.glob(os.path.join(TEAMS_BY_LEAGUE_DIR, f"{EXCEL_BASE_PREFIX}[0-9]*.csv"))):
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                r = csv.DictReader(f)
                for row in r:
                    c = str(row.get("国", "")).strip()
                    l = str(row.get("リーグ", "")).strip()
                    t = str(row.get("所属チーム", "")).strip()
                    n = str(row.get("選手名", "")).strip()
                    if c and l and t and n:
                        s.add((c, l, t, n))
        except Exception as e:
            print(f"[WARN] 既存CSV読込失敗: {path} / {e}")

    for path in sorted(glob.glob(os.path.join(base_dir, f"{EXCEL_BASE_PREFIX}[0-9]*.xlsx"))):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                s.add((str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()))
        except Exception as e:
            print(f"[WARN] 既存XLSX読込失敗: {path} / {e}")

    return s

def read_teams_from_file(path: str) -> List[Tuple[str, str, str, str]]:
    out: List[Tuple[str, str, str, str]] = []

    KEYMAP = {
        "国": ["国", "country", "Country"],
        "リーグ": ["リーグ", "league", "League"],
        "チーム": ["チーム", "team", "Team"],
        "チームリンク": ["チームリンク", "チームURL", "チームurl", "href", "link", "url", "チームリンクURL"],
    }

    def norm(s: str) -> str:
        return (s or "").strip()

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return out

            resolved: Dict[str, str] = {}
            fields = [h.strip() for h in reader.fieldnames if h]
            for canonical, aliases in KEYMAP.items():
                for a in aliases:
                    if a in fields:
                        resolved[canonical] = a
                        break

            use_fallback = any(k not in resolved for k in ("国", "リーグ", "チーム", "チームリンク"))
            if use_fallback:
                f.seek(0)
                r = csv.reader(f)
                _header = next(r, [])
                for row in r:
                    if not row:
                        continue
                    c = norm(row[0] if len(row) > 0 else "")
                    l = norm(row[1] if len(row) > 1 else "")
                    t = norm(row[2] if len(row) > 2 else "")
                    h = norm(row[3] if len(row) > 3 else "")
                    if c and l and t and h:
                        out.append((c, l, t, h))
                return out

            for row in reader:
                c = norm(row.get(resolved["国"], ""))
                l = norm(row.get(resolved["リーグ"], ""))
                t = norm(row.get(resolved["チーム"], ""))
                h = norm(row.get(resolved["チームリンク"], ""))
                if c and l and t and h:
                    out.append((c, l, t, h))

    except Exception as e:
        print(f"[WARN] チームCSV読込失敗: {path} / {e}")

    return out

def build_targets_from_contains() -> Dict[str, Set[str]]:
    out: Dict[str, Set[str]] = {}
    for ent in CONTAINS_LIST:
        c, l = [x.strip() for x in ent.split(":", 1)]
        out.setdefault(c, set()).add(l)
    return out

def extract_country_leagues_map(json_path: str) -> Dict[str, Set[str]]:
    p = Path(json_path).expanduser()
    if not p.exists():
        return {}

    data = json.loads(p.read_text(encoding="utf-8"))
    out: Dict[str, Set[str]] = {}

    def norm(s) -> str:
        return re.sub(r"\s+", " ", str(s or "")).strip()

    def add(country: str, league: str):
        c = norm(country)
        l = norm(league)
        if c and l:
            out.setdefault(c, set()).add(l)

    def walk(obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if isinstance(v, list) and all(isinstance(x, (str, int, float)) for x in v):
                    for lv in v:
                        add(k, lv)

            if "country" in obj and "league" in obj:
                add(obj.get("country"), obj.get("league"))

            for kk in ("items", "data", "results", "country_league"):
                if kk in obj:
                    walk(obj[kk])

            for v in obj.values():
                if isinstance(v, (dict, list)):
                    walk(v)

        elif isinstance(obj, list):
            for x in obj:
                walk(x)

    walk(data)
    return out

def list_teamdata_csv_paths() -> List[str]:
    paths = glob.glob(TEAMDATA_GLOB)
    if not paths:
        return []

    def extract_num(p: str) -> int:
        base = os.path.basename(p)
        m = re.search(rf"^{re.escape(TEAMDATA_PREFIX)}(\d+)\.csv$", base)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return 10**18
        return 10**18

    return sorted(paths, key=lambda p: (extract_num(p), os.path.basename(p)))

def group_by_league(rows: List[Tuple[str, str, str, str]]) -> Dict[Tuple[str, str], List[Tuple[str, str, str, str]]]:
    mp: Dict[Tuple[str, str], List[Tuple[str, str, str, str]]] = {}
    for r in rows:
        mp.setdefault((r[0], r[1]), []).append(r)
    return mp

# =========================
# Playwright ヘルパ
# =========================
global_nav_sema = asyncio.Semaphore(GLOBAL_NAV_CONCURRENCY)
playerdetail_sema = asyncio.Semaphore(GLOBAL_PLAYERDETAIL_CONCURRENCY)

_nav_gap_lock = asyncio.Lock()
_last_nav_ts = 0.0

async def _respect_nav_gap():
    global _last_nav_ts
    async with _nav_gap_lock:
        loop = asyncio.get_running_loop()
        now = loop.time()
        due = _last_nav_ts + (NAV_GAP_MS / 1000.0)
        if now < due:
            await asyncio.sleep(due - now)
        _last_nav_ts = loop.time()

async def make_context(browser):
    ctx = await browser.new_context(
        locale="ja-JP",
        timezone_id="Asia/Tokyo",
        user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"),
        viewport={"width": 1280, "height": 800},
        ignore_https_errors=True,
        java_script_enabled=True,
    )
    ctx.set_default_timeout(OP_TIMEOUT)
    ctx.set_default_navigation_timeout(NAV_TIMEOUT)

    async def _route(route):
        req = route.request
        rtype = req.resource_type
        url = req.url.lower()
        if rtype in {"image", "stylesheet", "font", "media", "beacon"}:
            return await route.abort()
        if any(k in url for k in BLOCKED_HOST_KEYWORDS):
            return await route.abort()
        return await route.continue_()

    await ctx.route("**/*", _route)
    return ctx

async def safe_goto(page, url: str, selector_to_wait: Optional[str],
                    nav_timeout=NAV_TIMEOUT, sel_timeout=SEL_TIMEOUT, tries=3) -> bool:
    last = None
    for i in range(tries):
        try:
            async with global_nav_sema:
                await _respect_nav_gap()
                await page.goto(url, wait_until="domcontentloaded", timeout=nav_timeout)
            if selector_to_wait:
                await page.wait_for_selector(selector_to_wait, timeout=sel_timeout, state="attached")
            return True
        except Exception as e:
            last = e
            await asyncio.sleep(0.5 * (i + 1))
    print(f"Page.goto 最終失敗: {url} / {last}")
    return False

# =========================
# Page プール（チーム内）
# =========================
class PagePool:
    def __init__(self, ctx, size: int):
        self.ctx = ctx
        self.sema = asyncio.Semaphore(size)
        self.cache: List = []
        self.lock = asyncio.Lock()

    async def acquire(self):
        await self.sema.acquire()
        async with self.lock:
            if self.cache:
                return self.cache.pop()
        return await self.ctx.new_page()

    async def release(self, page):
        async with self.lock:
            self.cache.append(page)
        self.sema.release()

    async def close(self):
        async with self.lock:
            while self.cache:
                p = self.cache.pop()
                try:
                    await p.close()
                except Exception:
                    pass

# =========================
# Excel Writer（専用タスク）※ここが重複回避の肝
# =========================
class ExcelWriter:
    """
    重要:
    - 既存ファイルからロードした existing_keys に加えて
      「Writer が実際に書いた瞬間」に existing_keys を更新する
    => 途中停止でも再実行時に確実に二重書きしない
    """
    def __init__(self, base_dir: str, prefix: str, max_records: int,
                 existing_keys: Set[Tuple[str, str, str, str]], keys_lock: asyncio.Lock):
        self.base_dir = base_dir
        self.prefix = prefix
        self.max_records = max_records
        self.existing_keys = existing_keys
        self.keys_lock = keys_lock

        self.q: asyncio.Queue = asyncio.Queue()
        self._task: Optional[asyncio.Task] = None

        self.wb = None
        self.ws = None
        self.seq = -1
        self.rows_in_current = 0
        self._dirty_since_last_save = 0

    def _scan_existing_max_seq(self) -> int:
        existing = sorted(glob.glob(os.path.join(self.base_dir, f"{self.prefix}[0-9]*.xlsx")))
        if not existing:
            return 0
        def num(p):
            m = re.search(rf"{re.escape(self.prefix)}(\d+)\.xlsx$", os.path.basename(p))
            return int(m.group(1)) if m else 0
        return max((num(p) for p in existing), default=0)

    def _open_initial(self):
        max_seq = self._scan_existing_max_seq()
        start_seq = 1 if max_seq <= 0 else max_seq

        for seq in range(start_seq, start_seq + 2):
            path = os.path.join(self.base_dir, f"{self.prefix}{seq}.xlsx")
            if os.path.exists(path):
                wb = load_workbook(path)
                ws = wb.active
                used = max(ws.max_row - 1, 0)
                if used < self.max_records:
                    self.wb, self.ws, self.seq = wb, ws, seq
                    self.rows_in_current = used
                    return

        new_seq = 1 if max_seq <= 0 else (max_seq + 1)
        self.seq = new_seq
        self.wb = None
        self.ws = None
        self.rows_in_current = 0

    async def _rotate_new(self, seq_new: int):
        path = os.path.join(self.base_dir, f"{self.prefix}{seq_new}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append([
            "国","リーグ","所属チーム","選手名","ポジション","背番号",
            "得点数","年齢","誕生日","市場価値","ローン保有元",
            "契約期限","顔写真","故障情報","データ取得時間"
        ])
        self.wb, self.ws, self.seq = wb, ws, seq_new
        self.rows_in_current = 0
        await self._save_now()
        print(f"[WRITER] Rotate -> {os.path.basename(path)}")

    async def _save_now(self):
        path = os.path.join(self.base_dir, f"{self.prefix}{self.seq}.xlsx")
        await asyncio.to_thread(self.wb.save, path)
        self._dirty_since_last_save = 0

    def enqueue_team_rows(self, rows: List[List[object]], label: str):
        self.q.put_nowait((rows, label))

    async def run(self):
        self._open_initial()
        if self.wb is None:
            await self._rotate_new(self.seq)

        while True:
            try:
                rows, label = await asyncio.wait_for(self.q.get(), timeout=SAVE_INTERVAL_SEC)
            except asyncio.TimeoutError:
                if self._dirty_since_last_save > 0:
                    await self._save_now()
                continue

            if rows is None:
                if self._dirty_since_last_save > 0:
                    await self._save_now()
                break

            # ---- ここで「実際に書く直前」にも既存キーを見て二重防止（最終防衛線） ----
            filtered: List[List[object]] = []
            async with self.keys_lock:
                for r in rows:
                    # ヘッダ: ["国","リーグ","所属チーム","選手名",...]
                    c = str(r[0]).strip()
                    l = str(r[1]).strip()
                    t = str(r[2]).strip()
                    n = str(r[3]).strip()
                    key = (c, l, t, n)
                    if key in self.existing_keys:
                        continue
                    # 「書く予定」として先に登録（これで直後に落ちても次回はスキップ）
                    self.existing_keys.add(key)
                    filtered.append(r)

            if not filtered:
                if LOG_EVERY_ENQUEUE:
                    print(f"[WRITER] SKIP {label}: 0 new rows (all existed)")
                continue

            offset = 0
            while offset < len(filtered):
                remain = self.max_records - self.rows_in_current
                chunk = filtered[offset: offset + remain]

                for r in chunk:
                    self.ws.append(r)

                self.rows_in_current += len(chunk)
                self._dirty_since_last_save += len(chunk)
                offset += len(chunk)

                if self.rows_in_current >= self.max_records:
                    await self._save_now()
                    await self._rotate_new(self.seq + 1)

            if LOG_EVERY_ENQUEUE:
                print(f"[WRITER] APPEND {label}: +{len(filtered)} rows -> file #{self.seq} (cur={self.rows_in_current})")

            if self._dirty_since_last_save >= SAVE_EVERY_N_ROWS:
                await self._save_now()

    async def start(self):
        if self._task is None:
            self._task = asyncio.create_task(self.run())

    async def stop(self):
        await self.q.put((None, "STOP"))
        if self._task:
            await self._task

# =========================
# スクレイピング
# =========================
async def scrape_player_detail_with_page(page, full_url: str) -> Dict[str, str]:
    age = birth = mv = contract = loan = img = inj = "N/A"

    ok = await safe_goto(page, full_url, "div.playerHeader__wrapper div.playerInfoItem")
    if not ok:
        return {
            "age_text": age, "birth_date": birth, "market_value": mv,
            "loan_info": loan, "contract_date": contract,
            "img_url": img, "injury_text": inj
        }

    try:
        el = await page.query_selector(
            'div[data-testid="wcl-assetContainerBoxed-xl"] img, div[data-testid="wcl-assetContainerBoxed-XL"] img'
        )
        if el:
            img = await el.get_attribute("src") or "N/A"
    except Exception:
        pass

    try:
        items = await page.query_selector_all("div.playerHeader__wrapper div.playerInfoItem")
        for it in items:
            spans = await it.query_selector_all("span")
            if not spans:
                continue

            label = (await spans[0].text_content() or "").strip()
            label = re.sub(r"\s+", " ", label).replace(":", "").strip()

            values = []
            for sp in spans[1:]:
                tx = (await sp.text_content() or "").strip()
                tx = re.sub(r"\s+", " ", tx)
                if tx:
                    values.append(tx)

            if "年齢" in label:
                if len(values) >= 1:
                    age = values[0]
                if len(values) >= 2:
                    birth = values[1].strip().strip("()")
            elif "市場価値" in label:
                if values:
                    mv = values[0]
            elif "契約期限" in label:
                if values:
                    contract = values[0]
            elif "ローン" in label or "ローン元" in label:
                if values:
                    loan = values[0]
                if len(values) >= 2 and contract == "N/A":
                    contract = values[1].replace("期限", "").replace(":", "").strip().strip("()")
    except Exception as e:
        print(f"[PLAYER DETAIL] parse失敗: {full_url} / {e}")

    try:
        inj_svg = await page.query_selector('svg[data-testid="wcl-icon-incidents-injury"]')
        if inj_svg:
            title_el = await inj_svg.query_selector("title")
            if title_el:
                t = (await title_el.text_content() or "").strip()
                if t:
                    inj = re.sub(r"\s+", " ", t)
    except Exception:
        pass

    return {
        "age_text": age, "birth_date": birth, "market_value": mv,
        "loan_info": loan, "contract_date": contract,
        "img_url": img, "injury_text": inj
    }

async def process_team(
    ctx,
    existing_keys: Set[Tuple[str, str, str, str]],
    claimed_keys: Set[Tuple[str, str, str, str]],
    keys_lock: asyncio.Lock,
    writer: ExcelWriter,
    country: str, league: str, team: str, href: str
):
    page = await ctx.new_page()
    pool = PagePool(ctx, size=max(PLAYER_PER_TEAM, 1))
    try:
        team_url  = href if href.startswith("http") else f"https://flashscore.co.jp{href}"
        squad_url = team_url.rstrip("/") + "/squad/"

        ok = await safe_goto(page, squad_url, "div.lineupTable.lineupTable--soccer")
        if not ok:
            print(f"[{country}:{league}:{team}] スカッド取得失敗: goto failed")
            return

        players: List[Tuple[str, str, str, str, str]] = []
        tables = await page.query_selector_all("div.lineupTable.lineupTable--soccer")
        for t in tables or []:
            pos_el = await t.query_selector("div.lineupTable__title")
            position = (await pos_el.text_content()).strip() if pos_el else "N/A"
            rows = await t.query_selector_all("div.lineupTable__row")
            for r in rows:
                jersey_el = await r.query_selector(".lineupTable__cell--jersey")
                jersey = (await jersey_el.text_content()).strip() if jersey_el else "N/A"

                name_el = await r.query_selector(".lineupTable__cell--player .lineupTable__cell--name")
                if not name_el:
                    continue
                name = (await name_el.text_content()).strip()
                href2 = await name_el.get_attribute("href")
                if not href2:
                    continue

                goal_el = await r.query_selector(".lineupTable__cell--goal")
                goals = (await goal_el.text_content()).strip() if goal_el else "N/A"

                players.append((position, jersey, name, href2, goals))

        if not players:
            print(f"[{country}:{league}:{team}] 選手0件")
            return

        BATCH = max(PLAYER_PER_TEAM * 4, 4)

        async def handle(p) -> Optional[List[object]]:
            position, jersey, name, href2, goals = p
            key = (country, league, team, name)

            # 早めに除外（Writer側でも最終防衛するので、ここは高速化用）
            async with keys_lock:
                if key in existing_keys or key in claimed_keys:
                    return None
                claimed_keys.add(key)

            purl = href2 if href2.startswith("http") else f"https://flashscore.co.jp{href2}"

            async with playerdetail_sema:
                tab = await pool.acquire()
                try:
                    detail = await scrape_player_detail_with_page(tab, purl)
                finally:
                    await pool.release(tab)

            return [
                country, league, team, name, position, jersey, goals,
                detail.get("age_text","N/A"), detail.get("birth_date","N/A"),
                detail.get("market_value","N/A"), detail.get("loan_info","N/A"),
                detail.get("contract_date","N/A"), detail.get("img_url","N/A"),
                detail.get("injury_text","N/A"), datetime.datetime.now()
            ]

        rows_to_write: List[List[object]] = []
        for i in range(0, len(players), BATCH):
            got = await asyncio.gather(*(handle(p) for p in players[i:i+BATCH]))
            rows_to_write.extend([r for r in got if r])

        if rows_to_write:
            if LOG_EVERY_ENQUEUE:
                print(f"[ENQUEUE] {country} / {league} / {team}: {len(rows_to_write)} 件をWriterへ")
            writer.enqueue_team_rows(rows_to_write, f"{country}/{league}/{team}")

        print(f"[TEAM-DONE] {country} / {league} / {team}")

    except Exception as e:
        print(f"[{country}:{league}:{team}] 処理中エラー: {e}")
    finally:
        try:
            await pool.close()
        except Exception:
            pass
        try:
            await page.close()
        except Exception:
            pass

async def process_league(
    ctx,
    existing_keys: Set[Tuple[str, str, str, str]],
    claimed_keys: Set[Tuple[str, str, str, str]],
    keys_lock: asyncio.Lock,
    writer: ExcelWriter,
    league_key: Tuple[str, str],
    teams: List[Tuple[str, str, str, str]],
):
    country, league = league_key
    print(f"[LEAGUE-START] {country} / {league} teams={len(teams)}")

    for i in range(0, len(teams), TEAMS_PER_LEAGUE_CONCURRENCY):
        chunk = teams[i:i + TEAMS_PER_LEAGUE_CONCURRENCY]
        results = await asyncio.gather(*[
            process_team(ctx, existing_keys, claimed_keys, keys_lock, writer, c, l, team, href)
            for (c, l, team, href) in chunk
        ], return_exceptions=True)

        for r in results:
            if isinstance(r, Exception):
                print(f"[WARN] チームタスク例外: {country}/{league} / {r}")

    print(f"[LEAGUE-DONE] {country} / {league}")

# =========================
# 変換
# =========================
def convert_xlsx_to_csv_and_delete_all(base_dir: str, prefix: str) -> None:
    xlsx_paths = sorted(glob.glob(os.path.join(base_dir, f"{prefix}[0-9]*.xlsx")))
    if not xlsx_paths:
        print("[CONVERT] 対象xlsxなし")
        return

    for xlsx_path in xlsx_paths:
        csv_path = re.sub(r"\.xlsx$", ".csv", xlsx_path)
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = wb.active

            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow([("" if v is None else v) for v in row])

            os.remove(xlsx_path)
            print(f"[CONVERT] {os.path.basename(xlsx_path)} -> {os.path.basename(csv_path)}（xlsx削除）")
        except Exception as e:
            print(f"[ERROR] 変換失敗: {xlsx_path} / {e}")

# =========================
# エントリポイント
# =========================
async def main():
    ensure_dirs()
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{now}  TeamMemberData: 開始（リーグ並列 + 10チーム並列 + 重複完全回避）")

    existing_keys = load_existing_player_keys(BASE_DIR)
    print(f"[INIT] 既取得キー: {len(existing_keys)} 件")

    team_rows: List[Tuple[str, str, str, str]] = []
    paths = list_teamdata_csv_paths()
    if not paths:
        print(f"[EXIT] teamData_*.csv が見つかりません: {TEAMDATA_GLOB}")
        return

    for p in paths:
        rows = read_teams_from_file(p)
        print(f"[LOAD] {os.path.basename(p)}: {len(rows)} 行")
        team_rows.extend(rows)

    seen: Set[Tuple[str, str, str, str]] = set()
    uniq_rows: List[Tuple[str, str, str, str]] = []
    for row in team_rows:
        if row not in seen:
            seen.add(row)
            uniq_rows.append(row)

    print(f"[PLAN] 処理チーム数（重複除去後）: {len(uniq_rows)}")

    p_json = Path(B001_JSON_PATH).expanduser()
    if p_json.exists():
        json_targets = extract_country_leagues_map(B001_JSON_PATH)
        before = len(uniq_rows)
        uniq_rows = [r for r in uniq_rows if (r[0] in json_targets and r[1] in json_targets[r[0]])]
        print(f"[FILTER] JSON(存在) / チーム {before} -> {len(uniq_rows)} (targets={sum(len(v) for v in json_targets.values())})")

        # JSONが存在するのに0件になったら、JSONの中身/形式が想定と違う可能性が高いので明示終了
        if not uniq_rows:
            print("[EXIT] b001_country_league.json は存在しますが、対象チームが0件です。JSONの国名/リーグ名の表記ゆれ、またはJSON形式を確認してください。")
            return
    else:
        contains_targets = build_targets_from_contains()
        before = len(uniq_rows)
        uniq_rows = [r for r in uniq_rows if (r[0] in contains_targets and r[1] in contains_targets[r[0]])]
        print(f"[FILTER] JSON(なし) / チーム {before} -> {len(uniq_rows)}")

    if not uniq_rows:
        print("[EXIT] フィルタ後の対象チームが0件のため終了します")
        return

    league_map = group_by_league(uniq_rows)
    league_items = list(league_map.items())
    print(f"[PLAN] 対象リーグ数: {len(league_items)}")

    # 既存キー/claimedキーは同じロックで守る
    keys_lock = asyncio.Lock()
    claimed_keys: Set[Tuple[str, str, str, str]] = set()

    # Writer（Writerが書いた瞬間に existing_keys を更新する）
    writer = ExcelWriter(BASE_DIR, EXCEL_BASE_PREFIX, EXCEL_MAX_RECORDS, existing_keys, keys_lock)
    await writer.start()

    league_sema = asyncio.Semaphore(LEAGUE_CONCURRENCY)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--disable-background-timer-throttling",
                "--disable-renderer-backgrounding",
                "--disable-backgrounding-occluded-windows",
                "--disable-breakpad",
                "--disable-features=TranslateUI,BackForwardCache,AcceptCHFrame,MediaRouter",
                "--blink-settings=imagesEnabled=false",
            ],
        )
        ctx = await make_context(browser)

        async def run_one_league(lk: Tuple[str, str], teams: List[Tuple[str, str, str, str]]):
            async with league_sema:
                await process_league(ctx, existing_keys, claimed_keys, keys_lock, writer, lk, teams)

        tasks = [run_one_league(lk, teams) for (lk, teams) in league_items]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in results:
            if isinstance(r, Exception):
                print(f"[WARN] リーグタスク例外: {r}")

        await ctx.close()
        await browser.close()

    await writer.stop()

    convert_xlsx_to_csv_and_delete_all(TEAMS_BY_LEAGUE_DIR, EXCEL_BASE_PREFIX)

    print("TeamMemberData: 完了（重複完全回避）")

if __name__ == "__main__":
    asyncio.run(main())

# -*- coding: utf-8 -*-
"""
Flashscore ライブ試合 -> (stats / summary meta / standings) を取得して Excel に逐次追記する全処理版（停止対策・ログ強化版）

✅ 重要仕様
- stats の値は「34%（31/90）」など“そのまま”取得（加工しない）
- Flashscore側: "セクション:ラベル" -> canonicalキー -> STAT_KEY_MAP -> HEADER列へ投入
- HEADERとスクレイプが正しい列に入っているか検証ログを出す（VERIFY）

✅ 今回あなたのコードに混入していた致命的バグを解消
- normalize_stats_pairs が二重定義で上書き → 1つに統合
- rb / mid / meta / final_category など未定義 → すべて定義して流れを成立
- scrape_stats_pairs を二重に呼ぶ → 1回に統一
"""

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import time, re, os, pickle, datetime, traceback
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit, urlparse, parse_qs
from typing import Optional, List, Dict, Tuple, Any
import multiprocessing as mp
import queue as pyqueue
import pandas as pd
import openpyxl


# =========================
# 設定
# =========================
SAVE_DIR = "/Users/shiraishitoshio/bookmaker/outputs"
SEQMAP_PATH = os.path.join(SAVE_DIR, "seqmap.pkl")
_MATCH_ROOT_RE = re.compile(r"^(/match/[^/]+/[^/]+/[^/]+/)")

WORKER_TIMEOUT_SEC = 180      # 子プロセス全体（1試合）
NAV_TIMEOUT_MS     = 20000    # safe_goto の 1st try
WAIT_TIMEOUT_MS    = 10000    # wait_for_selector

VERBOSE = True
def log(msg: str):
    if VERBOSE:
        print(msg, flush=True)

BOT_WALL_PAT = r"Just a moment|Access Denied|verify you are human|チェック|確認"


# =========================
# HEADER（Excel列）
# =========================
HEADER = [
    "ホーム順位","試合国及びカテゴリ","試合時間","ホームチーム","ホームスコア","アウェー順位","アウェーチーム",
    "アウェースコア","ホーム期待値","アウェー期待値","ホーム枠内ゴール期待値","アウェー枠内ゴール期待値",
    "ホームボール支配率","アウェーボール支配率","ホームシュート数","アウェーシュート数",
    "ホーム枠内シュート数","アウェー枠内シュート数","ホーム枠外シュート数","アウェー枠外シュート数",
    "ホームブロックシュート","アウェーブロックシュート","ホームビッグチャンス","アウェービッグチャンス",
    "ホームコーナーキック","アウェーコーナーキック","ホームボックス内シュート","アウェーボックス内シュート",
    "ホームボックス外シュート","アウェーボックス外シュート","ホームゴールポスト","アウェーゴールポスト","ホームヘディングゴール","アウェーヘディングゴール",
    "ホームキーパーセーブ","アウェーキーパーセーブ","ホームフリーキック","アウェーフリーキック",
    "ホームオフサイド","アウェーオフサイド","ホームファウル","アウェーファウル",
    "ホームイエローカード","アウェーイエローカード","ホームレッドカード","アウェーレッドカード","ホームスローイン","アウェースローイン",
    "ホーム相手ボックスタッチ","アウェー相手ボックスタッチ","ホームパス","アウェーパス","ホームロングパス","アウェーロングパス","ホームファイナルサードパス","アウェーファイナルサードパス",
    "ホームクロス","アウェークロス","ホームタックル","アウェータックル","ホームクリア","アウェークリア","ホームデュエル勝利数","アウェーデュエル勝利数",
    "ホームインターセプト","アウェーインターセプト",
    "スコア時間","天気","気温","湿度","審判名","ホーム監督名","アウェー監督名","ホームフォーメーション","アウェーフォーメーション",
    "スタジアム","収容人数","観客数","場所","ホームチーム最大得点取得者","アウェーチーム最大得点取得者","ホームチーム最大得点取得者出場状況","アウェーチーム最大得点取得者出場状況",
    "ホームチームホーム得点","ホームチームホーム失点","アウェーチームホーム得点","アウェーチームホーム失点","ホームチームアウェー得点","ホームチームアウェー失点",
    "アウェーチームアウェー得点","アウェーチームアウェー失点","通知フラグ","試合リンク文字列","ゴール時間","選手名","判定結果","ホームチームスタイル","アウェイチームスタイル",
    "ゴール確率","得点予想時間","試合ID","通番","ソート用秒"
]


# =========================
# 統計キー -> HEADER列対応（canonicalキー前提）
# =========================
STAT_KEY_MAP = {
    "アタック:期待値（xG）": ("ホーム期待値", "アウェー期待値"),
    "アタック:枠内ゴール期待値": ("ホーム枠内ゴール期待値", "アウェー枠内ゴール期待値"),
    "ポゼッション:ボール支配率": ("ホームボール支配率", "アウェーボール支配率"),
    "シュート:シュート数": ("ホームシュート数", "アウェーシュート数"),
    "シュート:枠内シュート数": ("ホーム枠内シュート数", "アウェー枠内シュート数"),
    "シュート:枠外シュート数": ("ホーム枠外シュート数", "アウェー枠外シュート数"),
    "シュート:ブロックシュート": ("ホームブロックシュート", "アウェーブロックシュート"),
    "アタック:ビッグチャンス": ("ホームビッグチャンス", "アウェービッグチャンス"),
    "セットプレー:コーナーキック": ("ホームコーナーキック", "アウェーコーナーキック"),
    "シュート:ボックス内シュート": ("ホームボックス内シュート", "アウェーボックス内シュート"),
    "シュート:ボックス外シュート": ("ホームボックス外シュート", "アウェーボックス外シュート"),
    "シュート:ポストヒット": ("ホームゴールポスト", "アウェーゴールポスト"),
    "シュート:ヘディングゴール": ("ホームヘディングゴール", "アウェーヘディングゴール"),

    "ディフェンス:キーパーセーブ": ("ホームキーパーセーブ", "アウェーキーパーセーブ"),
    "ディフェンス:フリーキック": ("ホームフリーキック", "アウェーフリーキック"),
    "ディフェンス:オフサイド": ("ホームオフサイド", "アウェーオフサイド"),
    "ディフェンス:ファウル": ("ホームファウル", "アウェーファウル"),
    "カード:イエローカード": ("ホームイエローカード", "アウェーイエローカード"),
    "カード:レッドカード": ("ホームレッドカード", "アウェーレッドカード"),
    "ディフェンス:スローイン": ("ホームスローイン", "アウェースローイン"),

    "パス:相手ボックスタッチ": ("ホーム相手ボックスタッチ", "アウェー相手ボックスタッチ"),
    "パス:総パス数": ("ホームパス", "アウェーパス"),
    "パス:ロングパス": ("ホームロングパス", "アウェーロングパス"),
    "パス:ファイナルサードパス": ("ホームファイナルサードパス", "アウェーファイナルサードパス"),
    "パス:クロス": ("ホームクロス", "アウェークロス"),

    "ディフェンス:タックル": ("ホームタックル", "アウェータックル"),
    "ディフェンス:クリア": ("ホームクリア", "アウェークリア"),
    "ディフェンス:デュエル勝利": ("ホームデュエル勝利数", "アウェーデュエル勝利数"),
    "ディフェンス:インターセプト": ("ホームインターセプト", "アウェーインターセプト"),
}


# =========================
# Flashscore「ラベル」→ canonicalキー
# =========================
LABEL_TO_CANON: Dict[str, str] = {
    "ゴール期待値（xG）": "アタック:期待値（xG）",
    "枠内ゴール期待値（xGOT）": "アタック:枠内ゴール期待値",

    "ボール支配率": "ポゼッション:ボール支配率",
    "合計シュート": "シュート:シュート数",
    "枠内シュート": "シュート:枠内シュート数",
    "枠外シュート": "シュート:枠外シュート数",
    "シュートブロック": "シュート:ブロックシュート",
    "ボックス内からのシュート": "シュート:ボックス内シュート",
    "ボックス外からのシュート": "シュート:ボックス外シュート",
    "ゴール枠に当たる": "シュート:ポストヒット",
    "ゴール枠に当たるシュート": "シュート:ポストヒット",

    "ビッグチャンス": "アタック:ビッグチャンス",
    "コーナーキック": "セットプレー:コーナーキック",

    "イエローカード": "カード:イエローカード",
    "レッドカード": "カード:レッドカード",
    "ファウル": "ディフェンス:ファウル",
    "オフサイド": "ディフェンス:オフサイド",
    "フリーキック": "ディフェンス:フリーキック",
    "スローイン": "ディフェンス:スローイン",
    "タックル": "ディフェンス:タックル",
    "デュエル勝利数": "ディフェンス:デュエル勝利",
    "クリアリング": "ディフェンス:クリア",
    "クリア": "ディフェンス:クリア",
    "インターセプト": "ディフェンス:インターセプト",
    "キーパーセーブ": "ディフェンス:キーパーセーブ",

    "相手ボックス内でのタッチ": "パス:相手ボックスタッチ",
    "相手ボックス内タッチ": "パス:相手ボックスタッチ",
    "パス": "パス:総パス数",
    "ロングパス": "パス:ロングパス",
    "ファイナルサードでのパス": "パス:ファイナルサードパス",
    "ファイナルサードのパス": "パス:ファイナルサードパス",
    "クロス": "パス:クロス",
}

# 同一ラベルが複数セクションに出る場合、優先順位（主なスタッツを優先）
SECTION_PREFER = ["主なスタッツ", "シュート", "アタック", "パス", "ディフェンス", "ゴールキーパー"]

def _section_rank(section: str) -> int:
    try:
        return SECTION_PREFER.index(section)
    except ValueError:
        return 999


# =========================
# 対象リーグフィルタ（親で使用）
# =========================
CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]
UNDER_LIST  = ["U17", "U18", "U19", "U20", "U21", "U22", "U23", "U24", "U25"]
GENDER_LIST = ["女子"]
EXP_LIST    = ["ポルトガル: リーガ・ポルトガル 2", "イングランド: プレミアリーグ 2", "イングランド: プレミアリーグ U18"]


# =========================
# SEQMAP
# =========================
SEQMAP: Dict[str, int] = {}

def load_seqmap():
    global SEQMAP
    if os.path.exists(SEQMAP_PATH):
        try:
            with open(SEQMAP_PATH, "rb") as f:
                SEQMAP = pickle.load(f) or {}
            log(f"🔁 [SEQ] 読み込み: {SEQMAP_PATH}（件数={len(SEQMAP)}）")
        except Exception as e:
            log(f"⚠️ [SEQ] load失敗: {e}")
            SEQMAP = {}
    else:
        SEQMAP = {}
        log(f"🆕 [SEQ] 新規（{SEQMAP_PATH} が存在しません）")

def save_seqmap():
    try:
        os.makedirs(os.path.dirname(SEQMAP_PATH), exist_ok=True)
        with open(SEQMAP_PATH, "wb") as f:
            pickle.dump(SEQMAP, f)
        log(f"💾 [SEQ] 保存: {SEQMAP_PATH}（件数={len(SEQMAP)}）")
    except Exception as e:
        log(f"⚠️ [SEQ] save失敗: {e}")


# =========================
# util
# =========================
def text_clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def parse_live_time_to_seconds(tstr: str) -> int:
    if not tstr:
        return 0
    t = tstr.strip()
    if "終了" in t or "FT" in t.upper():
        return 5400
    if "前半" in t:
        num = re.sub(r"[^0-9]", "", t)
        return int(num) * 60 if num else 0
    if "後半" in t:
        num = re.sub(r"[^0-9]", "", t)
        return 2700 + int(num) * 60 if num else 2700
    if re.match(r"^\d+\+\d+$", t):
        a, b = t.split("+")
        return (int(a) + int(b)) * 60
    if t.isdigit():
        return int(t) * 60
    return 0


# =========================
# VERIFY（マッピングと埋まりチェック）
# =========================
def verify_header_and_stat_map():
    bad = []
    for k, (hcol, acol) in STAT_KEY_MAP.items():
        if hcol not in HEADER:
            bad.append(("HOME", k, hcol))
        if acol not in HEADER:
            bad.append(("AWAY", k, acol))
    if bad:
        log("❌ [VERIFY] STAT_KEY_MAP が参照している列が HEADER に無い")
        for side, k, col in bad:
            log(f"   - {side} key='{k}' col='{col}'")
        raise RuntimeError("STAT_KEY_MAP column mismatch with HEADER")
    log("✅ [VERIFY] STAT_KEY_MAP の列名は HEADER と整合しています")

def verify_stats_mapping(keys_title: str, stats_pairs: Dict[str, Tuple[str, str]]):
    scraped_keys = set(stats_pairs.keys())
    mapped_keys  = set(STAT_KEY_MAP.keys())

    direct_hit   = sorted(scraped_keys & mapped_keys)
    miss_scraped = sorted(scraped_keys - mapped_keys)

    log(f"✅ [VERIFY:{keys_title}] 直接一致キー数: {len(direct_hit)}")
    for k in direct_hit[:40]:
        hv, av = stats_pairs.get(k, ("",""))
        log(f"   ✓ {k} = ({hv}, {av})")

    log(f"⚠️ [VERIFY:{keys_title}] マップに無いスクレイプキー数: {len(miss_scraped)}")
    for k in miss_scraped[:60]:
        hv, av = stats_pairs.get(k, ("",""))
        log(f"   ? {k} = ({hv}, {av})")

def verify_row_filled(rb: "RowBuilder", limit: int = 120):
    filled = []
    for col in HEADER:
        v = rb.d.get(col, "")
        if v not in ("", None):
            filled.append((col, v))
    log(f"✅ [VERIFY] row に値が入った列数: {len(filled)}")
    for col, v in filled[:limit]:
        log(f"   • {col} = {v}")


# =========================
# URL builders
# =========================
def extract_mid(any_url: str) -> Optional[str]:
    if not any_url:
        return None
    qs = parse_qs(urlparse(any_url).query)
    return qs.get("mid", [None])[0]

def _match_root_from_any_url(any_url: str) -> Tuple[Tuple[str, str, str], str]:
    parts = urlsplit(any_url)
    path = parts.path or ""
    m = _MATCH_ROOT_RE.search(path)
    if not m:
        raise ValueError(f"match root not found: {any_url}")
    match_root_path = m.group(1)
    mid = extract_mid(any_url)
    return (parts.scheme, parts.netloc, match_root_path), (mid or "")

def build_stats_url(any_url: str) -> str:
    (scheme, netloc, root), mid = _match_root_from_any_url(any_url)
    path = root + "summary/stats/0/"
    query = f"mid={mid}" if mid else ""
    return urlunsplit((scheme, netloc, path, query, ""))

def build_summary_url(any_url: str) -> str:
    (scheme, netloc, root), mid = _match_root_from_any_url(any_url)
    path = root + "summary/"
    query = f"mid={mid}" if mid else ""
    return urlunsplit((scheme, netloc, path, query, ""))

def build_live_standings_url(any_url: str) -> str:
    (scheme, netloc, root), mid = _match_root_from_any_url(any_url)
    path = root + "standings/live-standings/"
    query = f"mid={mid}" if mid else ""
    return urlunsplit((scheme, netloc, path, query, ""))

def build_overall_standings_url(any_url: str) -> str:
    (scheme, netloc, root), mid = _match_root_from_any_url(any_url)
    path = root + "standings/standings/overall/"
    query = f"mid={mid}" if mid else ""
    return urlunsplit((scheme, netloc, path, query, ""))


# =========================
# STOP対策：safe_goto（★最重要）
# =========================
def safe_goto(pg, url: str, timeout_ms: int = NAV_TIMEOUT_MS, tag: str = "") -> bool:
    ttag = f" {tag}" if tag else ""
    log(f"🧭 [GOTO]{ttag} try1 commit: {url}")
    try:
        pg.goto(url, timeout=timeout_ms, wait_until="commit")
        return True
    except PWTimeout as e:
        log(f"⏱️ [GOTO]{ttag} try1 timeout: {e}")
    except Exception as e:
        log(f"⚠️ [GOTO]{ttag} try1 error: {type(e).__name__}: {e}")

    try:
        log(f"🛑 [GOTO]{ttag} window.stop()")
        pg.evaluate("() => window.stop()")
    except Exception as e:
        log(f"⚠️ [GOTO]{ttag} window.stop error: {type(e).__name__}: {e}")

    log(f"🧭 [GOTO]{ttag} try2 domcontentloaded: {url}")
    try:
        pg.goto(url, timeout=8000, wait_until="domcontentloaded")
        return True
    except PWTimeout as e:
        log(f"⏱️ [GOTO]{ttag} try2 timeout: {e}")
    except Exception as e:
        log(f"⚠️ [GOTO]{ttag} try2 error: {type(e).__name__}: {e}")

    try:
        log(f"🧼 [GOTO]{ttag} about:blank")
        pg.goto("about:blank", timeout=3000, wait_until="commit")
    except Exception:
        pass

    log(f"🧭 [GOTO]{ttag} try3 domcontentloaded(after blank): {url}")
    try:
        pg.goto(url, timeout=8000, wait_until="domcontentloaded")
        return True
    except Exception as e:
        log(f"❌ [GOTO]{ttag} try3 failed: {type(e).__name__}: {e}")
        return False


# =========================
# consent / cmp
# =========================
def kill_onetrust(page):
    try:
        btn = page.locator("#onetrust-accept-btn-handler")
        if btn.count():
            btn.click(timeout=1200, force=True)
    except:
        pass
    try:
        page.evaluate("""
        () => {
          const ids = ["onetrust-consent-sdk", "onetrust-banner-sdk"];
          ids.forEach(id => document.getElementById(id)?.remove());
          document.querySelectorAll(".ot-sdk-container, .ot-sdk-row, .otOverlay, .ot-pc-footer, .ot-pc-header")
            .forEach(el => el.remove());
        }""", timeout=1200)
    except:
        pass

def kill_consent_banners(page):
    kill_onetrust(page)
    candidates = [
        "button:has-text('すべて拒否する')",
        "button:has-text('全て拒否する')",
        "button:has-text('拒否する')",
        "button:has-text('同意します')",
        "button:has-text('同意する')",
        "button:has-text('Reject all')",
        "button:has-text('Accept all')",
        "[role='button']:has-text('すべて拒否する')",
        "[role='button']:has-text('同意します')",
    ]
    for sel in candidates:
        try:
            b = page.locator(sel).first
            if b.count() and b.is_visible(timeout=500):
                b.click(timeout=1200, force=True)
                break
        except:
            pass

    try:
        page.evaluate("""
        () => {
          const kill = [
            "#qc-cmp2-container",
            "#didomi-host",
            "#sp_message_container_",
            ".fc-consent-root",
            ".message-component",
            ".pmConsentWall"
          ];
          kill.forEach(s => document.querySelectorAll(s).forEach(el => el.remove()));
        }""", timeout=1200)
    except:
        pass

def is_bot_wall(pg) -> bool:
    try:
        return pg.locator(f"text=/{BOT_WALL_PAT}/i").first.is_visible(timeout=900)
    except:
        return False


# =========================
# route blocking（軽量化）
# =========================
def setup_route_blocking(ctx):
    def _route(route):
        try:
            rtype = route.request.resource_type
            url = (route.request.url or "").lower()
            if rtype in ("image", "media", "font"):
                return route.abort()
            if any(x in url for x in ("doubleclick", "googlesyndication", "adservice", "adsystem", "taboola", "outbrain")):
                return route.abort()
        except:
            pass
        return route.continue_()
    ctx.route("**/*", _route)


# =========================
# 試合ページ：チーム名・スコア・時間
# =========================
def get_home_away_names(pg) -> Tuple[str, str]:
    try:
        cont = pg.locator("div.duelParticipant__container").first
        if not cont.count():
            cont = pg
        h = cont.locator(".duelParticipant__home a.participant__participantName").first
        a = cont.locator(".duelParticipant__away a.participant__participantName").first
        home = text_clean(h.text_content()) if h.count() else ""
        away = text_clean(a.text_content()) if a.count() else ""
        if not home:
            img = cont.locator(".duelParticipant__home img.participant__image").first
            home = text_clean(img.get_attribute("alt") or "")
        if not away:
            img = cont.locator(".duelParticipant__away img.participant__image").first
            away = text_clean(img.get_attribute("alt") or "")
        return home, away
    except:
        return "", ""

def get_scores(pg) -> Tuple[str, str]:
    def _clean(s: str) -> str:
        return re.sub(r"\s+", "", (s or "").replace("\u00A0", " ")).strip()

    def _from_container(el):
        try:
            spans = el.locator("span")
            vals = []
            for i in range(spans.count()):
                sp = spans.nth(i)
                cls = (sp.get_attribute("class") or "")
                if "divider" in cls:
                    continue
                txt = _clean(sp.text_content() or "")
                if re.fullmatch(r"\d+", txt):
                    vals.append(txt)
            if len(vals) >= 2:
                return vals[0], vals[-1]
        except:
            pass

        try:
            t = _clean(el.inner_text() or "")
            m = re.search(r"(\d+)\s*[\-\u2212\u2012\u2013\u2014\u2015]\s*(\d+)", t)
            if m:
                return m.group(1), m.group(2)
        except:
            pass
        return "", ""

    try:
        fx = pg.locator(".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .fixedScore").first
        if fx.count():
            h, a = _from_container(fx)
            if h and a:
                return h, a
    except:
        pass

    try:
        live = pg.locator("div.detailScore__wrapper.detailScore__live").first
        if live.count():
            h, a = _from_container(live)
            if h and a:
                return h, a
    except:
        pass

    try:
        wrap = pg.locator("div.detailScore__wrapper").first
        if wrap.count():
            h, a = _from_container(wrap)
            if h and a:
                return h, a
    except:
        pass

    return "", ""

def get_match_time_text(pg) -> str:
    sels = [
        ".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .fixedHeaderDuel__detailStatus",
        "div.detailScore__status .fixedHeaderDuel__detailStatus",
        "div.detailScore__status .eventAndAddedTime .eventTime",
        ".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .eventAndAddedTime .eventTime",
        "div.detailScore__status",
        "[data-testid='wcl-time']",
    ]
    for s in sels:
        try:
            el = pg.locator(f"{s} >> visible=true").first
            if el.count():
                txt = (el.text_content() or "").strip().replace("\u00A0", " ")
                if txt:
                    return txt
        except:
            pass
    return ""


# =========================
# RowBuilder
# =========================
class RowBuilder:
    def __init__(self, header: List[str], match_tag: str = ""):
        self.header = header
        self.d = {col: "" for col in header}
        self.match_tag = match_tag

    def put(self, key: str, value: Any):
        if key not in self.d:
            return
        self.d[key] = "" if value is None else value

    def put_pair(self, home_key: str, away_key: str, home_val: Any, away_val: Any):
        self.put(home_key, home_val)
        self.put(away_key, away_val)

def apply_stats_to_row(rb: RowBuilder, canon_pairs: Dict[str, Tuple[str, str]]):
    for stat_key, (hcol, acol) in STAT_KEY_MAP.items():
        if stat_key in canon_pairs:
            hv, av = canon_pairs[stat_key]
            rb.put_pair(hcol, acol, hv, av)


# =========================
# stats：raw取得（加工なし）
# =========================
STAT_ROW_SELECTOR = "[data-testid='wcl-statistics']"

def goto_statistics_page(pg) -> bool:
    stats_url = build_stats_url(pg.url)

    kill_consent_banners(pg)

    if "/summary/stats/0/" not in pg.url:
        log(f"➡️ [STATS] goto: {stats_url}")
        ok = safe_goto(pg, stats_url, timeout_ms=NAV_TIMEOUT_MS, tag="STATS")
        if not ok:
            log("⚠️ [STATS] goto失敗（safe_gotoでも）")
            return False
        kill_consent_banners(pg)

    if is_bot_wall(pg):
        log("🧱 [STATS] BOT壁っぽい → statsスキップ")
        return False

    log("⏳ [STATS] wait_for_selector wcl-statistics ...")
    for attempt in range(2):
        try:
            pg.wait_for_selector(STAT_ROW_SELECTOR, timeout=WAIT_TIMEOUT_MS)
            log("✅ [STATS] statistics selector appeared")
            return True
        except Exception as e:
            log(f"⚠️ [STATS] wait_for_selector failed({attempt+1}/2): {type(e).__name__}: {e}")
            kill_consent_banners(pg)

    log("⚠️ [STATS] 統計出ない（同意除去後も）")
    return False

def scrape_stats_raw(pg) -> Dict[str, List[str]]:
    """
    out: {"主なスタッツ:パス": ["34%（31/90）","51%（24/47）"], ...}
    ※ 値は加工しない
    """
    if not goto_statistics_page(pg):
        return {}

    log("📊 [STATS] JS extraction start")
    try:
        pg.wait_for_selector('[data-testid="wcl-statistics"]', timeout=WAIT_TIMEOUT_MS)
    except Exception:
        log("⚠️ [STATS] statistics selector not found")
        return {}

    stats = pg.evaluate("""
    () => {
      const out = {};
      const wrapper = document.querySelector('.sectionsWrapper');
      if (!wrapper) return out;

      const clean = (s) =>
        (s || '')
          .replace(/\\u00A0/g, ' ')
          .replace(/\\s+/g, ' ')
          .trim();

      const pickValue = (cell) => {
        if (!cell) return '';
        const spans = cell.querySelectorAll("span[data-testid='wcl-scores-simple-text-01']");
        if (!spans || spans.length === 0) return '';
        const parts = [];
        for (const sp of spans) {
          const t = clean(sp.textContent);
          if (t) parts.push(t);
        }
        // ★順番そのまま連結（例: "34%（31/90）"）
        return parts.join(' ');
      };

      const sections = wrapper.querySelectorAll('.section');
      for (const sec of sections) {
        const sectionTitle = clean(sec.querySelector('.sectionHeader')?.textContent);
        if (!sectionTitle) continue;

        const rows = sec.querySelectorAll('[data-testid="wcl-statistics"]');
        for (const row of rows) {
          const label = clean(
            row.querySelector('[data-testid="wcl-statistics-category"] span')?.textContent
          );
          if (!label) continue;

          const values = row.querySelectorAll('[data-testid="wcl-statistics-value"]');
          if (!values || values.length < 2) continue;

          const home = pickValue(values[0]);
          const away = pickValue(values[values.length - 1]);

          out[`${sectionTitle}:${label}`] = [home, away];
        }
      }
      return out;
    }
    """) or {}

    log(f"📊 [STATS] extracted items = {len(stats)}")
    return stats


# =========================
# stats：raw -> canonical（label→canon, セクション優先）
# =========================
def normalize_stats_raw_to_canon(stats_raw: Dict[str, Any]) -> Dict[str, Tuple[str, str]]:
    """
    stats_raw: {"主なスタッツ:合計シュート":[home,away], ...}
    -> {"シュート:シュート数":(home,away), ...}
    """
    if not stats_raw:
        return {}

    best_by_label: Dict[str, Tuple[int, Tuple[str, str]]] = {}

    for k, v in stats_raw.items():
        if not isinstance(v, (list, tuple)) or len(v) < 2:
            continue
        if ":" not in k:
            continue

        section, label = k.split(":", 1)
        section = (section or "").strip()
        label = (label or "").strip()

        canon = LABEL_TO_CANON.get(label)
        if not canon:
            continue

        hv, av = str(v[0]), str(v[1])
        rank = _section_rank(section)

        if (label not in best_by_label) or (rank < best_by_label[label][0]):
            best_by_label[label] = (rank, (hv, av))

    out: Dict[str, Tuple[str, str]] = {}
    for label, (_, pair) in best_by_label.items():
        canon = LABEL_TO_CANON.get(label)
        if canon:
            out[canon] = pair
    return out

def dump_unmapped_stats(stats_raw: Dict[str, Any]):
    labels = []
    for k in (stats_raw or {}).keys():
        if ":" in k:
            labels.append(k.split(":", 1)[1].strip())
    unmapped = sorted({lb for lb in labels if lb and lb not in LABEL_TO_CANON})
    if unmapped:
        log(f"🧪 [STATS] unmapped labels ({len(unmapped)}): {unmapped}")


# =========================
# meta（summary）
# =========================
def get_match_meta(pg) -> Dict[str, str]:
    meta: Dict[str, str] = {}
    summary_url = build_summary_url(pg.url)

    if "/summary/" not in pg.url or "/summary/stats" in pg.url:
        log(f"➡️ [META] goto: {summary_url}")
        ok = safe_goto(pg, summary_url, timeout_ms=NAV_TIMEOUT_MS, tag="META")
        if not ok:
            meta["試合時間"] = get_match_time_text(pg)
            meta["取得時刻"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return meta
        kill_consent_banners(pg)

    if is_bot_wall(pg):
        meta["試合時間"] = get_match_time_text(pg)
        meta["取得時刻"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return meta

    try:
        pg.wait_for_selector(
            "ol li span[itemprop='name'], [data-testid*='breadcrumbs'], .tournamentHeader, .duelParticipant",
            timeout=WAIT_TIMEOUT_MS
        )
    except:
        pass

    country = ""
    league = ""

    try:
        crumb_txts = []
        cand = pg.locator("ol li span[itemprop='name'], [data-testid*='breadcrumbs'] span[itemprop='name']")
        for i in range(cand.count()):
            t = text_clean(cand.nth(i).text_content() or "")
            if t:
                crumb_txts.append(t)
        if len(crumb_txts) >= 2:
            country = crumb_txts[1]
        if len(crumb_txts) >= 3:
            league = crumb_txts[2]
    except:
        pass

    if not country or not league:
        try:
            c = pg.locator(".tournamentHeader__country, .tournamentHeader__category, [class*='tournamentHeader__category']").first
            t = text_clean(c.text_content() or "")
            if t:
                if ":" in t and (not country or not league):
                    a, b = [x.strip() for x in t.split(":", 1)]
                    if not country:
                        country = a
                    if not league:
                        league = b
                elif not country:
                    country = t
        except:
            pass

        try:
            l = pg.locator(".tournamentHeader__name, [class*='tournamentHeader__name']").first
            t = text_clean(l.text_content() or "")
            if t and not league:
                league = t
        except:
            pass

    meta["国"] = country
    meta["リーグ"] = league

    label_aliases = {
        "レフェリー": "レフェリー",
        "審判": "レフェリー",
        "主審": "レフェリー",
        "開催地": "開催地",
        "スタジアム": "開催地",
        "会場": "開催地",
        "収容人数": "収容人数",
        "キャパシティ": "収容人数",
        "観客": "観客",
        "観客数": "観客",
        "参加": "観客",
    }
    want = set(label_aliases.keys())

    def put_meta(label: str, value: str):
        label = text_clean(label).replace(":", "").replace("：", "").strip()
        value = text_clean(value)
        if not label or not value:
            return
        norm = label_aliases.get(label, label)
        if norm not in meta:
            meta[norm] = value

    try:
        dts = pg.locator("dl dt")
        n = dts.count()
        for i in range(n):
            dt = dts.nth(i)
            lab = text_clean(dt.text_content() or "")
            if lab not in want:
                continue
            dd = dt.locator("xpath=following-sibling::dd[1]").first
            val = text_clean(dd.text_content() or "") if dd.count() else ""
            put_meta(lab, val)
    except:
        pass

    meta["試合時間"] = get_match_time_text(pg)
    meta["取得時刻"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return meta


# =========================
# standings（順位）
# =========================
def goto_standings_page(pg) -> Optional[str]:
    url1 = build_live_standings_url(pg.url)
    url2 = build_overall_standings_url(pg.url)

    log(f"➡️ [RANK] goto(live): {url1}")
    ok = safe_goto(pg, url1, timeout_ms=NAV_TIMEOUT_MS, tag="RANK-LIVE")
    if ok:
        kill_consent_banners(pg)
        try:
            pg.wait_for_selector(".ui-table__body .ui-table__row", timeout=WAIT_TIMEOUT_MS)
            return url1
        except Exception as e:
            log(f"⚠️ [RANK] live wait failed: {type(e).__name__}: {e}")

    log(f"➡️ [RANK] goto(overall): {url2}")
    ok = safe_goto(pg, url2, timeout_ms=NAV_TIMEOUT_MS, tag="RANK-OVERALL")
    if ok:
        kill_consent_banners(pg)
        try:
            pg.wait_for_selector(".ui-table__body .ui-table__row", timeout=WAIT_TIMEOUT_MS)
            return url2
        except Exception as e:
            log(f"⚠️ [RANK] overall wait failed: {type(e).__name__}: {e}")

    return None

def get_match_standings(pg, home_name: str, away_name: str) -> Dict[str, Any]:
    url = goto_standings_page(pg)
    if not url:
        return {}

    rows = pg.locator(".ui-table__body .ui-table__row")
    try:
        n = rows.count()
    except:
        n = 0
    if n == 0:
        return {}

    table = {}
    for i in range(n):
        r = rows.nth(i)
        rank_txt = text_clean(r.locator(".table__cell--rank .tableCellRank").first.text_content() or "")
        team_name = text_clean(r.locator(".table__cell--participant .tableCellParticipant__name").first.text_content() or "")
        pts_txt = text_clean(r.locator(".table__cell--value").last.text_content() or "")

        try:
            rank = int(rank_txt.strip().rstrip("."))
        except:
            rank = None
        try:
            pts = int(pts_txt)
        except:
            pts = None

        if team_name:
            table[team_name] = {"rank": rank, "pts": pts}

    h = text_clean(home_name)
    a = text_clean(away_name)

    home = next((v for k, v in table.items() if (h and (h in k or k in h))), None)
    away = next((v for k, v in table.items() if (a and (a in k or k in a))), None)

    return {
        "url": url,
        "home_rank": home["rank"] if home else None,
        "home_pts": home["pts"] if home else None,
        "away_rank": away["rank"] if away else None,
        "away_pts": away["pts"] if away else None,
    }


# =========================
# Excel：逐次追記
# =========================
MAX_ROWS_PER_FILE = 10
SHEET_NAME = "Sheet1"
FILE_PREFIX = "output_"
FILE_SUFFIX = ".xlsx"

def _existing_serials(output_dir: str) -> List[int]:
    p = Path(output_dir)
    nums = []
    for f in p.glob(f"{FILE_PREFIX}*{FILE_SUFFIX}"):
        m = re.match(rf"^{re.escape(FILE_PREFIX)}(\d+){re.escape(FILE_SUFFIX)}$", f.name)
        if m:
            nums.append(int(m.group(1)))
    return sorted(nums)

def _next_serial(output_dir: str) -> int:
    nums = _existing_serials(output_dir)
    return (max(nums) + 1) if nums else 1

def _current_file_path(output_dir: str) -> Path:
    p = Path(output_dir)
    nums = _existing_serials(output_dir)
    if not nums:
        return p / f"{FILE_PREFIX}{_next_serial(output_dir)}{FILE_SUFFIX}"
    return p / f"{FILE_PREFIX}{max(nums)}{FILE_SUFFIX}"

def _data_rows_in(path: Path) -> int:
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    total = ws.max_row or 0
    wb.close()
    return max(0, total - 1)

def _create_new_workbook(path: Path):
    df = pd.DataFrame(columns=HEADER)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=SHEET_NAME)
    log(f"🆕 [EXCEL] 新規作成: {path.name}")

def append_row_to_excel(row_dict: dict, output_dir: str, max_rows_per_file: int = MAX_ROWS_PER_FILE):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    cur = _current_file_path(output_dir)
    if not cur.exists():
        _create_new_workbook(cur)

    current_rows = _data_rows_in(cur)
    if current_rows >= max_rows_per_file:
        cur = output_dir / f"{FILE_PREFIX}{_next_serial(output_dir)}{FILE_SUFFIX}"
        _create_new_workbook(cur)
        current_rows = 0

    df = pd.DataFrame([row_dict])
    for col in HEADER:
        if col not in df.columns:
            df[col] = ""
    df = df[HEADER]

    with pd.ExcelWriter(cur, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
        startrow = current_rows + 1
        df.to_excel(w, index=False, header=False, sheet_name=SHEET_NAME, startrow=startrow)

    log(f"💾 [EXCEL] 追記完了: {cur.name} （{current_rows} → {current_rows+1}）")


# =========================
# 親：ライブURL列挙
# =========================
def collect_live_links_filtered() -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, slow_mo=70)
        ctx = browser.new_context(
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
            locale="ja-JP",
            timezone_id="Asia/Tokyo"
        )
        ctx.set_default_timeout(15000)
        ctx.set_default_navigation_timeout(15000)
        setup_route_blocking(ctx)
        page = ctx.new_page()

        log("🌐 Flashscoreトップを開きます...")
        ok = safe_goto(page, "https://www.flashscore.co.jp/", timeout_ms=45000, tag="TOP")
        if not ok:
            log("❌ トップページが開けませんでした")
            try:
                browser.close()
            except:
                pass
            return []

        kill_consent_banners(page)

        try:
            live_sel = (
                "div.filters__tab:has(div.filters__text--short:has-text('ライブ')),"
                " div.filters__tab:has(div.filters__text--long:has-text('開催中の試合'))"
            )
            page.locator(live_sel).first.click(timeout=4000)
            log("✅ [TOP] ライブタブ click")
        except Exception as e:
            log(f"⚠️ [TOP] ライブタブ click 失敗: {type(e).__name__}: {e}")

        try:
            page.wait_for_selector("div.event__match.event__match--live", timeout=20000)
        except:
            pass

        # アコーディオン展開（あれば）
        try:
            buttons = page.locator("button[data-testid='wcl-accordionButton']")
            n_btn = buttons.count()
            opened = 0
            for i in range(n_btn):
                btn = buttons.nth(i)
                aria = btn.get_attribute("aria-label") or ""
                if "非表示" in aria:
                    continue
                try:
                    btn.scroll_into_view_if_needed(timeout=1000)
                    btn.click(timeout=1500)
                    opened += 1
                    time.sleep(0.12)
                except:
                    pass
            if opened:
                time.sleep(0.2)
        except:
            pass

        items = page.evaluate("""
        () => {
          const results = [];
          const headers = Array.from(document.querySelectorAll("[data-testid='wcl-headerLeague']"));
          for (const h of headers) {
            const league =
              (h.querySelector("[data-testid='wcl-scores-simple-text-01']")?.textContent || "").trim() ||
              (h.querySelector("a.headerLeague__title")?.getAttribute("title") || "").trim();

            const country =
              (h.querySelector(".headerLeague__category-text")?.textContent || "").trim() ||
              (h.querySelector(".headerLeague__flag")?.getAttribute("title") || "").trim();

            const category = [country, league].filter(Boolean).join(": ").trim();

            const wrapper = h.closest(".headerLeague__wrapper");
            if (!wrapper) continue;

            let cur = wrapper.nextElementSibling;
            while (cur) {
              if (cur.querySelector?.("[data-testid='wcl-headerLeague']")) break;

              const links = cur.querySelectorAll(
                "div.event__match.event__match--live a.eventRowLink[href*='/match/'][href*='?mid=']"
              );
              for (const a of links) {
                results.push({ href: a.href, category });
              }
              cur = cur.nextElementSibling;
            }
          }
          return results;
        }
        """) or []

        log(f"🧱 headerLeagueブロック取得: {len(items)} 件")

        seen_mid = set()
        for it in items:
            href = it.get("href", "") or ""
            cat  = it.get("category", "") or ""
            mid  = extract_mid(href)

            if not mid or mid in seen_mid:
                continue
            seen_mid.add(mid)

            if not cat:
                continue

            if not any(c in cat for c in CONTAINS_LIST):
                continue
            if any(x in cat for x in UNDER_LIST) or any(x in cat for x in GENDER_LIST) or any(x in cat for x in EXP_LIST):
                continue

            out.append({"url": href, "category": cat})

        log(f"🎯 フィルタ後URL: {len(out)} 件")

        try:
            browser.close()
        except:
            pass

    return out


# =========================
# Worker：1試合処理（完全版）
# =========================
def process_one_match_in_worker(url: str, top_category: str = "") -> Dict[str, Any]:
    result: Dict[str, Any] = {"ok": False, "url": url, "top_category": top_category}
    mid = extract_mid(url) or ""

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
            locale="ja-JP",
            timezone_id="Asia/Tokyo"
        )
        ctx.set_default_timeout(15000)
        ctx.set_default_navigation_timeout(15000)
        setup_route_blocking(ctx)
        pg = ctx.new_page()

        try:
            rb = RowBuilder(HEADER, match_tag=mid)

            # 基本情報（row に先に入れる）
            rb.put("試合ID", mid)
            rb.put("試合リンク文字列", url)  # そのまま URL を入れる

            log("🧩 [WORKER] open match page")
            ok = safe_goto(pg, url, timeout_ms=NAV_TIMEOUT_MS, tag="MATCH")
            if not ok:
                return {"ok": False, "url": url, "top_category": top_category, "error": "match_goto_failed"}

            kill_consent_banners(pg)
            if is_bot_wall(pg):
                return {"ok": False, "url": url, "top_category": top_category, "error": "bot_wall"}

            log("🔎 [WORKER] read teams/scores/time")
            home, away = get_home_away_names(pg)
            hs, aw = get_scores(pg)
            ttxt = get_match_time_text(pg)

            rb.put("ホームチーム", home)
            rb.put("アウェーチーム", away)
            rb.put("ホームスコア", hs)
            rb.put("アウェースコア", aw)
            rb.put("試合時間", ttxt)

            # meta（国/リーグ/開催地/観客など）
            meta = get_match_meta(pg)
            meta_category = ""
            if meta.get("国") and meta.get("リーグ"):
                meta_category = f"{meta['国']}: {meta['リーグ']}".strip()
            final_category = meta_category or (top_category or "")

            rb.put("試合国及びカテゴリ", final_category)
            rb.put("スコア時間", meta.get("取得時刻", ""))  # 例: 2025-12-28 18:57:49

            # 代表的な meta を HEADER へ
            # ※ あなたの HEADER は「スタジアム」「収容人数」「観客数」「場所」などがある
            # meta 側のキーは「開催地」「収容人数」「観客」
            if meta.get("開催地"):
                rb.put("スタジアム", meta.get("開催地"))
            if meta.get("収容人数"):
                rb.put("収容人数", meta.get("収容人数"))
            if meta.get("観客"):
                rb.put("観客数", meta.get("観客"))
            if meta.get("レフェリー"):
                rb.put("審判名", meta.get("レフェリー"))

            # standings（順位）
            log("📌 [WORKER] standings")
            st = get_match_standings(pg, home, away)
            if st:
                if st.get("home_rank") is not None:
                    rb.put("ホーム順位", st.get("home_rank"))
                if st.get("away_rank") is not None:
                    rb.put("アウェー順位", st.get("away_rank"))

            # stats（raw -> canonical -> row）
            log("📈 [WORKER] scrape stats")
            stats_raw = scrape_stats_raw(pg)
            # dump_unmapped_stats(stats_raw)  # 必要ならオン

            canon_pairs = normalize_stats_raw_to_canon(stats_raw)

            # VERIFY
            verify_stats_mapping("CANON", canon_pairs)

            apply_stats_to_row(rb, canon_pairs)

            # 仕上げ
            rb.put("通番", "")  # 親で確定
            rb.put("ソート用秒", parse_live_time_to_seconds(rb.d.get("試合時間", "")))

            verify_row_filled(rb)

            result.update({
                "ok": True,
                "mid": mid,
                "meta_category": meta_category,
                "final_category": final_category,
                "row": rb.d,
            })
            return result

        except Exception as e:
            result["error"] = f"{type(e).__name__}: {e}"
            result["trace"] = traceback.format_exc(limit=7)
            log(f"💥 [WORKER] exception: {result['error']}")
            log(result["trace"])
            return result

        finally:
            try:
                pg.close()
            except:
                pass
            try:
                browser.close()
            except:
                pass


def _worker_entry(url: str, top_category: str, q: "mp.Queue"):
    res = process_one_match_in_worker(url, top_category=top_category)
    try:
        q.put(res)
    except:
        pass

def run_match_with_timeout(url: str, top_category: str, timeout_sec: int = WORKER_TIMEOUT_SEC) -> Dict[str, Any]:
    q: mp.Queue = mp.Queue(maxsize=1)
    p = mp.Process(target=_worker_entry, args=(url, top_category, q), daemon=True)
    p.start()
    p.join(timeout=timeout_sec)

    if p.is_alive():
        log(f"🧨 [TIMEOUT] 子プロセス強制終了: {url} ({timeout_sec}s)")
        try:
            p.terminate()
        except:
            pass
        p.join(3)
        return {"ok": False, "url": url, "top_category": top_category, "error": f"timeout({timeout_sec}s)"}

    try:
        return q.get(timeout=2)
    except pyqueue.Empty:
        return {"ok": False, "url": url, "top_category": top_category, "error": "no_result_from_worker"}


# =========================
# main
# =========================
def main():
    mp.set_start_method("spawn", force=True)

    verify_header_and_stat_map()
    load_seqmap()
    os.makedirs(SAVE_DIR, exist_ok=True)

    items = collect_live_links_filtered()
    log(f"🎯 ライブ試合: {len(items)} 件")

    for i, it in enumerate(items, 1):
        url = it.get("url", "") or ""
        top_category = it.get("category", "") or ""

        log(f"\n==============================")
        log(f"[{i}/{len(items)}] {url}")
        log(f"TOP category = {top_category}")
        log(f"==============================")

        res = run_match_with_timeout(url, top_category=top_category, timeout_sec=WORKER_TIMEOUT_SEC)
        if not res.get("ok"):
            log(f"⚠️ [WORKER] 失敗: {res.get('error','')} url={url}")
            continue

        mid = res.get("mid", "") or ""
        row = res.get("row", {}) or {}
        final_category = (res.get("final_category", "") or "").strip()

        # 親でも再フィルタ（最終カテゴリで判定）
        if not any(c in final_category for c in CONTAINS_LIST):
            log(f"⏭️ スキップ（リスト外）: {final_category}")
            continue
        if any(x in final_category for x in UNDER_LIST) or any(x in final_category for x in GENDER_LIST) or any(x in final_category for x in EXP_LIST):
            log(f"🚫 除外: {final_category}")
            continue

        # 通番確定
        last_seq = int(SEQMAP.get(mid, 0)) if mid else 0
        seq = last_seq + 1
        if mid:
            SEQMAP[mid] = seq

        row["通番"] = seq
        row["試合ID"] = mid

        append_row_to_excel(row, SAVE_DIR)

    save_seqmap()

if __name__ == "__main__":
    main()

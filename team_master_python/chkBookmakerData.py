from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os, unicodedata
import datetime
import sys
import smtplib
from email.mime.text import MIMEText
import json
import datetime
import inspect as isp
import csv
import time
from python_analytics import write_prediction_data
import pandas as pd
import re
from playwright.sync_api import TimeoutError as PTimeout


def execute(playwright):
    loop_bef_seq = -1
    # 対象時間を設定
    start_time = datetime.time(23, 30)
    end_time = datetime.time(23, 59)
    # 現在時刻
    now = datetime.datetime.today()
    now_info = datetime.datetime(now.year,now.month,now.day,now.hour,now.minute,now.second)

    if now_info.weekday() == 1 or now_info.weekday() == 2 or now_info.weekday() == 3 or now_info.weekday() == 4 or now_info.weekday() == 5:
        if now.hour == 12 or now.hour == 13 or now.hour == 14 or now.hour == 15 or now.hour == 16:
            print(f"{now_info}  データ取得対象外の時間です。")

    print(f"{now_info}  データ取得対象の時間です。")

    # Chromiumブラウザを起動
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    # Googleをひらく
    page.goto("https://www.google.com")

    # flashscoreに遷移
    page.goto("https://flashscore.co.jp/soccer/")
    # ライブメニュー遷移
    try:
        live_page = page.locator("div.filters__tab").nth(1)  # CSSセレクタを使用
        live_page.click()

        page.wait_for_load_state("networkidle")

        # 非表示要素を表示に変更
        alter = page.locator("button[data-testid='wcl-accordionButton']")
        alter_count = alter.count()
        for i in range(alter_count):
            alter_label = alter.nth(i).get_attribute("aria-label")
            if alter_label == None or alter_label == "リーグ全試合 表示":
                alter.nth(i).click()

        # ライブ試合取得
        live_game_list = []
        live_game = page.locator("a.eventRowLink[href*='/summary']")
        live_game_list = live_game.evaluate_all("els => els.map(e => e.href)")
    except Exception as e:
        print("ライブ試合リンクの取得失敗", e)
        live_game_list = []

    # cron実行通知用にホームチーム,アウェイチーム,カテゴリを入れたリストを保持
    #live_game_list.append("https://www.flashscore.co.jp/match/soccer/as-marsa-GfH6hHa2/js-kairouan-rVdSw2PP/?mid=6BdnXSaL")
    print(live_game_list)
    cron_notice_data_list = []
    for j in range(len(live_game_list)):
        main_stats = []  # 例: [{"name":"ボール支配率","home":"54%","away":"46%"}, ...]

        notice_data_list = []
        # 試合用文字列のみ抜く
        random_game_str = live_game_list[j]

        main_stats.append({"リンク": random_game_str})

        # FLASHSCOREに遷移
        page.goto(random_game_str)
        time.sleep(3)

        # 試合の国及びカテゴリを取得
        game_info_sub_info_text = ""
        try:
            team_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-2)
            game_info = page.locator("span[data-testid='wcl-scores-overline-03']").nth(-1)
            #game_info_sub_info = game_info.locator("span.tournamentHeader__country")
            team_info_text = team_info.text_content()
            game_info_text = game_info.text_content()
            print(team_info_text)
            print(game_info_text)
        except Exception as e:
            print("チーム名の取得失敗", e)
            continue

        if team_info_text == "" or game_info_text == "":
            print("チーム名の取得失敗: 空")
            continue

        game_info_sub_info_text = team_info_text + ": " + game_info_text
        main_stats.append({"国及びカテゴリ": game_info_sub_info_text})
        print(game_info_sub_info_text)

        #試合のカテゴリが以下の場合のみ取得
        contains_list = []
        contains_list.append("ケニア: プレミアリーグ")
        contains_list.append("コロンビア: プリメーラ A")
        contains_list.append("タンザニア: プレミアリーグ")
        contains_list.append("イングランド: プレミアリーグ")
        contains_list.append("イングランド: EFL チャンピオンシップ")
        contains_list.append("イングランド: EFL リーグ 1")
        contains_list.append("エチオピア: プレミアリーグ")
        contains_list.append("コスタリカ: リーガ FPD")
        contains_list.append("ジャマイカ: プレミアリーグ")
        contains_list.append("スペイン: ラ・リーガ")
        contains_list.append("ブラジル: セリエ A ベターノ")
        contains_list.append("ブラジル: セリエ B")
        contains_list.append("ドイツ: ブンデスリーガ")
        contains_list.append("韓国: K リーグ 1")
        contains_list.append("中国: 中国スーパーリーグ")
        contains_list.append("日本: J1 リーグ")
        contains_list.append("日本: J2 リーグ")
        contains_list.append("日本: J3 リーグ")
        contains_list.append("インドネシア: リーガ 1")
        contains_list.append("オーストラリア: A リーグ・メン")
        contains_list.append("チュニジア: チュニジア･プロリーグ")
        contains_list.append("ウガンダ: プレミアリーグ")
        contains_list.append("メキシコ: リーガ MX")
        contains_list.append("フランス: リーグ・アン")
        contains_list.append("スコットランド: プレミアシップ")
        contains_list.append("オランダ: エールディビジ")
        contains_list.append("アルゼンチン: トルネオ・ベターノ")
        contains_list.append("イタリア: セリエ A")
        contains_list.append("イタリア: セリエ B")
        contains_list.append("ポルトガル: リーガ・ポルトガル")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("セルビア: スーペルリーガ")
        contains_list.append("日本: WEリーグ")
        contains_list.append("ボリビア: LFPB")
        contains_list.append("ブルガリア: パルヴァ・リーガ")
        contains_list.append("カメルーン: エリート 1")
        contains_list.append("ペルー: リーガ 1")
        contains_list.append("エストニア: メスタリリーガ")
        contains_list.append("トルコ: スュペル・リグ")
        contains_list.append("ウクライナ: プレミアリーグ")
        contains_list.append("ベルギー: ジュピラー･プロリーグ")
        contains_list.append("エクアドル: リーガ・プロ")
        contains_list.append("日本: YBC ルヴァンカップ")
        contains_list.append("日本: 天皇杯")
       

        under_list = []
        under_list.append("U17")
        under_list.append("U18")
        under_list.append("U19")
        under_list.append("U20")
        under_list.append("U21")
        under_list.append("U22")
        under_list.append("U23")
        under_list.append("U24")
        under_list.append("U25")

        gender_list = []
        gender_list.append("女子")

        exp_list = []
        exp_list.append("ポルトガル: リーガ・ポルトガル 2")
        exp_list.append("イングランド: プレミアリーグ 2")
        exp_list.append("イングランド: プレミアリーグ U18")

        if any(item in game_info_sub_info_text for item in contains_list):
            pass
        else:
            print("pass1")
            continue

        if any(item in game_info_sub_info_text for item in under_list):
            print("pass2")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in gender_list):
            print("pass3")
            continue
        else:
            pass

        if any(item in game_info_sub_info_text for item in exp_list):
            print("pass4")
            continue
        else:
            pass

        # 試合ブロックが出るまで待機
        page.wait_for_selector("div.duelParticipant__container")
        container = page.locator("div.duelParticipant__container").first

        # キックオフ時刻（例: 11.09.2025 10:00）
        kickoff = container.locator(".duelParticipant__startTime > div").first.text_content().strip()

        # ライブ経過時間（例: 33:12）※あれば
        live_time = get_display_time(page, container)
        print("試合時間(表示用):", live_time)

        # チーム名（テキスト優先、なければ画像の alt をフォールバック）
        home_team = container.locator(".duelParticipant__home a.participant__participantName").first
        away_team = container.locator(".duelParticipant__away a.participant__participantName").first

        home_team_name = home_team.text_content().strip() if home_team.count() else \
        container.locator(".duelParticipant__home img.participant__image").first.get_attribute("alt")

        away_team_name = away_team.text_content().strip() if away_team.count() else \
        container.locator(".duelParticipant__away img.participant__image").first.get_attribute("alt")

        print("kickoff:", kickoff)
        print("home:", home_team_name)
        print("away:", away_team_name)
        print("live_time:", live_time)  # 進行中のみ値が入る

        # スコア（ホーム・アウェー）
        home_score = ""
        away_score = ""

        # まず固定ヘッダー側が可視なら優先
        fixed_score = page.locator(".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .fixedScore")
        if fixed_score.count() > 0:
            nums = fixed_score.locator("span:not(.fixedScore__divider)")
        else:
            # 通常のスコア表示を使用（dividerは除外）
            nums = container.locator(".detailScore__wrapper span:not(.detailScore__divider)")

        if nums.count() >= 2:
            home_score = nums.nth(0).text_content().strip()
            away_score = nums.nth(1).text_content().strip()

        main_stats.append({"試合時間": live_time, "ホームチーム": home_team_name, "アウェーチーム": away_team_name, "ホームスコア": home_score, "アウェースコア": away_score})

        #監督名(新規タブでスカッドに移動して取得)
        home_manager = ""
        away_manager = ""
        team_data_list = page.locator("a[class='participant__participantLink participant__participantLink--team']")
        tag_flg = 0
        tag_list = []
        manager_name_list = []
        for link in range(team_data_list.count()):
            links = team_data_list.nth(link).get_attribute("href")
            if not links:
                continue
            
            new_page = context.new_page()
            new_page.goto("https://flashscore.co.jp" + links)
            #スカッドに移動
            new_page.goto(new_page.url + "squad/")
            time.sleep(3)
            try:
                manager_name = new_page.locator("a[class='lineupTable__cell--name']")
                if manager_name.count() != 0:
                    if tag_flg == 0:
                        tag_list.append("home")
                    elif tag_flg == 1:
                        tag_list.append("away")
                        
                    name = str.strip(manager_name.nth(manager_name.count() - 1).text_content())
                    manager_name_list.append(name)
            except Exception as e:
                print("スカッドなし", e)

            tag_flg += 1

            new_page.close()

        if len(manager_name_list) != 0:
            if len(manager_name_list) == 1 and tag_list[0] == "home":
                home_manager = manager_name_list[0]
            elif len(manager_name_list) == 1 and tag_list[0] == "away":
                away_manager = manager_name_list[0]
            else:
                home_manager = manager_name_list[0]
                away_manager = manager_name_list[1]
            print(home_manager)
            print(away_manager)
            main_stats.append({"ホーム監督": home_manager, "アウェー監督": away_manager})

        #審判名,スタジアム,収容人数,観客数,場所
        judge_member = ""
        studium = ""
        capacity = ""
        audience = ""
        location = ""
        try:
            wrapper = page.locator("div[class*='wcl-infoLabelWrapper']")
            wrapperVal = page.locator("div[class*='wcl-infoValue']")
            for ind in range(wrapper.count()):
                key = wrapper.nth(ind).locator("span[data-testid='wcl-scores-overline-02']")
                value = wrapperVal.nth(ind)
                if key.text_content() == "レフェリー:":
                    judge_member = value.text_content()
                elif key.text_content() == "開催地:":
                    studium = str.strip(value.text_content())
                elif key.text_content() == "収容人数:":
                    capacity = value.text_content()
                elif key.text_content() == "参加:":
                    audience = value.text_content()

                prefix = studium.find("(")
                suffix = studium.find(")")
                location = studium[prefix + 1:suffix]

            print(judge_member)
            print(studium)
            print(capacity)
            print(audience)
            print(location)
        
            main_stats.append({"レフェリー": judge_member, "開催地": studium, "収容人数": capacity, "観客数": audience, "場所": location})
                
        except Exception as e:
            print("試合場所情報取得失敗", e)

        # スコア取得
        try:
            score = page.locator("div.detailScore__wrapper").text_content()
        except Exception as e:
            print("スコア取得失敗", e)
            continue
        
        data_list = []
        try:
            # ライブなら0、終了済なら1をnthに導入
            live_fin = page.locator("span[class='fixedHeaderDuel__detailStatus']").nth(1)
            live_fin_text = live_fin.text_content()
            if live_fin_text == '終了済':
                index = 1
            elif live_fin_text == 'ハーフタイム':
                index = 2
            else:
                index = 0

            # イベントタイムを取得
            if index == 0:
                page.wait_for_selector("div.eventAndAddedTime span.eventTime")
                live_fin_text = page.locator("div.eventAndAddedTime span.eventTime").first.text_content().strip()
            
            # --- 統計タブへ ---
            # A11yロールで堅く取得
            page.get_by_role("tab", name="統計").click()
            page.wait_for_selector("[data-analytics-context='tab-match-statistics']")
            # 選択状態になるまで待機（どちらかが満たされればOK）
            page.wait_for_selector(
                "[data-analytics-alias='match-statistics'].selected, "
                "a[href*='/summary/stats'] .wcl-tabSelected_rHdTM"
            )

            # --- 「もっと表示」を押す（存在する場合のみ） ---
            more_locator = page.locator("button:has-text('もっと'), a:has-text('もっと')")
            if more_locator.count() > 0:
                print("もっと表示クリック")
                more_locator.first.click()

            # --- 統計値の取得 ---
            

            #統計
            page.wait_for_selector("div.sectionHeader:has-text('主なスタッツ')")
            # （あれば）「試合」サブタブに寄せておく
            try:
                page.get_by_role("tab", name="試合").click()
            except:
                pass

            # 値 strong の両表記に対応
            VAL_STRONG = "strong[data-testid='wcl-scores-simple-text-01'], strong[data-testid='wcl-scores-simpleText-01']"

            # --- 主なスタッツ セクションを特定 ---
            section1 = page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('主なスタッツ')")
            ).first

            rows = section1.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})

            # --- 主なスタッツ セクションを特定 ---
            section2 = page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('シュート')")
            ).first

            rows = section2.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})
        
            # --- 主なスタッツ セクションを特定 ---
            section3 = page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('アタック')")
            ).first

            rows = section3.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})
        
            # --- 主なスタッツ セクションを特定 ---
            section4 = page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('ディフェンス')")
            ).first

            rows = section4.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})
            
            # --- 主なスタッツ セクションを特定 ---
            section5 = page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('ゴールキーパー')")
            ).first

            rows = section5.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})

            # --- 主なスタッツ セクションを特定 ---
            section6= page.locator(
                "div.section",
                has=page.locator("div.sectionHeader:has-text('パス')")
            ).first

            rows = section6.locator("[data-testid='wcl-statistics']")
            count = rows.count()
            print(count)

            for i in range(count):
                row = rows.nth(i)
                # 項目名
                name = row.locator("[data-testid='wcl-statistics-category'] " + VAL_STRONG).first.text_content().strip()
                # 値（home/away）
                home_cell = row.locator("[data-testid='wcl-statistics-value']").first
                away_cell = row.locator("[data-testid='wcl-statistics-value']").last
                home = combined_value(home_cell)  # 例: "78%(273/348)"
                away = combined_value(away_cell)
                main_stats.append({"ホーム" + name: home, "アウェー" + name: away})

            # --- 足りなければスキップ ---
            if len(main_stats) == 0:
                print("データ取得できなかったためスキップします")
                continue
        except Exception as e:
            print("選択失敗", e)

        # FLASHSCORE(順位表)に遷移
        rank = True
        try:
            comma = random_game_str.split('?')
            rankLink = comma[0] + "standings/live-standings/?" + comma[1]
            print(rankLink)
            page.goto(rankLink)
        except Exception as e:
            print("順位表に遷移できません")
            rank = False

        home_team_max_getting_scorer = ""
        away_team_max_getting_scorer = ""
        home_team_max_getting_scorer_participating = "未出場"
        away_team_max_getting_scorer_participating = "未出場"

        # 順位表メニューをクリック
        home_score = ""
        away_score = ""
        home_rank = ""
        away_rank = ""
        if rank == True:
            try:
                container = find_table_scope(page)  # ここで page か該当 frame が返る
                home_rank = str(get_team_rank(container, home_team_name))
                away_rank = str(get_team_rank(container, away_team_name))
            except Exception as e:
                print("順位表選択失敗", e)

            # ホームチーム,アウェーチームの順位を特定する
            main_stats.append({"ホーム順位": home_rank, "アウェー順位": away_rank})
            print("home_rank: " + str(home_rank))
            print("away_rank: " + str(away_rank))

            # スコア分割
            score_split = score.split("-")
            home_score = score_split[0]
            away_score = score_split[1]

            #チーム得点,失点取得
            home_team_home_score = ""
            home_team_home_lost = ""
            away_team_home_score = ""
            away_team_home_lost = ""
            home_team_away_score = ""
            home_team_away_lost = ""
            away_team_away_score = ""
            away_team_away_lost = ""
            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/home")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub1 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub2 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub1.count())
                print(team_info_sub2.count())
                home_total_1 = ""
                home_total_2 = ""
                for ind in range(team_info_sub1.count()):
                    if team_info_sub1.nth(ind).text_content() == home_team_name:
                        home_total_1 = team_info_sub2.nth(ind).text_content()
                    elif team_info_sub1.nth(ind).text_content() == away_team_name:
                        home_total_2 = team_info_sub2.nth(ind).text_content()

                    if home_total_1 != "" and home_total_2 != "":
                        break

                if home_total_1 != "":
                    home_total_sub_1 = home_total_1.split(':')
                    home_team_home_score = home_total_sub_1[0]
                    home_team_home_lost = home_total_sub_1[1]

                if home_total_2 != "":
                    away_total_sub_1 = home_total_2.split(':')
                    away_team_home_score = away_total_sub_1[0]
                    away_team_home_lost = away_total_sub_1[1]

            page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/standings/table/away")
            time.sleep(5)
            
            data = page.locator("div.table__row--selected")
            print(data.count())
            for inds in range(data.count()):
                team_info_sub3 = data.nth(inds).locator("div.table__cell--participant").locator("div.tableCellParticipant").locator("div.tableCellParticipant__block").locator("a[class='tableCellParticipant__name']")
                team_info_sub4 = data.nth(inds).locator("span[class=' table__cell table__cell--value  table__cell--score ']")
                print(team_info_sub3.count())
                print(team_info_sub4.count())
                away_total_1 = ""
                away_total_2 = ""
                for ind in range(team_info_sub3.count()):
                    if team_info_sub3.nth(ind).text_content() == home_team_name:
                        away_total_1 = team_info_sub4.nth(ind).text_content()
                    elif team_info_sub3.nth(ind).text_content() == away_team_name:
                        away_total_2 = team_info_sub4.nth(ind).text_content()

                    if away_total_1 != "" and away_total_2 != "":
                        break

                if away_total_1 != "":
                    away_total_sub_1 = away_total_1.split(':')
                    home_team_away_score = away_total_sub_1[0]
                    home_team_away_lost = away_total_sub_1[1]

                if away_total_2 != "":
                    away_total_sub_2 = away_total_2.split(':')
                    away_team_away_score = away_total_sub_2[0]
                    away_team_away_lost = away_total_sub_2[1]

            print(home_team_home_score)
            print(home_team_home_lost)
            print(away_team_home_score)
            print(away_team_home_lost)
            print(home_team_away_score)
            print(home_team_away_lost)
            print(away_team_away_score)
            print(away_team_away_lost)

            main_stats.append({"ホームチームホーム得点": home_team_home_score, "ホームチームホーム失点": home_team_home_lost})
            main_stats.append({"アウェーチームホーム得点": away_team_home_score, "アウェーチームホーム失点": away_team_home_lost})
            main_stats.append({"ホームチームアウェー得点": home_team_away_score, "ホームチームアウェー失点": home_team_away_lost})
            main_stats.append({"アウェーチームアウェー得点": away_team_away_score, "アウェーチームアウェー失点": away_team_away_lost})

            # チーム得点王取得
            print("チーム得点王取得")
            get_score_flg = True
            try:
                comma = random_game_str.split('?')
                rankLink = comma[0] + "standings/top_scorers/?" + comma[1]
                page.goto(rankLink) # CSSセレクタを使用

                #ページが存在しない場合スキップ(処理追加)
                dummy_page = "https://www.flashscore.co.jp/match/" + random_game_str + "/#/match-summary/match-statistics"
                if page.url != dummy_page:
            
                    top_scorer_member = ""
                    top_scorer_team = ""
                    home_scorer_flg = False
                    away_scorer_flg = False
                    #得点王表が完全表示されるまで待機
                    time.sleep(2)
                    top_scorer_graph = page.locator("div[class*='topScorers__row--selected']")

                    for graph_ind in range(top_scorer_graph.count()):
                        sub_top_scorer_graph = top_scorer_graph.nth(graph_ind).locator("div.topScorers__participantCell").locator("a[class*='topScorersParticipant']")
                        tmp = 0
                        for sub_graph_ind in range(sub_top_scorer_graph.count()):
                            if tmp == 0:
                                top_scorer_member = sub_top_scorer_graph.nth(sub_graph_ind).text_content()
                            elif tmp == 1:
                                top_scorer_team = sub_top_scorer_graph.nth(sub_graph_ind).text_content()

                            if top_scorer_team == home_team_name and home_team_max_getting_scorer == "":
                                home_team_max_getting_scorer = top_scorer_member
                                home_scorer_flg = True
                            if top_scorer_team == away_team_name and away_team_max_getting_scorer == "":
                                away_team_max_getting_scorer = top_scorer_member
                                away_scorer_flg = True

                            tmp += 1
                                           
                        if home_scorer_flg == True and away_scorer_flg == True:
                            break

                    print("home_team_max_getting_scorer:" + home_team_max_getting_scorer) #取得済み
                    print("away_team_max_getting_scorer:" + away_team_max_getting_scorer) #取得済み
                else:
                    get_score_flg = False

            except Exception as e:
                print("得点王選択失敗", e)
                get_score_flg = False

            # チーム得点王,フォーメーション取得
            home_formation = ""
            away_formation = ""
            if get_score_flg == True:
                page.goto("https://www.flashscore.co.jp/match/" + random_game_str + "/#/match-summary/lineups")
                #表示待ち
                time.sleep(2)

                try:
                    formation = page.locator("span[class='lf__headerPart']")
                    page.wait_for_selector("span[class='lf__headerPart']", state="visible", timeout=2000)
                            
                    print(formation)
                    home_formation = formation.nth(0).text_content()
                    away_formation = formation.nth(1).text_content()
                    print(home_formation)
                    print(away_formation)
                    
                except Exception as e:
                    print("フォーメーション取得失敗", e)
                
                home_lineup = page.locator("div.lf__side").nth(0)
                away_lineup = page.locator("div.lf__side").nth(1)
                if home_lineup.count() != 11 and away_lineup.count() != 11:
                    home_lineup = page.locator("div.lf__side").nth(2)
                    away_lineup = page.locator("div.lf__side").nth(3)
                    
                try:
                    home_lineup_box = home_lineup.locator("div[class='lf__participantNew']")
                    print(home_lineup_box.count())
                    for home_box_ind in range(home_lineup_box.count()):
                        try:
                            home_lineup_box_sub = home_lineup_box.nth(home_box_ind).locator("div[class='wcl-participant_QKIld wcl-hasNumber_g943f']")
                            home_lineup_box_sub_all = home_lineup_box_sub.locator("div[class='wcl-mainRow_Xi7Hi']")
                            home_lineup_box_sub_member = home_lineup_box_sub_all.locator("a[class='wcl-link_bLtj3 wcl-linkBase_CdaEq wcl-primary_C2HA0 wcl-nameWrapper_m-xnS']")
                            home_lineup_box_sub_member2 = home_lineup_box_sub_member.locator("strong[class='wcl-simpleText_Asp-0 wcl-scores-simpleText-01_pV2Wk wcl-bold_roH-0 wcl-name_IO3G1']")
                            
                            #出場しているか
                            print("**************************")
                            print(home_team_max_getting_scorer)
                            print(home_lineup_box_sub_member2.text_content())
                            print("**************************")
                            if home_team_max_getting_scorer == home_lineup_box_sub_member2.text_content() or home_team_max_getting_scorer in home_lineup_box_sub_member2.text_content():
                                home_team_max_getting_scorer_participating = "出場中"
                                break

                            if home_box_ind == 10:
                                break
                                    
                            #if home_team_max_getting_scorer_participating == "":
                                #交代選手の場合の取得
                            #    home_sub_lineup = page.locator("div.lf__side").nth(2)

                            #    home_lineup_subbox = home_sub_lineup.locator("div.lf__participantNew")
                            #    for home_box_ind in range(home_lineup_subbox.count()):
                            #        home_lineup_box_sub = home_lineup_subbox.nth(home_box_ind).locator("div[data-testid='wcl-lineupsParticipantGeneral-left']")
                            #        home_lineup_box_sub_all = home_lineup_box_sub.locator("div[class*='wcl-mainRow']")
                        
                            #        home_lineup_box_sub_member = home_lineup_box_sub_all.nth(0)
                            #        home_lineup_box_sub_member2 = home_lineup_box_sub_member.nth(0).locator("a[data-testid='wcl-textLink']").locator("strong[data-testid='wcl-scores-simpleText-01']")
                            
                                    #ベンチ入りしているか
                            #        if home_team_max_getting_scorer == home_lineup_box_sub_member2.nth(0).text_content() or home_team_max_getting_scorer in home_lineup_box_sub_member2.nth(0).text_content():
                            #            home_team_max_getting_scorer_participating = "交替済み"
                            #            break
                        except Exception as e:
                            print("取得エラー", e)
                            break

                    print("出場結果")
                    print(home_team_max_getting_scorer_participating)
                #except Exception as e:
                #    print("出場結果取得失敗", e)

                #try:
                    away_lineup_box = away_lineup.locator("div[class='lf__participantNew lf__isReversed']")
                    print(away_lineup_box.count())
                    for away_box_ind in range(away_lineup_box.count()):
                        try:
                            away_lineup_box_sub = away_lineup_box.nth(away_box_ind).locator("div[class='wcl-participant_QKIld wcl-rtl_245jA wcl-hasNumber_g943f']")
                            away_lineup_box_sub_all = away_lineup_box_sub.locator("div[class='wcl-mainRow_Xi7Hi']")
                            away_lineup_box_sub_member = away_lineup_box_sub_all.locator("a[class='wcl-link_bLtj3 wcl-linkBase_CdaEq wcl-primary_C2HA0 wcl-nameWrapper_m-xnS']")
                            away_lineup_box_sub_member2 = away_lineup_box_sub_member.locator("strong[class='wcl-simpleText_Asp-0 wcl-scores-simpleText-01_pV2Wk wcl-bold_roH-0 wcl-name_IO3G1']")

                            #出場しているか
                            print("**************************")
                            print(away_team_max_getting_scorer)
                            print(away_lineup_box_sub_member2.text_content())
                            print("**************************")
                            if away_team_max_getting_scorer == away_lineup_box_sub_member2.text_content() or away_team_max_getting_scorer in away_lineup_box_sub_member2.text_content():
                                away_team_max_getting_scorer_participating = "出場中"
                                break

                            if away_box_ind == 10:
                                break

                            #if away_team_max_getting_scorer_participating == "":
                                #交代選手の場合の取得
                            #    away_sub_lineup = page.locator("div.lf__side").nth(3)

                            #    away_lineup_subbox = away_sub_lineup.locator("div.lf__participantNew")
                            #    for away_box_ind in range(away_lineup_subbox.count()):
                            #        away_lineup_box_sub = away_lineup_subbox.nth(away_box_ind).locator("div[data-testid='wcl-lineupsParticipantGeneral-right']")
                            #        away_lineup_box_sub_all = away_lineup_box_sub.locator("div[class*='wcl-mainRow']")
                        
                            #        away_lineup_box_sub_member = away_lineup_box_sub_all.nth(0)
                            #        away_lineup_box_sub_member2 = away_lineup_box_sub_member.nth(0).locator("a[data-testid='wcl-textLink']").locator("strong[data-testid='wcl-scores-simpleText-01']")
                            
                                    #ベンチ入りしているか
                            #        if away_team_max_getting_scorer == away_lineup_box_sub_member2.nth(0).text_content() or away_team_max_getting_scorer in away_lineup_box_sub_member2.nth(0).text_content():
                            #            away_team_max_getting_scorer_participating = "交替済み"
                            #            break
                        except Exception as e:
                            print("取得エラー", e)
                            break

                    print("出場結果")
                    print(away_team_max_getting_scorer_participating)
                except Exception as e:
                    print("出場結果取得失敗", e)

        main_stats.append({"ホームチーム最大得点取得者": home_team_max_getting_scorer, "アウェーチーム最大得点取得者": away_team_max_getting_scorer})
        main_stats.append({"ホームチーム最大得点取得者出場状況": home_team_max_getting_scorer_participating, "アウェーチーム最大得点取得者出場状況": away_team_max_getting_scorer_participating})
        main_stats.append({"ホームフォーメーション": home_formation, "アウェーフォーメーション": away_formation})

        #天気,気温,湿度
        weather = ""
        temperature = ""
        humid = ""

        #country_sub = game_info_sub_info_text.split(":")
        #country = country_sub[0]
        #print(country)
        #print(location)

        #query = f"{country} {location} 天気"
        #url = f"https://www.google.com/search?={query}"
        #page.goto(url)

        #try:
        #    weather = page.locator("img#dimg_SvOeZ8_mDoOv0-kPt_Os4Qg_1").get_attribute("alt")
        #    temperature = page.locator("span#wob_tm").inner_text()
        #    humid = page.locator("span#wob_hm").inner_text()
        #    print(weather)
        #    print(temperature)
        #    print(humid)
        #except Exception as e:
        #    print("天気取得失敗", e)

        main_stats.append({"天気": weather, "気温": temperature, "湿度": humid})

        print(main_stats)

        # エクセルに表記
        main_stats_dict = merge_main_stats(main_stats)
        loop_bef_seq = makeExcelAndNotice(main_stats_dict, loop_bef_seq)

    #noticeCronMail(cron_notice_data_list)

    page.close()
    context.close()
    browser.close()


def makeExcelAndNotice(main_stats, loop_bef_seq):

    # ===== スタッツの取り出し =====
    home_exp, away_exp                       = gp(main_stats, "ゴール期待値（xG）")
    home_exp_goal_in, away_exp_goal_in       = gp(main_stats, "枠内ゴール期待値（xGOT）")
    home_domination, away_domination         = gp(main_stats, "ボール支配率")
    home_shoot_all, away_shoot_all           = gp(main_stats, "合計シュート")
    home_shoot_in, away_shoot_in             = gp(main_stats, "枠内シュート")
    home_shoot_out, away_shoot_out           = gp(main_stats, "枠外シュート")
    home_block_shoot, away_block_shoot       = gp(main_stats, "シュートブロック")
    home_big_chance, away_big_chance         = gp(main_stats, "ビッグチャンス")
    home_corner, away_corner                 = gp(main_stats, "コーナーキック")
    home_shoot_box_in, away_shoot_box_in     = gp(main_stats, "ボックス内からのシュート")
    home_shoot_box_out, away_shoot_box_out   = gp(main_stats, "ボックス外からのシュート")
    home_goal_post, away_goal_post           = gp(main_stats, "ゴール枠に当たる")
    home_head_goal, away_head_goal           = gp(main_stats, "ヘディングによるゴール")
    home_keeper_save, away_keeper_save       = gp(main_stats, "キーパーセーブ")
    home_free_kick, away_free_kick           = gp(main_stats, "フリーキック")
    home_offside, away_offside               = gp(main_stats, "オフサイド")
    home_foul, away_foul                     = gp(main_stats, "ファウル")
    home_yellow_card, away_yellow_card       = gp(main_stats, "イエローカード")
    home_red_card, away_red_card             = gp(main_stats, "レッドカード")
    home_slow_in, away_slow_in               = gp(main_stats, "スローイン")
    home_box_touch, away_box_touch           = gp(main_stats, "相手ボックス内でのタッチ")
    home_pass, away_pass                     = gp(main_stats, "パス")
    home_long_pass, away_long_pass           = gp(main_stats, "ロングパス")
    home_final_pass, away_final_pass         = gp(main_stats, "ファイナルサードでのパス")
    home_cross, away_cross                   = gp(main_stats, "クロス")
    home_tackle, away_tackle                 = gp(main_stats, "タックル")
    home_clear, away_clear                   = gp(main_stats, "クリアリング")
    home_duel, away_duel                     = gp(main_stats, "デュエル勝利数")
    home_intercept, away_intercept           = gp(main_stats, "インターセプト")

    # ===== メインデータ =====
    game_info_sub_info_text                  = g(main_stats, "国及びカテゴリ")
    home_rank                                = g(main_stats, "ホーム順位")
    print("ホーム順位です:" + str(home_rank))
    away_rank                                = g(main_stats, "アウェー順位")
    print("アウェー順位です:" + str(away_rank))
    home_score                               = g(main_stats, "ホームスコア")
    away_score                               = g(main_stats, "アウェースコア")
    home_team_name                           = g(main_stats, "ホームチーム")
    away_team_name                           = g(main_stats, "アウェーチーム")
    weather                                  = g(main_stats, "天気")
    temperature                              = g(main_stats, "気温")
    humid                                    = g(main_stats, "湿度")
    home_manager                             = g(main_stats, "ホーム監督")
    away_manager                             = g(main_stats, "アウェー監督")
    home_team_home_score                     = g(main_stats, "ホームチームホーム得点")
    home_team_home_lost                      = g(main_stats, "ホームチームホーム失点")
    away_team_home_score                     = g(main_stats, "アウェーチームホーム得点")
    away_team_home_lost                      = g(main_stats, "アウェーチームホーム失点")
    home_team_away_score                     = g(main_stats, "ホームチームアウェー得点")
    home_team_away_lost                      = g(main_stats, "ホームチームアウェー失点")
    away_team_away_score                     = g(main_stats, "アウェーチームアウェー得点")
    away_team_away_lost                      = g(main_stats, "アウェーチームアウェー失点")
    home_formation                           = g(main_stats, "ホームフォーメーション")
    away_formation                           = g(main_stats, "アウェーフォーメーション")
    judge_member                             = g(main_stats, "レフェリー")
    studium                                  = g(main_stats, "開催地")
    capacity                                 = g(main_stats, "収容人数")
    audience                                 = g(main_stats, "観客数")
    location                                 = g(main_stats, "場所")
    home_team_max_getting_scorer             = g(main_stats, "ホームチーム最大得点取得者")
    away_team_max_getting_scorer             = g(main_stats, "アウェーチーム最大得点取得者")
    home_team_max_getting_scorer_participating = g(main_stats, "ホームチーム最大得点取得者出場状況")
    away_team_max_getting_scorer_participating = g(main_stats, "アウェーチーム最大得点取得者出場状況")
    live_fin_text                            = g(main_stats, "試合時間")
    random_game_str                          = g(main_stats, "リンク")

    # ----- メール通知条件 -----
    chkResult = conditionChk(
        game_info_sub_info_text, home_rank, away_rank, home_score, away_score, home_exp, away_exp, weather, temperature, humid,
        home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in, live_fin_text,  home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
        home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
        home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating
    )
    print(chkResult)

    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    notice_flg = "未通知"
    if chkResult[0] != "" and chkResult[1] != "":
        subject = "試合データです"
        json_data = {
            "通知詳細:": chkResult[0],
            "試合カテゴリ": game_info_sub_info_text,
            "試合時間": live_fin_text,
            "ホーム": {
                "ホームチーム": home_team_name,
                "ホーム順位": home_rank,
                "ホームスコア": home_score,
                "ホーム支配率": home_domination,
                "ホームシュート": home_shoot_all,
                "ホーム枠内シュート": home_shoot_in,
            },
            "アウェイ": {
                "アウェイチーム": away_team_name,
                "アウェイ順位": away_rank,
                "アウェイスコア": away_score,
                "アウェイ支配率": away_domination,
                "アウェイシュート": away_shoot_all,
                "アウェイ枠内シュート": away_shoot_in,
            }
        }
        msg = MIMEText(f"試合データ:\n\n{json.dumps(json_data, ensure_ascii=False, indent=4)}")
        msg['Subject'] = subject
        msg['From'] = to_mail
        msg['To'] = to_mail
        try:
            mailobj = smtplib.SMTP(send_mail, 587)
            mailobj.ehlo()
            mailobj.starttls()
            mailobj.ehlo()
            mailobj.login(to_mail, password)
            mailobj.sendmail(to_mail, to_mail, msg.as_string())
            mailobj.quit()
            notice_flg = "通知済"
            print("送信成功")
        except Exception as e:
            print("送信失敗", e)

    # ----- 国/リーグの抽出（常に定義） -----
    country = ""
    league = ""
    if ":" in game_info_sub_info_text:
        country, rest = game_info_sub_info_text.split(":", 1)
        country = country.strip()
        league = rest.split("-")[0].strip()

    # ----- ブック/シートの準備（ここは例外に巻き込まない） -----
    OUTPUT_EXCEL_NAME = ""
    OUTPUT_EXCEL_PATH = ""
    write_bk = None
    sheet = None

    if loop_bef_seq < 0:
        for seq in range(1, 1200):
            p = f"/Users/shiraishitoshio/bookmaker/output_{seq}.xlsx"
            if os.path.exists(p):
                loop_bef_seq = seq
            else:
                write_bk = Workbook()
                sheet = write_bk.active
                loop_bef_seq = seq
                OUTPUT_EXCEL_NAME = f"output_{seq}.xlsx"
                OUTPUT_EXCEL_PATH = f"/Users/shiraishitoshio/bookmaker/{OUTPUT_EXCEL_NAME}"
                break

    if not OUTPUT_EXCEL_NAME:
        OUTPUT_EXCEL_NAME = f"output_{loop_bef_seq}.xlsx"
        OUTPUT_EXCEL_PATH = f"/Users/shiraishitoshio/bookmaker/{OUTPUT_EXCEL_NAME}"

    if write_bk is None:
        write_bk = load_workbook(OUTPUT_EXCEL_PATH) if os.path.exists(OUTPUT_EXCEL_PATH) else Workbook()
    sheet = write_bk.active

    # ----- ヘッダー定義 -----
    header_data = [
        "ホーム順位", "試合国及びカテゴリ", "試合時間", "ホームチーム", "ホームスコア", "アウェー順位", "アウェーチーム",
        "アウェースコア", "ホーム期待値", "アウェー期待値", "ホーム枠内ゴール期待値", "アウェー枠内ゴール期待値",
        "ホームボール支配率", "アウェーボール支配率", "ホームシュート数", "アウェーシュート数",
        "ホーム枠内シュート数", "アウェー枠内シュート数", "ホーム枠外シュート数", "アウェー枠外シュート数",
        "ホームブロックシュート", "アウェーブロックシュート", "ホームビッグチャンス", "アウェービッグチャンス",
        "ホームコーナーキック", "アウェーコーナーキック", "ホームボックス内シュート", "アウェーボックス内シュート",
        "ホームボックス外シュート", "アウェーボックス外シュート",
        "ホームゴールポスト", "アウェーゴールポスト", "ホームヘディングゴール", "アウェーヘディングゴール",
        "ホームキーパーセーブ", "アウェーキーパーセーブ", "ホームフリーキック", "アウェーフリーキック",
        "ホームオフサイド", "アウェーオフサイド", "ホームファウル", "アウェーファウル",
        "ホームイエローカード", "アウェーイエローカード", "ホームレッドカード", "アウェーレッドカード", "ホームスローイン", "アウェースローイン",
        "ホーム相手ボックスタッチ", "アウェー相手ボックスタッチ", "ホームパス", "アウェーパス", "ホームロングパス", "アウェーロングパス", "ホームファイナルサードパス", "アウェーファイナルサードパス",
        "ホームクロス", "アウェークロス", "ホームタックル", "アウェータックル", "ホームクリア", "アウェークリア", "ホームデュエル勝利数", "アウェーデュエル勝利数", 
        "ホームインターセプト", "アウェーインターセプト",
        "スコア時間", "天気", "気温", "湿度", "審判名", "ホーム監督名", "アウェー監督名", "ホームフォーメーション", "アウェーフォーメーション",
        "スタジアム", "収容人数", "観客数", "場所", "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者", "ホームチーム最大得点取得者出場状況", "アウェーチーム最大得点取得者出場状況",
        "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点", "ホームチームアウェー得点", "ホームチームアウェー失点",
        "アウェーチームアウェー得点", "アウェーチームアウェー失点", "通知フラグ", "試合リンク文字列", "ゴール時間", "選手名", "判定結果", "ホームチームスタイル", "アウェイチームスタイル",
        "ゴール確率", "得点予想時間"
    ]

    # ----- 取得時間など -----
    get_data_time = datetime.datetime.now()

    # ----- copydata に渡す行（xGOT含む） -----
    goal_time = ""   # 初期値
    goal_member = ""
    judge = ""
    row_data_for_copy = [
        home_rank, game_info_sub_info_text, live_fin_text, home_team_name, home_score, away_rank, away_team_name,
        away_score, home_exp, away_exp, home_exp_goal_in, away_exp_goal_in,
        home_domination, away_domination,
        home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in,
        home_shoot_out, away_shoot_out, home_block_shoot, away_block_shoot,
        home_big_chance, away_big_chance, home_corner, away_corner,
        home_shoot_box_in, away_shoot_box_in, home_shoot_box_out, away_shoot_box_out,
        home_goal_post, away_goal_post, home_head_goal, away_head_goal,
        home_keeper_save, away_keeper_save, home_free_kick, away_free_kick,
        home_offside, away_offside, home_foul, away_foul,
        home_yellow_card, away_yellow_card, home_red_card, away_red_card,
        home_slow_in, away_slow_in,
        home_box_touch, away_box_touch, home_pass, away_pass,
        home_long_pass, away_long_pass, home_final_pass, away_final_pass,
        home_cross, away_cross, home_tackle, away_tackle, home_clear, away_clear,
        home_duel, away_duel, home_intercept, away_intercept,
        get_data_time, weather, temperature, humid, judge_member, home_manager, away_manager,
        home_formation, away_formation, studium, capacity, audience, location,
        home_team_max_getting_scorer, away_team_max_getting_scorer,
        home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating,
        home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost,
        home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
        notice_flg, random_game_str, goal_time, goal_member, judge
    ]

    home_play_style = ""
    away_play_style = ""
    goal_prob = ""
    match_time = ""
    output_file_path = ""

    # ----- 予測系は try に分離（失敗してもシート書き込みは続行） -----
    try:
        home_play_style, away_play_style, goal_prob, match_time, output_file_path = \
            write_prediction_data.copydata(row_data_for_copy, country, league, home_team_name, away_team_name)
    except Exception as e:
        print("確率,時間,フォーメーション取得失敗", e)
        output_file_path = output_file_path or OUTPUT_EXCEL_PATH

    print("****確率,時間,フォーメーション****")
    print(goal_prob)
    print(match_time)
    print(home_play_style)
    print(away_play_style)
    print("****確率,時間,フォーメーション****")

    # ----- 実際に書き込む行（ヘッダと整合） -----
    row_data = [
        home_rank, game_info_sub_info_text, live_fin_text, home_team_name, home_score, away_rank, away_team_name,
        away_score, home_exp, away_exp, home_exp_goal_in, away_exp_goal_in,
        home_domination, away_domination,
        home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in,
        home_shoot_out, away_shoot_out, home_block_shoot, away_block_shoot,
        home_big_chance, away_big_chance, home_corner, away_corner,
        home_shoot_box_in, away_shoot_box_in, home_shoot_box_out, away_shoot_box_out,
        home_goal_post, away_goal_post, home_head_goal, away_head_goal,
        home_keeper_save, away_keeper_save, home_free_kick, away_free_kick,
        home_offside, away_offside, home_foul, away_foul,
        home_yellow_card, away_yellow_card, home_red_card, away_red_card, home_slow_in, away_slow_in,
        home_box_touch, away_box_touch, home_pass, away_pass, home_long_pass, away_long_pass, home_final_pass, away_final_pass,
        home_cross, away_cross, home_tackle, away_tackle, home_clear, away_clear, home_duel, away_duel, home_intercept, away_intercept,
        get_data_time, weather, temperature, humid, judge_member, home_manager, away_manager, home_formation, away_formation,
        studium, capacity, audience, location, home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating,
        home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score,
        away_team_away_lost, notice_flg, random_game_str, goal_time, goal_member, judge, home_play_style, away_play_style, str(goal_prob), str(match_time)
    ]

    # 初回のみヘッダー付与
    if sheet.max_row == 1:
        sheet.append(header_data)

    sheet.append(row_data)
    write_bk.save(OUTPUT_EXCEL_PATH)
    print("メイン書き込み完了")

    # 差分通知（ファイルパスのフォールバック）
    if not output_file_path:
        output_file_path = OUTPUT_EXCEL_PATH

    chk_diff_data_send_mail(
        output_file_path, game_info_sub_info_text, live_fin_text, home_team_name, away_team_name,
        home_score, away_score, header_data, row_data
    )

    # 条件式をCSVへ保存
    executeInspect()

    return loop_bef_seq

def conditionChk(game_info_sub_info_text, home_rank, away_rank, home_score, away_score, home_exp, away_exp, weather, temperature, humid,
                 home_shoot_all, away_shoot_all, home_shoot_in, away_shoot_in, live_fin_text,  home_formation, away_formation, judge_member, studium, capacity, audience, home_manager, away_manager,
                 home_team_home_score, home_team_home_lost, away_team_home_score, away_team_home_lost, home_team_away_score, home_team_away_lost, away_team_away_score, away_team_away_lost,
                 home_team_max_getting_scorer, away_team_max_getting_scorer, home_team_max_getting_scorer_participating, away_team_max_getting_scorer_participating):
    # --- 追加: None耐性 ---
    live_fin_text = s(live_fin_text)  # ここで必ず文字列化

    home_rank_rep = s(home_rank).replace(".", "") or "999"
    away_rank_rep = s(away_rank).replace(".", "") or "999"

    chkResult = []
    time_flg = ""

    # 90+ / 68' / 45:13 いずれでも落ちないように try をつける
    try:
        if "+" in live_fin_text:
            t = live_fin_text.replace("+", "").replace("'", "")
            if to_int(t, 99) < 20:
                time_flg = "OK"
        elif "'" in live_fin_text:
            t = live_fin_text.replace("'", "")
            if to_int(t, 99) < 20:
                time_flg = "OK"
        elif ":" in live_fin_text:
            t = live_fin_text.replace(":", "")
            if to_int(t, 9999) < 2000:
                time_flg = "OK"
    except Exception:
        pass

    print(live_fin_text)


    # アフリカの以下の国は順位がホームが上であること
    if "ケニア" in game_info_sub_info_text or "タンザニア" in game_info_sub_info_text or "ウガンダ" in game_info_sub_info_text:
        if (int(home_rank_rep) > int(away_rank_rep)) and (int(home_rank_rep) < 4 and int(away_rank_rep) > 9):
            chkResult.append("ケニア,タンザニア,ウガンダのいずれかです")
            chkResult.append(time_flg)
            return chkResult
        else:
            if (home_shoot_in != "" and home_shoot_all != "" and away_shoot_in != "" and away_shoot_all != "" and (
                    (float(int(home_shoot_in) / int(home_shoot_all)) >= 0.6) or (
                    float(int(away_shoot_in) / int(away_shoot_all)) >= 0.6))):
                chkResult.append("ケニア,タンザニア,ウガンダのいずれかで期待値は0.6以上です")
                chkResult.append(time_flg)
                return chkResult

    # 得点した方が得点数と枠内シュート数が同数（ただし1点）
    if time_flg == "OK" and (home_score == "1" or away_score == "1") and (home_score == home_shoot_all or away_score == away_shoot_all):
        chkResult.append("得点した方が得点数と枠内シュート数が同数（ただし1点）です")
        chkResult.append(time_flg)
        return chkResult

    # シュートが3本以上,枠内シュート2本以上,アウェーはシュートが2本以上,枠内シュート1本以上
    if (home_shoot_all != "" and home_shoot_in != "") and (away_shoot_all != "" and away_shoot_in != "") and (
            int(home_shoot_all) >= 3 and int(home_shoot_in) >= 2) and (
            int(away_shoot_all) >= 2 and int(away_shoot_in) >= 1):
        if ((float(int(home_shoot_in) / int(home_shoot_all)) >= 0.6) or (
                float(int(away_shoot_in) / int(away_shoot_all)) >= 0.6)):
            chkResult.append("シュートが3本以上,枠内シュート2本以上,アウェーはシュートが2本以上,枠内シュート1本以上です")
            chkResult.append(time_flg)
            return chkResult
        elif ((home_score == "1" or away_score == "1") and (
                float(int(home_shoot_in) / int(home_shoot_all)) >= 0.4) and (
                      float(int(away_shoot_in) / int(away_shoot_all)) >= 0.4)):
            chkResult.append("片方が1点取っており,どちらも4割以上の枠内シュートです。")
            chkResult.append(time_flg)
            return chkResult

    # ハーフタイム時に期待値が1.10以上,お互いシュート数が3本以上
    if live_fin_text == 'ハーフタイム' and home_exp != "" and float(home_exp) >= 1.1 and int(home_shoot_all) >= 3 and int(
            away_shoot_all) >= 3:
        chkResult.append("ハーフタイム時に期待値が1.10以上,お互いシュート数が3本以上です")
        chkResult.append(time_flg)
        return chkResult

    chkResult.append("")
    chkResult.append("")
    return chkResult


def noticeCronMail(cron_notice_data_list):
    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "cronの定期実行通知"
    body = "cronが実行されました。以下のライブデータに対してスクレイピングを行いました。"
    strings = ""
    for k in range(len(cron_notice_data_list)):
        stringstmp = ""
        data_list = cron_notice_data_list[k]
        chk = 0
        for l in range(len(data_list)):
            stringstmp += data_list[l]
            if chk == 0:
                stringstmp += ":"
            elif chk == 1:
                stringstmp += " vs "

            chk += 1
        if len(stringstmp) > 0:
            strings += "\n"
        strings += stringstmp

    body += "\n"
    body += "\n"
    body += strings

    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)

    print("cron実行通知を送信")

def executeInspect():
    source = isp.getsource(conditionChk)
    OUTPUT_CONDITION_PATH = "/Users/shiraishitoshio/bookmaker/conditiondata/conditiondata.csv"
    if os.path.exists(OUTPUT_CONDITION_PATH):
        os.remove(OUTPUT_CONDITION_PATH)
        
    with open(OUTPUT_CONDITION_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        for line in source.splitlines():
            writer.writerow([line])

def read_prediction_data(file_path):
    try:
        df = pd.read_excel(file_path)
        df.fillna(0, inplace=True)
        return df.tail(2) if len(df) >= 2 else df
    except Exception as e:
        print(f"Excelファイルの読み込みに失敗しました: {e}")
        return pd.DataFrame()

def diff_record(df, header_data):
    diff_list = []
    try:
        if len(df) < 2:
            print("データが2行未満のため、差分を取得できません。")
            return diff_list
    except Exception as e:
        print(f"Excelファイルの読み込みに失敗しました: {e}")
        return diff_list

    except_list = [
        "ホーム順位", "試合国及びカテゴリ", "試合時間", "ホームチーム", "ホームスコア", "アウェイ順位", "アウェイチーム",
        "アウェイスコア", "スコア時間", "天気", "気温", "湿度", "審判名", "ホーム監督名", "アウェイ監督名", "ホームフォーメーション", "アウェーフォーメーション",
        "スタジアム", "収容人数", "観客数", "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者", "ホームチーム最大得点取得者出場状況", "アウェーチーム最大得点取得者出場状況",
        "通知フラグ", "試合リンク文字列", "ゴール時間", "選手名", "判定結果", "ホームチームスタイル", "アウェイチームスタイル",
        "ゴール確率", "得点予想時間"
    ]

    three_data_list = [
        "ホームパス", "アウェイパス",
        "ホームファイナルサードパス", "アウェイファイナルサードパス",
        "ホームクロス", "アウェイクロス",
        "ホームタックル", "アウェイタックル"
    ]

    exp_list = [
        "ホーム期待値", "アウェイ期待値"
    ]

    percentage_list = [
        "ホームボール支配率", "アウェイボール支配率"
    ]

    
    for header in header_data:
        diff = ""
        if header in df.columns:  # カラムが存在するか確認
            if header in except_list:
                diff_list.append(diff)
                continue
            
            elif header in exp_list:
                bef = float(df.iloc[0][header])
                af = float(df.iloc[1][header])
                diff = af - bef
                if diff >= 0:
                    diff = "+" + str(diff)
                else:
                    diff = str(diff)

            elif header in percentage_list:
                bef = convert_feature_data(df.iloc[0][header])
                af = convert_feature_data(df.iloc[1][header])
                diff = af - bef
                diff *= 100
                if diff >= 0:
                    diff = "+" + str(diff) + "%"
                else:
                    diff = str(diff) + "%"

            elif header in three_data_list:
                success_bef, attempt_bef = extract_success_and_attempt(df.iloc[0][header])
                success_af, attempt_af = extract_success_and_attempt(df.iloc[1][header])
                success_diff = int(success_af) - int(success_bef)
                attempt_diff = int(attempt_af) - int(attempt_bef)
                success_diff_new = str(success_diff)
                attempt_diff_new = str(attempt_diff)
                if success_diff >= 0:
                    success_diff_new = "+" + success_diff_new
                if attempt_diff >= 0:
                    attempt_diff_new = "+" + attempt_diff_new
                    
                diff = success_diff_new + "/" + attempt_diff_new
            
            else:
                bef = int(df.iloc[0][header])
                af = int(df.iloc[1][header])
                diff = af - bef
                if diff >= 0:
                    diff = "+" + str(diff)
                else:
                    diff = str(diff)

        diff_list.append(diff)

    return diff_list

def convert_feature_data(data):
    print(data)
    data = str(data)
    if "%" in data:
        data = data.replace("%", "")
        return float(data) / 100
    else:
        return float(data)

def extract_success_and_attempt(value):
    """ 34%(20/60) の形式を (成功=20, 試行=60) に分割する """
    match = re.search(r"(\d+)/(\d+)", str(value))
    if match:
        success, attempt = map(int, match.groups())
        return success, attempt
    return 0, 0  # データがない場合は0を返す

def chk_diff_data_send_mail(file_path, game_info_sub_info_text, time, home, away, home_score, away_score, header_data, now_data):
    print("chk_diff_data_send_mail")
    df = read_prediction_data(file_path)
    if len(df) == 0:
        return
    
    diff_list = diff_record(df, header_data)
    if len(diff_list) == 0:
        return
    
    print(header_data)
    print(len(header_data))
    print(now_data)
    print(len(now_data))
    print(diff_list)
    print(len(diff_list))

    #通知するのは特定の特徴量が増加した時
    json_data = {
        key: f"{now} ({diff})"
        for key, now, diff in zip(header_data, now_data, diff_list)
        if str(diff) not in ["0", "+0", "+0/+0", "", "+0.0", "+0.0%"]
    }
    print(json_data)

    if not json_data:
        return

    send_mail = "smtp.gmail.com"
    to_mail = "hirokishiraishi73@gmail.com"
    password = "agqx lkpm dvnc bmua"

    subject = "前データとの差分データです"
    json_str = json.dumps(json_data, ensure_ascii=False, indent=4)

    # ボディ
    body = f"試合データ:{game_info_sub_info_text}\n{time}: {home}vs{away}({home_score}-{away_score})\n\n{json_str}"
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = to_mail
    msg['To'] = to_mail

    # 送信
    try:
        mailobj = smtplib.SMTP(send_mail, 587)
        mailobj.ehlo()
        mailobj.starttls()
        mailobj.ehlo()
        mailobj.login(to_mail, password)
        mailobj.sendmail(to_mail, to_mail, msg.as_string())
        mailobj.quit()
    except Exception as e:
        print("送信失敗", e)


    print("送信成功")


# 安全取得ヘルパ
def g(ms, key, default=""):
    v = ms.get(key) if isinstance(ms, dict) else None
    return default if v is None else v

def gp(ms, label, default=""):
    return g(ms, f"ホーム{label}", default), g(ms, f"アウェー{label}", default)

def merge_main_stats(lst):
    merged = {}
    for d in lst:
        merged.update(d)
    return merged

def s(x: object) -> str:
    """None→空文字、その他は文字列化"""
    return "" if x is None else str(x)

def to_int(x, default=0):
    try:
        return int(s(x).strip())
    except Exception:
        return default

def to_float(x, default=0.0):
    try:
        return float(s(x).strip())
    except Exception:
        return default

VAL_ANY = "[data-testid='wcl-scores-simple-text-01'], [data-testid='wcl-scores-simpleText-01']"

def combined_value(cell):
    # 例: ["78%", "(273/348)"] -> "78%(273/348)"
    parts = cell.locator(VAL_ANY).all_text_contents()
    parts = [p.strip() for p in parts if p and p.strip()]
    return "".join(parts) if parts else ""

import re

def get_display_time(page, container):
    """
    固定ヘッダーのステータス（終了済/ハーフタイム/延期/中断/延長/PK など）を優先。
    見つからなければ eventTime（33:12 / 72' / 90+2' 等）を返す。
    どれも無ければ空文字。
    """
    # 1) 可視の固定ヘッダーからステータスを読む
    status_loc = page.locator(
        ".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .fixedHeaderDuel__detailStatus"
    )
    if status_loc.count() > 0:
        status_text = (status_loc.first.text_content() or "").strip()
        # 「終了」「ハーフ」「延期」「中断」「延長」「PK」「ペナルティ」などを含むならそのまま採用
        if any(key in status_text for key in ["終了", "ハーフ", "延期", "中断", "延長", "PK", "ペナルティ", "試合前"]):
            return status_text

    # 2) 通常のタイム表示（本文側）
    et = container.locator(".eventAndAddedTime span.eventTime")
    if et.count() > 0:
        return (et.first.text_content() or "").strip()

    # 3) 固定ヘッダー側のタイム表示（可視のもの）
    et_fixed = page.locator(".fixedHeaderDuel:not(.fixedHeaderDuel--isHidden) .eventTime")
    if et_fixed.count() > 0:
        return (et_fixed.first.text_content() or "").strip()

    # 4) フォールバック：非可視ヘッダーにだけある場合（念のため）
    status_any = page.locator(".fixedHeaderDuel__detailStatus")
    if status_any.count() > 0:
        txt = (status_any.last.text_content() or "").strip()
        if txt:
            return txt

    return ""

def close_cookie_banners(page):
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('同意')", "button:has-text('許可')",
        "button:has-text('OK')", "button:has-text('Accept')",
        "[data-testid*='cookie'] button"
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click(timeout=1500)
        except Exception:
            pass

def find_table_scope(page, timeout_ms: int = 45000):
    page.wait_for_load_state("domcontentloaded")
    close_cookie_banners(page)

    # まずタブを有効化（ロケール揺れ対応）
    for sel in [
        "role=tab[name='ライブ順位表']",
        "role=tab[name='順位表']",
        "role=tab[name='Standings']",
        "text=ライブ順位表",
        "text=順位表",
        "text=Standings", "text=Table", "text=Tables",
    ]:
        try:
            t = page.locator(sel).first
            if t.count() and t.is_visible():
                t.click()
                # ライブはXHRが続くので短めスリープで十分
                time.sleep(0.8)
                break
        except Exception:
            pass

    # anti-bot壁などでDOMが生成されない場合の簡易検知
    wall = page.locator("text=/Just a moment|Access Denied|verify you are human|アクセス/i")
    if wall.count():
        raise RuntimeError("サイト側のボット対策でブロックされています。")

    # Shadow DOM/通常DOMの両方を許容する検知
    probe_selectors = [
        ":light(.ui-table__row)",               # Shadow DOM越しの行
        ":light(.tableCellParticipant__name)", # チーム名セル
        ":light(.tableCellRank)",              # 順位セル
        ".ui-table__row", ".tableCellParticipant__name", ".tableCellRank", # 通常DOM
    ]

    def has_table(scope) -> bool:
        for sel in probe_selectors:
            try:
                scope.locator(sel).first.wait_for(state="attached", timeout=2500)
                return True
            except Exception:
                continue
        return False

    # 本体
    if has_table(page):
        return page

    # すべてのiframe/子フレームを走査
    deadline = time.time() + timeout_ms/1000
    while time.time() < deadline:
        for fr in page.frames:
            try:
                if has_table(fr):
                    return fr
            except Exception:
                pass
        time.sleep(0.3)

    raise PTimeout("順位表のテーブルがどのフレームにも見つかりませんでした。")

def get_team_rank(container, team_name: str) -> int:
    # Shadow DOM越しも含めてアンカー列挙
    anchors = container.locator(":light(.ui-table__row) :light(.tableCellParticipant__name)")
    # ロード遅延に備えリトライ
    for _ in range(30):
        if anchors.count() > 0:
            break
        time.sleep(0.25)

    target = _norm(team_name)
    n = anchors.count()
    for i in range(n):
        a = anchors.nth(i)
        txt = (a.inner_text() or "").strip()
        if not txt:
            continue
        if _norm(txt) == target or target in _norm(txt):
            # 行（Shadow DOM越し）
            row = a.locator("xpath=ancestor::div[contains(@class,'ui-table__row')]").first
            # 順位セル（Shadow DOM越し + 通常DOM両対応）
            rank_el = row.locator(":light(.table__cell--rank) :light(.tableCellRank)").first
            rank_el.wait_for(state="attached", timeout=8000)

            rank_text = (rank_el.inner_text() or "").strip()
            if not rank_text:
                rank_text = (rank_el.get_attribute("title") or "").strip()
            m = re.search(r"\d+", rank_text)
            if not m:
                raise ValueError(f"順位数値が抽出できません: {rank_text!r}")
            return int(m.group())

    # 見つからなければサンプルを出してデバッグ
    samples = [anchors.nth(i).inner_text().strip() for i in range(min(15, anchors.count()))]
    raise ValueError(f"チームが見つかりません: {team_name!r} / 例: {samples}")

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("･", "・")
    return re.sub(r"\s+", "", s)

with sync_playwright() as playwright:
    execute(playwright)

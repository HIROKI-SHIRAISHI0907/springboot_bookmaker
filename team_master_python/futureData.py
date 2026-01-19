# -*- coding: utf-8 -*-
from playwright.sync_api import sync_playwright
import time
import re, os
import datetime
from typing import List, Dict, Optional
import pandas as pd
from pathlib import Path
import glob
import sys
try:
    import openpyxl
except ImportError:
    raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。")


# ============== 取得条件 ====================

HEADER_SCHEDULED = [
    "試合国及びカテゴリ", "試合予定時間", "ホーム順位", "アウェー順位", "ホームチーム", "アウェーチーム",
    "ホームチーム最大得点取得者", "アウェーチーム最大得点取得者",
    "ホームチームホーム得点", "ホームチームホーム失点", "アウェーチームホーム得点", "アウェーチームホーム失点",
    "ホームチームアウェー得点", "ホームチームアウェー失点",
    "アウェーチームアウェー得点", "アウェーチームアウェー失点", "試合リンク文字列", "データ取得時間"
]

# ===== Excel 出力関連 =====
# bmData.py と同様の outputs ディレクトリを想定
SAVE_DIR_SCHEDULED = "/Users/shiraishitoshio/bookmaker/future"

# ===== Excel 逐次書き込み（開催予定） =====
# 1ファイルあたりの最大データ行数（ヘッダ除く）
MAX_ROWS_PER_FILE_SCHEDULED = 10      # bmData.py と同じ値。必要なら後で増やしてください。
SHEET_NAME_SCHEDULED = "Sheet1"
FILE_PREFIX_SCHEDULED = "future_"
FILE_SUFFIX_SCHEDULED = ".xlsx"

def _existing_serials_scheduled(output_dir: str) -> List[int]:
    p = Path(output_dir)
    nums = []
    for f in p.glob(f"{FILE_PREFIX_SCHEDULED}*{FILE_SUFFIX_SCHEDULED}"):
        m = re.match(rf"^{re.escape(FILE_PREFIX_SCHEDULED)}(\d+){re.escape(FILE_SUFFIX_SCHEDULED)}$", f.name)
        if m:
            nums.append(int(m.group(1)))
    return sorted(nums)

def _next_serial_scheduled(output_dir: str) -> int:
    """既存の最大連番+1 を返す"""
    nums = _existing_serials_scheduled(output_dir)
    return (max(nums) + 1) if nums else 1

def _current_file_path_scheduled(output_dir: str) -> Path:
    """今使うべき future_*.xlsx のパスを返す。無ければ新規（連番）。"""
    p = Path(output_dir)
    nums = _existing_serials_scheduled(output_dir)
    if not nums:
        return p / f"{FILE_PREFIX_SCHEDULED}{_next_serial_scheduled(output_dir)}{FILE_SUFFIX_SCHEDULED}"
    # 直近のファイル（最大連番）
    return p / f"{FILE_PREFIX_SCHEDULED}{max(nums)}{FILE_SUFFIX_SCHEDULED}"

def _data_rows_in_scheduled(path: Path) -> int:
    """既存Excelのデータ行数（ヘッダ除く）を返す。無ければ0。"""
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME_SCHEDULED] if SHEET_NAME_SCHEDULED in wb.sheetnames else wb.active
    total = ws.max_row or 0
    wb.close()
    return max(0, total - 1)  # ヘッダ1行を除外

def _create_new_workbook_scheduled(path: Path):
    """HEADER_SCHEDULED 付きで新規 future_*.xlsx を作成"""
    df = pd.DataFrame(columns=HEADER_SCHEDULED)
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=SHEET_NAME_SCHEDULED)

def append_scheduled_row_to_excel(
    row_dict: Dict[str, str],
    output_dir: str = SAVE_DIR_SCHEDULED,
    max_rows_per_file: int = MAX_ROWS_PER_FILE_SCHEDULED
):
    """
    1 試合分の辞書 row_dict を
    - future_N.xlsx の末尾に追記
    - N ファイルごと（max_rows_per_file 行ごと）に新しい future_(N+1).xlsx を作る
    という形で保存する。
    """
    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)

    cur = _current_file_path_scheduled(output_dir_path)
    if not cur.exists():
        _create_new_workbook_scheduled(cur)

    # 現在のデータ行数（ヘッダ除く）
    current_rows = _data_rows_in_scheduled(cur)

    # 上限を超えたら次のファイルを作成
    if current_rows >= max_rows_per_file:
        cur = output_dir_path / f"{FILE_PREFIX_SCHEDULED}{_next_serial_scheduled(output_dir)}{FILE_SUFFIX_SCHEDULED}"
        _create_new_workbook_scheduled(cur)
        current_rows = 0

    # 追記用 DF を HEADER_SCHEDULED 順に整形
    df = pd.DataFrame([row_dict])
    for col in HEADER_SCHEDULED:
        if col not in df.columns:
            df[col] = ""
    df = df[HEADER_SCHEDULED]

    with pd.ExcelWriter(cur, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
        startrow = current_rows + 1  # ヘッダ1行あり
        df.to_excel(w, index=False, header=False, sheet_name=SHEET_NAME_SCHEDULED, startrow=startrow)

    print(f"💾 [FUTURE] 追記完了: {cur.name} （データ行 {current_rows} → {current_rows+1} 件目を追加）")

def save_scheduled_to_excel(match_results: List[Dict[str, str]], output_dir: str = SAVE_DIR_SCHEDULED):
    """
    開催予定データ（match_results）を Excel に保存する。
    bmData.py の append_row_to_excel と同じ考え方で、
    future_N.xlsx に逐次追記し、一定件数で次のファイルにローテーションする。
    """
    if not match_results:
        log("✋ Excel に書き込む開催予定データがありません（match_results が空）")
        return

    os.makedirs(output_dir, exist_ok=True)

    # 列ごとの非空件数をざっくりログ（まとめて）
    try:
        df_tmp = pd.DataFrame(match_results)
        for col in HEADER_SCHEDULED:
            if col not in df_tmp.columns:
                df_tmp[col] = ""
        df_tmp = df_tmp[HEADER_SCHEDULED]

        non_empty_counts = df_tmp.apply(lambda s: s.astype(str).str.strip().ne("").sum())
        log("📄 [EXCEL-SCHEDULED] 列ごとの非空件数（上位10列）:")
        top10 = non_empty_counts.sort_values(ascending=False).head(10)
        for col, cnt in top10.items():
            log(f"   - {col}: {cnt}")
        log(f"📄 [EXCEL-SCHEDULED] 総行数(今回追加分): {len(df_tmp)} / 総列数: {len(df_tmp.columns)}")
    except Exception as e:
        log(f"⚠️ [EXCEL-SCHEDULED] 非空件数計算で例外: {e}")

# ==========================================
# ✅ 対象リーグフィルタリング用リスト
# ==========================================
CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]
UNDER_LIST  = ["U17", "U18", "U19", "U20", "U21", "U22", "U23", "U24", "U25"]
GENDER_LIST = ["女子"]
EXP_LIST    = ["ポルトガル: リーガ・ポルトガル 2", "イングランド: プレミアリーグ 2", "イングランド: プレミアリーグ U18"]

# ============== 共通ユーティリティ ==============

VERBOSE = True

def log(msg: str):
    if VERBOSE:
        print(msg)

def text_clean(s: str) -> str:
    import re
    return re.sub(r"\s+", " ", (s or "")).strip()

def extract_mid(s: str) -> Optional[str]:
    """URLから mid を抽出"""
    if not s:
        return None
    s = str(s).strip()
    m = re.search(r"[?&#]mid=([A-Za-z0-9]+)", s)
    if m:
        return m.group(1)
    m = re.search(r"/match/([A-Za-z0-9]{6,20})(?:/|$)", s)
    if m:
        return m.group(1)
    return None

# ============== Flashscore「開催予定」ナビ ==============

def goto_football_top(page):
    """サッカー → 開催予定タブへ遷移"""
    log("🌐 Flashscore トップへアクセス...")
    page.goto("https://www.flashscore.co.jp/", timeout=45000, wait_until="domcontentloaded")

    # Cookieバナーなど
    try:
        page.locator("#onetrust-accept-btn-handler").click(timeout=2000)
        log("✅ Cookieバナーを閉じました")
    except:
        pass

    # サッカーが選ばれていない場合に備えて「サッカー」クリック（保険）
    try:
        soccer_btn = page.locator("a,button").filter(has_text="サッカー").first
        if soccer_btn and soccer_btn.count():
            soccer_btn.click(timeout=4000)
            time.sleep(0.8)
    except:
        pass

    # 「開催予定」タブをクリック
    try:
        tab = page.locator("div.filters__tab[data-analytics-alias='scheduled']").first
        if tab and tab.count():
            tab.click(timeout=4000)
        else:
            tab = page.locator("div.filters__tab").filter(
                has_text=re.compile(r"(開催予定)")
            ).first
            tab.click(timeout=4000)
        log("✅ 『開催予定』タブに切り替えました")
    except Exception as e:
        log(f"⚠️ 開催予定タブ切り替え失敗: {e}")

    # 試合行が描画されるまで少し待つ
    try:
        page.wait_for_timeout(1000)
        page.wait_for_load_state("networkidle", timeout=8000)
    except:
        pass

def expand_all_collapsed_leagues(page):
    """
    Flashscore『開催予定』タブで、
    折りたたまれているリーグ（＝リーグ全試合が非表示）をすべて展開する。
    """
    print("📂 折りたたみリーグ（非表示）を展開します...")

    btn_selector = (
        "button[data-testid='wcl-accordionButton']"
        "[aria-label='リーグ全試合 表示']"
    )

    max_loops = 200  # 念のため安全弁

    for _ in range(max_loops):
        btns = page.locator(btn_selector)
        count = btns.count()

        if count == 0:
            print("   ✅ すべての折りたたみリーグを展開しました")
            return

        print(f"   残り『表示』ボタン数: {count}")

        btn = btns.first

        try:
            btn.scroll_into_view_if_needed()
        except Exception:
            pass

        try:
            btn.click(timeout=2000)
        except Exception as e:
            print(f"   ⚠️ ボタンクリック失敗: {e}")
            break

        page.wait_for_timeout(200)

    print("   ⚠️ ループ上限に達しました。まだ非表示リーグが残っている可能性があります。")

# ============== 試合行の基本情報取得 ==============

def _get_match_row_teams_and_time(row):
    """
    1つの試合行から
      - 試合予定時間
      - ホームチーム名
      - アウェーチーム名
    を抽出（旧UI / 新UI 両対応）
    """
    # ===== 時間 =====
    ktime = ""

    # 旧UI: .event__time
    try:
        ktime = text_clean(row.locator(".event__time").first.text_content() or "")
    except Exception:
        pass

    # 新UI: data-testid ベース
    if not ktime:
        try:
            ktime = text_clean(
                row.locator(
                    "[data-testid='wcl-time'], "
                    "[data-testid='wcl-start-time'], "
                    "[data-testid='wcl-time-status']"
                ).first.text_content() or ""
            )
        except Exception:
            pass

    # ===== チーム名 =====
    home = ""
    away = ""

    # パターン①: 旧UI (.event__participant--home / --away)
    try:
        h = row.locator(".event__participant--home .event__participant--name").first
        a = row.locator(".event__participant--away .event__participant--name").first
        if h.count():
            home = text_clean(h.text_content() or "")
        if a.count():
            away = text_clean(a.text_content() or "")
    except Exception:
        pass

    # パターン②: 旧UI フォールバック (.event__participant)
    if not home or not away:
        try:
            ps = row.locator(".event__participant .event__participant--name")
            if ps.count() >= 2:
                if not home:
                    home = text_clean(ps.first.text_content() or "")
                if not away:
                    away = text_clean(ps.last.text_content() or "")
        except Exception:
            pass

    # パターン③: 新UI 明示的セレクタ
    if not home or not away:
        try:
            h = row.locator(
                ".event__homeParticipant span.wcl-name_jjfMf, "
                ".event__homeParticipant [data-testid='wcl-scores-simple-text-01']"
            ).first
            a = row.locator(
                ".event__awayParticipant span.wcl-name_jjfMf, "
                ".event__awayParticipant [data-testid='wcl-scores-simple-text-01']"
            ).first

            if h and h.count():
                if not home:
                    home = text_clean(h.text_content() or "")
            if a and a.count():
                if not away:
                    away = text_clean(a.text_content() or "")
        except Exception:
            pass

    # パターン④: 新UI ゆるめ
    if not home or not away:
        try:
            ps = row.locator(
                "[data-testid='wcl-matchRow-participant'] span.wcl-name_jjfMf, "
                "[data-testid='wcl-matchRow-participant'] [data-testid='wcl-scores-simple-text-01']"
            )
            n = ps.count()
            if n >= 2:
                if not home:
                    home = text_clean(ps.nth(0).text_content() or "")
                if not away:
                    away = text_clean(ps.nth(n - 1).text_content() or "")
        except Exception:
            pass

    # パターン⑤: <img alt="チーム名">
    if not home or not away:
        try:
            imgs = row.locator("[data-testid='wcl-matchRow-participant'] img[data-testid='wcl-participantLogo']")
            n_img = imgs.count()
            if n_img >= 2:
                if not home:
                    alt0 = imgs.nth(0).get_attribute("alt") or ""
                    home = text_clean(alt0)
                if not away:
                    alt1 = imgs.nth(n_img - 1).get_attribute("alt") or ""
                    away = text_clean(alt1)
        except Exception:
            pass

    if not home or not away:
        try:
            snippet = (row.inner_text() or "").strip().replace("\n", " ")[:200]
        except Exception:
            snippet = "<inner_text 取得失敗>"
        print(f"⚠️ チーム名取得失敗: time={ktime}, snippet={snippet}")

    return ktime, home, away

def _get_match_row_link(row) -> str:
    """試合行から matchURL を抽出"""
    try:
        a = row.locator("a.eventRowLink[href*='/match/'][href*='?mid=']").first
        if a and a.count():
            href = a.get_attribute("href") or ""
            if href.startswith("http"):
                return href
            return "https://www.flashscore.co.jp" + href
    except:
        pass
    return ""

# ============== 日付 ==============

def get_current_match_date(page) -> Optional[datetime.date]:
    """
    Flashscore トップの日付ボタン（例: '05/12 金'）から
    datetime.date を作って返す。
    ボタンが取れなかったりパースできなければ None を返す。
    """
    try:
        btn = page.locator("button[data-testid='wcl-dayPickerButton']").first
        if not btn or not btn.count():
            return None

        txt = text_clean(btn.inner_text() or "")
        # 例: "05/12 金" → day=05, month=12
        m = re.search(r"(\d{2})/(\d{2})", txt)
        if not m:
            return None

        day = int(m.group(1))
        month = int(m.group(2))

        # 年はとりあえず今年を採用（年またぎは簡易対応）
        year = datetime.datetime.now().year
        return datetime.date(year, month, day)

    except Exception:
        return None

# ============== 開催予定タブから基本情報だけ集める ==============

def collect_scheduled_matches_on_current_day(page) -> List[Dict[str, str]]:
    """
    現在表示中の日付（開催予定タブ）から試合情報を収集。
    ここでは「時間・ホーム・アウェー・リンク」だけを集め、
    国リーグ＆順位は後で試合ページから埋める。
    """
    try:
        page.wait_for_selector("div.event__match", timeout=12000)
    except:
        log("⚠️ event__match が見つからないまま続行")

    expand_all_collapsed_leagues(page)

    # 日付取得
    match_date = get_current_match_date(page)

    rows = page.locator("div.event__match.event__match--scheduled")
    if rows.count() == 0:
        rows = page.locator("div.event__match")
    n = rows.count()
    log(f"🎯 開催予定試合 行数: {n}")

    results: List[Dict[str, str]] = []
    seen_mids = set()

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i in range(n):
        row = rows.nth(i)
        try:
            ktime, home, away = _get_match_row_teams_and_time(row)
            link = _get_match_row_link(row)

            # 🔹 日付＋時間の文字列を作成
            if match_date and ktime:
                # 例: "2025-12-05 04:00"
                match_dt_str = f"{match_date.strftime('%Y-%m-%d')} {ktime}"
            else:
                match_dt_str = ktime  # フォールバック

            log(f"row:  {match_dt_str}, {home}, {away}, {link}")

            mid = extract_mid(link)
            if mid and mid in seen_mids:
                log(f"   ⏭️ 重複試合(mid={mid})をスキップ")
                continue
            if mid:
                seen_mids.add(mid)

            d = {k: "" for k in HEADER_SCHEDULED}
            d["試合国及びカテゴリ"] = ""  # ← 後で埋める
            d["試合予定時間"]       = match_dt_str
            d["ホームチーム"]       = home
            d["アウェーチーム"]     = away
            d["試合リンク文字列"]   = link
            d["データ取得時間"]     = now_str

            results.append(d)

            log(f"   [{i+1}/{n}] | {ktime} | {home} vs {away} | mid={mid}")
        except Exception as e:
            log(f"   ⚠️ 行{i}でエラー: {e}")
            continue

    log(f"✅ 当日分 取得件数: {len(results)}")
    return results

def click_next_day(page) -> bool:
    """
    「翌日」ボタンをクリックして日付を1日進める。
    成功したら True, 見つからなければ False。
    """
    try:
        btn = page.locator("button.wcl-arrow_YpdN4[data-day-picker-arrow='next']").first
        if not btn or not btn.count():
            log("⚠️ 翌日ボタンが見つかりませんでした")
            return False
        btn.click(timeout=3000)
        log("➡️ 『翌日』ボタンをクリックしました")
        time.sleep(1.0)
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except:
            pass
        return True
    except Exception as e:
        log(f"⚠️ 翌日ボタンクリック失敗: {e}")
        return False

# ============== 順位表 & 国リーグ取得 ==============

def build_standings_url_from_match_url(match_url: str) -> str:
    """
    試合ページURLを順位表タブのURLに変換する。
    """
    if not match_url:
        return ""
    if "/standings/" in match_url:
        return match_url
    return re.sub(r"/(\?mid=)", r"/standings/\1", match_url, count=1)

def get_team_ranks_from_standings_table(page, home_name: str, away_name: str):
    """
    すでに「順位表」タブ（オーバーオール）が表示されている page から、
    ホーム＆アウェーチームの順位（rank）を取得する。
    """
    home_name_norm = text_clean(home_name)
    away_name_norm = text_clean(away_name)

    home_rank = ""
    away_rank = ""

    rows = page.locator("div.ui-table__body > div.ui-table__row")
    n_rows = rows.count()

    for i in range(n_rows):
        row = rows.nth(i)
        try:
            team_elem = row.locator(".tableCellParticipant__name").first
            if not team_elem.count():
                continue
            team_name = text_clean(team_elem.text_content() or "")

            rank_elem = row.locator(".tableCellRank").first
            if not rank_elem.count():
                continue
            rank_text = text_clean(rank_elem.text_content() or "")
            rank_text = rank_text.rstrip(".").strip()

            if not home_rank and team_name == home_name_norm:
                home_rank = rank_text
            if not away_rank and team_name == away_name_norm:
                away_rank = rank_text

            if home_rank and away_rank:
                break
        except:
            continue

    return home_rank, away_rank

def get_country_and_league_from_match_page(page):
    """
    試合ページ（サマリー / 順位表タブ）のパンくずから
    - 国名
    - リーグ名（＋ラウンド）
    を取得する。
    """
    country = ""
    league = ""

    try:
        try:
            page.wait_for_selector(
                "nav[data-testid='wcl-breadcrumbs'] span[data-testid='wcl-scores-overline-03']",
                timeout=3000
            )
        except Exception:
            pass

        spans = page.locator(
            "nav[data-testid='wcl-breadcrumbs'] span[data-testid='wcl-scores-overline-03']"
        )
        count = spans.count()

        if count == 0:
            return "", ""

        texts = []
        for i in range(count):
            txt = text_clean(spans.nth(i).text_content() or "")
            if txt:
                texts.append(txt)

        # 期待パターン:
        #   0: サッカー
        #   1: イングランド（国）
        #   2: プレミアリーグ - ラウンド 14（リーグ）
        start_idx = 0
        if texts and texts[0] == "サッカー":
            start_idx = 1

        if len(texts) > start_idx:
            country = texts[start_idx]
        if len(texts) > start_idx + 1:
            league = texts[start_idx + 1]

    except Exception:
        pass

    return country, league

def fetch_ranks_for_match(page, match_url: str, home_name: str, away_name: str):
    """
    試合URLから順位表タブに飛び、
      - ホーム順位
      - アウェー順位
      - 国名
      - リーグ名（ラウンド付き）
    を取得する。
    """
    if not match_url:
        return "", "", "", ""

    standings_url = build_standings_url_from_match_url(match_url)

    try:
        log(f"   📊 順位表取得: {standings_url}")
        page.goto(standings_url, timeout=25000, wait_until="domcontentloaded")

        try:
            standings_tab = page.locator(
                "a[data-analytics-alias='stats-detail'] button, "
                "a[href*='/standings/'] button"
            ).first
            if standings_tab and standings_tab.count():
                selected = standings_tab.get_attribute("data-selected")
                if selected != "true":
                    standings_tab.click(timeout=3000)
                    page.wait_for_timeout(500)
        except:
            pass

        country, league = get_country_and_league_from_match_page(page)

        page.wait_for_selector("div.ui-table__body div.ui-table__row", timeout=12000)

        home_rank, away_rank = get_team_ranks_from_standings_table(page, home_name, away_name)
        log(f"      → rank: home={home_rank}, away={away_rank}, country={country}, league={league}")
        return home_rank, away_rank, country, league

    except Exception as e:
        log(f"   ⚠️ 順位表取得失敗: {e}")
        return "", "", "", ""

# ============== 全試合に対して 国リーグ & 順位 & フィルタ ==============

def fill_ranks_for_matches(ctx, matches: List[Dict[str, str]]):
    """
    「開催予定」で収集した試合リストに対して、
    各試合ページの順位表から
      - ホーム順位
      - アウェー順位
      - 試合国及びカテゴリ（国: リーグ）
    を埋め、かつ対象リーグだけ残す。
    """
    if not matches:
        return

    page = ctx.new_page()

    filtered: List[Dict[str, str]] = []

    for idx, m in enumerate(matches):
        url  = m.get("試合リンク文字列") or ""
        home = m.get("ホームチーム") or ""
        away = m.get("アウェーチーム") or ""

        if not url or not home or not away:
            log(f"⏭️ URL/チーム名不足のためスキップ: {home} vs {away}")
            continue

        log(f"=== 順位取得 {idx+1}/{len(matches)} ===")
        home_rank, away_rank, country, league = fetch_ranks_for_match(page, url, home, away)

        # 国＋リーグからカテゴリ文字列を組み立て
        game_category = ""
        if country and league:
            game_category = f"{country}: {league}"
        elif country or league:
            game_category = country or league

        if not game_category:
            log("⏭️ スキップ対象: （カテゴリ取得失敗）")
            continue

        # 1) CONTAINS_LIST に含まれるリーグだけ対象
        if not any(c in game_category for c in CONTAINS_LIST):
            log(f"⏭️ スキップ対象: {game_category}（リスト外）")
            continue

        # 2) 年代（Uxx）・女子・例外リーグを含む場合はスキップ
        if (any(x in game_category for x in UNDER_LIST) or
            any(x in game_category for x in GENDER_LIST) or
            any(x in game_category for x in EXP_LIST)):
            log(f"🚫 除外対象: {game_category}")
            continue

        # ここまで来た試合だけを採用
        if home_rank:
            m["ホーム順位"] = home_rank
        if away_rank:
            m["アウェー順位"] = away_rank
        m["試合国及びカテゴリ"] = game_category

        # 🔹 ここで即 Excel に1行追記
        append_scheduled_row_to_excel(m, output_dir=SAVE_DIR_SCHEDULED)

        filtered.append(m)
        log(f"✅ 対象試合: {game_category} | {home} vs {away}")

    page.close()

    # 元のリストを書き換え
    matches[:] = filtered


# ============== メイン入口 ==============

def fetch_scheduled_matches(days) -> List[Dict[str, str]]:
    """
    Flashscore 開催予定タブから、
      - 今日（表示中の日付）
      - 翌日以降（days-1回『翌日』をクリック）
    の試合情報を取得してリストで返す。

    days=2 → 今日＋翌日
    """
    all_results: List[Dict[str, str]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, slow_mo=70)
        ctx = browser.new_context(
            user_agent=("Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
            locale="ja-JP",
            timezone_id="Asia/Tokyo",
        )
        page = ctx.new_page()

        goto_football_top(page)

        # ① 開催予定タブから全試合を集める（カテゴリはまだ空）
        for day_idx in range(days):
            if day_idx > 0:
                ok = click_next_day(page)
                if not ok:
                    log("⏭️ 翌日への遷移ができなかったため、以降の取得はスキップします")
                    break
            log(f"==================== 日付オフセット {day_idx} 日目 ====================")
            day_results = collect_scheduled_matches_on_current_day(page)
            all_results.extend(day_results)

        page.close()

        # ② 各試合ページの「順位表」から
        #    ホーム順位 / アウェー順位 / 国＆リーグを埋めつつ、
        #    対象リーグだけにフィルタ
        fill_ranks_for_matches(ctx, all_results)

        browser.close()

    log(f"🎉 総取得件数: {len(all_results)}")
    return all_results

# ==========================================
# ✅ future_*.xlsx → future_*.csv 変換
# ==========================================

def get_existing_future_xlsx_seqs(base_dir: str = SAVE_DIR_SCHEDULED) -> List[int]:
    """
    future_*.xlsx の通番一覧を昇順で返す。
    例: future_1.xlsx, future_3.xlsx → [1, 3]
    """
    pattern = os.path.join(base_dir, "future_*.xlsx")
    xlsx_files = glob.glob(pattern)
    seqs: List[int] = []

    for path in xlsx_files:
        basename = os.path.basename(path)
        if not basename.startswith("future_") or not basename.endswith(".xlsx"):
            continue
        try:
            num = int(basename.replace("future_", "").replace(".xlsx", ""))
            seqs.append(num)
        except ValueError:
            continue

    if not seqs:
        print("変換対象の future_*.xlsx ファイルが見つかりません。")
        return []

    print("対象Excelファイル数:", len(seqs))
    return sorted(seqs)


def excel_to_csv(excel_file: str, csv_file: str):
    """
    1つの Excel (future_N.xlsx) を CSV (future_N.csv) に変換し、
    変換に成功したら元の Excel を削除する。
    """
    try:
        df = pd.read_excel(excel_file)
        df.to_csv(csv_file, index=False)
        print(f"✅ Excel -> CSV 変換完了: {os.path.basename(excel_file)} → {os.path.basename(csv_file)}")
        os.remove(excel_file)
        print(f"🗑 元Excel削除: {os.path.basename(excel_file)}")
    except Exception as e:
        print(f"⚠️ Excel -> CSV変換失敗: {excel_file} ({e})")


def convert_all_future_excels_to_csv(base_dir: str = SAVE_DIR_SCHEDULED):
    """
    base_dir 配下の future_*.xlsx をすべて future_*.csv に変換する。
    すでに同名の future_N.csv がある場合はスキップ。
    """
    seq_list_all = get_existing_future_xlsx_seqs(base_dir)
    if not seq_list_all:
        return

    for seq in seq_list_all:
        xlsx_name = f"future_{seq}.xlsx"
        csv_name  = f"future_{seq}.csv"
        xlsx_path = os.path.join(base_dir, xlsx_name)
        csv_path  = os.path.join(base_dir, csv_name)

        if not os.path.exists(xlsx_path):
            continue

        # 既に同じ番号のCSVがある場合は念のためスキップ
        if os.path.exists(csv_path):
            print(f"⏭️ 既にCSVが存在するためスキップ: {csv_name}")
            continue

        excel_to_csv(xlsx_path, csv_path)

if __name__ == "__main__":
    matches = fetch_scheduled_matches(days=7)
    print(f"総件数: {len(matches)}")

    # 🔹 Excel に保存
    save_scheduled_to_excel(matches, output_dir=SAVE_DIR_SCHEDULED)

    # 🔹 保存された future_*.xlsx を CSV に変換
    convert_all_future_excels_to_csv(base_dir=SAVE_DIR_SCHEDULED)

    if matches:
        from pprint import pprint
        pprint(matches[0])

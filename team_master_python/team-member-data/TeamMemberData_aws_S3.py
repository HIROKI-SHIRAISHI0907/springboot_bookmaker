# -*- coding: utf-8 -*-
"""
TeamMemberECS.py（ECS用）
Flashscore: teamData_*.csv（チーム一覧） -> teamMemberData_*.csv（選手詳細）を生成しS3へ

- 作業フォルダ: /tmp/bookmaker

【S3】
- 入力(チーム一覧): aws-s3-team-csv （バケット直下 teamData_*.csv）
- 入力JSON(任意): aws-s3-team-member-csv/json/b001_country_league.json
- 出力(選手CSV): aws-s3-team-member-csv （バケット直下 teamMemberData_*.csv）
"""

import json
from pathlib import Path
import os
import re
import glob
import csv
import asyncio
import datetime
from typing import List, Tuple, Dict, Set, Optional

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright
from playwright.async_api import Error as PwError

# =========================
# S3（固定）
# =========================
import boto3
from botocore.exceptions import ClientError

S3_BUCKET_TEAM = "aws-s3-team-csv"
S3_BUCKET_MEMBER = "aws-s3-team-member-csv"

# JSON（任意：member側に置く想定）
B001_S3_KEY = "json/b001_country_league.json"
B001_JSON_PATH = "/tmp/bookmaker/json/b001/b001_country_league.json"

# 入力 teamData（teamバケット直下）
TEAMDATA_S3_PREFIX = "teamData_"
TEAMDATA_S3_SUFFIX = ".csv"

_s3 = boto3.client("s3")


def s3_download_if_exists(bucket: str, key: str, local_path: str) -> bool:
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    try:
        _s3.download_file(bucket, key, local_path)
        print(f"[S3 DOWNLOAD] s3://{bucket}/{key} -> {local_path}")
        return True
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "Unknown")
        print(f"[S3] not found or no permission: s3://{bucket}/{key} (Code={code})")
        return False


def s3_upload_to_bucket_root(bucket: str, local_path: str) -> bool:
    key = os.path.basename(local_path)
    try:
        _s3.upload_file(local_path, bucket, key)
        print(f"[S3 UPLOAD] {local_path} -> s3://{bucket}/{key}")
        return True
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "Unknown")
        print(f"[S3 ERROR] upload failed: {local_path} -> s3://{bucket}/{key} (Code={code})")
        return False


def s3_list_teamdata_keys(bucket: str) -> List[str]:
    """
    バケット直下の teamData_*.csv を列挙
    """
    keys: List[str] = []
    token = None
    while True:
        kwargs = {"Bucket": bucket, "MaxKeys": 1000, "Prefix": TEAMDATA_S3_PREFIX}
        if token:
            kwargs["ContinuationToken"] = token
        resp = _s3.list_objects_v2(**kwargs)

        for obj in resp.get("Contents", []) or []:
            k = obj.get("Key", "")
            if not k:
                continue
            # 直下のみ（フォルダ配下は対象外）
            if "/" in k:
                continue
            if k.startswith(TEAMDATA_S3_PREFIX) and k.endswith(TEAMDATA_S3_SUFFIX):
                keys.append(k)

        if resp.get("IsTruncated"):
            token = resp.get("NextContinuationToken")
        else:
            break

    def _num(k: str) -> int:
        m = re.fullmatch(rf"{re.escape(TEAMDATA_S3_PREFIX)}(\d+)\.csv", os.path.basename(k))
        return int(m.group(1)) if m else 10**18

    keys.sort(key=lambda x: (_num(x), x))
    return keys


def s3_download_teamdata_all(bucket: str, local_dir: str) -> List[str]:
    os.makedirs(local_dir, exist_ok=True)
    keys = s3_list_teamdata_keys(bucket)
    if not keys:
        print(f"[S3] teamData_*.csv not found in bucket root: s3://{bucket}/")
        return []

    local_paths: List[str] = []
    for k in keys:
        lp = os.path.join(local_dir, os.path.basename(k))
        try:
            _s3.download_file(bucket, k, lp)
            local_paths.append(lp)
            print(f"[S3 DOWNLOAD] s3://{bucket}/{k} -> {lp}")
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code", "Unknown")
            print(f"[S3 ERROR] download failed: s3://{bucket}/{k} (Code={code})")
    return local_paths


def getenv_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, default))
    except Exception:
        return default


def getenv_str(name: str, default: str) -> str:
    v = os.getenv(name)
    if v is None or v == "":
        return default
    return str(v)


import sys
import time


def log(msg: str):
    # 1行単位で即時CloudWatchに流す
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{ts} {msg}", flush=True)
    # 念のため（stdoutが特殊なとき）
    try:
        sys.stdout.flush()
    except Exception:
        pass


HEARTBEAT_SEC = getenv_int("HEARTBEAT_SEC", 30)  # 例: 30秒ごと
PROGRESS_SEC = getenv_int("PROGRESS_SEC", 60)  # 例: 60秒ごと


class Progress:
    """
    全体の進捗を一定間隔で出す（ログが止まって見える問題を潰す）
    """

    def __init__(self):
        self.started = time.time()
        self.last_progress = 0.0
        self.counters = {
            "leagues_total": 0,
            "leagues_done": 0,
            "teams_total": 0,
            "teams_done": 0,
            "players_enqueued": 0,
            "player_detail_ok": 0,
            "player_detail_fail": 0,
            "goto_fail": 0,
        }

    def inc(self, key: str, n: int = 1):
        if key in self.counters:
            self.counters[key] += n

    def set(self, key: str, val: int):
        self.counters[key] = val

    def maybe_log(self, force: bool = False):
        now = time.time()
        if (not force) and (now - self.last_progress < PROGRESS_SEC):
            return
        self.last_progress = now

        elapsed = int(now - self.started)
        c = self.counters
        log(
            "[PROGRESS] "
            f"elapsed={elapsed}s "
            f"leagues={c['leagues_done']}/{c['leagues_total']} "
            f"teams={c['teams_done']}/{c['teams_total']} "
            f"enq={c['players_enqueued']} "
            f"detail_ok={c['player_detail_ok']} "
            f"detail_fail={c['player_detail_fail']} "
            f"goto_fail={c['goto_fail']}"
        )


progress = Progress()


async def heartbeat_task():
    try:
        while True:
            await asyncio.sleep(HEARTBEAT_SEC)
            progress.maybe_log(force=True)
    except asyncio.CancelledError:
        # 正常終了（キャンセルされたら静かに終わる）
        return

# =========================
# 設定（ECS固定）
# =========================
BASE_DIR = "/tmp/bookmaker"
TEAMS_BY_LEAGUE_DIR = os.path.join(BASE_DIR, "teams_by_league")

TEAMDATA_PREFIX = "teamData_"
TEAMDATA_GLOB = os.path.join(TEAMS_BY_LEAGUE_DIR, f"{TEAMDATA_PREFIX}*.csv")

EXCEL_BASE_PREFIX = "teamMemberData_"
EXCEL_MAX_RECORDS = 50

CONTAINS_LIST = [
    "ケニア: プレミアリーグ",
    "コロンビア: プリメーラ A",
    "タンザニア: プレミアリーグ",
    "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ",
    "イングランド: EFL リーグ 1",
    "エチオピア: プレミアリーグ",
    "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ",
    "スペイン: ラ・リーガ",
    "ブラジル: セリエ A ベターノ",
    "ブラジル: セリエ B",
    "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1",
    "中国: 中国スーパーリーグ",
    "日本: J1 リーグ",
    "日本: J2 リーグ",
    "日本: J3 リーグ",
    "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン",
    "チュニジア: チュニジア･プロリーグ",
    "ウガンダ: プレミアリーグ",
    "メキシコ: リーガ MX",
    "フランス: リーグ・アン",
    "スコットランド: プレミアシップ",
    "オランダ: エールディビジ",
    "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A",
    "イタリア: セリエ B",
    "ポルトガル: リーガ・ポルトガル",
    "トルコ: スュペル・リグ",
    "セルビア: スーペルリーガ",
    "日本: WEリーグ",
    "ボリビア: LFPB",
    "ブルガリア: パルヴァ・リーガ",
    "カメルーン: エリート 1",
    "ペルー: リーガ 1",
    "エストニア: メスタリリーガ",
    "ウクライナ: プレミアリーグ",
    "ベルギー: ジュピラー･プロリーグ",
    "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ",
    "日本: 天皇杯",
]

# ---- Timeouts / retries ----
OP_TIMEOUT = getenv_int("OP_TIMEOUT_MS", 15000)
NAV_TIMEOUT = getenv_int("NAV_TIMEOUT_MS", 60000)
SEL_TIMEOUT = getenv_int("SEL_TIMEOUT_MS", 60000)
SAFE_GOTO_TRIES = getenv_int("SAFE_GOTO_TRIES", 5)

# ---- Fast-fail for squad ----
SQUAD_NAV_TIMEOUT = getenv_int("SQUAD_NAV_TIMEOUT_MS", 15000)  # 15s
SQUAD_SEL_TIMEOUT = getenv_int("SQUAD_SEL_TIMEOUT_MS", 15000)  # 15s
SQUAD_TRIES = getenv_int("SQUAD_TRIES", 1)  # 1回で諦める

WAIT_UNTIL = getenv_str("WAIT_UNTIL", "commit")  # commit / domcontentloaded など

LEAGUE_CONCURRENCY = getenv_int("LEAGUE_CONCURRENCY", 1)
TEAMS_PER_LEAGUE_CONCURRENCY = getenv_int("TEAMS_PER_LEAGUE_CONCURRENCY", 2)

PLAYER_PER_TEAM = getenv_int("PLAYER_PER_TEAM", 1)
GLOBAL_PLAYERDETAIL_CONCURRENCY = getenv_int("GLOBAL_PLAYERDETAIL_CONCURRENCY", 2)

GLOBAL_NAV_CONCURRENCY = getenv_int("GLOBAL_NAV_CONCURRENCY", 1)
NAV_GAP_MS = getenv_int("NAV_GAP_MS", 700)

SAVE_INTERVAL_SEC = 2.0
SAVE_EVERY_N_ROWS = 80
LOG_EVERY_ENQUEUE = True

# pool size（ページ使い回し）
TEAM_PAGEPOOL_SIZE = getenv_int("TEAM_PAGEPOOL_SIZE", TEAMS_PER_LEAGUE_CONCURRENCY)
DETAIL_PAGEPOOL_SIZE = getenv_int("DETAIL_PAGEPOOL_SIZE", GLOBAL_PLAYERDETAIL_CONCURRENCY)

BLOCKED_HOST_KEYWORDS = [
    "googletagmanager",
    "google-analytics",
    "doubleclick",
    "googlesyndication",
    "scorecardresearch",
    "criteo",
    "adsystem",
    "mathtag",
    "quantserve",
    "taboola",
    "facebook",
    "twitter",
    "hotjar",
    "onetrust",
    "cookiebot",
    "trustarc",
    "ioam.de",
    "amazon-adsystem",
    "adservice",
]


def ensure_dirs():
    os.makedirs(BASE_DIR, exist_ok=True)
    os.makedirs(TEAMS_BY_LEAGUE_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(B001_JSON_PATH), exist_ok=True)


# =========================
# JSON（国×リーグ）読み取り（指定版）
# =========================
def extract_countries_and_leagues(json_path: str) -> tuple[list[str], list[tuple[str, str]]]:
    """
    対応形式:
    1) list形式: [{"country":"日本","leagues":["J1","J2"]}, ...]
    2) dict形式: {"日本":["J1","J2"], "ドイツ":"ブンデスリーガ"} も互換で対応
    """
    p = Path(json_path).expanduser()
    if not p.exists():
        return [], []

    data = json.loads(p.read_text(encoding="utf-8"))
    pairs: list[tuple[str, str]] = []

    # --- 1) list形式（今回の実データ） ---
    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            country = str(item.get("country", "")).strip()
            if not country:
                continue

            leagues_val = item.get("leagues", [])
            if isinstance(leagues_val, list):
                for lv in leagues_val:
                    league = str(lv).strip()
                    if league:
                        pairs.append((country, league))
            else:
                league = str(leagues_val).strip()
                if league:
                    pairs.append((country, league))

    # --- 2) dict形式（互換） ---
    elif isinstance(data, dict):
        for c, v in data.items():
            country = str(c).strip()
            if not country:
                continue
            if isinstance(v, list):
                for lv in v:
                    league = str(lv).strip()
                    if league:
                        pairs.append((country, league))
            else:
                league = str(v).strip()
                if league:
                    pairs.append((country, league))

    countries = sorted({c for c, _ in pairs})
    return countries, pairs


def load_existing_player_keys(base_dir: str) -> Set[Tuple[str, str, str, str]]:
    s: Set[Tuple[str, str, str, str]] = set()

    for path in sorted(glob.glob(os.path.join(base_dir, f"{EXCEL_BASE_PREFIX}[0-9]*.csv"))):
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                r = csv.DictReader(f)
                for row in r:
                    c = str(row.get("国", "")).strip()
                    l = str(row.get("リーグ", "")).strip()
                    t = str(row.get("所属チーム", "")).strip()
                    n = str(row.get("選手名", "")).strip()
                    if c and l and t and n:
                        s.add((c, l, t, n))
        except Exception as e:
            print(f"[WARN] 既存CSV読込失敗: {path} / {e}")

    for path in sorted(glob.glob(os.path.join(base_dir, f"{EXCEL_BASE_PREFIX}[0-9]*.xlsx"))):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                s.add((str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip(), str(row[3]).strip()))
        except Exception as e:
            print(f"[WARN] 既存XLSX読込失敗: {path} / {e}")

    return s


def read_teams_from_file(path: str) -> List[Tuple[str, str, str, str]]:
    out: List[Tuple[str, str, str, str]] = []

    KEYMAP = {
        "国": ["国", "country", "Country"],
        "リーグ": ["リーグ", "league", "League"],
        "チーム": ["チーム", "team", "Team"],
        "チームリンク": ["チームリンク", "チームURL", "チームurl", "href", "link", "url", "チームリンクURL"],
    }

    def norm(s: str) -> str:
        return (s or "").strip()

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return out

            resolved: Dict[str, str] = {}
            fields = [h.strip() for h in reader.fieldnames if h]
            for canonical, aliases in KEYMAP.items():
                for a in aliases:
                    if a in fields:
                        resolved[canonical] = a
                        break

            use_fallback = any(k not in resolved for k in ("国", "リーグ", "チーム", "チームリンク"))
            if use_fallback:
                f.seek(0)
                r = csv.reader(f)
                _header = next(r, [])
                for row in r:
                    if not row:
                        continue
                    c = norm(row[0] if len(row) > 0 else "")
                    l = norm(row[1] if len(row) > 1 else "")
                    t = norm(row[2] if len(row) > 2 else "")
                    h = norm(row[3] if len(row) > 3 else "")
                    if c and l and t and h:
                        out.append((c, l, t, h))
                return out

            for row in reader:
                c = norm(row.get(resolved["国"], ""))
                l = norm(row.get(resolved["リーグ"], ""))
                t = norm(row.get(resolved["チーム"], ""))
                h = norm(row.get(resolved["チームリンク"], ""))
                if c and l and t and h:
                    out.append((c, l, t, h))

    except Exception as e:
        print(f"[WARN] チームCSV読込失敗: {path} / {e}")

    return out


def build_targets_from_contains() -> Dict[str, Set[str]]:
    out: Dict[str, Set[str]] = {}
    for ent in CONTAINS_LIST:
        c, l = [x.strip() for x in ent.split(":", 1)]
        out.setdefault(c, set()).add(l)
    return out


def group_by_league(rows: List[Tuple[str, str, str, str]]) -> Dict[Tuple[str, str], List[Tuple[str, str, str, str]]]:
    mp: Dict[Tuple[str, str], List[Tuple[str, str, str, str]]] = {}
    for r in rows:
        mp.setdefault((r[0], r[1]), []).append(r)
    return mp


# =========================
# Playwright（修正版）
# =========================
global_nav_sema = asyncio.Semaphore(GLOBAL_NAV_CONCURRENCY)
playerdetail_sema = asyncio.Semaphore(GLOBAL_PLAYERDETAIL_CONCURRENCY)

_nav_gap_lock = asyncio.Lock()
_last_nav_ts = 0.0


def is_driver_dead(e: Exception) -> bool:
    s = str(e)
    return (
        "Connection closed while reading from the driver" in s
        or "pipe closed by peer" in s
        or "Browser has been closed" in s
        or "Target crashed" in s
        or "BrowserType.launch: Target closed" in s
    )


async def safe_close(obj):
    try:
        if obj is not None:
            await obj.close()
    except Exception:
        pass


async def _respect_nav_gap():
    global _last_nav_ts
    async with _nav_gap_lock:
        loop = asyncio.get_running_loop()
        now = loop.time()
        due = _last_nav_ts + (NAV_GAP_MS / 1000.0)
        if now < due:
            await asyncio.sleep(due - now)
        _last_nav_ts = loop.time()


async def make_context(browser):
    ctx = await browser.new_context(
        locale="ja-JP",
        timezone_id="Asia/Tokyo",
        user_agent=(
            "Mozilla/5.0 (X11; Linux x86_64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1280, "height": 800},
        ignore_https_errors=True,
        java_script_enabled=True,
    )
    ctx.set_default_timeout(OP_TIMEOUT)
    ctx.set_default_navigation_timeout(NAV_TIMEOUT)

    async def _route(route):
        req = route.request
        rtype = req.resource_type
        url = req.url.lower()
        if rtype in {"image", "font", "media", "beacon"}:
            return await route.abort()
        if any(k in url for k in BLOCKED_HOST_KEYWORDS):
            return await route.abort()
        return await route.continue_()

    await ctx.route("**/*", _route)
    return ctx


async def safe_goto(
    page,
    url: str,
    selector_to_wait: Optional[str],
    nav_timeout=NAV_TIMEOUT,
    sel_timeout=SEL_TIMEOUT,
    tries=SAFE_GOTO_TRIES,
) -> bool:
    last = None
    for i in range(tries):
        try:
            async with global_nav_sema:
                await _respect_nav_gap()
                resp = await page.goto(url, wait_until=WAIT_UNTIL, timeout=nav_timeout)

            # HTTPエラーは即失敗（selector待ちに行かない）
            if resp is not None:
                st = resp.status
                if st >= 400:
                    raise RuntimeError(f"HTTP {st}")

            if selector_to_wait:
                await page.wait_for_selector(selector_to_wait, timeout=sel_timeout, state="attached")
            return True
        except Exception as e:
            last = e
            # driver死はここで粘っても戻らないので投げる
            if is_driver_dead(e):
                raise
            await asyncio.sleep(0.8 * (i + 1))

    progress.inc("goto_fail", 1)
    log(f"Page.goto 最終失敗: {url} / {last}")
    return False


class PagePool:
    """
    ctx.new_page() 連発を避ける（ページを使い回す）
    """

    def __init__(self, ctx, size: int):
        self.ctx = ctx
        self.sema = asyncio.Semaphore(size)
        self.cache: List = []
        self.lock = asyncio.Lock()

    async def acquire(self):
        await self.sema.acquire()
        async with self.lock:
            if self.cache:
                return self.cache.pop()
        return await self.ctx.new_page()

    async def release(self, page):
        # 次の利用に備えて軽くリセット（重くしない）
        try:
            await page.goto("about:blank", wait_until="commit", timeout=3000)
        except Exception:
            pass

        async with self.lock:
            self.cache.append(page)
        self.sema.release()

    async def close(self):
        async with self.lock:
            while self.cache:
                p = self.cache.pop()
                try:
                    await p.close()
                except Exception:
                    pass


class ExcelWriter:
    def __init__(
        self,
        base_dir: str,
        prefix: str,
        max_records: int,
        existing_keys: Set[Tuple[str, str, str, str]],
        keys_lock: asyncio.Lock,
    ):
        self.base_dir = base_dir
        self.prefix = prefix
        self.max_records = max_records
        self.existing_keys = existing_keys
        self.keys_lock = keys_lock

        self.q: asyncio.Queue = asyncio.Queue()
        self._task: Optional[asyncio.Task] = None

        self.wb = None
        self.ws = None
        self.seq = -1
        self.rows_in_current = 0
        self._dirty_since_last_save = 0

    def _scan_existing_max_seq(self) -> int:
        existing = sorted(glob.glob(os.path.join(self.base_dir, f"{self.prefix}[0-9]*.xlsx")))
        if not existing:
            return 0

        def num(p):
            m = re.search(rf"{re.escape(self.prefix)}(\d+)\.xlsx$", os.path.basename(p))
            return int(m.group(1)) if m else 0

        return max((num(p) for p in existing), default=0)

    def _open_initial(self):
        max_seq = self._scan_existing_max_seq()
        start_seq = 1 if max_seq <= 0 else max_seq

        for seq in range(start_seq, start_seq + 2):
            path = os.path.join(self.base_dir, f"{self.prefix}{seq}.xlsx")
            if os.path.exists(path):
                wb = load_workbook(path)
                ws = wb.active
                used = max(ws.max_row - 1, 0)
                if used < self.max_records:
                    self.wb, self.ws, self.seq = wb, ws, seq
                    self.rows_in_current = used
                    return

        new_seq = 1 if max_seq <= 0 else (max_seq + 1)
        self.seq = new_seq
        self.wb = None
        self.ws = None
        self.rows_in_current = 0

    async def _rotate_new(self, seq_new: int):
        path = os.path.join(self.base_dir, f"{self.prefix}{seq_new}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "国",
                "リーグ",
                "所属チーム",
                "選手名",
                "ポジション",
                "背番号",
                "得点数",
                "年齢",
                "誕生日",
                "市場価値",
                "ローン保有元",
                "契約期限",
                "顔写真",
                "故障情報",
                "データ取得時間",
            ]
        )
        self.wb, self.ws, self.seq = wb, ws, seq_new
        self.rows_in_current = 0
        await self._save_now()
        print(f"[WRITER] Rotate -> {os.path.basename(path)}")

    async def _save_now(self):
        path = os.path.join(self.base_dir, f"{self.prefix}{self.seq}.xlsx")
        await asyncio.to_thread(self.wb.save, path)
        self._dirty_since_last_save = 0

    def enqueue_team_rows(self, rows: List[List[object]], label: str):
        self.q.put_nowait((rows, label))

    async def run(self):
        self._open_initial()
        if self.wb is None:
            await self._rotate_new(self.seq)

        while True:
            try:
                rows, label = await asyncio.wait_for(self.q.get(), timeout=SAVE_INTERVAL_SEC)
            except asyncio.TimeoutError:
                if self._dirty_since_last_save > 0:
                    await self._save_now()
                continue

            if rows is None:
                if self._dirty_since_last_save > 0:
                    await self._save_now()
                break

            filtered: List[List[object]] = []
            async with self.keys_lock:
                for r in rows:
                    c = str(r[0]).strip()
                    l = str(r[1]).strip()
                    t = str(r[2]).strip()
                    n = str(r[3]).strip()
                    key = (c, l, t, n)
                    if key in self.existing_keys:
                        continue
                    self.existing_keys.add(key)
                    filtered.append(r)

            if not filtered:
                if LOG_EVERY_ENQUEUE:
                    print(f"[WRITER] SKIP {label}: 0 new rows (all existed)")
                continue

            offset = 0
            while offset < len(filtered):
                remain = self.max_records - self.rows_in_current
                chunk = filtered[offset : offset + remain]

                for r in chunk:
                    self.ws.append(r)

                self.rows_in_current += len(chunk)
                self._dirty_since_last_save += len(chunk)
                offset += len(chunk)

                if self.rows_in_current >= self.max_records:
                    await self._save_now()
                    await self._rotate_new(self.seq + 1)

            if LOG_EVERY_ENQUEUE:
                print(
                    f"[WRITER] APPEND {label}: +{len(filtered)} rows -> file #{self.seq} (cur={self.rows_in_current})"
                )

            if self._dirty_since_last_save >= SAVE_EVERY_N_ROWS:
                await self._save_now()

    async def start(self):
        if self._task is None:
            self._task = asyncio.create_task(self.run())

    async def stop(self):
        await self.q.put((None, "STOP"))
        if self._task:
            await self._task


async def scrape_player_detail_with_page(page, full_url: str) -> Dict[str, str]:
    age = birth = mv = contract = loan = img = inj = "N/A"

    ok = await safe_goto(page, full_url, "div.playerHeader__wrapper div.playerInfoItem")
    if not ok:
        return {
            "age_text": age,
            "birth_date": birth,
            "market_value": mv,
            "loan_info": loan,
            "contract_date": contract,
            "img_url": img,
            "injury_text": inj,
        }

    try:
        el = await page.query_selector(
            'div[data-testid="wcl-assetContainerBoxed-xl"] img, div[data-testid="wcl-assetContainerBoxed-XL"] img'
        )
        if el:
            img = await el.get_attribute("src") or "N/A"
    except Exception:
        pass

    try:
        items = await page.query_selector_all("div.playerHeader__wrapper div.playerInfoItem")
        for it in items:
            spans = await it.query_selector_all("span")
            if not spans:
                continue

            label = (await spans[0].text_content() or "").strip()
            label = re.sub(r"\s+", " ", label).replace(":", "").strip()

            values = []
            for sp in spans[1:]:
                tx = (await sp.text_content() or "").strip()
                tx = re.sub(r"\s+", " ", tx)
                if tx:
                    values.append(tx)

            if "年齢" in label:
                if len(values) >= 1:
                    age = values[0]
                if len(values) >= 2:
                    birth = values[1].strip().strip("()")
            elif "市場価値" in label:
                if values:
                    mv = values[0]
            elif "契約期限" in label:
                if values:
                    contract = values[0]
            elif "ローン" in label or "ローン元" in label:
                if values:
                    loan = values[0]
                if len(values) >= 2 and contract == "N/A":
                    contract = values[1].replace("期限", "").replace(":", "").strip().strip("()")
    except Exception as e:
        # driver死は上に投げてリーグ再起動させる
        if is_driver_dead(e):
            raise
        print(f"[PLAYER DETAIL] parse失敗: {full_url} / {e}")

    try:
        inj_svg = await page.query_selector('svg[data-testid="wcl-icon-incidents-injury"]')
        if inj_svg:
            title_el = await inj_svg.query_selector("title")
            if title_el:
                t = (await title_el.text_content() or "").strip()
                if t:
                    inj = re.sub(r"\s+", " ", t)
    except Exception:
        pass

    return {
        "age_text": age,
        "birth_date": birth,
        "market_value": mv,
        "loan_info": loan,
        "contract_date": contract,
        "img_url": img,
        "injury_text": inj,
    }


async def process_team(
    ctx,
    team_pool: PagePool,
    detail_pool: PagePool,
    existing_keys,
    claimed_keys,
    keys_lock,
    writer,
    country,
    league,
    team,
    href,
):
    page = None
    try:
        page = await team_pool.acquire()
    except Exception as e:
        log(f"[WARN] new_page failed: {country}/{league}/{team} / {e}")
        progress.inc("teams_done", 1)
        progress.maybe_log()
        return

    try:
        team_url = href if href.startswith("http") else f"https://flashscore.co.jp{href}"
        squad_url = team_url.rstrip("/") + "/squad/"

        ok = await safe_goto(
            page,
            squad_url,
            "div.lineupTable.lineupTable--soccer",
            nav_timeout=SQUAD_NAV_TIMEOUT,
            sel_timeout=SQUAD_SEL_TIMEOUT,
            tries=SQUAD_TRIES,
        )
        if not ok:
            log(f"[{country}:{league}:{team}] スカッド取得失敗（fast-skip）")
            progress.inc("teams_done", 1)
            progress.maybe_log()
            return

        players: List[Tuple[str, str, str, str, str]] = []
        tables = await page.query_selector_all("div.lineupTable.lineupTable--soccer")
        for t in tables or []:
            pos_el = await t.query_selector("div.lineupTable__title")
            position = (await pos_el.text_content()).strip() if pos_el else "N/A"
            rows = await t.query_selector_all("div.lineupTable__row")
            for r in rows:
                jersey_el = await r.query_selector(".lineupTable__cell--jersey")
                jersey = (await jersey_el.text_content()).strip() if jersey_el else "N/A"

                name_el = await r.query_selector(".lineupTable__cell--player .lineupTable__cell--name")
                if not name_el:
                    continue
                name = (await name_el.text_content()).strip()
                href2 = await name_el.get_attribute("href")
                if not href2:
                    continue

                goal_el = await r.query_selector(".lineupTable__cell--goal")
                goals = (await goal_el.text_content()).strip() if goal_el else "N/A"

                players.append((position, jersey, name, href2, goals))

        if not players:
            log(f"[{country}:{league}:{team}] 選手0件")
            progress.inc("teams_done", 1)
            progress.maybe_log()
            return

        # gatherの塊（大きすぎるとメモリ/負荷が跳ねるので控えめ）
        BATCH = max(PLAYER_PER_TEAM * 4, 4)

        async def handle(p) -> Optional[List[object]]:
            position, jersey, name, href2, goals = p
            key = (country, league, team, name)

            async with keys_lock:
                if key in existing_keys or key in claimed_keys:
                    return None
                claimed_keys.add(key)

            purl = href2 if href2.startswith("http") else f"https://flashscore.co.jp{href2}"

            async with playerdetail_sema:
                tab = None
                try:
                    tab = await detail_pool.acquire()
                    detail = await scrape_player_detail_with_page(tab, purl)
                    progress.inc("player_detail_ok", 1)
                except Exception as e:
                    progress.inc("player_detail_fail", 1)
                    # driver死はリーグごと作り直しのため上へ
                    if is_driver_dead(e):
                        raise
                    return None
                finally:
                    if tab is not None:
                        try:
                            await detail_pool.release(tab)
                        except Exception:
                            pass

            return [
                country,
                league,
                team,
                name,
                position,
                jersey,
                goals,
                detail.get("age_text", "N/A"),
                detail.get("birth_date", "N/A"),
                detail.get("market_value", "N/A"),
                detail.get("loan_info", "N/A"),
                detail.get("contract_date", "N/A"),
                detail.get("img_url", "N/A"),
                detail.get("injury_text", "N/A"),
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]

        rows_to_write: List[List[object]] = []
        for i in range(0, len(players), BATCH):
            got = await asyncio.gather(*(handle(p) for p in players[i : i + BATCH]), return_exceptions=True)
            for r in got:
                if isinstance(r, list):
                    rows_to_write.append(r)
                elif isinstance(r, Exception):
                    # driver死だけは上に投げる（リーグ再起動）
                    if is_driver_dead(r):
                        raise r

        if rows_to_write:
            writer.enqueue_team_rows(rows_to_write, f"{country}/{league}/{team}")
            progress.inc("players_enqueued", len(rows_to_write))

        log(f"[TEAM-DONE] {country} / {league} / {team}")
        progress.inc("teams_done", 1)
        progress.maybe_log()

    except Exception as e:
        # driver死は上位でリーグ再起動させるため投げる
        if is_driver_dead(e):
            raise
        log(f"[WARN] team failed: {country}/{league}/{team} / {e}")
        progress.inc("teams_done", 1)
        progress.maybe_log()
    finally:
        if page is not None:
            try:
                await team_pool.release(page)
            except Exception:
                pass


async def process_league(
    ctx,
    team_pool: PagePool,
    detail_pool: PagePool,
    existing_keys,
    claimed_keys,
    keys_lock,
    writer,
    league_key,
    teams,
):
    country, league = league_key
    log(f"[LEAGUE-START] {country} / {league} teams={len(teams)}")

    for i in range(0, len(teams), TEAMS_PER_LEAGUE_CONCURRENCY):
        chunk = teams[i : i + TEAMS_PER_LEAGUE_CONCURRENCY]
        results = await asyncio.gather(
            *[
                process_team(ctx, team_pool, detail_pool, existing_keys, claimed_keys, keys_lock, writer, c, l, team, href)
                for (c, l, team, href) in chunk
            ],
            return_exceptions=True,
        )

        for r in results:
            if isinstance(r, Exception):
                # driver死は上位へ
                if is_driver_dead(r):
                    raise r
                print(f"[WARN] チーム例外: {country}/{league} / {r}")

    log(f"[LEAGUE-DONE] {country} / {league}")
    progress.inc("leagues_done", 1)
    progress.maybe_log()


def convert_xlsx_to_csv_upload_and_delete_all(base_dir: str, prefix: str) -> None:
    xlsx_paths = sorted(glob.glob(os.path.join(base_dir, f"{prefix}[0-9]*.xlsx")))
    if not xlsx_paths:
        print("[CONVERT] 対象xlsxなし")
        return

    for xlsx_path in xlsx_paths:
        csv_path = re.sub(r"\.xlsx$", ".csv", xlsx_path)
        try:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = wb.active

            with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow([("" if v is None else v) for v in row])

            # 出力は member バケット直下へ
            s3_upload_to_bucket_root(S3_BUCKET_MEMBER, csv_path)

            os.remove(xlsx_path)
            print(f"[CONVERT] {os.path.basename(xlsx_path)} -> {os.path.basename(csv_path)}（xlsx削除 & S3アップロード）")
        except Exception as e:
            print(f"[ERROR] 変換失敗: {xlsx_path} / {e}")


async def main():
    ensure_dirs()
    log("TeamMemberData(ECS): 開始")

    hb = asyncio.create_task(heartbeat_task())

    # JSON（任意）は member バケットから
    s3_download_if_exists(S3_BUCKET_MEMBER, B001_S3_KEY, B001_JSON_PATH)

    # teamData_*.csv は team バケットから /tmp/bookmaker/teams_by_league へ
    s3_download_teamdata_all(S3_BUCKET_TEAM, TEAMS_BY_LEAGUE_DIR)

    existing_keys = load_existing_player_keys(TEAMS_BY_LEAGUE_DIR)
    log(f"[INIT] 既取得キー: {len(existing_keys)} 件")

    paths = glob.glob(os.path.join(TEAMS_BY_LEAGUE_DIR, "teamData_*.csv"))
    if not paths:
        print(f"[EXIT] teamData_*.csv が見つかりません（DL失敗の可能性）: s3://{S3_BUCKET_TEAM}/teamData_*.csv")
        hb.cancel()
        return

    team_rows: List[Tuple[str, str, str, str]] = []
    for p in sorted(paths):
        rows = read_teams_from_file(p)
        print(f"[LOAD] {os.path.basename(p)}: {len(rows)} 行")
        team_rows.extend(rows)

    # チーム行の重複除去
    seen: Set[Tuple[str, str, str, str]] = set()
    uniq_rows: List[Tuple[str, str, str, str]] = []
    for row in team_rows:
        if row not in seen:
            seen.add(row)
            uniq_rows.append(row)

    log(f"[PLAN] 処理チーム数（重複除去後）: {len(uniq_rows)}")

    # JSONがあれば読み取り、無ければ CONTAINS_LIST フィルタ
    p_json = Path(B001_JSON_PATH).expanduser()
    if p_json.exists():
        _countries, country_league_pairs = extract_countries_and_leagues(B001_JSON_PATH)
        if country_league_pairs:
            json_targets: Dict[str, Set[str]] = {}
            for c, l in country_league_pairs:
                c2 = re.sub(r"\s+", " ", str(c or "")).strip()
                l2 = re.sub(r"\s+", " ", str(l or "")).strip()
                if c2 and l2:
                    json_targets.setdefault(c2, set()).add(l2)

            before = len(uniq_rows)
            uniq_rows = [r for r in uniq_rows if (r[0] in json_targets and r[1] in json_targets[r[0]])]
            print(f"[FILTER] JSON(存在) / チーム {before} -> {len(uniq_rows)} (pairs={len(country_league_pairs)})")

            if not uniq_rows:
                print("[EXIT] JSONはあるが対象チームが0件（表記ゆれ/JSON形式を確認）")
                hb.cancel()
                return
        else:
            # JSONファイルはあるが空/形式不一致 => contains fallback
            contains_targets = build_targets_from_contains()
            before = len(uniq_rows)
            uniq_rows = [r for r in uniq_rows if (r[0] in contains_targets and r[1] in contains_targets[r[0]])]
            print(f"[FILTER] JSON(存在だが空) -> CONTAINS / チーム {before} -> {len(uniq_rows)}")
    else:
        contains_targets = build_targets_from_contains()
        before = len(uniq_rows)
        uniq_rows = [r for r in uniq_rows if (r[0] in contains_targets and r[1] in contains_targets[r[0]])]
        print(f"[FILTER] JSON(なし) / チーム {before} -> {len(uniq_rows)}")

    if not uniq_rows:
        print("[EXIT] フィルタ後の対象チームが0件")
        hb.cancel()
        return

    league_map = group_by_league(uniq_rows)
    league_items = list(league_map.items())
    progress.set("leagues_total", len(league_items))
    progress.set("teams_total", len(uniq_rows))
    progress.maybe_log(force=True)
    log(f"[PLAN] 対象リーグ数: {len(league_items)}")

    keys_lock = asyncio.Lock()
    claimed_keys: Set[Tuple[str, str, str, str]] = set()

    writer = ExcelWriter(TEAMS_BY_LEAGUE_DIR, EXCEL_BASE_PREFIX, EXCEL_MAX_RECORDS, existing_keys, keys_lock)
    await writer.start()

    league_sema = asyncio.Semaphore(LEAGUE_CONCURRENCY)

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--disable-gpu",
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--no-zygote",
                "--disable-dev-shm-usage",
                "--blink-settings=imagesEnabled=false",
            ],
        )

        async def run_one_league(lk, teams):
            async with league_sema:
                # リーグ単位で context を作って捨てる（長時間のメモリ累積を切る）
                for attempt in range(2):  # 0:通常 / 1:driver死んだ時の1回だけ再試行
                    ctx = None
                    team_pool = None
                    detail_pool = None
                    try:
                        ctx = await make_context(browser)
                        team_pool = PagePool(ctx, size=max(1, TEAM_PAGEPOOL_SIZE))
                        detail_pool = PagePool(ctx, size=max(1, DETAIL_PAGEPOOL_SIZE))

                        await process_league(
                            ctx,
                            team_pool,
                            detail_pool,
                            existing_keys,
                            claimed_keys,
                            keys_lock,
                            writer,
                            lk,
                            teams,
                        )
                        return
                    except Exception as e:
                        if is_driver_dead(e) and attempt == 0:
                            log(f"[LEAGUE RETRY] driver closed: {lk} / retry with fresh context")
                            # 次のattemptで作り直し
                        else:
                            log(f"[WARN] league failed: {lk} / {e}")
                            return
                    finally:
                        if team_pool is not None:
                            await team_pool.close()
                        if detail_pool is not None:
                            await detail_pool.close()
                        await safe_close(ctx)

        tasks = [run_one_league(lk, teams) for (lk, teams) in league_items]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for r in results:
            if isinstance(r, Exception):
                print(f"[WARN] リーグ例外: {r}")

        await safe_close(browser)

    await writer.stop()

    # xlsx -> csv & memberバケットへupload & xlsx delete
    convert_xlsx_to_csv_upload_and_delete_all(TEAMS_BY_LEAGUE_DIR, EXCEL_BASE_PREFIX)

    log("TeamMemberData(ECS): 完了（teamバケットから入力 -> memberバケットへ出力）")

    hb.cancel()
    with contextlib.suppress(asyncio.CancelledError):
        await hb


if __name__ == "__main__":
    asyncio.run(main())

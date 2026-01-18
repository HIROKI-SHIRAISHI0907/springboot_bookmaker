# -*- coding: utf-8 -*-
"""
TeamExcel.py
Flashscore: 国×リーグ -> チームExcel(SEED 専用)

- 左メニューを「もっと表示」連打＋無限スクロールで最大までロード
- すべての国ブロックを opened にしてから対象リーグの a を抽出
- 抽出した各リーグページの standings へ遷移（#/compId/table/overall を優先）
- standings のテーブルから「チーム名 / チームリンク(/team/...)」を抽出
- 国×リーグごとに /tmp/bookmaker/teams_by_league/<国>__<リーグ>.xlsx へ保存（追記・重複除外）
- 最後に teamData_*.xlsx -> teamData_*.csv へ変換し、CSVのみを S3 に保存

✅ 要件（今回の修正）
- 作業フォルダ: /tmp/bookmaker
- S3バケット: aws-s3-team-csv
- S3には CSV を「バケット直下」に格納（フォルダ無し）
  例: s3://aws-s3-team-csv/teamData_1.csv
- 入力JSONがS3にある場合は取得して利用
  s3://aws-s3-team-csv/json/b001/b001_country_league.json -> /tmp/bookmaker/json/b001/b001_country_league.json
"""

import json
from pathlib import Path
import glob
import pandas as pd
import os
import re
import time
import asyncio
import datetime
from typing import List, Tuple, Dict, Set, Optional

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

# =========================
# S3（固定）
# =========================
import boto3
from botocore.exceptions import ClientError

S3_BUCKET = "aws-s3-team-csv"
B001_S3_KEY = "json/b001_country_league.json"
B001_JSON_PATH = "/tmp/bookmaker/json/b001/b001_country_league.json"

_s3 = boto3.client("s3")

def s3_download_if_exists(bucket: str, key: str, local_path: str) -> bool:
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    try:
        _s3.download_file(bucket, key, local_path)
        print(f"[S3 DOWNLOAD] s3://{bucket}/{key} -> {local_path}")
        return True
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "Unknown")
        print(f"[S3] input json not found or no permission: s3://{bucket}/{key} (Code={code})")
        return False

def s3_upload_csv_to_bucket_root(local_csv_path: str) -> bool:
    """
    CSV をバケット直下に保存する（フォルダなし）
      /tmp/bookmaker/teams_by_league/teamData_12.csv
        -> s3://aws-s3-team-csv/teamData_12.csv
    """
    key = os.path.basename(local_csv_path)
    try:
        _s3.upload_file(local_csv_path, S3_BUCKET, key)
        print(f"[S3 UPLOAD] {local_csv_path} -> s3://{S3_BUCKET}/{key}")
        return True
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "Unknown")
        print(f"[S3 ERROR] upload failed: {local_csv_path} -> s3://{S3_BUCKET}/{key} (Code={code})")
        return False
    
# =========================
# 設定（固定）
# =========================
EXCEL_HEADER = "teamData_"
BASE_OUTPUT_URL = "/tmp/bookmaker"

# 対象（国: リーグ）
B001_JSON_PATH = str(Path(BASE_OUTPUT_URL) / "json/b001/b001_country_league.json")
os.makedirs(BASE_OUTPUT_URL, exist_ok=True)

TEAMS_BY_LEAGUE_DIR = os.path.join(BASE_OUTPUT_URL, "teams_by_league")
os.makedirs(TEAMS_BY_LEAGUE_DIR, exist_ok=True)

# タイムアウト(ms)
OP_TIMEOUT = 5000
NAV_TIMEOUT = 10000
SEL_TIMEOUT = 20000

CONTAINS_LIST = [
    "ケニア: プレミアリーグ", "コロンビア: プリメーラ A", "タンザニア: プレミアリーグ", "イングランド: プレミアリーグ",
    "イングランド: EFL チャンピオンシップ", "イングランド: EFL リーグ 1", "エチオピア: プレミアリーグ", "コスタリカ: リーガ FPD",
    "ジャマイカ: プレミアリーグ", "スペイン: ラ・リーガ", "ブラジル: セリエ A ベターノ", "ブラジル: セリエ B", "ドイツ: ブンデスリーガ",
    "韓国: K リーグ 1", "中国: 中国スーパーリーグ", "日本: J1 リーグ", "日本: J2 リーグ", "日本: J3 リーグ", "インドネシア: スーパーリーグ",
    "オーストラリア: A リーグ・メン", "チュニジア: チュニジア･プロリーグ", "ウガンダ: プレミアリーグ", "メキシコ: リーガ MX",
    "フランス: リーグ・アン", "スコットランド: プレミアシップ", "オランダ: エールディビジ", "アルゼンチン: トルネオ・ベターノ",
    "イタリア: セリエ A", "イタリア: セリエ B", "ポルトガル: リーガ・ポルトガル", "トルコ: スュペル・リグ", "セルビア: スーペルリーガ",
    "日本: WEリーグ", "ボリビア: LFPB", "ブルガリア: パルヴァ・リーガ", "カメルーン: エリート 1", "ペルー: リーガ 1",
    "エストニア: メスタリリーガ", "ウクライナ: プレミアリーグ", "ベルギー: ジュピラー･プロリーグ", "エクアドル: リーガ・プロ",
    "日本: YBC ルヴァンカップ", "日本: 天皇杯"
]
UNDER_LIST  = ["U17", "U18", "U19", "U20", "U21", "U22", "U23", "U24", "U25"]
GENDER_LIST = ["女子"]
EXP_LIST    = ["ポルトガル: リーガ・ポルトガル 2", "イングランド: プレミアリーグ 2", "イングランド: プレミアリーグ U18"]

# 国→{リーグ名,...} へ変換
TARGETS: Dict[str, Set[str]] = {}
for ent in CONTAINS_LIST:
    country, league = [x.strip() for x in ent.split(":", 1)]
    TARGETS.setdefault(country, set()).add(league)

# =========================
# 共通ユーティリティ（Excel）
# =========================
def sanitize_filename(name: str) -> str:
    s = re.sub(r"[\\/:*?\"<>|]", "_", name)
    s = re.sub(r"\s+", "_", s.strip())
    return s

def teams_excel_path(country: str, league: str) -> str:
    filename = f"{EXCEL_HEADER}{sanitize_filename(country)}__{sanitize_filename(league)}.xlsx"
    return os.path.join(TEAMS_BY_LEAGUE_DIR, filename)

def ensure_header(ws, headers: List[str]):
    if ws.max_row == 0 or str(ws["A1"].value or "") != headers[0]:
        ws.delete_rows(1, ws.max_row if ws.max_row else 1)
        ws.append(headers)

def write_teams_excel(country: str, league: str, pairs: List[Tuple[str, str]]):
    """
    pairs: [(team_name, team_href('/team/.../')), ...]
    """
    path = teams_excel_path(country, league)
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        ensure_header(ws, ["国","リーグ","チーム","チームリンク"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "teams"
        ws.append(["国","リーグ","チーム","チームリンク"])

    existing = {(str(r[2]), str(r[3])) for r in ws.iter_rows(min_row=2, values_only=True) if r}
    added = 0
    for team, href in pairs:
        key = (team, href)
        if key in existing:
            continue
        ws.append([country, league, team, href])
        added += 1

    wb.save(path)
    print(f"[WRITE] {path} に {added} 件追加（合計 {ws.max_row-1} 件）")

# =========================
# Playwright ヘルパ
# =========================
async def make_context(browser, block_css: bool):
    ctx = await browser.new_context(
        locale="ja-JP", timezone_id="Asia/Tokyo",
        user_agent=("Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"),
        viewport={"width": 1440, "height": 1000},
        ignore_https_errors=True,
    )
    ctx.set_default_timeout(OP_TIMEOUT)
    ctx.set_default_navigation_timeout(NAV_TIMEOUT)

    async def _route(route):
        rtype = route.request.resource_type
        block = {"image","font","media"} | ({"stylesheet"} if block_css else set())
        if rtype in block:
            await route.abort()
        else:
            await route.continue_()
    await ctx.route("**/*", _route)
    return ctx

# =========================
# メニュー → 全ブロックロード & Open
# =========================
async def load_all_country_blocks(page) -> int:
    for _ in range(60):
        try:
            if await page.is_visible("span.lmc__itemMore"):
                await page.click("span.lmc__itemMore")
                await asyncio.sleep(0.1)
            else:
                break
        except Exception:
            break

    prev = -1
    stagnant = 0
    while True:
        blocks = await page.query_selector_all("div.lmc__block")
        count = len(blocks)
        if count == prev:
            stagnant += 1
            if stagnant >= 6:
                break
        else:
            stagnant = 0
            prev = count
        try:
            await page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
        except Exception:
            pass
        await asyncio.sleep(0.2)

    return max(prev, 0)

async def open_all_blocks_fast(page, overall_ms: int = 180000, batch_size: int = 120, settle_ms: int = 120) -> int:
    total = await load_all_country_blocks(page)
    print(f"country_blocks(before open): {total}")

    for sel in [
        "#onetrust-accept-btn-handler",
        "button#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll",
        "button[aria-label='同意']",
        ".fc-consent-root button:has-text('同意')",
    ]:
        try:
            if await page.is_visible(sel):
                await page.click(sel, timeout=1000)
        except Exception:
            pass

    start = time.monotonic()
    opened_prev = -1
    no_progress = 0

    while True:
        opened = await page.locator("div.lmc__block.lmc__blockOpened").count()
        if opened >= total:
            break
        if (time.monotonic() - start) * 1000 > overall_ms:
            print(f"[WARN] open_all_blocks timeout: opened {opened}/{total}")
            break
        if opened == opened_prev:
            no_progress += 1
            if no_progress >= 12:
                print(f"[WARN] open_all_blocks no progress: {opened}/{total}")
                break
        else:
            no_progress = 0
            opened_prev = opened

        clicked = await page.evaluate(
            """
            (batch) => {
              const blocks = Array.from(document.querySelectorAll('div.lmc__block:not(.lmc__blockOpened)'));
              let n = 0;
              for (const b of blocks) {
                const a = b.querySelector('a.lmc__element.lmc__item');
                if (!a) continue;
                try {
                  a.scrollIntoView({block: 'center'});
                  a.click();
                  n++;
                  if (n >= batch) break;
                } catch(e) {}
              }
              return n;
            }
            """,
            batch_size
        )
        await page.wait_for_timeout(settle_ms)

        try:
            await page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
        except Exception:
            pass

        new_opened = await page.locator("div.lmc__block.lmc__blockOpened").count()
        print(f"open progress: {new_opened}/{total} (+{new_opened - opened}, clicked {clicked})")
        if clicked == 0 and new_opened == opened:
            no_progress += 3
            if no_progress >= 12:
                print(f"[WARN] open_all_blocks stuck: {new_opened}/{total}")
                break

    opened_final = await page.locator("div.lmc__block.lmc__blockOpened").count()
    print(f"opened_blocks: {opened_final}/{total}")
    return opened_final

# =========================
# standings 到達＆待機（柔軟）
# =========================
STANDINGS_WAIT_SELECTORS = [
    "div.tableWrapper",
    "div[data-testid='wcl-tabs']",
    "div#tournamentTable",
    "a.tableCellParticipant__image",
    "a.tableCellParticipant__name",
    ".ui-table__body .ui-table__row",
    ".table__cell--participant a",
    ".tableWrapper .ui-table",
]

async def wait_any_selector(page, selectors, timeout_ms=SEL_TIMEOUT) -> str:
    deadline = time.time() + (timeout_ms / 1000)
    for sel in selectors:
        remaining = int(max(0.0, deadline - time.time()) * 1000)
        if remaining <= 0:
            break
        try:
            await page.wait_for_selector(sel, state="attached", timeout=remaining)
            return sel
        except Exception:
            continue
    raise TimeoutError("wait_any_selector: timeout")

def extract_comp_id(s: str) -> Optional[str]:
    if not s:
        return None
    token = s.split("_")[-1].strip()
    if re.fullmatch(r"[A-Za-z0-9]{6,}", token):
        return token
    m = re.search(r"([A-Za-z0-9]{6,})", s)
    return m.group(1) if m else None

async def goto_standings_with_redirect(page, base_href: str, comp_id: Optional[str]) -> bool:
    base_href = (base_href or "").rstrip("/") + "/"
    base = "https://flashscore.co.jp" + base_href + "standings/"

    async def _goto(url, tries=2) -> bool:
        for i in range(tries):
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
                hit = await wait_any_selector(page, STANDINGS_WAIT_SELECTORS, timeout_ms=SEL_TIMEOUT)
                print(f"[NAV OK] {url} (waited: {hit})")
                return True
            except Exception as e:
                print(f"[NAV RETRY {i+1}/{tries}] {url} ({e.__class__.__name__})")
                await asyncio.sleep(0.8*(i+1))
        return False

    if comp_id:
        urlA = f"{base}#/{comp_id}/standings/overall/"
        if await _goto(urlA, tries=2):
            return True
        urlB = f"{base}#/{comp_id}/table/overall"
        if await _goto(urlB, tries=2):
            return True

    if await _goto(base, tries=2):
        return True

    try:
        await page.goto(base, wait_until="commit", timeout=NAV_TIMEOUT)
        loc = page.locator("a[href*='/standings/#/']")
        if await loc.count() > 0:
            href2 = await loc.first.get_attribute("href")
            if href2:
                abs_url = href2 if href2.startswith("http") else "https://flashscore.co.jp" + href2
                if await _goto(abs_url, tries=2):
                    return True
    except Exception:
        pass

    hack = base + "#/table/overall"
    if await _goto(hack, tries=1):
        return True

    try:
        await page.click('a[data-analytics-alias="stats-detail_table"]', timeout=2000)
        hit = await wait_any_selector(page, STANDINGS_WAIT_SELECTORS, timeout_ms=SEL_TIMEOUT)
        print(f"[NAV OK] via stats-detail_table (waited: {hit})")
        return True
    except Exception:
        pass

    return False

async def get_wcl_tabs(page) -> dict[str, str]:
    tabs = {}
    try:
        items = page.locator('div[data-testid="wcl-tabs"] a[href^="#/"]')
        n = await items.count()
        for i in range(n):
            a = items.nth(i)
            href = (await a.get_attribute("href")) or ""
            m = re.match(r"^#\/([A-Za-z0-9]+)\/?$", href.strip())
            if not m:
                continue
            stage_id = m.group(1)
            label = (await a.inner_text()) or ""
            label = re.sub(r"\s+", " ", label).strip()
            if label:
                tabs[label] = stage_id
    except Exception:
        pass
    return tabs

async def force_main_stage_overall(page, base_href: str) -> bool:
    tabs = await get_wcl_tabs(page)
    stage_id = tabs.get("本大会")
    if not stage_id:
        return False

    try:
        await page.evaluate("""(sid) => { location.hash = `#/${sid}/standings/overall/`; }""", stage_id)
        hit = await wait_any_selector(page, STANDINGS_WAIT_SELECTORS, timeout_ms=SEL_TIMEOUT)
        print(f"[MAIN OK] #/{stage_id}/standings/overall/ (waited: {hit})")
        return True
    except Exception as e:
        print(f"[MAIN FAIL] stage={stage_id} ({e.__class__.__name__})")
        return False

async def goto_first_wcl_tab(page, base_href: str) -> bool:
    try:
        tabs = page.locator("div[data-testid='wcl-tabs'][role='tablist'] a[href^='#/']")
        if await tabs.count() == 0:
            return False

        first_a = tabs.first
        href = (await first_a.get_attribute("href")) or ""
        m = re.match(r"^#\/([A-Za-z0-9]+)\/?$", href.strip())
        if not m:
            return False
        sid = m.group(1)

        async def _try_hash(hash_value: str) -> bool:
            try:
                await page.evaluate("(h) => { location.hash = h; }", hash_value)
                await page.wait_for_timeout(250)
                await wait_any_selector(page, [
                    ".ui-table__body a[href^='/team/']",
                    "a.tableCellParticipant__name[href^='/team/']",
                    "a.tableCellParticipant__image[href^='/team/']",
                    ".ui-table__body .ui-table__row",
                    "div.tableWrapper",
                ], timeout_ms=SEL_TIMEOUT)
                return True
            except Exception:
                return False

        if await _try_hash(f"#/{sid}/standings/overall/"):
            print(f"[WCL] switched to first tab: #/{sid}/standings/overall/")
            return True
        if await _try_hash(f"#/{sid}/table/overall"):
            print(f"[WCL] switched to first tab: #/{sid}/table/overall")
            return True
        if await _try_hash(f"#/{sid}/"):
            print(f"[WCL] switched to first tab: #/{sid}/")
            return True

        try:
            await first_a.click(timeout=1200)
            await page.wait_for_timeout(300)
            await wait_any_selector(page, [
                ".ui-table__body a[href^='/team/']",
                ".ui-table__body .ui-table__row",
                "div.tableWrapper",
            ], timeout_ms=SEL_TIMEOUT)
            print("[WCL] clicked first tab anchor")
            return True
        except Exception:
            pass

        return False

    except Exception as e:
        print(f"[WCL WARN] goto_first_wcl_tab failed: {e.__class__.__name__}")
        return False

async def fully_render_standings(page, max_cycles: int = 40, pause_ms: int = 150):
    try:
        await page.click('a[data-analytics-alias="stats-detail_table"]', timeout=1200)
    except Exception:
        pass

    prev_rows = -1
    stable = 0
    for _ in range(max_cycles):
        try:
            rows = await page.locator(".ui-table__body .ui-table__row").count()
        except Exception:
            rows = 0
        if rows == prev_rows:
            stable += 1
            if stable >= 3:
                break
        else:
            stable = 0
        prev_rows = rows

        try:
            await page.evaluate("""
                () => {
                  window.scrollBy(0, document.body.scrollHeight);
                  const tw = document.querySelector('.tableWrapper');
                  if (tw) tw.scrollTop = tw.scrollHeight;
                }
            """)
        except Exception:
            pass
        await asyncio.sleep(pause_ms/1000)

# =========================
# リーグ → チーム
# =========================
async def scrape_league_to_teams(
    ctx,
    country: str,
    league: str,
    href: str,
    comp_id: Optional[str]
) -> List[Tuple[str, str]]:
    page = await ctx.new_page()
    teams: List[Tuple[str, str]] = []

    try:
        ok = await goto_standings_with_redirect(page, href, comp_id)
        if not ok:
            print(f"[WARN] standings到達失敗: {country} / {league} ({href}) -> 救済ルート継続")

        try:
            await force_main_stage_overall(page, href)
        except Exception:
            pass

        try:
            await goto_first_wcl_tab(page, href)
        except Exception:
            pass

        try:
            await wait_any_selector(
                page,
                [
                    ".ui-table__body a[href^='/team/']",
                    "a.tableCellParticipant__name[href^='/team/']",
                    "a.tableCellParticipant__image[href^='/team/']",
                    ".table__cell--participant a[href^='/team/']",
                ],
                timeout_ms=SEL_TIMEOUT
            )
        except Exception:
            print(f"[WARN] {country}:{league} /team/ が見つからず終了 ({page.url})")
            return []

        await fully_render_standings(page)

        anchors = await page.query_selector_all(
            ".ui-table__body a[href^='/team/'], "
            "a.tableCellParticipant__image, "
            "a.tableCellParticipant__name"
        )

        tmp: List[Tuple[str, str]] = []
        for a in anchors:
            href2 = (await a.get_attribute("href")) or ""
            if not href2.startswith("/team/"):
                continue
            name = (await a.get_attribute("title")) or (await a.text_content()) or ""
            name = name.strip()
            if not name:
                continue
            tmp.append((name, href2))

        if not tmp:
            html = await page.content()
            for m in re.finditer(r'<a[^>]+href="(/team/[^"]+/)"[^>]*title="([^"]+)"', html):
                tmp.append((m.group(2).strip(), m.group(1)))
            for m in re.finditer(r'<a[^>]+href="(/team/[^"]+/)"[^>]*>([^<]+)</a>', html):
                nm = re.sub(r"\s+", " ", m.group(2)).strip()
                if nm:
                    tmp.append((nm, m.group(1)))

        seen: Set[Tuple[str, str]] = set()
        for t in tmp:
            if t in seen:
                continue
            seen.add(t)
            teams.append(t)

        print(f"[{country}:{league}] チーム抽出: {len(teams)}件")
        return teams

    finally:
        await page.close()

# =========================
# 国×リーグの抽出
# =========================
async def collect_target_leagues(page, allowed_countries: Optional[Set[str]] = None):
    await page.goto("https://flashscore.co.jp/soccer/", wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
    await page.wait_for_selector("div.lmc__block", state="attached", timeout=SEL_TIMEOUT)

    opened = await open_all_blocks_fast(page, overall_ms=180000, batch_size=120, settle_ms=120)
    print(f"opened_blocks: {opened}")

    leagues: List[Tuple[str,str,str,Optional[str]]] = []
    blocks = await page.query_selector_all("div.lmc__block.lmc__blockOpened")
    for b in blocks:
        cspan = await b.query_selector("span.lmc__elementName")
        country = (await cspan.inner_text()).strip() if cspan else ""

        if allowed_countries is not None and country not in allowed_countries:
            continue

        if TARGETS is not None and country not in TARGETS:
            continue

        links = await b.query_selector_all("span.lmc__template a.lmc__templateHref") \
             or await b.query_selector_all("span.lmc__template a")

        for a in links:
            league_name = (await a.inner_text()).strip()

            if TARGETS is not None:
                if league_name not in TARGETS.get(country, set()):
                    continue

            key = f"{country}: {league_name}"
            if any(key.endswith(x) for x in EXP_LIST) or any(g in key for g in GENDER_LIST):
                continue

            href = await a.get_attribute("href")
            if not href:
                continue

            comp_id: Optional[str] = None
            try:
                parent = await a.evaluate_handle("el => el.parentElement")
                if parent:
                    pin = await parent.query_selector("span.pin")
                    if pin:
                        label_key = await pin.get_attribute("data-label-key") or ""
                        classes = await pin.get_attribute("class") or ""
                        comp_id = extract_comp_id(label_key) or extract_comp_id(classes)
            except Exception:
                pass

            print(f"   - {league_name} -> {href} (compId={comp_id or '-'})")
            leagues.append((country, league_name, href, comp_id))

    return leagues

# =========================
# JSON（国×リーグ）読み取り
# =========================
def extract_countries_and_leagues(json_path: str) -> tuple[list[str], list[tuple[str, str]]]:
    """
    対応形式:
    1) list形式: [{"country":"日本","leagues":["J1","J2"]}, ...]
    2) dict形式: {"日本":["J1","J2"], "ドイツ":"ブンデスリーガ"} も互換で対応
    """
    p = Path(json_path).expanduser()
    if not p.exists():
        return [], []

    data = json.loads(p.read_text(encoding="utf-8"))
    pairs: list[tuple[str, str]] = []

    # --- 1) list形式（今回の実データ） ---
    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            country = str(item.get("country", "")).strip()
            if not country:
                continue

            leagues_val = item.get("leagues", [])
            if isinstance(leagues_val, list):
                for lv in leagues_val:
                    league = str(lv).strip()
                    if league:
                        pairs.append((country, league))
            else:
                league = str(leagues_val).strip()
                if league:
                    pairs.append((country, league))

    # --- 2) dict形式（互換） ---
    elif isinstance(data, dict):
        for c, v in data.items():
            country = str(c).strip()
            if not country:
                continue
            if isinstance(v, list):
                for lv in v:
                    league = str(lv).strip()
                    if league:
                        pairs.append((country, league))
            else:
                league = str(v).strip()
                if league:
                    pairs.append((country, league))

    countries = sorted({c for c, _ in pairs})
    return countries, pairs

# =========================
# teamData_*.xlsx -> teamData_*.csv（最後に実行）
# =========================
TEAM_MEMBER_PREFIX = EXCEL_HEADER

def get_team_member_xlsx_paths(base_dir: str) -> list[str]:
    return sorted(glob.glob(os.path.join(base_dir, f"{TEAM_MEMBER_PREFIX}*.xlsx")))

def next_team_member_csv_seq(base_dir: str) -> int:
    max_seq = 0
    for path in glob.glob(os.path.join(base_dir, f"{TEAM_MEMBER_PREFIX}*.csv")):
        m = re.fullmatch(rf"{TEAM_MEMBER_PREFIX}(\d+)\.csv", os.path.basename(path))
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return max_seq + 1

def validate_team_member_xlsx(excel_file: str) -> bool:
    required_cols = {"国", "リーグ", "チーム", "チームリンク"}
    try:
        df = pd.read_excel(excel_file)
    except Exception:
        print(f"[WARN] 読み込み失敗/破損: {excel_file} -> 削除")
        try:
            os.remove(excel_file)
        except Exception:
            pass
        return False

    if not required_cols.issubset(set(df.columns)):
        print(f"[WARN] 必要カラム不足: {excel_file} -> 変換スキップ（削除しない）")
        return False

    return True

def excel_to_csv_and_delete(excel_file: str, csv_file: str) -> bool:
    """
    ✅ CSVを作ったら、S3(バケット直下)へアップロードしてから xlsx を削除
    """
    try:
        df = pd.read_excel(excel_file)
        df.to_csv(csv_file, index=False)

        # ✅ CSVをS3バケット直下へ
        s3_upload_csv_to_bucket_root(csv_file)

        os.remove(excel_file)
        print(f"[CONVERT] {excel_file} -> {csv_file}（xlsx削除 & S3アップロード）")
        return True
    except Exception as e:
        print(f"[ERROR] Excel->CSV失敗: {excel_file} ({e})")
        return False

def convert_team_member_xlsx_to_csv(base_dir: str, limit: int = 50) -> None:
    xlsx_paths = get_team_member_xlsx_paths(base_dir)
    if not xlsx_paths:
        print("変換対象の teamData_*.xlsx がありません。")
        return

    done = 0
    for xlsx_path in xlsx_paths:
        if done >= limit:
            break

        if not validate_team_member_xlsx(xlsx_path):
            done += 1
            print(f"{done}/{len(xlsx_paths)}番目")
            continue

        out_seq = next_team_member_csv_seq(base_dir)
        csv_path = os.path.join(base_dir, f"{TEAM_MEMBER_PREFIX}{out_seq}.csv")
        excel_to_csv_and_delete(xlsx_path, csv_path)

        done += 1
        print(f"{done}/{len(xlsx_paths)}番目")

# =========================
# 実行
# =========================
async def main():
    global TARGETS

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"{now}  データ取得対象の時間です。（SEED）")

    # ✅ 入力JSONをS3から復元（あれば）
    s3_download_if_exists(S3_BUCKET, B001_S3_KEY, B001_JSON_PATH)

    # --- JSON を読む（あれば country+league 指定、なければ CONTAINS_LIST） ---
    _, country_league_pairs = extract_countries_and_leagues(B001_JSON_PATH)
    json_has_pairs = bool(country_league_pairs)

    if json_has_pairs:
        json_targets: Dict[str, Set[str]] = {}
        for c, l in country_league_pairs:
            c2 = re.sub(r"\s+", " ", str(c or "")).strip()
            l2 = re.sub(r"\s+", " ", str(l or "")).strip()
            if c2 and l2:
                json_targets.setdefault(c2, set()).add(l2)

        TARGETS = json_targets
        allowed_countries: Optional[Set[str]] = set(TARGETS.keys())

        print(f"[FILTER] JSONあり: countries={len(allowed_countries)} pairs={sum(len(v) for v in TARGETS.values())}")
        print("[TARGETS]", {k: sorted(list(v)) for k, v in TARGETS.items()})
    else:
        allowed_countries = set(TARGETS.keys())
        print("[FILTER] JSONなし/空: CONTAINS_LIST のみ運用")
        print("[TARGETS]", {k: sorted(list(v)) for k, v in TARGETS.items()})

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--disable-gpu",
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--no-zygote",
                "--disable-dev-shm-usage",
            ],
        )

        menu_ctx = await make_context(browser, block_css=False)
        menu_page = await menu_ctx.new_page()
        leagues = await collect_target_leagues(menu_page, allowed_countries=allowed_countries)
        await menu_page.close()
        await menu_ctx.close()
        print(f"対象リーグ数: {len(leagues)}")

        work_ctx = await make_context(browser, block_css=True)
        for (country, league, href, comp_id) in leagues:
            teams = await scrape_league_to_teams(work_ctx, country, league, href, comp_id)
            if teams:
                write_teams_excel(country, league, teams)  # xlsxはローカルのみ
            else:
                print(f"[WARN] {country}:{league} チーム取得ゼロ")
        await work_ctx.close()

        await browser.close()

    # ✅ xlsx -> csv 変換（csvのみS3へアップロード、xlsxは削除）
    convert_team_member_xlsx_to_csv(TEAMS_BY_LEAGUE_DIR, limit=50)

# Lambda互換
def handler(event=None, context=None):
    try:
        asyncio.run(main())
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(main())
    return {"statusCode": 200, "body": "ok"}

if __name__ == "__main__":
    asyncio.run(main())

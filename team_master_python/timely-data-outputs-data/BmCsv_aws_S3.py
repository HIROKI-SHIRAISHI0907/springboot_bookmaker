# -*- coding: utf-8 -*-
"""
bmCsv_s3.py — ゴール情報補完（スコア一致） & CSV 変換 → S3に保存（ECS向け）

S3想定フロー:
1) s3://{S3_BUCKET}/{S3_INPUT_PREFIX} 配下から output_*.xlsx を列挙
2) 各xlsxを /tmp にDL
3) Playwrightでゴール時間/選手名を補完（無ければ '---' 埋め）
4) CSVへ変換（通番/ソート秒 付与も維持）
5) CSVを s3://{S3_BUCKET}/{S3_OUTPUT_PREFIX} に Put
6) 成功したxlsxは（任意で）S3から削除
7) match_sequence.json（試合ごとの通番）はS3で永続化

ENV:
- S3_BUCKET=aws-s3-outputs-csv
- S3_INPUT_PREFIX=   (例: "" / "xlsx/")
- S3_OUTPUT_PREFIX=  (例: "" / "csv/")
- SEQ_JSON_S3_KEY=csv_seq/match_sequence.json
- DELETE_SOURCE_XLSX=0 or 1
- BM_DEBUG=0/1  (ECSでは通常0推奨)
- HEADLESS=1/0
- MAX_FILES=0 (0=無制限)
"""

from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PTimeout

from openpyxl import load_workbook
import pandas as pd

import os
import re
import time
import sys
import json
import io
from typing import List, Tuple, Optional, Dict

import boto3
from botocore.exceptions import ClientError


# ==========================
# ENV / Settings
# ==========================
S3_BUCKET = "aws-s3-outputs-csv"
S3_INPUT_PREFIX = ""
S3_OUTPUT_PREFIX = ""
SEQ_JSON_S3_KEY = "csv_seq/match_sequence.json"

DELETE_SOURCE_XLSX = 1
MAX_FILES = 0

DEBUG = 0
HEADLESS = 1

# Sheet / file naming
FILE_PREFIX = "output_"
SHEET_NAME = "Sheet1"

# Excel headers
COL_LINK   = "試合リンク文字列"
COL_HOME   = "ホームスコア"
COL_AWAY   = "アウェースコア"
COL_TIME   = "ゴール時間"
COL_MEMBER = "選手名"

# Added columns
COL_MATCH_ID = "試合ID"
COL_SEQ      = "通番"
COL_SORT_SEC = "ソート用秒"

# Local temp
TMP_DIR = "/tmp/bm_csv"
os.makedirs(TMP_DIR, exist_ok=True)

# Log to S3 (optional)
LOG_PREFIX = (S3_OUTPUT_PREFIX.rstrip("/") + "/log/") if S3_OUTPUT_PREFIX else "log/"


def log(msg: str):
    print(msg, flush=True)


# ==========================
# S3 helpers
# ==========================
def s3_client():
    return boto3.client("s3")

def s3_list_keys(prefix: str) -> List[str]:
    s3 = s3_client()
    keys: List[str] = []
    token = None
    while True:
        kwargs = {"Bucket": S3_BUCKET, "Prefix": prefix}
        if token:
            kwargs["ContinuationToken"] = token
        resp = s3.list_objects_v2(**kwargs)
        for it in resp.get("Contents", []):
            k = it.get("Key")
            if k:
                keys.append(k)
        if resp.get("IsTruncated"):
            token = resp.get("NextContinuationToken")
        else:
            break
    return keys

def s3_get_bytes(key: str) -> Optional[bytes]:
    s3 = s3_client()
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=key)
        return obj["Body"].read()
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code", "")
        if code in ("NoSuchKey", "404", "NotFound"):
            return None
        raise

def s3_put_bytes(key: str, data: bytes, content_type: str) -> None:
    s3 = s3_client()
    s3.put_object(Bucket=S3_BUCKET, Key=key, Body=data, ContentType=content_type)
    log(f"✅ S3 put: s3://{S3_BUCKET}/{key}")

def s3_delete(key: str) -> None:
    s3 = s3_client()
    s3.delete_object(Bucket=S3_BUCKET, Key=key)
    log(f"🗑️ S3 delete: s3://{S3_BUCKET}/{key}")

def s3_download_to_tmp(key: str) -> str:
    s3 = s3_client()
    fname = os.path.basename(key)
    local_path = os.path.join(TMP_DIR, fname)
    s3.download_file(S3_BUCKET, key, local_path)
    return local_path

def s3_upload_file(local_path: str, key: str, content_type: str) -> None:
    with open(local_path, "rb") as f:
        data = f.read()
    s3_put_bytes(key, data, content_type=content_type)

def s3_write_text(key: str, text: str) -> None:
    s3_put_bytes(key, text.encode("utf-8"), content_type="text/plain; charset=utf-8")


# ==========================
# DataFrame utilities
# ==========================
def _normalize_and_drop_blank_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df2 = df.replace(r"^\s*$", pd.NA, regex=True)
    df2 = df2.dropna(how="all")
    return df2

def read_excel_df(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"ブックが存在しません: {path}")
    df = pd.read_excel(path, sheet_name=SHEET_NAME)
    return df


# ==========================
# match_sequence.json (S3)
# ==========================
def load_seq_dict_s3() -> dict:
    b = s3_get_bytes(SEQ_JSON_S3_KEY)
    if not b:
        log(f"🆕 [SEQJSON] S3に既存なし: s3://{S3_BUCKET}/{SEQ_JSON_S3_KEY}")
        return {}
    try:
        return json.loads(b.decode("utf-8"))
    except Exception as e:
        log(f"⚠️ [SEQJSON] 読み込み失敗 → 空で開始: {e}")
        return {}

def save_seq_dict_s3(d: dict):
    body = json.dumps(d, ensure_ascii=False, indent=2).encode("utf-8")
    s3_put_bytes(SEQ_JSON_S3_KEY, body, content_type="application/json; charset=utf-8")


# ==========================
# ID / seq / seconds
# ==========================
def extract_mid(link: str) -> Optional[str]:
    if not link:
        return None
    url = link.strip()
    m_q = re.search(r"[?&#]mid=([A-Za-z0-9]+)", url)
    if m_q:
        mid = m_q.group(1)
        return mid if 8 <= len(mid) <= 16 else None
    m_p = re.search(r"/match/([A-Za-z0-9]+)/", url)
    if m_p:
        mid = m_p.group(1)
        if mid.lower() == "soccer":
            return None
        return mid if 8 <= len(mid) <= 16 else None
    return None

def get_match_id_from_link(link: str) -> str:
    mid = extract_mid(link)
    return f"mid_{mid}" if mid else "mid_unknown"

def mmss_to_seconds(txt: str) -> int:
    txt = (txt or "").strip()
    if not txt:
        return 0
    if ":" in txt:
        try:
            m, s = [int(x) for x in txt.split(":")[:2]]
            return m * 60 + s
        except Exception:
            pass
    m = re.search(r"(\d+)\s*(\+\s*\d+)?", txt)
    if m:
        base = int(m.group(1))
        extra = int(re.sub(r"[^\d]", "", m.group(2))) if m.group(2) else 0
        return (base + extra) * 60
    if txt.isdigit():
        return int(txt) * 60
    return 0

def add_sequence_and_sort_seconds(df: pd.DataFrame) -> pd.DataFrame:
    seq_dict = load_seq_dict_s3()
    df = df.copy()
    if COL_LINK not in df.columns:
        df[COL_LINK] = ""
    df[COL_MATCH_ID] = df[COL_LINK].astype(str).apply(get_match_id_from_link)

    new_seq_dict = seq_dict.copy()
    seq_col, sort_col = [], []
    for _, row in df.iterrows():
        mid = row.get(COL_MATCH_ID, "mid_unknown")
        new_seq = new_seq_dict.get(mid, 0) + 1
        new_seq_dict[mid] = new_seq
        seq_col.append(new_seq)
        time_txt = str(row.get(COL_TIME) or "").strip()
        sort_col.append(mmss_to_seconds(time_txt))
    df[COL_SEQ] = seq_col
    df[COL_SORT_SEC] = sort_col

    save_seq_dict_s3(new_seq_dict)
    return df


# ==========================
# URL normalize
# ==========================
def to_full_url(link: str) -> str:
    if not link:
        return ""
    url = link.strip()

    if url.startswith("http"):
        full = url
    elif url.startswith("/"):
        full = "https://www.flashscore.co.jp" + url
    else:
        full = "https://www.flashscore.co.jp/" + url

    full = re.sub(r"/standings(?:/[^?#]*)*", "", full)
    full = full.split("#")[0]
    return full

def build_candidate_urls(link: str) -> List[str]:
    cands: List[str] = []
    base = to_full_url(link)
    if base:
        cands.append(base)

    mid = extract_mid(base)
    if mid:
        root = f"https://www.flashscore.co.jp/match/{mid}/"
        cands.append(root)

    if base:
        cands.append(base + "#/")
        cands.append(base + "#/match-summary")
    if mid:
        cands.append(root + "#/")
        cands.append(root + "#/match-summary")

    return list(dict.fromkeys(cands))


# ==========================
# Playwright goal extraction
# ==========================
def close_cookie_banners(page):
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('同意')",
        "button:has-text('許可')",
        "button:has-text('OK')",
        "button:has-text('Accept')",
        "[data-testid*='cookie'] button",
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click(timeout=1500)
                time.sleep(0.2)
        except Exception:
            pass

def _score_text_match(txt: str, home: int, away: int) -> bool:
    hy = r"[\-\u2212\u2012\u2013\u2014\u2015]"
    pat = rf"\b{home}\s*{hy}\s*{away}\b"
    return re.search(pat, txt) is not None

def _extract_goal_for_score(page, home: int, away: int) -> Tuple[str, str]:
    try:
        icons = page.locator("svg[data-testid='wcl-icon-soccer']")
        n = icons.count()
        for i in range(n):
            ic = icons.nth(i)
            row = ic.locator("xpath=ancestor::*[contains(@class,'smv__incident')][1]")
            if not row or row.count() == 0:
                row = ic.locator("xpath=ancestor::*[self::div or self::li][1]")

            # score check
            score_ok = False
            for ssel in [
                ".smv__incidentHomeScore",
                "[class*='incidentHomeScore']",
                "[data-testid*='score']",
            ]:
                try:
                    loc = row.locator(ssel).first
                    if loc and loc.count():
                        txt = loc.inner_text(timeout=400)
                        if txt and _score_text_match(txt, home, away):
                            score_ok = True
                            break
                except Exception:
                    pass
            if not score_ok:
                try:
                    txt = row.inner_text(timeout=500)
                    if not _score_text_match(txt, home, away):
                        continue
                except Exception:
                    continue

            # time
            time_txt = ""
            for tsel in [".smv__timeBox", "[data-testid*='incident-time']", "[class*='time']"]:
                try:
                    loc = row.locator(tsel).first
                    if loc and loc.count():
                        time_txt = loc.inner_text(timeout=500)
                        if time_txt:
                            break
                except Exception:
                    pass
            if not time_txt:
                try:
                    txt = row.inner_text(timeout=500)
                    m = re.search(r"(\d{1,3}(?::\d{2})?(?:\+\d+)?|^\d{1,3}'(?:\+\d{1,2})?)", txt or "")
                    if m:
                        time_txt = m.group(0)
                except Exception:
                    pass

            # player
            player = ""
            for psel in [".smv__playerName", "a[href*='/player/']", "[data-testid*='participant']", "[class*='participant'], [class*='player']"]:
                try:
                    loc = row.locator(psel).first
                    if loc and loc.count():
                        player = loc.inner_text(timeout=700)
                        if player:
                            break
                except Exception:
                    pass
            if not player:
                try:
                    txt = row.inner_text(timeout=500)
                    t = re.sub(r"(\d{1,3}(?::\d{2})?(?:\+\d+)?|ゴール|Goal|PK|ペナルティ|オウン|OG|Own)", "", txt or "", flags=re.IGNORECASE)
                    t = re.sub(r"\d+\s*[-−–]\s*\d+", "", t)
                    t = re.sub(r"\s{2,}", " ", t).strip()
                    player = t.splitlines()[0].strip() if t else ""
                except Exception:
                    pass

            if time_txt and player:
                return (time_txt.strip(), player.strip())
    except Exception:
        pass

    return ("---", "---")

def _ensure_columns(ws, needed: List[str]):
    header = [cell.value for cell in ws[1]]
    for name in needed:
        if name not in header:
            ws.cell(row=1, column=len(header) + 1, value=name)
            header.append(name)

def _write_back_excel(df: pd.DataFrame, xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]
    _ensure_columns(ws, list(df.columns))
    header_pos = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for r_idx, (_, s) in enumerate(df.iterrows(), start=2):
        for col_name, val in s.items():
            cidx = header_pos[col_name]
            ws.cell(r_idx, cidx, value=val)
    tmp = xlsx_path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, xlsx_path)

def fill_missing_goals_with_playwright(xlsx_path: str, pw) -> bool:
    df = read_excel_df(xlsx_path)
    df = _normalize_and_drop_blank_rows(df)
    if df is None or df.empty:
        return False

    for col in [COL_LINK, COL_HOME, COL_AWAY, COL_TIME, COL_MEMBER]:
        if col not in df.columns:
            log(f"必要な列が不足のため補完不可: {xlsx_path}")
            return False

    browser = pw.chromium.launch(
        headless=HEADLESS and (not DEBUG),
        args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage"],
    )
    context = browser.new_context(user_agent=(
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    ))
    page = context.new_page()

    changed = False
    page_cache: Dict[str, Dict] = {}

    for idx, row in df.iterrows():
        cur_time   = str(row.get(COL_TIME) or "")
        cur_member = str(row.get(COL_MEMBER) or "")
        if cur_time.strip() and cur_member.strip():
            continue

        link = str(row.get(COL_LINK, "") or "").strip()
        if not link:
            if not cur_time.strip():
                df.at[idx, COL_TIME] = "---"; changed = True
            if not cur_member.strip():
                df.at[idx, COL_MEMBER] = "---"; changed = True
            continue

        try:
            home = int(str(row.get(COL_HOME)).strip())
            away = int(str(row.get(COL_AWAY)).strip())
        except Exception:
            if not cur_time.strip():
                df.at[idx, COL_TIME] = "---"; changed = True
            if not cur_member.strip():
                df.at[idx, COL_MEMBER] = "---"; changed = True
            continue

        attempted: List[str] = []
        ok_url = None

        cached = page_cache.get(link, {})
        if cached.get("ok_url"):
            try:
                page.goto(cached["ok_url"], timeout=15000, wait_until="domcontentloaded")
                close_cookie_banners(page)
                ok_url = cached["ok_url"]
            except Exception:
                ok_url = None

        if not ok_url:
            for url in build_candidate_urls(link):
                attempted.append(url)
                try:
                    page.goto(url, timeout=30000, wait_until="domcontentloaded")
                    close_cookie_banners(page)
                    ok_url = url
                    break
                except PTimeout:
                    log(f"timeout: {url}")
                    continue
                except Exception as e:
                    log(f"visit error: {url} / {e}")
                    continue
            page_cache[link] = {"attempted": attempted, "ok_url": ok_url}

        if not ok_url:
            log("到達不可（試行URL・正規化後）:")
            for u in attempted:
                log("  - " + u)
            if not cur_time.strip():
                df.at[idx, COL_TIME] = "---"; changed = True
            if not cur_member.strip():
                df.at[idx, COL_MEMBER] = "---"; changed = True
            continue

        time_txt, player = _extract_goal_for_score(page, home, away)
        if not time_txt or not player or time_txt == "---" or player == "---":
            attempted2 = page_cache.get(link, {}).get("attempted", attempted)
            log("ゴール抽出失敗（試行URL・正規化後）:")
            for u in attempted2:
                log("  - " + u)
            if not cur_time.strip():
                df.at[idx, COL_TIME] = "---"; changed = True
            if not cur_member.strip():
                df.at[idx, COL_MEMBER] = "---"; changed = True
            continue

        if not cur_time.strip():
            df.at[idx, COL_TIME] = time_txt; changed = True
        if not cur_member.strip():
            df.at[idx, COL_MEMBER] = player; changed = True

    # Close
    try:
        context.close()
    except Exception:
        pass
    try:
        browser.close()
    except Exception:
        pass

    if changed:
        _write_back_excel(df, xlsx_path)
        return True

    # ensure no blanks remain
    filled_any = False
    for idx, row in df.iterrows():
        if not str(row.get(COL_TIME) or "").strip():
            df.at[idx, COL_TIME] = "---"; filled_any = True
        if not str(row.get(COL_MEMBER) or "").strip():
            df.at[idx, COL_MEMBER] = "---"; filled_any = True
    if filled_any:
        _write_back_excel(df, xlsx_path)
        return True

    return False


# ==========================
# Excel -> CSV (local) then S3 upload
# ==========================
def excel_to_csv_local(excel_file: str, csv_file: str) -> bool:
    try:
        df = pd.read_excel(excel_file, sheet_name=SHEET_NAME)
        df = _normalize_and_drop_blank_rows(df)
        if df is None or df.empty:
            log(f"空ブックのためスキップ: {excel_file}")
            return False

        df = add_sequence_and_sort_seconds(df)
        df.to_csv(csv_file, index=False)
        log(f"CSV 出力完了: {csv_file}")
        return True
    except Exception as e:
        log(f"CSV 変換失敗: {excel_file} -> {csv_file} / {e}")
        return False

def output_key_for_csv(src_xlsx_key: str) -> str:
    # 入力キーのファイル番号を維持してCSVにする（例: output_12.xlsx -> output_12.csv）
    base = os.path.basename(src_xlsx_key)
    csv_name = base.replace(".xlsx", ".csv")
    out_prefix = S3_OUTPUT_PREFIX
    if out_prefix and not out_prefix.endswith("/"):
        out_prefix += "/"
    return f"{out_prefix}{csv_name}"

def should_process_xlsx_key(key: str) -> bool:
    base = os.path.basename(key)
    if not base.startswith(FILE_PREFIX):
        return False
    return base.endswith(".xlsx")

def list_input_xlsx_keys() -> List[str]:
    prefix = S3_INPUT_PREFIX
    if prefix and not prefix.endswith("/"):
        prefix += "/"
    keys = s3_list_keys(prefix)
    out = [k for k in keys if should_process_xlsx_key(k)]
    # output_<num>.xlsx を数値順に
    def sort_key(k: str) -> int:
        b = os.path.basename(k)
        s = b.replace(FILE_PREFIX, "").replace(".xlsx", "")
        return int(s) if s.isdigit() else 10**9
    out.sort(key=sort_key)
    return out


# ==========================
# Main
# ==========================
def main():
    log(f"S3_BUCKET={S3_BUCKET}, INPUT_PREFIX='{S3_INPUT_PREFIX}', OUTPUT_PREFIX='{S3_OUTPUT_PREFIX}', SEQ_JSON_S3_KEY={SEQ_JSON_S3_KEY}")
    xlsx_keys = list_input_xlsx_keys()
    if not xlsx_keys:
        log("処理対象の XLSX がS3に見つかりません。終了します。")
        return

    if MAX_FILES > 0:
        xlsx_keys = xlsx_keys[:MAX_FILES]

    log(f"処理対象 XLSX: {len(xlsx_keys)} 件")

    with sync_playwright() as pw:
        processed = 0
        for key in xlsx_keys:
            log(f"\n=== 処理対象: s3://{S3_BUCKET}/{key} ===")
            try:
                local_xlsx = s3_download_to_tmp(key)
            except Exception as e:
                log(f"❌ DL失敗: {e}")
                continue

            # 欠損補完（見つからない場合でも '---' を埋めるロジック）
            try:
                _ = fill_missing_goals_with_playwright(local_xlsx, pw)
            except Exception as e:
                log(f"⚠️ 補完処理で例外（続行）: {e}")

            local_csv = os.path.join(TMP_DIR, os.path.basename(key).replace(".xlsx", ".csv"))

            if not excel_to_csv_local(local_xlsx, local_csv):
                log(f"CSV 変換失敗: {key}")
                continue

            # Upload CSV to S3 (same bucket)
            out_key = output_key_for_csv(key)
            try:
                s3_upload_file(local_csv, out_key, content_type="text/csv; charset=utf-8")
            except Exception as e:
                log(f"❌ CSVのS3アップロード失敗: {e}")
                continue

            # Optionally delete source xlsx in S3
            if DELETE_SOURCE_XLSX:
                try:
                    s3_delete(key)
                except Exception as e:
                    log(f"⚠️ xlsx削除失敗（続行）: {e}")

            processed += 1

    log(f"\n処理件数: {processed}")


if __name__ == "__main__":
    main()

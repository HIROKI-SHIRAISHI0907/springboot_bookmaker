# -*- coding: utf-8 -*-
"""
bmCsv.py — ゴール情報補完（スコア一致） & CSV 変換（正規化URLログ / DEBUG対応）
---------------------------------------------------------------------------
処理フロー:
1) /Users/shiraishitoshio/bookmaker/outputs 配下の output_*.xlsx を通番順に処理
2) 行ごとに「ホームスコア/アウェースコア」を読み、試合ページから
   ⚽️ アイコン(data-testid='wcl-icon-soccer')の行で「スコア 1-0」「時間 64'」「選手名」を抽出
   - 合致スコアが見つかれば「ゴール時間」「選手名」を埋める
   - 見つからなければ両方 '---' を入れる（空欄は残さない）
3) 全行が「値 or ---」で埋まっていれば CSV 変換 → 成功で XLSX を削除
4) ログは必ず「正規化後に試行した URL」を出す（元の standings URL ではなく）

DEBUG:
- 環境変数 BM_DEBUG=1 でブラウザ表示＆終了待ち（Enter）にする
- 通常はヘッドレスで自動クローズ
"""

from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PTimeout
from openpyxl import load_workbook
import pandas as pd
import os
import re
import glob
import time
import sys
import json
from typing import List, Tuple, Optional

# ===== スイッチ =====
DEBUG = os.getenv("BM_DEBUG", "0") in ("1", "true", "True")
DEBUG = False
FORCE_CONVERT = True  # ← 追加：常に CSV 変換する

# ===== パス設定 =====
BASE_DIR = "/Users/shiraishitoshio/bookmaker/outputs"
LOG_DIR = os.path.join(BASE_DIR, "log")
os.makedirs(LOG_DIR, exist_ok=True)

FILE_PREFIX = "output_"
SHEET_NAME = "Sheet1"

# エクセル列ヘッダー
COL_LINK   = "試合リンク文字列"  # CO
COL_HOME   = "ホームスコア"
COL_AWAY   = "アウェースコア"
COL_TIME   = "ゴール時間"        # CP
COL_MEMBER = "選手名"            # CQ

# 付与列
COL_MATCH_ID = "試合ID"
COL_SEQ      = "通番"
COL_SORT_SEC = "ソート用秒"

SEQ_FILE = os.path.join(BASE_DIR, "match_sequence.json")

# ==========================
# ロック
# ==========================
class FileLock:
    def __init__(self, path: str, timeout: int = 0):
        self.path = path
        self.timeout = timeout
    def acquire(self) -> bool:
        start = time.time()
        while True:
            try:
                fd = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.write(fd, str(os.getpid()).encode())
                os.close(fd)
                return True
            except FileExistsError:
                if self.timeout and (time.time() - start) < self.timeout:
                    time.sleep(0.2)
                    continue
                return False
    def release(self):
        try:
            os.remove(self.path)
        except FileNotFoundError:
            pass

def atomic_replace(src_tmp: str, dest_path: str):
    os.replace(src_tmp, dest_path)

# ==========================
# ユーティリティ
# ==========================
def fmt(path: str) -> str:
    return os.path.abspath(path)

def safe_remove(path: str, label: str = "") -> bool:
    try:
        if os.path.exists(path):
            os.remove(path)
            print(f"{label}削除: {path}")
            return True
        return False
    except Exception as e:
        print(f"{label}削除失敗: {path} / {e}")
        return False

def list_all_xlsx() -> List[str]:
    paths = glob.glob(os.path.join(BASE_DIR, f"{FILE_PREFIX}*.xlsx"))
    def key(p):
        s = os.path.basename(p).replace(FILE_PREFIX, "").replace(".xlsx", "")
        return int(s) if s.isdigit() else 10**9
    return sorted(paths, key=key)

def next_csv_seq() -> int:
    seqs = []
    for path in glob.glob(os.path.join(BASE_DIR, f"{FILE_PREFIX}*.csv")):
        s = os.path.basename(path).replace(FILE_PREFIX, "").replace(".csv", "")
        if s.isdigit():
            seqs.append(int(s))
    return (max(seqs) + 1) if seqs else 1

# ==========================
# JSON: matchごとの通番
# ==========================
def load_seq_dict() -> dict:
    if os.path.exists(SEQ_FILE):
        try:
            with open(SEQ_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_seq_dict(d: dict):
    tmp = SEQ_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
    atomic_replace(tmp, SEQ_FILE)

# ==========================
# DataFrame処理
# ==========================
def _normalize_and_drop_blank_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    df2 = df.replace(r"^\s*$", pd.NA, regex=True)
    df2 = df2.dropna(how='all')
    return df2

def read_excel_df(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"ブックが存在しません: {path}")
    try:
        df = pd.read_excel(path, sheet_name=SHEET_NAME)
        return df
    except Exception as e:
        print(f"ファイルが破損しています。削除します: {path} / {e}")
        safe_remove(path, label="[XLSX]")
        raise

# ==========================
# 試合ID/通番/ソート秒
# ==========================
def extract_mid(link: str) -> Optional[str]:
    """?mid= を最優先。/match/<MID>/ は 'soccer' を除外し長さもチェック。"""
    if not link:
        return None
    url = link.strip()

    m_q = re.search(r"[?&#]mid=([A-Za-z0-9]+)", url)
    if m_q:
        mid = m_q.group(1)
        return mid if 8 <= len(mid) <= 16 else None

    m_p = re.search(r"/match/([A-Za-z0-9]+)/", url)
    if m_p:
        mid = m_p.group(1)
        if mid.lower() == "soccer":
            return None
        return mid if 8 <= len(mid) <= 16 else None
    return None

def get_match_id_from_link(link: str) -> str:
    mid = extract_mid(link)
    return f"mid_{mid}" if mid else "mid_unknown"

def mmss_to_seconds(txt: str) -> int:
    txt = (txt or "").strip()
    if not txt:
        return 0
    if ":" in txt:
        try:
            m, s = [int(x) for x in txt.split(":")[:2]]
            return m * 60 + s
        except Exception:
            pass
    m = re.search(r"(\d+)\s*(\+\s*\d+)?", txt)
    if m:
        base = int(m.group(1))
        extra = int(re.sub(r"[^\d]", "", m.group(2))) if m.group(2) else 0
        return (base + extra) * 60
    if txt.isdigit():
        return int(txt) * 60
    return 0

def add_sequence_and_sort_seconds(df: pd.DataFrame) -> pd.DataFrame:
    seq_dict = load_seq_dict()
    df = df.copy()
    if COL_LINK not in df.columns:
        df[COL_LINK] = ""
    df[COL_MATCH_ID] = df[COL_LINK].astype(str).apply(get_match_id_from_link)

    new_seq_dict = seq_dict.copy()
    seq_col, sort_col = [], []
    for _, row in df.iterrows():
        mid = row.get(COL_MATCH_ID, "mid_unknown")
        new_seq = new_seq_dict.get(mid, 0) + 1
        new_seq_dict[mid] = new_seq
        seq_col.append(new_seq)
        time_txt = str(row.get(COL_TIME) or "").strip()
        sort_col.append(mmss_to_seconds(time_txt))
    df[COL_SEQ] = seq_col
    df[COL_SORT_SEC] = sort_col
    save_seq_dict(new_seq_dict)
    return df

# ==========================
# Excel → CSV 変換
# ==========================
def excel_to_csv(excel_file: str, csv_file: str) -> bool:
    """空行だけ除外し、CSVへ。'---' はそのまま出力。成功で XLSX 削除。"""
    try:
        df = pd.read_excel(excel_file, sheet_name=SHEET_NAME)
        df = _normalize_and_drop_blank_rows(df)
        if df is None or df.empty:
            print(f"空ブックのためスキップ: {excel_file}")
            return False

        df = add_sequence_and_sort_seconds(df)
        df.to_csv(csv_file, index=False)
        print(f"CSV 出力完了: {csv_file}")

        safe_remove(excel_file, label="[XLSX]")
        return True
    except Exception as e:
        print(f"CSV 変換失敗: {excel_file} -> {csv_file} / {e}")
        return False

# ==========================
# URL 正規化（素の ?mid=... を最優先）
# ==========================
def to_full_url(link: str) -> str:
    """standings サブパスを除去し、フラグメント(#/...)は付けない。"""
    if not link:
        return ""
    url = link.strip()

    if url.startswith("http"):
        full = url
    elif url.startswith("/"):
        full = "https://www.flashscore.co.jp" + url
    else:
        full = "https://www.flashscore.co.jp/" + url

    # standings 系を根こそぎ除去
    full = re.sub(r"/standings(?:/[^?#]*)*", "", full)
    # フラグメントは落とす
    full = full.split("#")[0]
    return full

def build_candidate_urls(link: str) -> List[str]:
    """正規化後 URL を返す（素のURL最優先）。"""
    cands: List[str] = []
    base = to_full_url(link)  # 例: https://.../?mid=U3COLpOk
    if base:
        cands.append(base)  # 最優先

    mid = extract_mid(base)
    if mid:
        root = f"https://www.flashscore.co.jp/match/{mid}/"
        cands.append(root)  # 一部で有効なルート

    # フォールバック（最後に）
    if base:
        cands.append(base + "#/")
        cands.append(base + "#/match-summary")
    if mid:
        cands.append(root + "#/")
        cands.append(root + "#/match-summary")

    # ユニーク化
    return list(dict.fromkeys(cands))

# ==========================
# Playwright 補完（スコア一致）
# ==========================
def close_cookie_banners(page):
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('同意')",
        "button:has-text('許可')",
        "button:has-text('OK')",
        "button:has-text('Accept')",
        "[data-testid*='cookie'] button",
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click(timeout=1500)
                time.sleep(0.2)
        except Exception:
            pass

def _score_text_match(txt: str, home: int, away: int) -> bool:
    # 例: "1 - 0" / "1-0" / "1–0"（ハイフンゆらぎ対応）
    hy = r"[\-\u2212\u2012\u2013\u2014\u2015]"  # -, −, ‒, –, — など
    pat = rf"\b{home}\s*{hy}\s*{away}\b"
    return re.search(pat, txt) is not None

def _extract_goal_for_score(page, home: int, away: int) -> Tuple[str, str]:
    """指定スコア(home-away)に一致する ⚽️ 行を探し、(time, player) を返す。無ければ ('---','---')"""
    try:
        icons = page.locator("svg[data-testid='wcl-icon-soccer']")
        n = icons.count()
        for i in range(n):
            ic = icons.nth(i)
            row = ic.locator("xpath=ancestor::*[contains(@class,'smv__incident')][1]")
            if not row or row.count() == 0:
                row = ic.locator("xpath=ancestor::*[self::div or self::li][1]")

            # スコア
            score_ok = False
            # まずは専用スコア要素を探す
            for ssel in [
                ".smv__incidentHomeScore",
                "[class*='incidentHomeScore']",
                "[data-testid*='score']",
            ]:
                try:
                    loc = row.locator(ssel).first
                    if loc and loc.count():
                        txt = loc.inner_text(timeout=400)
                        if txt and _score_text_match(txt, home, away):
                            score_ok = True
                            break
                except Exception:
                    pass
            # 見つからなければ行テキスト全体で判定
            if not score_ok:
                try:
                    txt = row.inner_text(timeout=500)
                    if not _score_text_match(txt, home, away):
                        continue
                except Exception:
                    continue

            # 時間
            time_txt = ""
            for tsel in [".smv__timeBox", "[data-testid*='incident-time']", "[class*='time']"]:
                try:
                    loc = row.locator(tsel).first
                    if loc and loc.count():
                        time_txt = loc.inner_text(timeout=500)
                        if time_txt: break
                except Exception:
                    pass
            if not time_txt:
                try:
                    txt = row.inner_text(timeout=500)
                    m = re.search(r"(\d{1,3}(?::\d{2})?(?:\+\d+)?|^\d{1,3}'(?:\+\d{1,2})?)", txt or "")
                    if m: time_txt = m.group(0)
                except Exception:
                    pass

            # 選手
            player = ""
            for psel in [".smv__playerName", "a[href*='/player/']", "[data-testid*='participant']", "[class*='participant'], [class*='player']"]:
                try:
                    loc = row.locator(psel).first
                    if loc and loc.count():
                        player = loc.inner_text(timeout=700)
                        if player: break
                except Exception:
                    pass
            if not player:
                try:
                    txt = row.inner_text(timeout=500)
                    t = re.sub(r"(\d{1,3}(?::\d{2})?(?:\+\d+)?|ゴール|Goal|PK|ペナルティ|オウン|OG|Own)", "", txt or "", flags=re.IGNORECASE)
                    t = re.sub(r"\d+\s*[-−–]\s*\d+", "", t)
                    t = re.sub(r"\s{2,}", " ", t).strip()
                    player = t.splitlines()[0].strip() if t else ""
                except Exception:
                    pass

            if time_txt and player:
                return (time_txt.strip(), player.strip())
    except Exception:
        pass

    return ("---", "---")

def _ensure_columns(ws, needed: List[str]):
    header = [cell.value for cell in ws[1]]
    for name in needed:
        if name not in header:
            ws.cell(row=1, column=len(header)+1, value=name)
            header.append(name)

def _write_back_excel(df: pd.DataFrame, xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME]
    _ensure_columns(ws, list(df.columns))
    header_pos = {ws.cell(1, c).value: c for c in range(1, ws.max_column+1)}
    for r_idx, (_, s) in enumerate(df.iterrows(), start=2):
        for col_name, val in s.items():
            cidx = header_pos[col_name]
            ws.cell(r_idx, cidx, value=val)
    tmp = xlsx_path + ".tmp"
    wb.save(tmp)
    atomic_replace(tmp, xlsx_path)

def fill_missing_goals_with_playwright(xlsx_path: str, pw) -> bool:
    """
    行ごとにホーム/アウェースコアを読み、そのスコアに一致するゴール行から
    「ゴール時間」「選手名」を埋める。無ければ '---' を入れる。何かしら書き込めば True。
    """
    try:
        df = read_excel_df(xlsx_path)
        df = _normalize_and_drop_blank_rows(df)
        if df is None or df.empty:
            return False

        for col in [COL_LINK, COL_HOME, COL_AWAY, COL_TIME, COL_MEMBER]:
            if col not in df.columns:
                print(f"必要な列が不足のため補完不可: {xlsx_path}")
                return False

        browser = pw.chromium.launch(
            headless=not DEBUG,
            args=["--disable-gpu"]
        )
        context = browser.new_context(user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ))
        page = context.new_page()

        changed = False
        # {orig_link: {"attempted":[正規化後URL...], "ok_url":str}}
        page_cache: dict[str, dict] = {}

        # 欠損のある行だけ対象
        targets = df[(df[COL_TIME].astype(str).str.strip() == "") | (df[COL_MEMBER].astype(str).str.strip() == "")]
        if DEBUG and targets.empty and len(df) > 0:
            # 欠損が無い場合でも1件だけ開いて確認（任意）
            sample = str(df.iloc[0].get(COL_LINK) or "").strip()
            for url in build_candidate_urls(sample):
                try:
                    page.goto(url, timeout=15000, wait_until="domcontentloaded")
                    close_cookie_banners(page)
                    break
                except Exception:
                    continue

        for idx, row in df.iterrows():
            cur_time   = str(row.get(COL_TIME) or "")
            cur_member = str(row.get(COL_MEMBER) or "")
            if cur_time.strip() and cur_member.strip():
                continue

            link = str(row.get(COL_LINK, "") or "").strip()
            if not link:
                if not cur_time.strip():
                    df.at[idx, COL_TIME] = "---"; changed = True
                if not cur_member.strip():
                    df.at[idx, COL_MEMBER] = "---"; changed = True
                continue

            # スコア
            try:
                home = int(str(row.get(COL_HOME)).strip())
                away = int(str(row.get(COL_AWAY)).strip())
            except Exception:
                if not cur_time.strip():
                    df.at[idx, COL_TIME] = "---"; changed = True
                if not cur_member.strip():
                    df.at[idx, COL_MEMBER] = "---"; changed = True
                continue

            # 到達（正規化後URLの試行ログも保存）
            attempted: List[str] = []
            ok_url = None
            if link in page_cache and "ok_url" in page_cache[link] and page_cache[link]["ok_url"]:
                try:
                    page.goto(page_cache[link]["ok_url"], timeout=15000, wait_until="domcontentloaded")
                    close_cookie_banners(page)
                    ok_url = page_cache[link]["ok_url"]
                except Exception:
                    ok_url = None
            if not ok_url:
                for url in build_candidate_urls(link):
                    attempted.append(url)
                    try:
                        page.goto(url, timeout=30000, wait_until="domcontentloaded")
                        close_cookie_banners(page)
                        ok_url = url
                        break
                    except PTimeout:
                        print(f"timeout: {url}")
                        continue
                    except Exception as e:
                        if "has been closed" in str(e):
                            # ブラウザ/コンテキスト閉鎖 → 再生成
                            try:
                                context.close()
                            except Exception:
                                pass
                            context = pw.new_context()
                            page = context.new_page()
                        print(f"visit error: {url} / {e}")
                        continue
                page_cache[link] = {"attempted": attempted, "ok_url": ok_url}

            if not ok_url:
                print("到達不可（試行URL・正規化後）:")
                for u in attempted: print("  -", u)
                if not cur_time.strip():
                    df.at[idx, COL_TIME] = "---"; changed = True
                if not cur_member.strip():
                    df.at[idx, COL_MEMBER] = "---"; changed = True
                continue

            # === スコア一致抽出 ===
            time_txt, player = _extract_goal_for_score(page, home, away)

            if not time_txt or not player or time_txt == "---" or player == "---":
                attempted = page_cache.get(link, {}).get("attempted", attempted)
                print("ゴール抽出失敗（試行URL・正規化後）:")
                for u in attempted: print("  -", u)
                if not cur_time.strip():
                    df.at[idx, COL_TIME] = "---"; changed = True
                if not cur_member.strip():
                    df.at[idx, COL_MEMBER] = "---"; changed = True
                continue

            # 成功 → エクセルに書き込み
            if not cur_time.strip():
                df.at[idx, COL_TIME] = time_txt; changed = True
            if not cur_member.strip():
                df.at[idx, COL_MEMBER] = player; changed = True

        # ブラウザのクローズ/保持
        if DEBUG:
            try:
                print("\n[BM_DEBUG] ブラウザは開いたままです。Enter で閉じます。")
                input()
            except EOFError:
                time.sleep(5)
        else:
            context.close(); browser.close()

        # 何か書いたら保存。書いてなくても空欄が残っていれば '---' で埋めて保存
        if changed:
            _write_back_excel(df, xlsx_path)
        else:
            filled_any = False
            for idx, row in df.iterrows():
                if not str(row.get(COL_TIME) or "").strip():
                    df.at[idx, COL_TIME] = "---"; filled_any = True
                if not str(row.get(COL_MEMBER) or "").strip():
                    df.at[idx, COL_MEMBER] = "---"; filled_any = True
            if filled_any:
                _write_back_excel(df, xlsx_path)
                changed = True

        return changed

    except Exception as e:
        print(f"補完中に失敗: {xlsx_path} / {e}")
        return False

def is_all_filled_or_dashed(xlsx_path: str) -> bool:
    """
    シートに COL_TIME/COL_MEMBER が存在し、NaN/空文字が無いことを確認。
    値が '---' で埋まっているのもOK（空欄さえ無ければ CSV 化可）。
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name=SHEET_NAME)
        df = _normalize_and_drop_blank_rows(df)
        if df is None or df.empty:
            return False
        if COL_TIME not in df.columns or COL_MEMBER not in df.columns:
            return False
        empty_time   = df[COL_TIME].isna() | (df[COL_TIME].astype(str).str.strip() == "")
        empty_member = df[COL_MEMBER].isna() | (df[COL_MEMBER].astype(str).str.strip() == "")
        return not (empty_time.any() or empty_member.any())
    except Exception:
        return False

# ==========================
# メイン
# ==========================
def main():
    xlsx_list = list_all_xlsx()
    if not xlsx_list:
        print("処理対象の XLSX が見つかりません。終了します。")
        return

    with sync_playwright() as pw:
        processed = 0
        for xlsx in map(fmt, xlsx_list):
            print(f"\n=== 処理対象: {xlsx} ===")

            lock_path = xlsx + ".lock"
            if os.path.exists(lock_path):
                print("収集中のためスキップ:", xlsx)
                continue

            # 欠損をスコア一致で補完（見つからなければ '---' を埋める）
            _ = fill_missing_goals_with_playwright(xlsx, pw)

            # ここで「空欄が残っていないか」を厳密チェック
            #if is_all_filled_or_dashed(xlsx):
            csv_seq = next_csv_seq()
            csv = fmt(os.path.join(BASE_DIR, f"{FILE_PREFIX}{csv_seq}.csv"))
            if excel_to_csv(xlsx, csv):
                print(f"変換成功 → XLSX 削除済: {csv}")
            else:
                print("CSV 変換失敗:", xlsx)
            #else:
            #    print("未完了のため CSV 変換を保留（空欄が残っています）:", xlsx)

            processed += 1

    print(f"\n処理件数: {processed}")

if __name__ == "__main__":
    proc_lock = FileLock(os.path.join(BASE_DIR, ".updBookmakerData.lock"))
    if not proc_lock.acquire():
        holder = ''
        try:
            with open(os.path.join(BASE_DIR, ".updBookmakerData.lock"), 'r', encoding='utf-8') as f:
                holder = f.read().strip()
        except Exception:
            pass
        try:
            mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(os.path.join(BASE_DIR, ".updBookmakerData.lock"))))
        except Exception:
            mtime = 'unknown'
        print(f"already running; exit (lock by PID={holder} since {mtime})")
        sys.exit(0)
    try:
        main()
    finally:
        proc_lock.release()

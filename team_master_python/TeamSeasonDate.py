# -*- coding: utf-8 -*-
# TeamSeasonDate.py
import os
import re
import asyncio
import unicodedata
from typing import Tuple, List, Iterable, Optional, Dict, Set
from urllib.parse import urljoin
import argparse
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright, TimeoutError as PwTimeoutError, Page

SEASON_XLSX = "season_data.xlsx"
BASE_URL = "https://www.flashscore.co.jp"
ICON_HEADER = "リーグアイコン"
RESULTS_HEADER = "結果URL"
STANDINGS_HEADER = "順位表URL"

MAX_CLICKS = 80  # もっと表示の上限

# ====== 取得対象（「国: リーグ」）======
CONTAINS_LIST = [
    "ケニア: プレミアリーグ","コロンビア: プリメーラ A","タンザニア: プレミアリーグ",
    "イングランド: プレミアリーグ","イングランド: EFL チャンピオンシップ","イングランド: EFL リーグ 1",
    "エチオピア: プレミアリーグ","コスタリカ: リーガ FPD","ジャマイカ: プレミアリーグ",
    "スペイン: ラ・リーガ","ブラジル: セリエ A ベターノ","ブラジル: セリエ B",
    "ドイツ: ブンデスリーガ","韓国: K リーグ 1","中国: 中国スーパーリーグ",
    "日本: J1 リーグ","日本: J2 リーグ","日本: J3 リーグ","日本: WEリーグ",
    "インドネシア: リーガ 1","オーストラリア: A リーグ・メン","チュニジア: チュニジア･プロリーグ",
    "ウガンダ: プレミアリーグ","メキシコ: リーガ MX","フランス: リーグ・アン","スコットランド: プレミアシップ",
    "オランダ: エールディビジ","アルゼンチン: トルネオ・ベターノ","イタリア: セリエ A","イタリア: セリエ B",
    "ポルトガル: リーガ・ポルトガル","トルコ: スュペル・リグ","セルビア: スーペルリーガ",
    "ボリビア: LFPB","ブルガリア: パルヴァ・リーガ","カメルーン: エリート 1","ペルー: リーガ 1",
    "エストニア: メスタリリーガ","ウクライナ: プレミアリーグ","ベルギー: ジュピラー･プロリーグ","エクアドル: リーガ・プロ",
]

# =========================
# ターゲット判定（表記ゆれ吸収）
# =========================
SPONSOR_WORDS = [
    "ベターノ", "ベタノ", "EAスポーツ", "EA スポーツ", "EA SPORTS",
    "Presented by", "presented by", "powered by", "Powered by"
]

def _strip_sponsors(s: str) -> str:
    if not s:
        return s
    s = re.sub(r"[（(].*?[)）]", "", s)  # 括弧内スポンサー
    for w in SPONSOR_WORDS:
        s = s.replace(w, "")
    return s

def _canon(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = _strip_sponsors(s)
    s = (s.replace("･", "").replace("・", "").replace("·", "")
           .replace("—", "").replace("–", "").replace("‐", "").replace("-", ""))
    s = re.sub(r"[A-Za-z]", "", s)      # EFL / EA SPORTS などを吸収
    s = re.sub(r"\s+", "", s).strip(" .，、")
    return s

def _split_pair(text: str) -> Tuple[str, str]:
    if text is None:
        return "", ""
    parts = re.split(r"\s*[:：]\s*", text, maxsplit=1)
    if len(parts) != 2:
        return "", ""
    return parts[0], parts[1]

RAW_TARGETS: List[Tuple[str, str]] = []
for item in CONTAINS_LIST:
    c, l = _split_pair(item)
    if c and l:
        RAW_TARGETS.append((c, l))
TARGETS: Set[Tuple[str, str]] = {(_canon(c), _canon(l)) for c, l in RAW_TARGETS}

def is_target(country_name: str, league_name: str) -> bool:
    cc = _canon(country_name)
    ll = _canon(league_name)
    for tc, tl in TARGETS:
        if cc == tc or tc in cc or cc in tc:
            if ll == tl or tl in ll or ll in tl:
                return True
    return False

# =========================
# Excel ユーティリティ
# =========================
def open_or_init_season_book(file_path: str = SEASON_XLSX):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "season"
        ws.append(["国", "リーグ", "シーズン開始", "シーズン終了", "ラウンド数", "パス", ICON_HEADER, RESULTS_HEADER, STANDINGS_HEADER])
        wb.save(file_path)
        return wb, ws

    wb = load_workbook(file_path)
    ws = wb.active

    header = [c.value for c in ws[1]]
    expected = ["国", "リーグ", "シーズン開始", "シーズン終了", "ラウンド数", "パス"]
    if not header or len(header) < len(expected) or header[:len(expected)] != expected:
        ws.delete_rows(1, ws.max_row if ws.max_row else 1)
        ws.append(expected + [ICON_HEADER, RESULTS_HEADER, STANDINGS_HEADER])
        wb.save(file_path)
        return wb, ws

    changed = False
    if ICON_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=ICON_HEADER); header.append(ICON_HEADER); changed = True
    if RESULTS_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=RESULTS_HEADER); header.append(RESULTS_HEADER); changed = True
    if STANDINGS_HEADER not in header:
        ws.cell(row=1, column=len(header) + 1, value=STANDINGS_HEADER); header.append(STANDINGS_HEADER); changed = True
    if changed:
        wb.save(file_path)

    return wb, ws

def get_colmap(ws) -> Dict[str, int]:
    header = [c.value for c in ws[1]]
    return {name: idx + 1 for idx, name in enumerate(header)}

# =========================
# URL 正規化
# =========================
def normalize_to_fixtures(url_or_path: str) -> Optional[str]:
    p = (url_or_path or "").strip()
    if not p:
        return None
    if p.startswith("http://") or p.startswith("https://"):
        base = p
    else:
        if not p.startswith("/"):
            p = "/" + p
        base = urljoin(BASE_URL, p)

    if not base.endswith("/"):
        base += "/"
    base = re.sub(r'/(fixtures|results|standings|archive)/$', '/', base, flags=re.IGNORECASE)
    full = urljoin(base, "fixtures/")
    return full

def to_results_url(fixtures_url: str) -> str:
    if not fixtures_url:
        return ""
    if not fixtures_url.endswith("/"):
        fixtures_url += "/"
    return re.sub(r"/fixtures/?$", "/results/", fixtures_url)

def to_standings_url(fixtures_url: str) -> str:
    if not fixtures_url:
        return ""
    if not fixtures_url.endswith("/"):
        fixtures_url += "/"
    return re.sub(r"/fixtures/?$", "/standings/", fixtures_url)

def read_existing_paths(ws) -> Set[str]:
    col = get_colmap(ws)
    paths: Set[str] = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col["パス"]).value
        if not v:
            continue
        norm = normalize_to_fixtures(str(v))
        if norm:
            paths.add(norm)
    return paths

def append_row_if_missing(ws, existing_paths: Set[str],
                          country: str, league: str, path_or_url: str,
                          rounds: Optional[int] = None) -> bool:
    fixtures_url = normalize_to_fixtures(path_or_url)
    if not fixtures_url:
        return False
    if fixtures_url in existing_paths:
        return False
    ws.append([country, league, "", "", rounds if rounds is not None else "", fixtures_url, ""])
    r = ws.max_row
    col = get_colmap(ws)
    if RESULTS_HEADER in col:
        ws.cell(row=r, column=col[RESULTS_HEADER], value=to_results_url(fixtures_url))
    if STANDINGS_HEADER in col:
        ws.cell(row=r, column=col[STANDINGS_HEADER], value=to_standings_url(fixtures_url))
    existing_paths.add(fixtures_url)
    return True

def yield_targets(ws) -> Iterable[Tuple[int, str]]:
    col = get_colmap(ws)
    for r in range(2, ws.max_row + 1):
        country = (ws.cell(row=r, column=col["国"]).value or "").strip()
        league  = (ws.cell(row=r, column=col["リーグ"]).value or "").strip()
        if not is_target(country, league):
            continue
        path = (ws.cell(row=r, column=col["パス"]).value or "").strip()
        if not path:
            continue
        start = ws.cell(row=r, column=col["シーズン開始"]).value
        end   = ws.cell(row=r, column=col["シーズン終了"]).value
        rounds = ws.cell(row=r, column=col["ラウンド数"]).value if "ラウンド数" in col else None
        if (start not in (None, "")) and (end not in (None, "")) and rounds not in (None, ""):
            continue
        yield r, path

# =========================
# 左メニュー → 国 → 各国ページ → ターゲットリーグ収集
# =========================
async def expand_left_menu_all(page: Page, verbose: bool = False):
    rounds = 0
    while True:
        btns = page.locator("#category-left-menu .lmc__itemMore")
        n = await btns.count()
        clicked_any = False
        for i in range(n):
            b = btns.nth(i)
            try:
                if await b.is_visible():
                    await b.scroll_into_view_if_needed()
                    await b.click()
                    clicked_any = True
                    await page.wait_for_timeout(200)
            except Exception:
                pass
        rounds += 1
        if verbose:
            print(f"[debug] expand round={rounds}, clicked_any={clicked_any}")
        if not clicked_any or rounds > 20:
            break

async def collect_countries_from_left_menu(page: Page, verbose: bool = False) -> List[Tuple[str, str]]:
    await page.goto(urljoin(BASE_URL, "/soccer/"), wait_until="domcontentloaded")
    await page.wait_for_selector("#category-left-menu", timeout=10000)
    await expand_left_menu_all(page, verbose=verbose)

    loc = page.locator('#category-left-menu .lmc__block a.lmc__element.lmc__item[href^="/soccer/"]')
    count = await loc.count()
    results: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    patt = re.compile(r"^/soccer/[^/]+/?$", re.IGNORECASE)

    for i in range(count):
        a = loc.nth(i)
        href = await a.get_attribute("href")
        if not href or not patt.match(href):
            continue
        name_el = a.locator(".lmc__elementName").first
        try:
            if await name_el.count() > 0:
                name = (await name_el.inner_text()).strip()
            else:
                name = (await a.inner_text()).strip()
        except Exception:
            continue
        if href not in seen and name:
            seen.add(href)
            results.append((name, href))
    if verbose:
        print(f"[debug] countries collected: {len(results)}")
    return results

async def collect_leagues_in_country(page: Page, country_name: str, country_href: str, verbose: bool = False) -> List[Tuple[str, str, str]]:
    url = urljoin(BASE_URL, country_href)
    await page.goto(url, wait_until="domcontentloaded")
    m = re.match(r"^/soccer/([^/]+)/?$", country_href)
    if not m:
        return []
    slug = m.group(1)

    candidates: Dict[str, str] = {}
    def _push(href: str, text: str):
        if not href.endswith("/"):
            href += "/"
        if not re.match(rf"^/soccer/{re.escape(slug)}/[^/]+/$", href):
            return
        if re.search(r"/(fixtures|results|standings|archive)/", href, re.IGNORECASE):
            return
        if href not in candidates or len(text) > len(candidates[href]):
            candidates[href] = text

    # 左メニュー
    try:
        await page.wait_for_selector("#category-left-menu", timeout=5000)
        await expand_left_menu_all(page, verbose=verbose)
        left_links = page.locator(f'#category-left-menu a[href^="/soccer/{slug}/"]')
        n = await left_links.count()
        for i in range(n):
            a = left_links.nth(i)
            href = await a.get_attribute("href") or ""
            text = (await a.inner_text() or "").strip()
            if not text:
                el = a.locator("span,div").first
                if await el.count() > 0:
                    text = (await el.inner_text() or "").strip()
            if href and text:
                _push(href, text)
    except Exception:
        pass

    # 本文保険
    try:
        body_links = page.locator(f'a[href^="/soccer/{slug}/"]')
        n = await body_links.count()
        for i in range(min(n, 2000)):
            a = body_links.nth(i)
            href = await a.get_attribute("href") or ""
            text = (await a.inner_text() or "").strip()
            if not text:
                el = a.locator("span,div").first
                if await el.count() > 0:
                    text = (await el.inner_text() or "").strip()
            if href and text:
                _push(href, text)
    except Exception:
        pass

    out: List[Tuple[str, str, str]] = []
    for href, text in candidates.items():
        league_name = text if text else href.rstrip("/").split("/")[-1]
        if is_target(country_name, league_name):
            out.append((country_name, league_name, href))
    if verbose:
        print(f"[debug] {country_name}: found={len(candidates)} / matched targets={len(out)}")
    return out

# =========================
# もっと表示
# =========================
# 置き換え: click_show_more_until_end()
async def click_show_more_until_end(
    page: Page,
    button_text: str = "もっと試合を表示する",
    max_clicks: Optional[int] = None,
    wait_timeout: int = 15000,
) -> int:
    clicks = 0
    if max_clicks is None:
        max_clicks = MAX_CLICKS
    # 日本語の表記ゆれを全部許容
    import re
    text_re = re.compile(r"(もっと|さらに).*(表示)", re.I)

    while clicks < max_clicks:
        link = page.locator("a[data-testid='wcl-buttonLink']").filter(has_text=text_re).first
        if await link.count() == 0 or not await link.is_visible():
            break
        before = await page.locator(".event__match").count()
        await link.scroll_into_view_if_needed()
        await link.click()
        clicks += 1
        try:
            await page.wait_for_function(
                """
                ({sel, before, linkSel}) => {
                  const now = document.querySelectorAll(sel).length;
                  const link = document.querySelector(linkSel);
                  return (now > before) || !link || link.offsetParent === null;
                }
                """,
                arg={"sel": ".event__match", "before": before, "linkSel": "a[data-testid='wcl-buttonLink']"},
                timeout=wait_timeout,
            )
        except PwTimeoutError:
            break
    return clicks


# =========================
# 進捗抽出（開始/終了/ラウンド数/アイコン）
# =========================
def _ts_to_ddmm(ts: int) -> str:
    dt = datetime.fromtimestamp(ts, tz=timezone.utc)
    return dt.strftime("%d.%m")

async def _extract_all_timestamps(page: Page) -> List[int]:
    """
    data-time / data-dt / data-utc / data-epoch / data-timestamp を収集して epoch の配列で返す
    """
    sel_list = ["[data-time]", "[data-dt]", "[data-utc]", "[data-epoch]", "[data-timestamp]"]
    ts_list: List[int] = []
    for sel in sel_list:
        try:
            vals = await page.eval_on_selector_all(sel, "els => els.map(e => e.getAttribute(arguments[0]))", sel[1:-1])
        except Exception:
            vals = []
        for v in vals or []:
            if not v:
                continue
            v = v.strip()
            if not re.fullmatch(r"\d{10,13}", v):
                continue
            x = int(v)
            if len(v) == 13:
                x //= 1000
            if 946684800 <= x <= 4102444800:  # 2000-01-01 .. 2100-01-01
                ts_list.append(x)
    return ts_list

async def _extract_round_count(page: Page) -> Optional[int]:
    """
    ページのテキストや属性からラウンド数の最大値を推定。
    """
    try:
        body = await page.inner_text("body")
    except Exception:
        body = ""

    patterns = [
        r"第\s*(\d+)\s*節",
        r"ラウンド\s*(\d+)",
        r"Round\s*(\d+)",
        r"Matchday\s*(\d+)",
        r"Jornada\s*(\d+)",
        r"Giornata\s*(\d+)",
    ]
    nums: List[int] = []
    for p in patterns:
        for m in re.finditer(p, body, flags=re.IGNORECASE):
            try:
                nums.append(int(m.group(1)))
            except Exception:
                pass

    # data-round 属性
    try:
        data_rounds = await page.eval_on_selector_all("[data-round]", "els => els.map(e => e.getAttribute('data-round'))")
    except Exception:
        data_rounds = []
    for v in data_rounds or []:
        if v and v.isdigit():
            nums.append(int(v))

    # select/option から推定
    try:
        option_texts = await page.eval_on_selector_all("select option", "els => els.map(e => e.textContent)")
    except Exception:
        option_texts = []
    for t in option_texts or []:
        for p in patterns:
            for m in re.finditer(p, t or "", flags=re.IGNORECASE):
                try:
                    nums.append(int(m.group(1)))
                except Exception:
                    pass
        # 単に数字だけのオプション（「1」「2」…）が並ぶ場合
        if t and t.strip().isdigit():
            nums.append(int(t.strip()))

    if nums:
        mx = max(nums)
        print(f"[debug] round candidates={sorted(set(nums))} -> max={mx}")
        return mx
    print("[debug] round candidates=0 -> None")
    return None

async def open_and_extract(page: Page, url: str) -> Tuple[str, str, str, Optional[int]]:
    """
    指定URLから (icon, start_ddmm, end_ddmm, rounds) を抽出。
    """
    icon_url, start_ddmm, end_ddmm, rounds = "", "", "", None

    await page.goto(url, wait_until="domcontentloaded")
    # open_and_extract() 冒頭の goto 直後に追加
    print(f"[debug] open {url}")
    # アイコン
    try:
        icon = await page.query_selector("img.heading__logo, img[class*='heading__logo--']")
        if not icon:
            icon = await page.query_selector("[data-testid='wcl-headerLeague'] img.heading__logo")
        if icon:
            src = await icon.get_attribute("src")
            if src:
                icon_url = src
    except PwTimeoutError:
        pass

    # 試合ロード＆もっと表示
    try:
        await page.wait_for_selector(".event__match, .event__time", timeout=8000)
        try:
            await click_show_more_until_end(page, max_clicks=MAX_CLICKS)
        except PwTimeoutError:
            pass
    except PwTimeoutError:
        pass

    # 1) テキスト dd.mm フォールバック（ページ全体から拾う & ログ）
    if not start_ddmm or not end_ddmm:
        try:
            body_text = await page.inner_text("body")
        except Exception:
            body_text = ""
        ddmms = _ddmm_pairs_from_text(body_text)
        if ddmms:
            start_ddmm = start_ddmm or min(ddmms, key=lambda x: x[0])[1]
            end_ddmm   = end_ddmm   or max(ddmms, key=lambda x: x[0])[1]
        print(f"[debug] ddmm_fallback: count={len(ddmms)} start={start_ddmm or '-'} end={end_ddmm or '-'}")

    # 2) ラウンド数推定
    rounds = await _extract_round_count(page)

    return icon_url, start_ddmm, end_ddmm, rounds

async def fetch_season_info(page: Page, fixtures_url: str) -> Tuple[str, str, str, Optional[int], str]:
    """
    results → fixtures の順で試行し、(icon, start, end, rounds, source) を返す。
    source は 'results' / 'fixtures' / '' のいずれか。
    """
    # 2) fixtures
    try:
        icon, s, e, r = await open_and_extract(page, fixtures_url)
        if s or e or r:
            return icon, s, e, r, "fixtures"
        return icon, s, e, r, ""
    except Exception:
        return "", "", "", None, ""

# =========================
# メイン処理
# =========================
async def main():
    parser = argparse.ArgumentParser(description="Flashscore リーグ一覧 → Excel（シーズン開始/終了/ラウンド数/アイコン）")
    parser.add_argument("--mode", choices=["scan", "fill", "both"], default="both",
                        help="scan: 国→リーグ収集のみ / fill: 開始・終了・ラウンド・アイコン埋め / both: 両方")
    parser.add_argument("--headful", action="store_true", help="ヘッドレス解除（デバッグ用）")
    parser.add_argument("--max-clicks", type=int, default=80, help="『もっと試合を表示する』の最大クリック回数")
    parser.add_argument("--verbose", action="store_true", help="収集状況を詳細表示")
    args = parser.parse_args()

    global MAX_CLICKS
    MAX_CLICKS = args.max_clicks

    wb, ws = open_or_init_season_book(SEASON_XLSX)
    col = get_colmap(ws)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not args.headful)
        context = await browser.new_context()
        page = await context.new_page()

        existing_paths = read_existing_paths(ws)

        if args.mode in ("scan", "both"):
            countries = await collect_countries_from_left_menu(page, verbose=args.verbose)
            print(f"[info] 国件数: {len(countries)}")

            added_count = 0
            total_matched = 0
            for idx, (country_name, country_href) in enumerate(countries, 1):
                try:
                    leagues = await collect_leagues_in_country(page, country_name, country_href, verbose=args.verbose)
                    total_matched += len(leagues)
                    for c, league_name, league_href in leagues:
                        if append_row_if_missing(ws, existing_paths, c, league_name, league_href):
                            added_count += 1
                except Exception as e:
                    print(f"[warn] {country_name} 取得中に例外: {e}")

                if idx % 5 == 0:
                    wb.save(SEASON_XLSX)
                await page.wait_for_timeout(200)

            wb.save(SEASON_XLSX)
            print(f"[info] ターゲット一致リーグ数: {total_matched}")
            print(f"[info] 追加行（新規リーグ）: {added_count}")

            # 既存行にも結果/順位URL補完
            col = get_colmap(ws)
            if RESULTS_HEADER in col or STANDINGS_HEADER in col:
                filled = 0
                for r in range(2, ws.max_row + 1):
                    path = (ws.cell(row=r, column=col["パス"]).value or "").strip()
                    furl = normalize_to_fixtures(path)
                    if not furl:
                        continue
                    if RESULTS_HEADER in col and not (ws.cell(row=r, column=col[RESULTS_HEADER]).value):
                        ws.cell(row=r, column=col[RESULTS_HEADER], value=to_results_url(furl)); filled += 1
                    if STANDINGS_HEADER in col and not (ws.cell(row=r, column=col[STANDINGS_HEADER]).value):
                        ws.cell(row=r, column=col[STANDINGS_HEADER], value=to_standings_url(furl)); filled += 1
                if filled:
                    wb.save(SEASON_XLSX)
                    print(f"[info] 追加列補完: {filled} セル更新")

        if args.mode in ("fill", "both"):
            updated_rows = 0
            touched = 0
            for row, path in yield_targets(ws):
                full_url = normalize_to_fixtures(path)
                if not full_url:
                    print(f"[skip] row={row}: URL正規化に失敗 path={path!r}")
                    continue

                print(f"[対象] row={row} -> {full_url}")
                wrote = False
                try:
                    icon_url, start_ddmm, end_ddmm, rounds, src = await fetch_season_info(page, full_url)

                    # ★ ログ（ここで“途中経過”として必ず出力）
                    print(f"[data] row={row} src={src or '-'} start={start_ddmm or '-'} end={end_ddmm or '-'} rounds={rounds if rounds is not None else '-'} icon={'OK' if icon_url else '-'}")

                    if (not ws.cell(row=row, column=col["シーズン開始"]).value) and start_ddmm:
                        ws.cell(row=row, column=col["シーズン開始"], value=start_ddmm); wrote = True
                    if (not ws.cell(row=row, column=col["シーズン終了"]).value) and end_ddmm:
                        ws.cell(row=row, column=col["シーズン終了"], value=end_ddmm); wrote = True
                    if (not ws.cell(row=row, column=col["ラウンド数"]).value) and rounds:
                        ws.cell(row=row, column=col["ラウンド数"], value=rounds); wrote = True
                    if (not ws.cell(row=row, column=col[ICON_HEADER]).value) and icon_url:
                        ws.cell(row=row, column=col[ICON_HEADER], value=icon_url); wrote = True

                    if not (start_ddmm or end_ddmm or rounds):
                        print(f"[warn] row={row}: シーズン情報を抽出できませんでした（ページ: {src or 'fixtures/resultsともに不可'}）")
                except Exception as e:
                    print(f"[err ] row={row}: {e}")

                if wrote:
                    updated_rows += 1
                touched += 1
                if touched % 10 == 0:
                    wb.save(SEASON_XLSX)
                await page.wait_for_timeout(150)

            wb.save(SEASON_XLSX)
            print(f"[info] 更新行数: {updated_rows}")

        await context.close()
        await browser.close()

# 追加: ユーティリティ
def _ddmm_pairs_from_text(text: str) -> List[Tuple[int, str]]:
    import re
    pat = re.compile(r"(?<!\d)(\d{1,2})\.(\d{1,2})\.(?!\d)")
    out = []
    for m in pat.finditer(text or ""):
        dd, mm = int(m.group(1)), int(m.group(2))
        if 1 <= dd <= 31 and 1 <= mm <= 12:
            out.append((mm * 100 + dd, f"{dd:02d}.{mm:02d}"))
    return out


if __name__ == "__main__":
    asyncio.run(main())

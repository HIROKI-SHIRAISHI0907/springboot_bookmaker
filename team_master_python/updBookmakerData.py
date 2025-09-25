# -*- coding: utf-8 -*-
"""
更新処理 → CSV 変換 → 成功したら元の XLSX を削除
- /Users/shiraishitoshio/bookmaker/ 配下の output_*.xlsx を順に処理（欠番OK）
- 「ゴール時間」「選手名」が欠けている行を Playwright で補完
- 全行埋まったら CSV 出力し、成功時に 元の XLSX を削除
"""

import os
import re
import glob
import time
from typing import List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PTimeout


# ===== 設定 =====
BASE_DIR = "/Users/shiraishitoshio/bookmaker"
FILE_PREFIX = "output_"
SHEET_NAME = "Sheet"
COL_LINK = "試合リンク文字列"
COL_HOME = "ホームスコア"
COL_AWAY = "アウェースコア"
COL_TIME = "ゴール時間"
COL_MEMBER = "選手名"


# ===== ユーティリティ =====
def fmt(path: str) -> str:
    return os.path.abspath(path)

def safe_remove(path: str, label: str = "") -> bool:
    try:
        if os.path.exists(path):
            os.remove(path)
            print(f"{label}削除: {path}")
            return True
        print(f"{label}削除対象なし: {path}")
        return False
    except Exception as e:
        print(f"{label}削除失敗: {path} / {e}")
        return False

def list_all_xlsx() -> List[str]:
    """欠番があってもディレクトリ内の output_*.xlsx を通番順に並べる"""
    paths = glob.glob(os.path.join(BASE_DIR, f"{FILE_PREFIX}*.xlsx"))
    def key(p):
        s = os.path.basename(p).replace(FILE_PREFIX, "").replace(".xlsx", "")
        return int(s) if s.isdigit() else 10**9
    return sorted(paths, key=key)

def next_csv_seq() -> int:
    """次に採番する CSV の通番（既存最大 + 1。存在しなければ 1）"""
    seqs = []
    for path in glob.glob(os.path.join(BASE_DIR, f"{FILE_PREFIX}*.csv")):
        s = os.path.basename(path).replace(FILE_PREFIX, "").replace(".csv", "")
        if s.isdigit():
            seqs.append(int(s))
    return (max(seqs) + 1) if seqs else 1


# ===== Excel 読み・判定 =====
def read_excel_df(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"ブックが存在しません: {path}")
    try:
        xls = pd.ExcelFile(path)
        df = xls.parse(SHEET_NAME)
        return df
    except Exception as e:
        print(f"ファイルが破損しています。削除します: {path} / {e}")
        safe_remove(path, label="[XLSX]")
        raise

def excel_is_fully_filled(path: str) -> bool:
    try:
        df = read_excel_df(path)
    except Exception:
        return False
    if COL_TIME not in df.columns or COL_MEMBER not in df.columns:
        print("必要な列が不足しています。")
        return False
    empties = df[COL_TIME].isna() | (df[COL_TIME] == "") | df[COL_MEMBER].isna() | (df[COL_MEMBER] == "")
    cnt = int(empties.sum())
    if cnt > 0:
        print(f"未記入 {cnt} 行あり: {path}")
        return False
    print(f"全行記入済み: {path}")
    return True

def get_pending_rows(path: str) -> Tuple[List[int], List[str], List[str], List[str]]:
    """
    未記入の行の ‘連番(1始まり)’ と 試合リンク、ホーム/アウェースコア を返す
    """
    df = read_excel_df(path).copy()
    df["連番"] = range(1, len(df) + 1)

    need = {COL_LINK, COL_HOME, COL_AWAY, COL_TIME, COL_MEMBER}
    if not need.issubset(df.columns):
        print("必要な列が不足しています。")
        return [], [], [], []

    seq_list, link_list, home_list, away_list = [], [], [], []
    for i in range(len(df)):
        if (pd.isna(df.loc[i, COL_TIME]) or df.loc[i, COL_TIME] == "") or \
           (pd.isna(df.loc[i, COL_MEMBER]) or df.loc[i, COL_MEMBER] == ""):
            seq_list.append(int(df.loc[i, "連番"]))
            link_list.append("" if pd.isna(df.loc[i, COL_LINK]) else str(df.loc[i, COL_LINK]))
            home_list.append("" if pd.isna(df.loc[i, COL_HOME]) else str(df.loc[i, COL_HOME]))
            away_list.append("" if pd.isna(df.loc[i, COL_AWAY]) else str(df.loc[i, COL_AWAY]))
    return seq_list, link_list, home_list, away_list


# ===== Excel 書き込み（ヘッダ名で列を特定） =====
def _find_header_columns(sh, headers: List[str]) -> dict:
    found = {}
    max_col = sh.max_column
    for c in range(1, max_col + 1):
        v = sh.cell(row=1, column=c).value
        if v in headers and v not in found:
            found[v] = c
        if len(found) == len(headers):
            break
    return found

def write_update_to_excel(row_1based: int, path: str, score_time: str, player_name: str) -> None:
    """
    ヘッダ「ゴール時間」「選手名」の列を動的に見つけて書き込む。
    row_1based はデータ行の 1 始まり（ヘッダは 1 行目）→ 実Excel行は +1。
    """
    print(f"処理開始 write_update_to_excel: {path} / 行={row_1based}")
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            sh = wb.active
        else:
            wb = Workbook()
            sh = wb.active

        r = row_1based + 1
        score_time = score_time if score_time else "---"
        player_name = player_name if player_name else "---"

        col_map = _find_header_columns(sh, [COL_TIME, COL_MEMBER])
        if COL_TIME not in col_map or COL_MEMBER not in col_map:
            raise RuntimeError("ヘッダー『ゴール時間』『選手名』の列が見つかりません。")

        c_time = col_map[COL_TIME]
        c_mem = col_map[COL_MEMBER]

        sh.cell(row=r, column=c_time, value=score_time)
        sh.cell(row=r, column=c_mem, value=player_name)
        wb.save(path)

        print(
            f"記入完了: {path} / "
            f"{get_column_letter(c_time)}{r}={score_time} / "
            f"{get_column_letter(c_mem)}{r}={player_name}"
        )
    except Exception as e:
        print(f"書き込み失敗: {path} / {e}")
        raise
    finally:
        print(f"処理終了 write_update_to_excel: {path}")


# ===== CSV 変換 =====
def excel_to_csv(excel_file: str, csv_file: str, require_all_filled: bool = True) -> bool:
    """
    Excel を CSV に変換。require_all_filled=True のときは
    全行埋まっている場合のみ変換し、成功したら元の XLSX を削除。
    """
    try:
        if require_all_filled and not excel_is_fully_filled(excel_file):
            return False
        df = pd.read_excel(excel_file, sheet_name=SHEET_NAME)
        df.to_csv(csv_file, index=False)
        print(f"CSV 出力完了: {csv_file}")
        safe_remove(excel_file, label="[XLSX]")  # 成功時のみ削除
        return True
    except Exception as e:
        print(f"CSV 変換失敗: {excel_file} -> {csv_file} / {e}")
        return False


# ===== Playwright 補助 =====
def close_cookie_banners(page):
    for sel in [
        "#onetrust-accept-btn-handler",
        "button:has-text('同意')", "button:has-text('許可')",
        "button:has-text('OK')", "button:has-text('Accept')",
        "[data-testid*='cookie'] button",
    ]:
        try:
            btn = page.locator(sel).first
            if btn.is_visible():
                btn.click(timeout=1500)
                time.sleep(0.2)
        except Exception:
            pass

def to_full_url(link: str) -> str:
    """
    ★ 変換しない★
    受け取ったリンクを絶対URLにするだけ（/match/soccer/.../?mid=XXXX をそのまま使う）
    """
    if not link:
        return ""
    if link.startswith("http://") or link.startswith("https://"):
        return link
    if link.startswith("/"):
        return "https://www.flashscore.co.jp" + link
    return "https://www.flashscore.co.jp/" + link

def build_candidate_urls(link: str) -> List[str]:
    """
    1) まず『/match/soccer/.../?mid=XXXX』など“元のリンク(絶対化)”を使う
    2) 取れない場合のみ mid ベースの #/summary → #/match-summary → 旧式 でフォールバック
    """
    cands: list[str] = []
    # ① 元のリンク（最優先）
    orig = to_full_url(link)
    if orig:
        cands.append(orig)

    # ② mid があるならフォールバック候補も
    mid = extract_mid(link)
    if mid:
        base = f"https://www.flashscore.co.jp/match/{mid}/#/"
        cands += [
            base + "summary",                      # 新
            base + "match-summary",                # 準新
            base + "match-summary/match-summary",  # 旧
        ]

    # 重複除去
    out, seen = [], set()
    for u in cands:
        if u and u not in seen:
            out.append(u); seen.add(u)
    return out

def extract_mid(s: str) -> str | None:
    if not s:
        return None
    s = str(s).strip()

    # ?mid=XXXX を最優先
    m = re.search(r"[?&#]mid=([A-Za-z0-9]+)", s)
    if m:
        return m.group(1)

    # /match/{mid}/… 形式
    m = re.search(r"/match/([A-Za-z0-9]{6,20})(?:/|$)", s)
    if m:
        return m.group(1)
    return None

def fetch_goal_info_for_score(page, link: str, target_home: str, target_away: str) -> Tuple[str, str]:
    """
    該当スコア ‘X-Y’ の事象の時間と選手名を返す（見つからなければ "---","---"）
    SVGへのinner_text()は呼ばず、インシデントDIVごとに抽出して安全化。
    """
    url = to_full_url(link)
    if not url:
        return "---", "---"

    print(f"アクセス: {url}")
    try:
        page.goto(url, timeout=60000)
        page.wait_for_load_state("domcontentloaded")
        close_cookie_banners(page)

        # “試合経過/Timeline/Summary” のタブへ
        try:
            tab = page.get_by_role("tab", name=re.compile(r"(試合経過|経過|Summary|Timeline|Match summary)", re.I))
            if tab.count():
                tab.first.click()
                page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        # 事象DIVの出現待ち
        page.wait_for_selector("div.smv__incident", timeout=20000)
        incidents = page.locator("div.smv__incident")
        n = incidents.count()

        target_score = f"{int(target_home)}-{int(target_away)}"
        for i in range(n):
            inc = incidents.nth(i)
            raw = (inc.text_content() or "").strip()
            # スコア表記抽出
            m = re.search(r"(\d+)\s*-\s*(\d+)", raw)
            if not m:
                continue
            sc = f"{m.group(1)}-{m.group(2)}"
            if sc != target_score:
                continue

            time_txt = (inc.locator(".smv__timeBox").first.text_content() or "").strip()
            player_txt = ""
            for sel in ["a.smv__playerName div", "div.smv__playerName"]:
                loc = inc.locator(sel).first
                if loc.count():
                    player_txt = (loc.text_content() or "").strip()
                    if player_txt:
                        break

            return time_txt if time_txt else "---", player_txt if player_txt else "---"

        return "---", "---"

    except Exception as e:
        print(f"スコア取得失敗: {url} / {e}")
        return "取得エラー", "取得エラー"

def fetch_goal_info_for_score(page, link: str, target_home: str, target_away: str) -> Tuple[str, str]:
    """
    指定スコア X-Y のインシデント(時間・選手)を抽出。
    新→旧の順でURLを試す。SVGに inner_text() は当てない。
    """
    candidates = build_candidate_urls(link)
    if not candidates:
        return "---", "---"

    target_score = f"{int(target_home)}-{int(target_away)}"

    def try_once(url: str) -> Tuple[str, str] | None:
        print(f"アクセス: {url}")
        page.goto(url, timeout=60000)
        page.wait_for_load_state("domcontentloaded")
        close_cookie_banners(page)

        # サマリ/タイムライン系タブに寄せる（出ていれば）
        try:
            tab = page.get_by_role("tab", name=re.compile(r"(試合経過|経過|Summary|Timeline|Match summary)", re.I))
            if tab.count():
                tab.first.click()
                page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        # 事象DIV待ち
        try:
            page.wait_for_selector("div.smv__incident", timeout=20000)
        except Exception:
            return None

        incidents = page.locator("div.smv__incident")
        n = incidents.count()
        if n == 0:
            return None

        for i in range(n):
            inc = incidents.nth(i)
            raw = (inc.text_content() or "").strip()
            m = re.search(r"(\d+)\s*-\s*(\d+)", raw)
            if not m:
                continue
            sc = f"{m.group(1)}-{m.group(2)}"
            if sc != target_score:
                continue

            time_txt = (inc.locator(".smv__timeBox").first.text_content() or "").strip()
            player_txt = ""
            for sel in ("a.smv__playerName div", "div.smv__playerName"):
                loc = inc.locator(sel).first
                if loc.count():
                    player_txt = (loc.text_content() or "").strip()
                    if player_txt:
                        break
            return (time_txt or "---", player_txt or "---")
        return ("---", "---")  # 事象はあるが該当スコアなし

    for url in candidates:
        try:
            res = try_once(url)
            if res:
                return res
        except Exception:
            # 次の候補へ
            continue

    print(f"スコア取得失敗（全候補NG）: {link}")
    return "取得エラー", "取得エラー"


def process_updates_with_playwright(
    playwright,
    seq_list: List[int],
    link_list: List[str],
    home_list: List[str],
    away_list: List[str],
    excel_path: str
) -> None:
    """未記入行をブラウザで補完して Excel に反映"""
    print(f"処理開始 process_updates_with_playwright: {excel_path}")

    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    try:
        for i in range(len(seq_list)):
            seq = seq_list[i]
            link = link_list[i]
            h = home_list[i]
            a = away_list[i]

            # スコアが空/非数ならスキップ
            try:
                int(h); int(a)
            except Exception:
                print(f"スキップ(不正スコア): seq={seq}, h={h}, a={a}")
                continue

            t, p = fetch_goal_info_for_score(page, link, h, a)
            try:
                write_update_to_excel(seq, excel_path, t, p)
            except Exception:
                pass

    finally:
        page.close()
        context.close()
        browser.close()

    print(f"処理終了 process_updates_with_playwright: {excel_path}")


# ===== メイン処理 =====
def main():
    xlsx_list = list_all_xlsx()
    if not xlsx_list:
        print("処理対象の XLSX が見つかりません。終了します。")
        return

    with sync_playwright() as pw:
        processed = 0
        for xlsx in map(fmt, xlsx_list):
            print(f"\n=== 処理対象: {xlsx} ===")

            # 既に全行埋まっていれば即CSV化 → 成功ならXLSX削除
            if excel_is_fully_filled(xlsx):
                csv_seq = next_csv_seq()
                csv = fmt(os.path.join(BASE_DIR, f"{FILE_PREFIX}{csv_seq}.csv"))
                if excel_to_csv(xlsx, csv, require_all_filled=True):
                    print(f"変換成功 → XLSX 削除済: {csv}")
                else:
                    print("変換失敗（スキップ）")
                processed += 1
                continue

            # 未記入行を補完
            try:
                seq_list, link_list, home_list, away_list = get_pending_rows(xlsx)
            except Exception:
                processed += 1
                continue

            if seq_list:
                process_updates_with_playwright(pw, seq_list, link_list, home_list, away_list, xlsx)
            else:
                print("未記入行なし。変換を再試行します。")

            # 補完後チェック → OKなら CSV 変換して成功時にXLSX削除
            if excel_is_fully_filled(xlsx):
                csv_seq = next_csv_seq()
                csv = fmt(os.path.join(BASE_DIR, f"{FILE_PREFIX}{csv_seq}.csv"))
                if excel_to_csv(xlsx, csv, require_all_filled=True):
                    print(f"変換成功 → XLSX 削除済: {csv}")
                else:
                    print("変換失敗（スキップ）")

            processed += 1

    print(f"\n処理件数: {processed}")

if __name__ == "__main__":
    main()